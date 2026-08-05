from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def build_initial_credentials_workbook(credentials, generated_at):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "初始登录凭据"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A7"

    headers = [
        "序号",
        "用户名",
        "初始密码",
        "姓名",
        "角色",
        "所属站点",
        "所属片区",
        "账号状态",
        "登录说明",
    ]
    last_column = get_column_letter(len(headers))
    sheet.merge_cells(f"A1:{last_column}1")
    sheet["A1"] = "业务督导中心数智管理平台·初始登录凭据"
    sheet["A1"].font = Font(name="Microsoft YaHei", size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="163E62")
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 32

    sheet.merge_cells(f"A2:{last_column}2")
    sheet["A2"] = f"生成时间：{generated_at}    账号数：{len(credentials)}"
    sheet["A2"].font = Font(name="Microsoft YaHei", size=10, color="465569")
    sheet["A2"].alignment = Alignment(horizontal="left", vertical="center")

    notices = [
        "本文件仅包含本次新生成的一次性初始密码，系统不保存密码明文，无法再次查看。",
        "请通过可靠渠道单独发送给对应用户，用户首次登录后必须立即修改密码。",
        "本次操作已使旧密码和现有登录会话全部失效；已暂停账号仍保持暂停状态。",
    ]
    for row_number, notice in enumerate(notices, start=3):
        sheet.merge_cells(f"A{row_number}:{last_column}{row_number}")
        cell = sheet.cell(row=row_number, column=1, value=notice)
        cell.font = Font(name="Microsoft YaHei", size=10, color="8A4B08")
        cell.fill = PatternFill("solid", fgColor="FFF7E1")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        sheet.row_dimensions[row_number].height = 25

    header_row = 6
    thin_border = Border(
        left=Side(style="thin", color="D8E1E8"),
        right=Side(style="thin", color="D8E1E8"),
        top=Side(style="thin", color="D8E1E8"),
        bottom=Side(style="thin", color="D8E1E8"),
    )
    for column, title in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=column, value=title)
        cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="087C71")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    sheet.row_dimensions[header_row].height = 25

    for index, credential in enumerate(credentials, start=1):
        values = [
            index,
            credential.get("username") or "",
            credential.get("initial_password") or "",
            credential.get("real_name") or "-",
            credential.get("role_label") or credential.get("role") or "-",
            credential.get("station_name") or "-",
            credential.get("station_region") or "-",
            credential.get("account_status_label") or "-",
            "使用初始密码登录后，请按页面提示设置新密码",
        ]
        row_number = header_row + index
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_number, column=column, value=value)
            if isinstance(value, str):
                # Treat all database text as literal cell content, never as an Excel formula.
                cell.data_type = "s"
            cell.font = Font(name="Microsoft YaHei", size=10, color="243247")
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="center" if column in {1, 3, 5, 8} else "left",
                vertical="center",
                wrap_text=True,
            )
            if column == 3:
                cell.font = Font(name="Consolas", size=10, bold=True, color="9A3412")
            if index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F7FAFC")
        sheet.row_dimensions[row_number].height = 29

    widths = [8, 18, 25, 15, 18, 24, 18, 13, 42]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.auto_filter.ref = f"A{header_row}:{last_column}{header_row + len(credentials)}"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
