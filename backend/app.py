from flask import Flask, g, jsonify, request, send_file, send_from_directory
from flask_cors import CORS
import fcntl
import hashlib
import json
import logging
import os
import re
import shutil
import subprocess
import tempfile
import threading
import time
import uuid
import zipfile
from collections import OrderedDict
from datetime import datetime, timedelta
from urllib.parse import quote_plus
from zoneinfo import ZoneInfo
from io import BytesIO
import psycopg2
from psycopg2 import sql
from psycopg2.extras import RealDictCursor
from flask_migrate import Migrate
from flask_sqlalchemy import SQLAlchemy
from itsdangerous import BadSignature, URLSafeTimedSerializer
from werkzeug.utils import secure_filename
from PIL import Image
from ai_utils import generate_feedback_title, generate_standard_recommendations
from ai_usage import get_ai_pricing_table

try:
    from qcloud_cos import CosConfig, CosS3Client
except Exception:
    CosConfig = None
    CosS3Client = None

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get(
    "APP_SECRET_KEY",
    os.environ.get("SECRET_KEY", "ywddzx-smart-platform-dev-secret"),
)
CORS(app)
db = SQLAlchemy()
migrate = Migrate()

BASE_DIR = os.path.dirname(__file__)
STORAGE_ROOT = os.path.join(BASE_DIR, "storage")
MAX_IMAGE_BYTES = 500 * 1024
MAX_PDF_BYTES = 50 * 1024 * 1024
MAX_VIDEO_BYTES = 500 * 1024 * 1024
ALLOWED_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".heic", ".heif"}
ALLOWED_TRAINING_MATERIAL_EXTENSIONS = {".pdf", ".mp4", ".mov", ".m4v", ".webm"}

ISSUES_STORAGE_DIR = os.path.join(STORAGE_ROOT, "issues")
RECTIFICATIONS_STORAGE_DIR = os.path.join(STORAGE_ROOT, "rectifications")
SIGNATURES_STORAGE_DIR = os.path.join(STORAGE_ROOT, "signatures")
INSPECTION_ORIGINALS_STORAGE_DIR = os.path.join(STORAGE_ROOT, "inspection_originals")
TRAINING_MATERIALS_STORAGE_DIR = os.path.join(STORAGE_ROOT, "training_materials")
FEEDBACK_SCREENSHOTS_STORAGE_DIR = os.path.join(STORAGE_ROOT, "feedback_screenshots")
ISSUE_EXPORTS_STORAGE_DIR = os.path.join(STORAGE_ROOT, "issue_exports")
STATION_SCORE_EXPORTS_STORAGE_DIR = os.path.join(STORAGE_ROOT, "station_score_exports")
BACKUP_CONFIG_PATH = os.path.join(STORAGE_ROOT, "backup_config.json")
DEFAULT_BACKUP_DIR = os.path.join(STORAGE_ROOT, "backups")
BACKUP_PREFIX = "ywddzx_full_backup"
LOCAL_BACKUP_FILENAME = f"{BACKUP_PREFIX}_latest.zip"
AUTO_BACKUP_FILENAME = LOCAL_BACKUP_FILENAME
COS_BACKUP_PREFIX = os.environ.get("COS_BACKUP_PREFIX", "ywddzx-full-backups/").strip().strip("/")
COS_BACKUP_RETENTION_COUNT = 3
BEIJING_TZ = ZoneInfo("Asia/Shanghai")
DEFAULT_INITIAL_PASSWORD = "123456"
WORK_ANNIVERSARY_START_DATE = datetime(2026, 4, 15).date()
DEFAULT_USER_BIRTHDAYS = [
    ("徐佳仪", 1, 11),
    ("王昕怡", 1, 16),
    ("宋辞", 2, 5),
    ("吴杰", 2, 11),
    ("程镇林", 2, 14),
    ("王涛", 3, 26),
    ("彭思宇", 4, 18),
    ("束紫荆", 5, 8),
    ("左翰承", 7, 3),
    ("李泊汛", 7, 3),
    ("袁姝慧", 7, 30),
    ("刘文喆", 8, 16),
    ("王子玥", 8, 20),
    ("吕雪儿", 9, 19),
    ("赵萌", 10, 22),
    ("徐晃", 11, 18),
    ("魏九发", 12, 1),
    ("姜傲云", 12, 7),
    ("侯明敖", 12, 10),
    ("葛心玉", 12, 18),
    ("陈中磊", 12, 21),
]
ISSUE_EXPORT_RETENTION_DAYS = 7
ISSUE_EXPORT_CLEANUP_INTERVAL_SECONDS = 60 * 60
PASSWORD_MIN_LENGTH = 8
PASSWORD_MAX_LENGTH = 32
AUTH_TOKEN_NORMAL_MAX_AGE_SECONDS = int(
    os.environ.get("AUTH_TOKEN_NORMAL_MAX_AGE_SECONDS", str(8 * 60 * 60))
)
AUTH_TOKEN_PRIVILEGED_MAX_AGE_SECONDS = int(
    os.environ.get("AUTH_TOKEN_PRIVILEGED_MAX_AGE_SECONDS", str(4 * 60 * 60))
)
AUTH_TOKEN_SALT = "ywddzx-auth-token-v1"
PRIVILEGED_AUTH_ROLES = {"root", "supervisor"}


def normalize_frontend_app_version(value):
    raw_value = str(value or "").strip()
    if not raw_value:
        raw_value = "3.7.1"
    if raw_value.lower().startswith("v"):
        raw_value = raw_value[1:]
    parts = raw_value.split(".")
    try:
        major = int(parts[0])
    except (TypeError, ValueError):
        major = 1
    try:
        minor = int(parts[1]) if len(parts) > 1 else 0
    except (TypeError, ValueError):
        minor = 0
    try:
        patch = int(parts[2]) if len(parts) > 2 else 0
    except (TypeError, ValueError):
        patch = 0
    base_version = f"{major}.{minor}"
    return f"{base_version}.{patch}" if patch > 0 else base_version


FRONTEND_APP_VERSION = normalize_frontend_app_version(os.environ.get("APP_FRONTEND_VERSION", "3.7.1"))
FRONTEND_VERSION_EXPIRED_CODE = "FRONTEND_VERSION_EXPIRED"
FRONTEND_VERSION_EXPIRED_MESSAGE = "页面版本已过期，请刷新页面后继续使用"
DISPLAY_REMOVED_STATION_PHRASE = "\u52a0\u6cb9\u7ad9"
DISPLAY_OIL_STATION_TYPE = "油站"
AUTH_SERVER_CACHE_TTL_SECONDS = max(1, int(os.environ.get("AUTH_SERVER_CACHE_TTL_SECONDS", "30")))
AUTH_TOKEN_CACHE_MAX_ENTRIES = max(128, int(os.environ.get("AUTH_TOKEN_CACHE_MAX_ENTRIES", "2048")))
SERVER_RESOURCE_SAMPLE_PATH = os.environ.get(
    "SERVER_RESOURCE_SAMPLE_PATH",
    "/tmp/ywddzx_server_resource_sample.json",
)
SERVER_ONLINE_USERS_PATH = os.environ.get(
    "SERVER_ONLINE_USERS_PATH",
    "/tmp/ywddzx_online_users.json",
)
SERVER_ONLINE_USER_TTL_SECONDS = max(60, int(os.environ.get("SERVER_ONLINE_USER_TTL_SECONDS", "90")))
SERVER_ONLINE_TOUCH_INTERVAL_SECONDS = max(
    5,
    int(os.environ.get("SERVER_ONLINE_TOUCH_INTERVAL_SECONDS", "15")),
)
LEGACY_FRONTEND_API_PATHS = {
    "/api/feedbacks/unread-count",
    "/api/assessment/peer-reviews/pending-count",
    "/api/my-issues/pending-rectification-count",
    "/api/inspection-plan-assignments/my-pending",
}
QUIET_ACCESS_LOG_PATHS = LEGACY_FRONTEND_API_PATHS | {
    "/api/auth/me",
    "/api/notifications/summary",
    "/api/system/resources",
}


CHECKLIST_PHYSICAL_TABLE_PREFIX = "inspection_table_"
CHECKLIST_FIELD_KEY_MAX_LENGTH = 63
CHECKLIST_STANDARD_ID_BLOCK_SIZE = 1000
SCHEMA_MIGRATION_LOCK_KEY = 2026051601
RESERVED_CHECKLIST_FIELD_KEYS = {
    "id",
    "standard_id",
    "created_at",
    "updated_at",
}


def acquire_schema_migration_lock(cur):
    cur.execute("SELECT pg_advisory_xact_lock(%s);", (SCHEMA_MIGRATION_LOCK_KEY,))


def frontend_version_expired_response():
    response = jsonify(
        {
            "success": False,
            "code": FRONTEND_VERSION_EXPIRED_CODE,
            "message": FRONTEND_VERSION_EXPIRED_MESSAGE,
        }
    )
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    response.headers["Clear-Site-Data"] = '"cache", "storage"'
    return response, 426


def sanitize_display_string(value):
    if not isinstance(value, str) or DISPLAY_REMOVED_STATION_PHRASE not in value:
        return value
    if value.strip() == DISPLAY_REMOVED_STATION_PHRASE:
        return DISPLAY_OIL_STATION_TYPE
    return value.replace(DISPLAY_REMOVED_STATION_PHRASE, "")


def sanitize_display_payload(value):
    if isinstance(value, dict):
        return {key: sanitize_display_payload(item) for key, item in value.items()}
    if isinstance(value, list):
        return [sanitize_display_payload(item) for item in value]
    if isinstance(value, tuple):
        return [sanitize_display_payload(item) for item in value]
    if isinstance(value, str):
        return sanitize_display_string(value)
    return value


@app.after_request
def sanitize_json_display_text(response):
    if response.direct_passthrough:
        return response
    if "application/json" not in str(response.content_type or "").lower():
        return response

    payload = response.get_json(silent=True)
    if payload is None:
        return response

    sanitized_payload = sanitize_display_payload(payload)
    response.set_data(json.dumps(sanitized_payload, ensure_ascii=False, default=str))
    response.headers["Content-Length"] = str(len(response.get_data()))
    return response


class QuietAccessLogFilter(logging.Filter):
    def filter(self, record):
        message = record.getMessage()
        return not any(path in message for path in QUIET_ACCESS_LOG_PATHS)


logging.getLogger("werkzeug").addFilter(QuietAccessLogFilter())


def read_proc_cpu_times():
    try:
        with open("/proc/stat", "r", encoding="utf-8") as stat_file:
            line = stat_file.readline().strip()
        if not line.startswith("cpu "):
            return None
        values = [int(value) for value in line.split()[1:]]
        if len(values) < 5:
            return None
        idle = values[3] + values[4]
        total = sum(values)
        return {"total": total, "idle": idle}
    except Exception:
        return None


def read_proc_memory_info():
    try:
        meminfo = {}
        with open("/proc/meminfo", "r", encoding="utf-8") as meminfo_file:
            for line in meminfo_file:
                key, value = line.split(":", 1)
                meminfo[key] = int(value.strip().split()[0])
        total_kb = meminfo.get("MemTotal", 0)
        available_kb = meminfo.get("MemAvailable", 0)
        used_kb = max(0, total_kb - available_kb)
        percent = (used_kb / total_kb * 100) if total_kb else 0
        return {
            "memory_percent": round(percent, 1),
            "memory_used_mb": round(used_kb / 1024, 1),
            "memory_total_mb": round(total_kb / 1024, 1),
        }
    except Exception:
        return {
            "memory_percent": None,
            "memory_used_mb": None,
            "memory_total_mb": None,
        }


def read_proc_network_bytes():
    rx_bytes = 0
    tx_bytes = 0
    try:
        with open("/proc/net/dev", "r", encoding="utf-8") as net_file:
            for line in net_file.readlines()[2:]:
                if ":" not in line:
                    continue
                interface, data = line.split(":", 1)
                if interface.strip() == "lo":
                    continue
                fields = data.split()
                if len(fields) < 16:
                    continue
                rx_bytes += int(fields[0])
                tx_bytes += int(fields[8])
    except Exception:
        pass
    return {"rx_bytes": rx_bytes, "tx_bytes": tx_bytes}


def swap_server_resource_sample(current_sample):
    with server_resource_sample_lock:
        fallback_previous = dict(server_resource_last_sample)
        server_resource_last_sample.update(current_sample)

    try:
        with open(SERVER_RESOURCE_SAMPLE_PATH, "a+", encoding="utf-8") as sample_file:
            fcntl.flock(sample_file.fileno(), fcntl.LOCK_EX)
            sample_file.seek(0)
            raw_previous = sample_file.read().strip()
            previous = json.loads(raw_previous) if raw_previous else {}
            sample_file.seek(0)
            sample_file.truncate()
            json.dump(current_sample, sample_file)
            sample_file.flush()
            os.fsync(sample_file.fileno())
            fcntl.flock(sample_file.fileno(), fcntl.LOCK_UN)
            return previous or fallback_previous
    except Exception:
        return fallback_previous


def build_online_user_entry(user, now=None):
    now = now or time.time()
    user_id = str((user or {}).get("id") or "").strip()
    role = str((user or {}).get("role") or "").strip()
    display_name = (
        normalize_text((user or {}).get("real_name"), 80)
        or normalize_text((user or {}).get("username"), 80)
        or f"用户{user_id}"
    )
    return {
        "id": user_id,
        "username": normalize_text((user or {}).get("username"), 80),
        "real_name": normalize_text((user or {}).get("real_name"), 80),
        "display_name": display_name,
        "role": role,
        "role_label": ROLE_LABELS.get(role, role or "未知角色"),
        "station_name": normalize_text((user or {}).get("station_name"), 120),
        "region": normalize_text((user or {}).get("region"), 120),
        "last_seen": now,
        "last_seen_label": datetime.fromtimestamp(now, BEIJING_TZ).strftime("%H:%M:%S"),
    }


def get_online_users_snapshot(current_user=None, include_users=False, touch=False):
    now = time.time()
    with server_online_users_lock:
        try:
            with open(SERVER_ONLINE_USERS_PATH, "a+", encoding="utf-8") as online_file:
                fcntl.flock(online_file.fileno(), fcntl.LOCK_EX)
                online_file.seek(0)
                raw_payload = online_file.read().strip()
                online_users = json.loads(raw_payload) if raw_payload else {}
                if not isinstance(online_users, dict):
                    online_users = {}

                cutoff = now - SERVER_ONLINE_USER_TTL_SECONDS
                pruned_users = {}
                for user_key, entry in online_users.items():
                    if not str(user_key).strip() or not isinstance(entry, dict):
                        continue
                    try:
                        last_seen = float(entry.get("last_seen") or 0)
                    except (TypeError, ValueError):
                        continue
                    if last_seen < cutoff:
                        continue
                    normalized_entry = dict(entry)
                    normalized_entry["id"] = str(normalized_entry.get("id") or user_key)
                    normalized_entry["last_seen"] = last_seen
                    normalized_entry["last_seen_label"] = datetime.fromtimestamp(
                        last_seen,
                        BEIJING_TZ,
                    ).strftime("%H:%M:%S")
                    pruned_users[str(user_key)] = normalized_entry
                online_users = pruned_users

                current_user_id = str((current_user or {}).get("id") or "").strip()
                if touch and current_user_id:
                    online_users[current_user_id] = build_online_user_entry(current_user, now)

                online_file.seek(0)
                online_file.truncate()
                json.dump(online_users, online_file)
                online_file.flush()
                fcntl.flock(online_file.fileno(), fcntl.LOCK_UN)
                users = sorted(
                    online_users.values(),
                    key=lambda item: float(item.get("last_seen") or 0),
                    reverse=True,
                )
                if not include_users:
                    users = []
                return {"count": len(online_users), "users": users}
        except Exception:
            return {"count": 0, "users": []}


def touch_online_user_presence(user):
    if not user:
        return
    user_id = str(user.get("id") or "").strip()
    if not user_id:
        return
    now = time.time()
    with server_online_touch_lock:
        last_touch = float(server_online_touch_cache.get(user_id) or 0)
        if now - last_touch < SERVER_ONLINE_TOUCH_INTERVAL_SECONDS:
            return
        server_online_touch_cache[user_id] = now
    get_online_users_snapshot(user, include_users=False, touch=True)


def remove_online_user_presence(user_id):
    user_key = str(user_id or "").strip()
    if not user_key:
        return
    with server_online_users_lock:
        try:
            with open(SERVER_ONLINE_USERS_PATH, "a+", encoding="utf-8") as online_file:
                fcntl.flock(online_file.fileno(), fcntl.LOCK_EX)
                online_file.seek(0)
                raw_payload = online_file.read().strip()
                online_users = json.loads(raw_payload) if raw_payload else {}
                if not isinstance(online_users, dict):
                    online_users = {}
                if user_key in online_users:
                    online_users.pop(user_key, None)
                    online_file.seek(0)
                    online_file.truncate()
                    json.dump(online_users, online_file)
                    online_file.flush()
                fcntl.flock(online_file.fileno(), fcntl.LOCK_UN)
        except Exception:
            return


def build_server_resource_snapshot():
    now = time.time()
    cpu_times = read_proc_cpu_times()
    net_bytes = read_proc_network_bytes()
    memory_info = read_proc_memory_info()
    current_sample = {
        "timestamp": now,
        "cpu_total": cpu_times["total"] if cpu_times else None,
        "cpu_idle": cpu_times["idle"] if cpu_times else None,
        "rx_bytes": net_bytes["rx_bytes"],
        "tx_bytes": net_bytes["tx_bytes"],
    }
    previous = swap_server_resource_sample(current_sample)

    cpu_percent = None
    if (
        cpu_times
        and previous.get("cpu_total") is not None
        and previous.get("cpu_idle") is not None
    ):
        total_delta = cpu_times["total"] - previous["cpu_total"]
        idle_delta = cpu_times["idle"] - previous["cpu_idle"]
        if total_delta > 0:
            cpu_percent = round(max(0, min(100, (1 - idle_delta / total_delta) * 100)), 1)
    if cpu_percent is None:
        try:
            load_1m = os.getloadavg()[0]
            cpu_percent = round(max(0, min(100, load_1m / max(1, os.cpu_count() or 1) * 100)), 1)
        except Exception:
            cpu_percent = None

    interval = max(0, now - previous["timestamp"]) if previous.get("timestamp") else 0
    rx_kbps = 0
    tx_kbps = 0
    if interval > 0 and previous.get("rx_bytes") is not None and previous.get("tx_bytes") is not None:
        rx_kbps = max(0, (net_bytes["rx_bytes"] - previous["rx_bytes"]) / 1024 / interval)
        tx_kbps = max(0, (net_bytes["tx_bytes"] - previous["tx_bytes"]) / 1024 / interval)
    current_user = getattr(g, "current_user", None)
    online_snapshot = get_online_users_snapshot(
        current_user,
        include_users=is_root_user(current_user),
        touch=True,
    )

    return {
        "cpu_percent": cpu_percent,
        "cpu_core_count": os.cpu_count() or 0,
        "network_rx_kbps": round(rx_kbps, 1),
        "network_tx_kbps": round(tx_kbps, 1),
        "online_user_count": online_snapshot["count"],
        "online_users": online_snapshot["users"],
        "sampled_at": datetime.now(BEIJING_TZ).strftime("%Y-%m-%d %H:%M:%S"),
        **memory_info,
    }

# === Permission constants ===
ROLE_OPTIONS = {
    "root",
    "supervisor",
    "station_manager",
    "quality_safety",
    "development_plan",
    "oil_gas",
    "non_oil",
    "finance",
    "area_account",
}
ROLE_LABELS = {
    "root": "系统管理员",
    "supervisor": "督导组账号",
    "station_manager": "站点账号",
    "quality_safety": "质安部账号",
    "development_plan": "发展计划部账号",
    "oil_gas": "油气事业部账号",
    "non_oil": "非油事业部账号",
    "finance": "财务部账号",
    "area_account": "片区账号",
}
QUALITY_SAFETY_DEFAULT_CHECKLIST_SCOPE = [
    ("计量稽查检查表", "online"),
    ("计量稽查检查表", "offline"),
    ("环境无异味管理检查表", "offline"),
    ("质量安全环保检查表", "online"),
    ("质量安全环保检查表", "offline"),
]
DEVELOPMENT_PLAN_DEFAULT_CHECKLIST_SCOPE = [
    ("设备设施检查表", "offline"),
]
OIL_GAS_DEFAULT_CHECKLIST_SCOPE = [
    ("环境无异味管理检查表", "offline"),
    ("现场检查明细表", "online"),
    ("现场检查明细表", "offline"),
]
NON_OIL_DEFAULT_CHECKLIST_SCOPE = [
    ("非油合规性检查（团购）", "online"),
    ("非油合规性检查（团购）", "offline"),
    ("非油检查表", "offline"),
]
FINANCE_DEFAULT_CHECKLIST_SCOPE = [
    ("财务检查表", "offline"),
]
ROLE_DEFAULT_CHECKLIST_SCOPES = {
    "quality_safety": QUALITY_SAFETY_DEFAULT_CHECKLIST_SCOPE,
    "development_plan": DEVELOPMENT_PLAN_DEFAULT_CHECKLIST_SCOPE,
    "oil_gas": OIL_GAS_DEFAULT_CHECKLIST_SCOPE,
    "non_oil": NON_OIL_DEFAULT_CHECKLIST_SCOPE,
    "finance": FINANCE_DEFAULT_CHECKLIST_SCOPE,
}
INSPECTION_TABLE_SCOPE_PERMISSION_KEYS = (
    "limit_issue_inspection_table_scope",
    "limit_record_inspection_table_scope",
    "limit_plan_inspection_table_scope",
)
INSPECTION_TABLE_SCOPE_CONTEXT_MAP = {
    "issues": "limit_issue_inspection_table_scope",
    "records": "limit_record_inspection_table_scope",
    "plans": "limit_plan_inspection_table_scope",
}
STATION_REGION_SCOPE_PERMISSION_KEYS = (
    "limit_issue_station_region_scope",
    "limit_record_station_region_scope",
    "limit_plan_station_region_scope",
    "limit_certificate_station_region_scope",
)
PERMISSION_CATALOG = [
    {
        "key": "view_station_map",
        "name": "查看页面",
        "category": "站点地图",
        "description": "访问地图中心的站点地图页面。",
        "defaults": {"root": True, "supervisor": True, "station_manager": False, "quality_safety": True},
    },
    {
        "key": "submit_inspections",
        "name": "录入巡检问题",
        "category": "巡检登记",
        "description": "访问巡检登记页面，并提交巡检记录和问题。",
        "defaults": {"root": True, "supervisor": True, "station_manager": False, "quality_safety": False},
    },
    {
        "key": "view_inspection_standards",
        "name": "查看页面",
        "category": "巡检规范库",
        "description": "访问业务督导中心自建的内部巡检规范库。",
        "defaults": {"root": True, "supervisor": True, "station_manager": True, "quality_safety": True},
    },
    {
        "key": "view_checklist_originals",
        "name": "查看页面",
        "category": "检查表原件库",
        "description": "查看各检查表原始 PDF 文件。",
        "defaults": {"root": True, "supervisor": True, "station_manager": True, "quality_safety": True},
    },
    {
        "key": "manage_checklist_originals",
        "name": "上传/更新 PDF",
        "category": "检查表原件库",
        "description": "上传检查表原始 PDF，或替换为新版 PDF。",
        "defaults": {"root": True, "supervisor": False, "station_manager": False, "quality_safety": False},
    },
    {
        "key": "limit_issue_inspection_table_scope",
        "name": "限定检查表范围",
        "category": "巡检问题列表",
        "description": "启用后，巡检问题列表只显示选定检查表的问题数据。",
        "defaults": {"root": False, "supervisor": False, "station_manager": False, "quality_safety": True},
    },
    {
        "key": "limit_issue_station_region_scope",
        "name": "查看片区站点数据",
        "category": "巡检问题列表",
        "description": "只查看选定片区/归属地的站点问题数据。与“查看本站数据”“查看全部站点数据”三选一。",
        "defaults": {"root": False, "supervisor": False, "station_manager": False, "quality_safety": False},
    },
    {
        "key": "view_own_inspection_issues",
        "name": "查看本站数据",
        "category": "巡检问题列表",
        "description": "只查看当前账号所属站点的巡检问题数据。",
        "defaults": {"root": True, "supervisor": False, "station_manager": True, "quality_safety": False},
    },
    {
        "key": "view_all_inspection_issues",
        "name": "查看全部站点数据",
        "category": "巡检问题列表",
        "description": "查看所有站点的巡检问题数据。与“查看本站数据”“查看片区站点数据”三选一。",
        "defaults": {"root": True, "supervisor": True, "station_manager": False, "quality_safety": True},
    },
    {
        "key": "edit_inspection_issues",
        "name": "编辑所选范围内问题",
        "category": "巡检问题列表",
        "description": "在已选择的本站/全部站点范围内，编辑巡检问题；已闭环问题仍只有 root 可改。",
        "defaults": {"root": True, "supervisor": False, "station_manager": False, "quality_safety": False},
    },
    {
        "key": "delete_inspection_issues",
        "name": "删除所选范围内问题",
        "category": "巡检问题列表",
        "description": "在已选择的本站/全部站点范围内删除巡检问题，并自动回算关联记录与计划状态。",
        "defaults": {"root": True, "supervisor": False, "station_manager": False, "quality_safety": False},
    },
    {
        "key": "audit_inspection_issues",
        "name": "审核巡检问题",
        "category": "巡检问题列表",
        "description": "对巡检问题判定审核通过或否决；否决后不参与巡检记录统计和问题流转。",
        "defaults": {"root": True, "supervisor": False, "station_manager": False, "quality_safety": False},
    },
    {
        "key": "change_issue_inspector",
        "name": "调整问题检查人归属",
        "category": "巡检问题列表",
        "description": "把单条巡检问题改挂到其他具备巡检登记能力的检查人名下。",
        "defaults": {"root": True, "supervisor": False, "station_manager": False, "quality_safety": False},
    },
    {
        "key": "export_issue_photos",
        "name": "导出巡检照片",
        "category": "巡检问题列表",
        "description": "导出巡检问题 Excel 时，可选择把问题照片、整改照片、复核照片嵌入表格。",
        "defaults": {"root": True, "supervisor": False, "station_manager": False, "quality_safety": False},
    },
    {
        "key": "hide_inspector_contact_info",
        "name": "隐藏检查人信息",
        "category": "巡检信息显示",
        "description": "勾选后，巡检问题列表和巡检记录不显示检查人姓名与检查人员手机号。",
        "defaults": {"root": False, "supervisor": False, "station_manager": True, "quality_safety": False},
    },
    {
        "key": "view_own_inspection_records",
        "name": "查看本站数据",
        "category": "巡检记录",
        "description": "只查看当前账号所属站点的巡检记录数据。",
        "defaults": {"root": True, "supervisor": False, "station_manager": True, "quality_safety": False},
    },
    {
        "key": "view_all_inspection_records",
        "name": "查看全部站点数据",
        "category": "巡检记录",
        "description": "查看所有站点的巡检记录数据。与“查看本站数据”“查看片区站点数据”三选一。",
        "defaults": {"root": True, "supervisor": True, "station_manager": False, "quality_safety": True},
    },
    {
        "key": "limit_record_inspection_table_scope",
        "name": "限定检查表范围",
        "category": "巡检记录",
        "description": "启用后，巡检记录只显示选定检查表的记录数据。",
        "defaults": {"root": False, "supervisor": False, "station_manager": False, "quality_safety": True},
    },
    {
        "key": "limit_record_station_region_scope",
        "name": "查看片区站点数据",
        "category": "巡检记录",
        "description": "只查看选定片区/归属地的站点记录数据。与“查看本站数据”“查看全部站点数据”三选一。",
        "defaults": {"root": False, "supervisor": False, "station_manager": False, "quality_safety": False},
    },
    {
        "key": "delete_inspection_records",
        "name": "删除巡检记录",
        "category": "巡检记录",
        "description": "在已选择的站点范围内删除巡检记录，并同步删除本记录下的问题。",
        "defaults": {"root": True, "supervisor": False, "station_manager": False, "quality_safety": False},
    },
    {
        "key": "reset_inspection_signature",
        "name": "重置巡检记录流程",
        "category": "巡检记录",
        "description": "按当前流程状态回退巡检记录，并清空本记录下的问题审核结果；已站经理签名验收的记录不可重置。",
        "defaults": {"root": True, "supervisor": False, "station_manager": False, "quality_safety": False},
    },
    {
        "key": "view_inspection_plans",
        "name": "查看页面",
        "category": "巡检计划",
        "description": "访问巡检计划页面和计划完成情况。",
        "defaults": {"root": True, "supervisor": True, "station_manager": False, "quality_safety": True},
    },
    {
        "key": "limit_plan_inspection_table_scope",
        "name": "限定检查表范围",
        "category": "巡检计划",
        "description": "启用后，巡检计划只显示选定检查表的计划数据。",
        "defaults": {"root": False, "supervisor": False, "station_manager": False, "quality_safety": True},
    },
    {
        "key": "limit_plan_station_region_scope",
        "name": "查看片区站点数据",
        "category": "巡检计划",
        "description": "只统计和展示选定片区/归属地内站点的巡检计划。",
        "defaults": {"root": False, "supervisor": False, "station_manager": False, "quality_safety": False},
    },
    {
        "key": "manage_inspection_plans",
        "name": "管理巡检计划",
        "category": "巡检计划",
        "description": "新建、编辑、删除检查表巡检计划。",
        "defaults": {"root": True, "supervisor": False, "station_manager": False, "quality_safety": False},
    },
    {
        "key": "view_own_certificates",
        "name": "查看本站数据",
        "category": "站点证照有效期管理",
        "description": "查看当前账号所属站点的证照有效期和到期提醒。",
        "defaults": {"root": True, "supervisor": False, "station_manager": True, "quality_safety": False},
    },
    {
        "key": "limit_certificate_station_region_scope",
        "name": "查看片区站点数据",
        "category": "站点证照有效期管理",
        "description": "只查看选定片区/归属地的站点证照有效期和到期提醒。与“查看本站数据”“查看全部站点数据”三选一。",
        "defaults": {"root": False, "supervisor": False, "station_manager": False, "quality_safety": False},
    },
    {
        "key": "edit_own_certificates",
        "name": "编辑本站证照",
        "category": "站点证照有效期管理",
        "description": "仅在选择“查看本站数据”时可用，用于录入、修改、删除当前账号所属站点证照。",
        "defaults": {"root": True, "supervisor": False, "station_manager": True, "quality_safety": False},
    },
    {
        "key": "view_all_certificates",
        "name": "查看全部站点数据",
        "category": "站点证照有效期管理",
        "description": "查看所有站点的证照有效期和到期提醒。与“查看本站数据”“查看片区站点数据”三选一。",
        "defaults": {"root": True, "supervisor": True, "station_manager": False, "quality_safety": True},
    },
    {
        "key": "view_assessment",
        "name": "查看页面",
        "category": "考核系统",
        "description": "访问考核系统页面。",
        "defaults": {"root": True, "supervisor": True, "station_manager": False, "quality_safety": False},
    },
    {
        "key": "view_attendance",
        "name": "查看人员出勤",
        "category": "人员出勤",
        "description": "访问考核系统里的人员出勤统计页面。",
        "defaults": {"root": True, "supervisor": True, "station_manager": False, "quality_safety": False},
    },
    {
        "key": "view_station_scores",
        "name": "查看站点评分",
        "category": "站点评分",
        "description": "访问考核系统里的站点评分页面。",
        "defaults": {"root": True, "supervisor": True, "station_manager": False, "quality_safety": True},
    },
    {
        "key": "adjust_station_scores",
        "name": "手动调整站点评分",
        "category": "站点评分",
        "description": "在站点评分页面对系统自动评分进行人工调整。",
        "defaults": {"root": True, "supervisor": False, "station_manager": False, "quality_safety": False},
    },
    {
        "key": "view_peer_reviews",
        "name": "查看成员互评",
        "category": "成员互评",
        "description": "访问考核系统里的成员互评页面，参与填写并查看自己的评价记录。",
        "defaults": {"root": True, "supervisor": True, "station_manager": False, "quality_safety": False},
    },
    {
        "key": "manage_peer_review_tasks",
        "name": "管理成员互评任务",
        "category": "成员互评",
        "description": "配置互评模板、发起互评任务，并查看所有人的评价内容和完成情况。",
        "defaults": {"root": True, "supervisor": False, "station_manager": False, "quality_safety": False},
    },
    {
        "key": "view_training",
        "name": "查看页面",
        "category": "督导组内部培训系统",
        "description": "访问督导组内部培训系统页面。",
        "defaults": {"root": True, "supervisor": True, "station_manager": False, "quality_safety": False},
    },
    {
        "key": "view_training_materials",
        "name": "查看页面",
        "category": "培训材料库",
        "description": "访问培训材料库并查看材料目录与预览。",
        "defaults": {"root": True, "supervisor": True, "station_manager": False, "quality_safety": True},
    },
    {
        "key": "upload_training_materials",
        "name": "上传/更新自己的材料",
        "category": "培训材料库",
        "description": "上传培训材料，并编辑或删除自己上传的材料。",
        "defaults": {"root": True, "supervisor": True, "station_manager": False, "quality_safety": False},
    },
    {
        "key": "manage_stations",
        "name": "管理站点数据",
        "category": "站点数据管理",
        "description": "访问站点数据管理页面，并新增、编辑、删除、导入导出站点主数据。",
        "defaults": {"root": True, "supervisor": False, "station_manager": False, "quality_safety": False},
    },
    {
        "key": "reset_station_account_password",
        "name": "重置站点账号密码",
        "category": "站点数据管理",
        "description": "在站点数据管理页面把绑定站点账号的密码重置为 123456。",
        "defaults": {"root": True, "supervisor": False, "station_manager": False, "quality_safety": False},
    },
    {
        "key": "manage_checklists",
        "name": "管理检查表数据",
        "category": "检查表数据管理",
        "description": "访问检查表数据管理页面，并维护外部检查表、字段结构和外部规范数据。",
        "defaults": {"root": True, "supervisor": False, "station_manager": False, "quality_safety": False},
    },
    {
        "key": "manage_internal_standards",
        "name": "管理内部巡检规范",
        "category": "巡检规范库数据管理",
        "description": "访问巡检规范库数据管理页面，并维护内部规范字段配置和外部规范挂载关系。",
        "defaults": {"root": True, "supervisor": False, "station_manager": False, "quality_safety": False},
    },
    {
        "key": "manage_ai_usage",
        "name": "查看 AI 调用统计",
        "category": "AI调用统计",
        "description": "查看系统内 DeepSeek AI 调用次数、使用位置、字符量、估算 token 和费用。",
        "defaults": {"root": True, "supervisor": False, "station_manager": False, "quality_safety": False},
    },
]
for permission_item in PERMISSION_CATALOG:
    defaults = permission_item.setdefault("defaults", {})
    defaults.setdefault("development_plan", bool(defaults.get("quality_safety", False)))
    defaults.setdefault("oil_gas", bool(defaults.get("quality_safety", False)))
    defaults.setdefault("non_oil", bool(defaults.get("quality_safety", False)))
    defaults.setdefault("finance", bool(defaults.get("quality_safety", False)))
    defaults.setdefault("area_account", bool(defaults.get("quality_safety", False)))

AREA_ACCOUNT_PERMISSION_OVERRIDES = {
    "view_all_inspection_issues": False,
    "view_all_inspection_records": False,
    "limit_issue_inspection_table_scope": False,
    "limit_record_inspection_table_scope": False,
    "limit_plan_inspection_table_scope": False,
    "limit_issue_station_region_scope": True,
    "limit_record_station_region_scope": True,
    "limit_plan_station_region_scope": True,
    "limit_certificate_station_region_scope": True,
    "view_all_certificates": False,
    "hide_inspector_contact_info": True,
}
for permission_item in PERMISSION_CATALOG:
    if permission_item["key"] in AREA_ACCOUNT_PERMISSION_OVERRIDES:
        permission_item["defaults"]["area_account"] = AREA_ACCOUNT_PERMISSION_OVERRIDES[permission_item["key"]]

PERMISSION_KEYS = {item["key"] for item in PERMISSION_CATALOG}
PERMISSION_EXCLUSIVE_GROUPS = [
    ("view_own_inspection_issues", "limit_issue_station_region_scope", "view_all_inspection_issues"),
    ("view_own_inspection_records", "limit_record_station_region_scope", "view_all_inspection_records"),
    ("view_own_certificates", "limit_certificate_station_region_scope", "view_all_certificates"),
]
PERMISSION_DEPENDENCIES = {
    "edit_own_certificates": "view_own_certificates",
    "adjust_station_scores": "view_station_scores",
    "reset_station_account_password": "manage_stations",
    "manage_peer_review_tasks": "view_peer_reviews",
}
PERMISSION_ANY_DEPENDENCIES = {
    "edit_inspection_issues": (
        "view_all_inspection_issues",
    ),
    "delete_inspection_issues": (
        "view_all_inspection_issues",
    ),
    "audit_inspection_issues": (
        "view_all_inspection_issues",
    ),
    "change_issue_inspector": (
        "view_all_inspection_issues",
    ),
    "export_issue_photos": (
        "view_all_inspection_issues",
        "view_own_inspection_issues",
    ),
    "delete_inspection_records": (
        "view_all_inspection_records",
    ),
    "reset_inspection_signature": (
        "view_all_inspection_records",
    ),
}
STATION_TYPE_OPTIONS = {DISPLAY_OIL_STATION_TYPE, "充电站"}
STATION_ASSET_TYPE_OPTIONS = {"全资", "股权"}
STATION_CONSOLIDATED_OPTIONS = {"是", "否"}
STATION_ONLINE_3_STATUS_OPTIONS = {"上线", "上线参股模式", "未上线"}
STATION_MONITORING_STATUS_OPTIONS = {"运行中", "未运行"}
STATION_STATUS_OPTIONS = {"营业中", "停业"}
INSPECTION_COMPLETION_PENDING = "待检查人确认"
INSPECTION_COMPLETION_DONE = "已确认完成"
INSPECTION_COMPLETION_SOURCES = {"manual", "manual_all", "auto", "admin", "signature", "admin_reopen"}
DEFAULT_INSPECTION_AUTO_COMPLETE_DAYS = 7
DEFAULT_INSPECTION_RECORD_UNIQUENESS_PERIOD = "month"
INSPECTION_RECORD_UNIQUENESS_PERIODS = {"week", "month", "quarter", "year"}
INSPECTION_CHECKLIST_MANAGEMENT_SCHEMA_READY = False
USER_SECURITY_SCHEMA_READY = False
STATION_MANAGEMENT_SCHEMA_READY = False
ISSUE_INSPECTOR_SCHEMA_READY = False
INSPECTION_COMPLETION_SCHEMA_READY = False
INSPECTION_PLAN_ASSIGNMENT_SCHEMA_READY = False
FEEDBACK_SCHEMA_READY = False
SYSTEM_PAGE_VISIBILITY_SCHEMA_READY = False
PLAN_COMPLETION_SYNC_LOCK_NAMESPACE = 2026060401
auth_token_cache_lock = threading.Lock()
auth_token_cache = {}
auth_me_response_cache_lock = threading.Lock()
auth_me_response_cache = {}
server_resource_sample_lock = threading.Lock()
server_resource_last_sample = {
    "timestamp": None,
    "cpu_total": None,
    "cpu_idle": None,
    "rx_bytes": None,
    "tx_bytes": None,
}
server_online_users_lock = threading.Lock()
server_online_touch_lock = threading.Lock()
server_online_touch_cache = {}
ISSUE_STATUS_OPTIONS = {"待整改", "待复核", "已闭环", "站经无法整改"}
ISSUE_RESULT_OPTIONS = {"已整改", "站经无法整改"}
ISSUE_AUDIT_STATUS_OPTIONS = {"pending", "approved", "rejected"}
ISSUE_AUDIT_STATUS_LABELS = {
    "pending": "待审核",
    "approved": "审核通过",
    "rejected": "审核否决",
}
ISSUE_STATUS_ALIASES = {"已整改": "已闭环"}
ISSUE_RESULT_ALIASES = {
    "站级无法整改": "站经无法整改",
    "站级无法完成整改": "站经无法整改",
    "站经理无法整改": "站经无法整改",
}
FEEDBACK_TYPE_OPTIONS = {"功能建议", "Bug反馈", "界面优化", "流程建议", "其他"}
FEEDBACK_MODULE_OPTIONS = {
    "巡检系统",
    "巡检规范库",
    "检查表原件库",
    "巡检计划",
    "考核系统",
    "培训系统",
    "培训材料库",
    "证照管理",
    "车辆管理系统",
    "数据备份管理",
    "用户数据管理",
    "站点数据管理",
    "检查表数据管理",
    "巡检规范库数据管理",
    "AI调用统计",
    "管理系统",
    "公共功能",
    "登录与账号",
    "其他",
}
COVERAGE_TYPE_LABELS = {
    "monthly": "月度覆盖",
    "quarterly": "季度覆盖",
    "yearly": "年度覆盖",
}

CERTIFICATE_TYPES = [
    {
        "code": "dangerous_chemicals_permit",
        "name": "危险化学品经营许可证",
        "note": "危化证继续经营应在有效期满90天前申请延期。",
        "recommended_reminder_days": 150,
        "legal_reminder_days": 90,
        "recommended_label": "到期前 150天",
        "legal_label": "到期前 90天",
        "rule": "150天进入推荐提醒；90天内进入法定提醒。危化证继续经营应在有效期满90天前申请延期。",
    },
    {
        "code": "oil_retail_permit",
        "name": "成品油零售经营批准证书",
        "note": "成品油零售经营批准证书继续经营应在届满30日前申请延续。",
        "recommended_reminder_days": 90,
        "legal_reminder_days": 30,
        "recommended_label": "到期前 90天",
        "legal_label": "到期前 30天",
        "rule": "90天进入推荐提醒；30天内进入法定提醒。成品油零售经营批准证书继续经营应在届满30日前申请延续。",
    },
    {
        "code": "pollutant_discharge_permit",
        "name": "排污许可证",
        "note": "排污许可证继续排污应在届满60日前申请延续。",
        "recommended_reminder_days": 120,
        "legal_reminder_days": 60,
        "recommended_label": "到期前 120天",
        "legal_label": "到期前 60天",
        "rule": "120天进入推荐提醒；60天内进入法定提醒。排污许可证继续排污应在届满60日前申请延续。",
    },
    {
        "code": "lightning_protection_report",
        "name": "防雷检测报告",
        "note": "爆炸、火灾危险环境场所防雷装置一般每半年检测一次。",
        "recommended_reminder_days": 30,
        "legal_reminder_days": 7,
        "recommended_label": "到期前 30天",
        "legal_label": "到期前 7天",
        "rule": "30天进入推荐提醒；7天内进入法定提醒。爆炸、火灾危险环境场所防雷装置一般每半年检测一次。",
    },
    {
        "code": "tobacco_monopoly_permit",
        "name": "烟草专卖零售许可证",
        "note": "烟草专卖许可证继续经营应在届满30日前申请延续。",
        "recommended_reminder_days": 60,
        "legal_reminder_days": 30,
        "recommended_label": "到期前 60天",
        "legal_label": "到期前 30天",
        "rule": "60天进入推荐提醒；30天内进入法定提醒。烟草专卖许可证继续经营应在届满30日前申请延续。",
    },
    {
        "code": "business_license",
        "name": "工商营业执照",
        "note": "只有存在经营期限时才提醒。",
        "recommended_reminder_days": 90,
        "legal_reminder_days": 30,
        "recommended_label": "到期前 90天",
        "legal_label": "到期前 30天",
        "rule": "新版营业执照照面通常不再记载营业期限，但电子营业执照或企业信用公示系统可能仍显示经营期限；因此只有存在经营期限时才提醒。",
    },
]

CERTIFICATE_TYPE_BY_CODE = {item["code"]: item for item in CERTIFICATE_TYPES}


def beijing_now():
    return datetime.now(BEIJING_TZ)


def beijing_today():
    return beijing_now().date()


os.makedirs(ISSUES_STORAGE_DIR, exist_ok=True)
os.makedirs(RECTIFICATIONS_STORAGE_DIR, exist_ok=True)
os.makedirs(SIGNATURES_STORAGE_DIR, exist_ok=True)
os.makedirs(INSPECTION_ORIGINALS_STORAGE_DIR, exist_ok=True)
os.makedirs(TRAINING_MATERIALS_STORAGE_DIR, exist_ok=True)
os.makedirs(FEEDBACK_SCREENSHOTS_STORAGE_DIR, exist_ok=True)
os.makedirs(DEFAULT_BACKUP_DIR, exist_ok=True)


def get_db_config():
    return {
        "host": os.environ.get("DB_HOST", "db"),
        "port": str(os.environ.get("DB_PORT", 5432)),
        "dbname": os.environ.get("DB_NAME", "ywddzx"),
        "user": os.environ.get("DB_USER", "postgres"),
        "password": os.environ.get("DB_PASSWORD", "postgres"),
    }


def get_sqlalchemy_database_uri():
    db_config = get_db_config()
    user = quote_plus(db_config["user"])
    password = quote_plus(db_config["password"])
    host = db_config["host"]
    port = db_config["port"]
    dbname = quote_plus(db_config["dbname"])
    return f"postgresql+psycopg2://{user}:{password}@{host}:{port}/{dbname}"


app.config["SQLALCHEMY_DATABASE_URI"] = get_sqlalchemy_database_uri()
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
db.init_app(app)
migrate.init_app(app, db)


def get_db_connection():
    db_config = get_db_config()
    return psycopg2.connect(
        host=db_config["host"],
        port=db_config["port"],
        dbname=db_config["dbname"],
        user=db_config["user"],
        password=db_config["password"],
        cursor_factory=RealDictCursor,
        options="-c timezone=Asia/Shanghai",
    )


def close_db_resources(cur=None, conn=None):
    if cur:
        cur.close()
    if conn:
        conn.close()


def ensure_storage_subdir(base_dir):
    now = beijing_now()
    year_dir = os.path.join(base_dir, now.strftime("%Y"))
    month_dir = os.path.join(year_dir, now.strftime("%m"))
    os.makedirs(month_dir, exist_ok=True)
    return month_dir, now


def save_uploaded_file(file_storage, category):
    if not file_storage or not file_storage.filename:
        return None

    if category == "issues":
        base_dir = ISSUES_STORAGE_DIR
        prefix = "issue"
    elif category == "rectifications":
        base_dir = RECTIFICATIONS_STORAGE_DIR
        prefix = "rectification"
    elif category == "feedback_screenshots":
        base_dir = FEEDBACK_SCREENSHOTS_STORAGE_DIR
        prefix = "feedback"
    else:
        raise ValueError("不支持的上传分类。")

    original_name = secure_filename(file_storage.filename)
    ext = os.path.splitext(original_name)[1].lower()
    if ext and ext not in ALLOWED_IMAGE_EXTENSIONS:
        raise ValueError("仅支持上传 JPG、JPEG、PNG、WEBP、HEIC、HEIF 格式图片。")

    target_dir, now = ensure_storage_subdir(base_dir)
    filename = f"{prefix}_{now.strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex}.jpg"
    save_path = os.path.join(target_dir, filename)

    image_bytes = file_storage.read()
    if not image_bytes:
        raise ValueError("上传图片内容为空。")

    try:
        image = Image.open(BytesIO(image_bytes))
    except Exception as exc:
        raise ValueError("上传文件不是可识别的图片。") from exc

    if image.mode not in ("RGB", "L"):
        image = image.convert("RGB")
    elif image.mode == "L":
        image = image.convert("RGB")

    max_width = 1600
    if image.width > max_width:
        ratio = max_width / float(image.width)
        new_size = (max_width, max(1, int(image.height * ratio)))
        image = image.resize(new_size, Image.LANCZOS)

    quality = 82
    output = BytesIO()
    image.save(output, format="JPEG", quality=quality, optimize=True)

    while output.tell() > MAX_IMAGE_BYTES and quality > 42:
        quality -= 6
        output = BytesIO()
        image.save(output, format="JPEG", quality=quality, optimize=True)

    if output.tell() > MAX_IMAGE_BYTES:
        temp_image = image
        while output.tell() > MAX_IMAGE_BYTES and temp_image.width > 960:
            new_width = int(temp_image.width * 0.9)
            new_height = max(1, int(temp_image.height * 0.9))
            temp_image = temp_image.resize((new_width, new_height), Image.LANCZOS)
            output = BytesIO()
            temp_image.save(output, format="JPEG", quality=70, optimize=True)
        image = temp_image

    with open(save_path, "wb") as f:
        f.write(output.getvalue())

    relative_dir = os.path.relpath(target_dir, STORAGE_ROOT).replace("\\", "/")
    return f"/{relative_dir}/{filename}"


# Helper for saving signature PNGs
def save_signature_file(file_storage):
    if not file_storage or not file_storage.filename:
        return None

    target_dir, now = ensure_storage_subdir(SIGNATURES_STORAGE_DIR)
    filename = f"signature_{now.strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex}.png"
    save_path = os.path.join(target_dir, filename)

    image_bytes = file_storage.read()
    if not image_bytes:
        raise ValueError("签名图片内容为空。")

    try:
        image = Image.open(BytesIO(image_bytes))
    except Exception as exc:
        raise ValueError("上传文件不是可识别的签名图片。") from exc

    if image.mode not in ("RGBA", "LA"):
        image = image.convert("RGBA")
    else:
        image = image.convert("RGBA")

    max_width = 800
    max_height = 320
    if image.width > max_width or image.height > max_height:
        ratio = min(max_width / float(image.width), max_height / float(image.height))
        new_size = (
            max(1, int(image.width * ratio)),
            max(1, int(image.height * ratio)),
        )
        image = image.resize(new_size, Image.LANCZOS)

    image.save(save_path, format="PNG", optimize=True)

    relative_dir = os.path.relpath(target_dir, STORAGE_ROOT).replace("\\", "/")
    return f"/{relative_dir}/{filename}"


def normalize_upload_display_name(filename, fallback="检查表原件.pdf"):
    raw_name = str(filename or "").replace("\\", "/").split("/")[-1].strip()
    return raw_name[:180] or fallback


def save_inspection_original_pdf(file_storage):
    if not file_storage or not file_storage.filename:
        raise ValueError("请选择需要上传的 PDF 文件。")

    original_name = normalize_upload_display_name(file_storage.filename)
    ext = os.path.splitext(original_name)[1].lower()
    if ext != ".pdf":
        raise ValueError("仅支持上传 PDF 文件。")

    pdf_bytes = file_storage.read()
    if not pdf_bytes:
        raise ValueError("上传 PDF 内容为空。")

    if len(pdf_bytes) > MAX_PDF_BYTES:
        raise ValueError("PDF 文件不能超过 50MB。")

    if not pdf_bytes.startswith(b"%PDF"):
        raise ValueError("上传文件不是有效的 PDF 文件。")

    target_dir, now = ensure_storage_subdir(INSPECTION_ORIGINALS_STORAGE_DIR)
    filename = f"checklist_original_{now.strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex}.pdf"
    save_path = os.path.join(target_dir, filename)

    with open(save_path, "wb") as f:
        f.write(pdf_bytes)

    relative_dir = os.path.relpath(target_dir, STORAGE_ROOT).replace("\\", "/")
    return f"/{relative_dir}/{filename}", original_name, len(pdf_bytes)


def get_training_material_file_type(extension):
    return "pdf" if extension == ".pdf" else "video"


def save_training_material_file(file_storage):
    if not file_storage or not file_storage.filename:
        raise ValueError("请选择需要上传的培训材料。")

    original_name = normalize_upload_display_name(file_storage.filename)
    ext = os.path.splitext(original_name)[1].lower()
    if ext not in ALLOWED_TRAINING_MATERIAL_EXTENSIONS:
        raise ValueError("仅支持上传 PDF 或 MP4、MOV、M4V、WEBM 视频文件。")

    file_bytes = file_storage.read()
    if not file_bytes:
        raise ValueError("上传文件内容为空。")

    file_type = get_training_material_file_type(ext)
    max_size = MAX_PDF_BYTES if file_type == "pdf" else MAX_VIDEO_BYTES
    if len(file_bytes) > max_size:
        size_label = "50MB" if file_type == "pdf" else "500MB"
        raise ValueError(f"{'PDF' if file_type == 'pdf' else '视频'}文件不能超过 {size_label}。")

    if file_type == "pdf" and not file_bytes.startswith(b"%PDF"):
        raise ValueError("上传文件不是有效的 PDF 文件。")

    target_dir, now = ensure_storage_subdir(TRAINING_MATERIALS_STORAGE_DIR)
    filename = (
        f"training_material_{now.strftime('%Y%m%d_%H%M%S')}_"
        f"{uuid.uuid4().hex}{ext}"
    )
    save_path = os.path.join(target_dir, filename)

    with open(save_path, "wb") as f:
        f.write(file_bytes)

    relative_dir = os.path.relpath(target_dir, STORAGE_ROOT).replace("\\", "/")
    return f"/{relative_dir}/{filename}", original_name, len(file_bytes), file_type


def resolve_storage_abs_path(relative_path):
    value = str(relative_path or "").strip()
    if value.startswith("/storage/"):
        value = value[len("/storage/") :]
    elif value.startswith("/"):
        value = value[1:]

    storage_root_abs = os.path.abspath(STORAGE_ROOT)
    target_path = os.path.abspath(os.path.join(storage_root_abs, value))
    if os.path.commonpath([storage_root_abs, target_path]) != storage_root_abs:
        return None
    return target_path


def remove_storage_file(relative_path):
    target_path = resolve_storage_abs_path(relative_path)
    try:
        if target_path and os.path.isfile(target_path):
            os.remove(target_path)
    except OSError:
        pass


def get_storage_file_size(relative_path):
    target_path = resolve_storage_abs_path(relative_path)
    if not target_path or not os.path.isfile(target_path):
        return 0
    try:
        return os.path.getsize(target_path)
    except OSError:
        return 0


def format_file_size(size_bytes):
    try:
        size = float(size_bytes or 0)
    except (TypeError, ValueError):
        size = 0
    if size <= 0:
        return ""
    units = ["B", "KB", "MB", "GB"]
    unit_index = 0
    while size >= 1024 and unit_index < len(units) - 1:
        size /= 1024
        unit_index += 1
    if unit_index == 0:
        return f"{int(size)} {units[unit_index]}"
    return f"{size:.1f} {units[unit_index]}"


BACKUP_FREQUENCY_INTERVALS = {
    "off": None,
    "hourly": timedelta(hours=1),
    "daily": timedelta(days=1),
    "weekly": timedelta(days=7),
    "monthly": timedelta(days=30),
}
backup_scheduler_started = False
backup_scheduler_lock = threading.Lock()
backup_job_lock = threading.Lock()
issue_export_cleanup_lock = threading.Lock()
issue_export_cleanup_last_run = 0
station_score_export_cleanup_lock = threading.Lock()
station_score_export_cleanup_last_run = 0


def isoformat_or_none(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value.astimezone(BEIJING_TZ).isoformat()
    return str(value)


def parse_backup_time(value):
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value))
    except ValueError:
        return None


def normalize_backup_scheduled_time(value):
    text = str(value or "").strip()
    if not text:
        return "02:00"
    matched = re.fullmatch(r"(\d{1,2}):([0-5]\d)", text)
    if not matched:
        raise ValueError("自动备份执行时间必须是 HH:MM 格式。")
    hour = int(matched.group(1))
    if hour > 23:
        raise ValueError("自动备份执行时间小时必须在 00-23 之间。")
    return f"{hour:02d}:{matched.group(2)}"


def parse_backup_scheduled_time(value):
    scheduled_time = normalize_backup_scheduled_time(value)
    hour, minute = scheduled_time.split(":")
    return int(hour), int(minute)


def normalize_backup_destination_path(value):
    raw_value = str(value or "").strip()
    if not raw_value:
        return os.path.abspath(DEFAULT_BACKUP_DIR)
    if os.path.isabs(raw_value):
        return os.path.abspath(raw_value)
    return os.path.abspath(os.path.join(DEFAULT_BACKUP_DIR, raw_value))


def get_default_backup_config():
    return {
        "destination_path": os.path.abspath(DEFAULT_BACKUP_DIR),
        "frequency": "off",
        "scheduled_time": "02:00",
        "next_auto_run_at": None,
        "last_auto_export_at": None,
        "last_backup_path": None,
        "last_backup_size": None,
        "last_cos_status": "not_configured",
        "last_cos_key": None,
        "last_cos_uploaded_at": None,
        "last_cos_retained_count": 0,
        "last_cos_error": "",
        "last_status": "idle",
        "last_error": "",
        "updated_at": None,
    }


def read_backup_config():
    config = get_default_backup_config()
    if os.path.isfile(BACKUP_CONFIG_PATH):
        try:
            with open(BACKUP_CONFIG_PATH, "r", encoding="utf-8") as f:
                saved_config = json.load(f)
            if isinstance(saved_config, dict):
                config.update(saved_config)
        except (OSError, json.JSONDecodeError):
            pass

    config["destination_path"] = normalize_backup_destination_path(
        config.get("destination_path")
    )
    if config.get("frequency") not in BACKUP_FREQUENCY_INTERVALS:
        config["frequency"] = "off"
    config["scheduled_time"] = normalize_backup_scheduled_time(config.get("scheduled_time"))
    if config.get("frequency") == "off":
        config["next_auto_run_at"] = None
    return config


def write_backup_config(config):
    next_config = get_default_backup_config()
    next_config.update(config or {})
    next_config["destination_path"] = normalize_backup_destination_path(
        next_config.get("destination_path")
    )
    if next_config.get("frequency") not in BACKUP_FREQUENCY_INTERVALS:
        next_config["frequency"] = "off"
    next_config["scheduled_time"] = normalize_backup_scheduled_time(next_config.get("scheduled_time"))
    if next_config.get("frequency") == "off":
        next_config["next_auto_run_at"] = None
    next_config["updated_at"] = beijing_now().isoformat()
    os.makedirs(os.path.dirname(BACKUP_CONFIG_PATH), exist_ok=True)
    with open(BACKUP_CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(next_config, f, ensure_ascii=False, indent=2)
    return next_config


def calculate_next_auto_run_at(config, from_time=None):
    frequency = config.get("frequency")
    interval = BACKUP_FREQUENCY_INTERVALS.get(frequency)
    if not interval:
        return None
    now = (from_time or beijing_now()).astimezone(BEIJING_TZ)

    if frequency == "hourly":
        return (now + interval).replace(second=0, microsecond=0).isoformat()

    hour, minute = parse_backup_scheduled_time(config.get("scheduled_time"))
    candidate = now.replace(hour=hour, minute=minute, second=0, microsecond=0)
    if frequency == "daily":
        if candidate <= now:
            candidate += timedelta(days=1)
        return candidate.isoformat()

    if frequency == "weekly":
        while candidate <= now:
            candidate += timedelta(days=7)
        return candidate.isoformat()

    if frequency == "monthly":
        while candidate <= now:
            candidate += timedelta(days=30)
        return candidate.isoformat()

    return None


def get_backup_next_run_at(config):
    if config.get("frequency") == "off":
        return None
    saved_next_run = parse_backup_time(config.get("next_auto_run_at"))
    if saved_next_run:
        return saved_next_run.astimezone(BEIJING_TZ).isoformat()
    return calculate_next_auto_run_at(config)


def list_backup_files(destination_path):
    backup_dir = normalize_backup_destination_path(destination_path)
    if not os.path.isdir(backup_dir):
        return []
    rows = []
    for filename in os.listdir(backup_dir):
        if not (filename.startswith(BACKUP_PREFIX) and filename.endswith(".zip")):
            continue
        file_path = os.path.join(backup_dir, filename)
        if not os.path.isfile(file_path):
            continue
        stat = os.stat(file_path)
        rows.append(
            {
                "filename": filename,
                "path": file_path,
                "size": stat.st_size,
                "updated_at": datetime.fromtimestamp(stat.st_mtime, BEIJING_TZ).isoformat(),
            }
        )
    return sorted(rows, key=lambda item: item["updated_at"], reverse=True)[:20]


def get_cos_env_config():
    secret_id = os.environ.get("COS_SECRET_ID", "").strip()
    secret_key = os.environ.get("COS_SECRET_KEY", "").strip()
    region = os.environ.get("COS_REGION", "").strip()
    bucket = os.environ.get("COS_BUCKET", "").strip()
    prefix = f"{COS_BACKUP_PREFIX}/" if COS_BACKUP_PREFIX else ""
    configured = all((secret_id, secret_key, region, bucket))
    return {
        "secret_id": secret_id,
        "secret_key": secret_key,
        "region": region,
        "bucket": bucket,
        "prefix": prefix,
        "configured": configured,
        "sdk_available": bool(CosConfig and CosS3Client),
        "retention_count": COS_BACKUP_RETENTION_COUNT,
    }


def get_cos_public_status(config=None, error=""):
    cos_config = get_cos_env_config()
    if not cos_config["configured"]:
        status = "not_configured"
        message = "未配置 COS 环境变量，当前仅保留本地备份。"
    elif not cos_config["sdk_available"]:
        status = "error"
        message = "后端缺少腾讯云 COS Python SDK，请重新安装依赖。"
    elif error:
        status = "error"
        message = error
    else:
        status = "ready"
        message = "COS 环境变量已配置，备份会自动上传对象存储。"

    return {
        "status": status,
        "message": message,
        "configured": cos_config["configured"],
        "sdk_available": cos_config["sdk_available"],
        "bucket": cos_config["bucket"] if cos_config["configured"] else "",
        "region": cos_config["region"] if cos_config["configured"] else "",
        "prefix": cos_config["prefix"],
        "retention_count": cos_config["retention_count"],
        "last_cos_status": (config or {}).get("last_cos_status"),
        "last_cos_key": (config or {}).get("last_cos_key"),
        "last_cos_uploaded_at": (config or {}).get("last_cos_uploaded_at"),
        "last_cos_error": (config or {}).get("last_cos_error") or "",
        "last_cos_retained_count": (config or {}).get("last_cos_retained_count") or 0,
    }


def get_cos_client():
    cos_config = get_cos_env_config()
    if not cos_config["configured"]:
        raise RuntimeError("未配置 COS_SECRET_ID、COS_SECRET_KEY、COS_REGION 或 COS_BUCKET。")
    if not cos_config["sdk_available"]:
        raise RuntimeError("后端缺少腾讯云 COS Python SDK，请重新安装依赖。")
    config = CosConfig(
        Region=cos_config["region"],
        SecretId=cos_config["secret_id"],
        SecretKey=cos_config["secret_key"],
        Scheme="https",
    )
    return CosS3Client(config), cos_config


def parse_cos_last_modified(value):
    text = str(value or "").strip()
    if not text:
        return datetime.min.replace(tzinfo=BEIJING_TZ)
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).astimezone(BEIJING_TZ)
    except ValueError:
        try:
            return datetime.strptime(text, "%Y-%m-%dT%H:%M:%S.%fZ").replace(tzinfo=ZoneInfo("UTC")).astimezone(BEIJING_TZ)
        except ValueError:
            return datetime.min.replace(tzinfo=BEIJING_TZ)


def list_cos_backup_objects():
    client, cos_config = get_cos_client()
    rows = []
    marker = ""
    while True:
        response = client.list_objects(
            Bucket=cos_config["bucket"],
            Prefix=cos_config["prefix"],
            Marker=marker,
            MaxKeys=1000,
        )
        contents = response.get("Contents") or []
        if isinstance(contents, dict):
            contents = [contents]
        for item in contents:
            key = item.get("Key") or ""
            filename = os.path.basename(key)
            if not (filename.startswith(BACKUP_PREFIX) and filename.endswith(".zip")):
                continue
            last_modified = item.get("LastModified") or ""
            rows.append(
                {
                    "key": key,
                    "filename": filename,
                    "size": int(item.get("Size") or 0),
                    "updated_at": parse_cos_last_modified(last_modified).isoformat(),
                    "etag": str(item.get("ETag") or "").strip('"'),
                }
            )
        if str(response.get("IsTruncated", "")).lower() not in {"true", "1"}:
            break
        marker = response.get("NextMarker") or response.get("Marker") or ""
        if not marker:
            break
    return sorted(rows, key=lambda item: item["updated_at"], reverse=True)


def prune_cos_backup_objects(objects=None):
    client, cos_config = get_cos_client()
    rows = objects if objects is not None else list_cos_backup_objects()
    stale_objects = rows[COS_BACKUP_RETENTION_COUNT:]
    for item in stale_objects:
        client.delete_object(Bucket=cos_config["bucket"], Key=item["key"])
    return rows[:COS_BACKUP_RETENTION_COUNT], stale_objects


def upload_backup_archive_to_cos(local_path, object_filename):
    cos_config = get_cos_env_config()
    if not cos_config["configured"]:
        return {
            "status": "not_configured",
            "message": "未配置 COS 环境变量，已跳过云端上传。",
        }
    if not cos_config["sdk_available"]:
        return {
            "status": "error",
            "message": "后端缺少腾讯云 COS Python SDK，请重新安装依赖。",
        }

    try:
        client, cos_config = get_cos_client()
        object_key = f"{cos_config['prefix']}{object_filename}"
        client.put_object_from_local_file(
            Bucket=cos_config["bucket"],
            Key=object_key,
            LocalFilePath=local_path,
        )
        retained, removed = prune_cos_backup_objects()
        return {
            "status": "success",
            "message": f"COS 上传成功，云端已保留最近 {len(retained)} 个备份。",
            "key": object_key,
            "bucket": cos_config["bucket"],
            "region": cos_config["region"],
            "uploaded_at": beijing_now().isoformat(),
            "retained_count": len(retained),
            "removed_count": len(removed),
        }
    except Exception as exc:
        return {
            "status": "error",
            "message": f"COS 上传失败：{str(exc)}",
        }


def get_cos_backup_overview(config=None):
    status = get_cos_public_status(config)
    backups = []
    if status["configured"] and status["sdk_available"]:
        try:
            backups = list_cos_backup_objects()[:COS_BACKUP_RETENTION_COUNT]
        except Exception as exc:
            status = get_cos_public_status(config, f"COS 备份列表读取失败：{str(exc)}")
    return status, backups


def build_pg_environment():
    db_config = get_db_config()
    env = os.environ.copy()
    env["PGPASSWORD"] = db_config["password"]
    return db_config, env


def run_backup_subprocess(command, env):
    try:
        subprocess.run(command, env=env, check=True, capture_output=True, text=True)
    except FileNotFoundError as exc:
        raise RuntimeError(
            "数据库备份工具未安装，请确认后端环境已安装 postgresql-client。"
        ) from exc
    except subprocess.CalledProcessError as exc:
        error_text = (exc.stderr or exc.stdout or "").strip()
        raise RuntimeError(error_text or "数据库备份工具执行失败。") from exc


def get_database_backup_member_name(names):
    if "database.dump" in names:
        return "database.dump"
    if "database.sql" in names:
        return "database.sql"
    raise ValueError("备份文件缺少 database.dump 或 database.sql。")


def should_skip_storage_backup_path(path, excluded_dirs):
    abs_path = os.path.abspath(path)
    for excluded_dir in excluded_dirs:
        try:
            if os.path.commonpath([excluded_dir, abs_path]) == excluded_dir:
                return True
        except ValueError:
            continue
    return False


def get_storage_backup_excluded_dirs(destination_path):
    storage_root_abs = os.path.abspath(STORAGE_ROOT)
    excluded_dirs = []
    for candidate in (destination_path, DEFAULT_BACKUP_DIR):
        candidate_abs = os.path.abspath(candidate)
        try:
            if (
                candidate_abs != storage_root_abs
                and os.path.commonpath([storage_root_abs, candidate_abs]) == storage_root_abs
                and candidate_abs not in excluded_dirs
            ):
                excluded_dirs.append(candidate_abs)
        except ValueError:
            continue
    return excluded_dirs


def add_storage_to_backup(zip_file, destination_path):
    storage_root_abs = os.path.abspath(STORAGE_ROOT)
    excluded_dirs = get_storage_backup_excluded_dirs(destination_path)
    for root, dirs, files in os.walk(storage_root_abs):
        dirs[:] = [
            name
            for name in dirs
            if not should_skip_storage_backup_path(os.path.join(root, name), excluded_dirs)
        ]
        if should_skip_storage_backup_path(root, excluded_dirs):
            continue
        for filename in files:
            file_path = os.path.join(root, filename)
            if file_path == os.path.abspath(BACKUP_CONFIG_PATH):
                continue
            if (
                os.path.abspath(root) == storage_root_abs
                and filename.startswith(BACKUP_PREFIX)
                and filename.endswith(".zip")
            ):
                continue
            if should_skip_storage_backup_path(file_path, excluded_dirs):
                continue
            relative_path = os.path.relpath(file_path, storage_root_abs).replace("\\", "/")
            zip_file.write(file_path, f"storage/{relative_path}")


def publish_backup_archive(temp_zip_path, final_path):
    final_dir = os.path.dirname(final_path)
    os.makedirs(final_dir, exist_ok=True)
    staging_path = os.path.join(
        final_dir,
        f".{os.path.basename(final_path)}.{uuid.uuid4().hex}.tmp",
    )

    try:
        shutil.copy2(temp_zip_path, staging_path)
        os.replace(staging_path, final_path)
    finally:
        if os.path.exists(staging_path):
            os.remove(staging_path)


def cleanup_local_backup_files(backup_dir, keep_path=None):
    backup_dir = normalize_backup_destination_path(backup_dir)
    keep_abs = os.path.abspath(keep_path) if keep_path else ""
    if not os.path.isdir(backup_dir):
        return
    for filename in os.listdir(backup_dir):
        if not (filename.startswith(BACKUP_PREFIX) and filename.endswith(".zip")):
            continue
        file_path = os.path.abspath(os.path.join(backup_dir, filename))
        try:
            if os.path.commonpath([backup_dir, file_path]) != backup_dir:
                continue
        except ValueError:
            continue
        if keep_abs and file_path == keep_abs:
            continue
        if os.path.isfile(file_path):
            os.remove(file_path)


def create_full_backup_archive(destination_path=None, reason="manual"):
    backup_dir = normalize_backup_destination_path(destination_path)
    os.makedirs(backup_dir, exist_ok=True)

    now = beijing_now()
    filename = LOCAL_BACKUP_FILENAME
    download_filename = f"{BACKUP_PREFIX}_{reason}_{now.strftime('%Y%m%d_%H%M%S')}.zip"
    final_path = os.path.join(backup_dir, filename)

    db_config, env = build_pg_environment()
    with tempfile.TemporaryDirectory() as temp_dir:
        database_dump_path = os.path.join(temp_dir, "database.dump")
        pg_dump_command = [
            "pg_dump",
            "-h",
            db_config["host"],
            "-p",
            db_config["port"],
            "-U",
            db_config["user"],
            "-d",
            db_config["dbname"],
            "-Fc",
            "--no-owner",
            "--no-acl",
            "-f",
            database_dump_path,
        ]
        run_backup_subprocess(pg_dump_command, env)

        manifest = {
            "backup_type": "ywddzx_full_backup",
            "version": 2,
            "created_at": now.isoformat(),
            "reason": reason,
            "database": db_config["dbname"],
            "database_member": "database.dump",
            "database_format": "custom",
            "restore_strategy": "drop_user_schemas_then_restore_sql",
            "storage_root": STORAGE_ROOT,
        }

        temp_zip_path = os.path.join(temp_dir, filename)
        with zipfile.ZipFile(temp_zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zip_file:
            zip_file.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))
            zip_file.write(database_dump_path, "database.dump")
            add_storage_to_backup(zip_file, backup_dir)

        publish_backup_archive(temp_zip_path, final_path)
        cleanup_local_backup_files(backup_dir, keep_path=final_path)
        if os.path.abspath(DEFAULT_BACKUP_DIR) != os.path.abspath(backup_dir):
            cleanup_local_backup_files(DEFAULT_BACKUP_DIR)

    cos_result = upload_backup_archive_to_cos(final_path, download_filename)
    return {
        "path": final_path,
        "filename": filename,
        "download_filename": download_filename,
        "size": os.path.getsize(final_path),
        "created_at": now.isoformat(),
        "cos": cos_result,
    }


def validate_backup_archive(zip_path):
    try:
        with zipfile.ZipFile(zip_path, "r") as zip_file:
            names = set(zip_file.namelist())
            if "manifest.json" not in names:
                raise ValueError("备份文件缺少 manifest.json。")
            database_member_name = get_database_backup_member_name(names)
            manifest = json.loads(zip_file.read("manifest.json").decode("utf-8"))
            if manifest.get("backup_type") != "ywddzx_full_backup":
                raise ValueError("备份文件类型不正确。")
            manifest["database_member"] = database_member_name
            return manifest
    except zipfile.BadZipFile as exc:
        raise ValueError("备份文件不是有效的 ZIP 文件。") from exc
    except json.JSONDecodeError as exc:
        raise ValueError("备份文件清单格式不正确。") from exc


def should_skip_restore_sql_line(line):
    normalized = line.strip().lower()
    return (
        normalized.startswith("set transaction_timeout ")
        or normalized == "create schema public;"
        or normalized == "create schema if not exists public;"
    )


def build_restore_sql_file(source_sql_path, restore_sql_path):
    with open(source_sql_path, "r", encoding="utf-8", errors="replace") as source:
        with open(restore_sql_path, "w", encoding="utf-8") as target:
            target.write("BEGIN;\n")
            target.write(
                """
DO $ywddzx_restore$
DECLARE
    schema_record RECORD;
BEGIN
    FOR schema_record IN
        SELECT nspname
        FROM pg_namespace
        WHERE nspname NOT LIKE 'pg_%'
          AND nspname <> 'information_schema'
    LOOP
        EXECUTE format('DROP SCHEMA IF EXISTS %I CASCADE', schema_record.nspname);
    END LOOP;
END
$ywddzx_restore$;
CREATE SCHEMA IF NOT EXISTS public;
GRANT ALL ON SCHEMA public TO PUBLIC;
"""
            )
            for line in source:
                if should_skip_restore_sql_line(line):
                    continue
                target.write(line)
            target.write("\nCOMMIT;\n")


def convert_database_backup_to_sql(database_backup_path, database_member_name, temp_dir):
    _db_config, env = build_pg_environment()
    if database_member_name == "database.sql":
        return database_backup_path

    database_sql_path = os.path.join(temp_dir, "database.sql")
    pg_restore_command = [
        "pg_restore",
        "--no-owner",
        "--no-acl",
        "-f",
        database_sql_path,
        database_backup_path,
    ]
    run_backup_subprocess(pg_restore_command, env)
    return database_sql_path


def restore_database_from_backup(database_backup_path, database_member_name, temp_dir):
    db_config, env = build_pg_environment()
    database_sql_path = convert_database_backup_to_sql(
        database_backup_path,
        database_member_name,
        temp_dir,
    )
    restore_sql_path = os.path.join(temp_dir, "restore.sql")
    build_restore_sql_file(database_sql_path, restore_sql_path)
    psql_command = [
        "psql",
        "-h",
        db_config["host"],
        "-p",
        db_config["port"],
        "-U",
        db_config["user"],
        "-d",
        db_config["dbname"],
        "-v",
        "ON_ERROR_STOP=1",
        "-q",
        "-f",
        restore_sql_path,
    ]
    run_backup_subprocess(psql_command, env)


def get_required_storage_dirs():
    return (
        ISSUES_STORAGE_DIR,
        RECTIFICATIONS_STORAGE_DIR,
        SIGNATURES_STORAGE_DIR,
        INSPECTION_ORIGINALS_STORAGE_DIR,
        TRAINING_MATERIALS_STORAGE_DIR,
        FEEDBACK_SCREENSHOTS_STORAGE_DIR,
        ISSUE_EXPORTS_STORAGE_DIR,
        STATION_SCORE_EXPORTS_STORAGE_DIR,
        DEFAULT_BACKUP_DIR,
    )


def ensure_required_storage_dirs():
    for directory in get_required_storage_dirs():
        os.makedirs(directory, exist_ok=True)


def clear_storage_root_for_restore():
    storage_root_abs = os.path.abspath(STORAGE_ROOT)
    os.makedirs(storage_root_abs, exist_ok=True)
    for name in os.listdir(storage_root_abs):
        target_path = os.path.abspath(os.path.join(storage_root_abs, name))
        if os.path.commonpath([storage_root_abs, target_path]) != storage_root_abs:
            raise ValueError("storage 目录中存在非法路径，无法安全清空。")
        if os.path.isdir(target_path) and not os.path.islink(target_path):
            shutil.rmtree(target_path)
        else:
            os.remove(target_path)
    ensure_required_storage_dirs()


def validate_storage_backup_members(zip_file):
    storage_root_abs = os.path.abspath(STORAGE_ROOT)
    for member in zip_file.infolist():
        name = member.filename
        if not name.startswith("storage/") or name.endswith("/"):
            continue
        relative_path = name[len("storage/") :]
        if not relative_path:
            continue
        target_path = os.path.abspath(os.path.join(storage_root_abs, relative_path))
        if os.path.commonpath([storage_root_abs, target_path]) != storage_root_abs:
            raise ValueError("备份文件中存在非法 storage 路径。")


def restore_storage_from_backup(zip_path):
    storage_root_abs = os.path.abspath(STORAGE_ROOT)
    with zipfile.ZipFile(zip_path, "r") as zip_file:
        validate_storage_backup_members(zip_file)
        clear_storage_root_for_restore()
        for member in zip_file.infolist():
            name = member.filename
            if not name.startswith("storage/") or name.endswith("/"):
                continue
            relative_path = name[len("storage/") :]
            target_path = os.path.abspath(os.path.join(storage_root_abs, relative_path))
            if os.path.commonpath([storage_root_abs, target_path]) != storage_root_abs:
                raise ValueError("备份文件中存在非法 storage 路径。")
            os.makedirs(os.path.dirname(target_path), exist_ok=True)
            with zip_file.open(member) as source, open(target_path, "wb") as target:
                shutil.copyfileobj(source, target)
        ensure_required_storage_dirs()


def restore_full_backup_archive(file_storage):
    if not file_storage or not file_storage.filename:
        raise ValueError("请选择需要导入的完整备份文件。")
    if not str(file_storage.filename).lower().endswith(".zip"):
        raise ValueError("完整备份只支持导入 ZIP 文件。")

    with tempfile.TemporaryDirectory() as temp_dir:
        backup_path = os.path.join(temp_dir, "backup.zip")
        file_storage.save(backup_path)
        manifest = validate_backup_archive(backup_path)
        with zipfile.ZipFile(backup_path, "r") as zip_file:
            database_member_name = manifest.get("database_member") or get_database_backup_member_name(
                set(zip_file.namelist())
            )
            database_backup_path = os.path.join(temp_dir, os.path.basename(database_member_name))
            with zip_file.open(database_member_name) as source, open(database_backup_path, "wb") as target:
                shutil.copyfileobj(source, target)
            validate_storage_backup_members(zip_file)
        restore_database_from_backup(database_backup_path, database_member_name, temp_dir)
        restore_storage_from_backup(backup_path)
        return manifest


def mark_auto_backup_result(success, result=None, error=""):
    config = read_backup_config()
    next_auto_run_at = calculate_next_auto_run_at(config)
    if success:
        cos_result = (result or {}).get("cos") or {}
        cos_status = cos_result.get("status") or "not_configured"
        config.update(
            {
                "last_auto_export_at": beijing_now().isoformat(),
                "next_auto_run_at": next_auto_run_at,
                "last_backup_path": result.get("path") if result else None,
                "last_backup_size": result.get("size") if result else None,
                "last_cos_status": cos_status,
                "last_cos_key": cos_result.get("key"),
                "last_cos_uploaded_at": cos_result.get("uploaded_at"),
                "last_cos_retained_count": cos_result.get("retained_count") or 0,
                "last_cos_error": cos_result.get("message") if cos_status == "error" else "",
                "last_status": "warning" if cos_status == "error" else "success",
                "last_error": cos_result.get("message") if cos_status == "error" else "",
            }
        )
    else:
        config.update(
            {
                "next_auto_run_at": next_auto_run_at,
                "last_status": "error",
                "last_error": str(error or "自动备份失败。")[:500],
            }
        )
    write_backup_config(config)


def maybe_run_scheduled_backup():
    config = read_backup_config()
    frequency = config.get("frequency")
    interval = BACKUP_FREQUENCY_INTERVALS.get(frequency)
    if not interval:
        return

    next_run_at = parse_backup_time(config.get("next_auto_run_at"))
    if not next_run_at:
        config["next_auto_run_at"] = calculate_next_auto_run_at(config)
        write_backup_config(config)
        return

    if beijing_now() < next_run_at.astimezone(BEIJING_TZ):
        return

    if not backup_job_lock.acquire(blocking=False):
        return
    try:
        try:
            result = create_full_backup_archive(config.get("destination_path"), reason="auto")
            mark_auto_backup_result(True, result=result)
        except Exception as exc:
            mark_auto_backup_result(False, error=exc)
    finally:
        backup_job_lock.release()


def backup_scheduler_loop():
    while True:
        try:
            maybe_run_scheduled_backup()
            maybe_cleanup_expired_issue_exports()
        except Exception:
            pass
        time.sleep(60)


def start_backup_scheduler_once():
    global backup_scheduler_started
    if backup_scheduler_started:
        return
    with backup_scheduler_lock:
        if backup_scheduler_started:
            return
        thread = threading.Thread(target=backup_scheduler_loop, daemon=True)
        thread.start()
        backup_scheduler_started = True


def get_auth_serializer():
    return URLSafeTimedSerializer(app.config["SECRET_KEY"], salt=AUTH_TOKEN_SALT)


def get_password_fingerprint(password):
    return hashlib.sha256(str(password or "").encode("utf-8")).hexdigest()[:24]


def get_auth_token_ttl_seconds(user):
    if str(user.get("role") or "") in PRIVILEGED_AUTH_ROLES:
        return AUTH_TOKEN_PRIVILEGED_MAX_AGE_SECONDS
    return AUTH_TOKEN_NORMAL_MAX_AGE_SECONDS


def current_epoch_seconds():
    return int(time.time())


def fetch_auth_user_by_id(cur, user_id):
    cur.execute(
        """
        SELECT
            u.id,
            u.username,
            u.password,
            u.role,
            u.real_name,
            u.phone,
            u.station_id,
            s.station_name,
            s.region,
            s.address
        FROM users u
        LEFT JOIN stations s ON u.station_id = s.id
        WHERE u.id = %s
        LIMIT 1;
        """,
        (user_id,),
    )
    return cur.fetchone()


def create_auth_token(user):
    issued_at = current_epoch_seconds()
    expires_at = issued_at + get_auth_token_ttl_seconds(user)
    payload = {
        "uid": int(user["id"]),
        "username": user["username"],
        "role": user["role"],
        "pwd": get_password_fingerprint(user.get("password")),
        "iat": issued_at,
        "exp": expires_at,
    }
    return get_auth_serializer().dumps(payload)


def get_auth_payload_expires_in(payload):
    try:
        expires_at = int(payload.get("exp") or 0)
    except (TypeError, ValueError):
        return 0
    return max(0, expires_at - current_epoch_seconds())


def build_auth_user_payload(cur, user):
    permissions = get_effective_permissions(cur, user)
    return {
        "id": user["id"],
        "username": user["username"],
        "role": user["role"],
        "real_name": user["real_name"],
        "phone": user["phone"],
        "station_id": user["station_id"],
        "station_name": user.get("station_name"),
        "region": user.get("region"),
        "address": user.get("address"),
        "permissions": permissions,
        "must_change_password": user.get("password") == DEFAULT_INITIAL_PASSWORD,
        "birthday_event": get_user_birthday_event(cur, user),
    }


def extract_bearer_token():
    header = str(request.headers.get("Authorization", "")).strip()
    if header.lower().startswith("bearer "):
        return header[7:].strip()
    return ""


def get_auth_cache_key(token):
    return hashlib.sha256(str(token or "").encode("utf-8")).hexdigest()


def prune_auth_token_cache_locked(now):
    if len(auth_token_cache) <= AUTH_TOKEN_CACHE_MAX_ENTRIES:
        return
    expired_keys = [
        key
        for key, item in auth_token_cache.items()
        if now - item.get("cached_at", 0) >= AUTH_SERVER_CACHE_TTL_SECONDS
        or int(item.get("payload", {}).get("exp") or 0) <= now
    ]
    for key in expired_keys:
        auth_token_cache.pop(key, None)
    if len(auth_token_cache) <= AUTH_TOKEN_CACHE_MAX_ENTRIES:
        return
    sorted_items = sorted(auth_token_cache.items(), key=lambda item: item[1].get("cached_at", 0))
    for key, _ in sorted_items[: max(1, len(auth_token_cache) - AUTH_TOKEN_CACHE_MAX_ENTRIES)]:
        auth_token_cache.pop(key, None)


def get_cached_auth_token_user(token, payload, now):
    cache_key = get_auth_cache_key(token)
    with auth_token_cache_lock:
        cached = auth_token_cache.get(cache_key)
        if not cached:
            return None
        if cached.get("user_id") != payload.get("uid"):
            auth_token_cache.pop(cache_key, None)
            return None
        if now - cached.get("cached_at", 0) >= AUTH_SERVER_CACHE_TTL_SECONDS:
            auth_token_cache.pop(cache_key, None)
            return None
        if int(cached.get("payload", {}).get("exp") or 0) <= now:
            auth_token_cache.pop(cache_key, None)
            return None
        return dict(cached["user"]), dict(cached["payload"])


def set_cached_auth_token_user(token, user, payload):
    now = current_epoch_seconds()
    cache_key = get_auth_cache_key(token)
    with auth_token_cache_lock:
        prune_auth_token_cache_locked(now)
        auth_token_cache[cache_key] = {
            "user_id": int(user["id"]),
            "user": dict(user),
            "payload": dict(payload),
            "cached_at": now,
        }


def get_auth_me_cache_key(user_id, token):
    return f"{user_id}:{get_auth_cache_key(token)}"


def get_cached_auth_me_payload(user_id, token):
    now = current_epoch_seconds()
    cache_key = get_auth_me_cache_key(user_id, token)
    with auth_me_response_cache_lock:
        cached = auth_me_response_cache.get(cache_key)
        if not cached:
            return None
        if now - cached.get("cached_at", 0) >= AUTH_SERVER_CACHE_TTL_SECONDS:
            auth_me_response_cache.pop(cache_key, None)
            return None
        return dict(cached["payload"])


def set_cached_auth_me_payload(user_id, token, payload):
    cache_key = get_auth_me_cache_key(user_id, token)
    with auth_me_response_cache_lock:
        auth_me_response_cache[cache_key] = {
            "payload": dict(payload),
            "cached_at": current_epoch_seconds(),
        }


def invalidate_auth_caches_for_user(user_id=None):
    target_user_id = None
    try:
        target_user_id = int(user_id) if user_id is not None else None
    except (TypeError, ValueError):
        target_user_id = None

    with auth_token_cache_lock:
        if target_user_id is None:
            auth_token_cache.clear()
        else:
            for key in list(auth_token_cache.keys()):
                if auth_token_cache[key].get("user_id") == target_user_id:
                    auth_token_cache.pop(key, None)

    with auth_me_response_cache_lock:
        if target_user_id is None:
            auth_me_response_cache.clear()
        else:
            prefix = f"{target_user_id}:"
            for key in list(auth_me_response_cache.keys()):
                if key.startswith(prefix):
                    auth_me_response_cache.pop(key, None)


def verify_auth_token(token):
    try:
        payload = get_auth_serializer().loads(token)
    except BadSignature as exc:
        raise PermissionError("登录已过期，请重新登录。") from exc

    user_id = payload.get("uid")
    if not user_id:
        raise PermissionError("登录已过期，请重新登录。")

    try:
        expires_at = int(payload.get("exp") or 0)
    except (TypeError, ValueError) as exc:
        raise PermissionError("登录已过期，请重新登录。") from exc

    now = current_epoch_seconds()
    if expires_at <= now:
        raise PermissionError("登录已过期，请重新登录。")

    cached_auth = get_cached_auth_token_user(token, payload, now)
    if cached_auth:
        return cached_auth

    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        user = fetch_auth_user_by_id(cur, user_id)
    finally:
        close_db_resources(cur, conn)

    if not user:
        raise PermissionError("登录已过期，请重新登录。")
    if payload.get("username") != user["username"]:
        raise PermissionError("登录已过期，请重新登录。")
    if payload.get("role") != user["role"]:
        raise PermissionError("登录已过期，请重新登录。")
    if payload.get("pwd") != get_password_fingerprint(user.get("password")):
        raise PermissionError("登录已过期，请重新登录。")
    set_cached_auth_token_user(token, user, payload)
    return dict(user), dict(payload)


def normalize_identity_for_compare(value):
    text = str(value or "").strip()
    if not text:
        return ""
    try:
        parsed = int(text)
    except (TypeError, ValueError):
        return text
    return str(parsed) if parsed > 0 else text


def get_authenticated_request_user_id(fallback=None):
    current_user = getattr(g, "current_user", None)
    if current_user:
        return str(current_user["id"])
    return str(fallback or "").strip()


def iter_request_identity_values():
    for key in ("user_id", "inspector_id"):
        for value in request.args.getlist(key):
            yield key, value
        for value in request.form.getlist(key):
            yield key, value

    data = request.get_json(silent=True) if request.is_json else None
    if isinstance(data, dict):
        for key in ("user_id", "inspector_id"):
            if key in data:
                yield key, data.get(key)


def is_public_api_path(path):
    if not path.startswith("/api/"):
        return True
    if path in LEGACY_FRONTEND_API_PATHS:
        return True
    return path in {
        "/api/health",
        "/api/db-test",
        "/api/login",
        "/api/version",
    }


@app.before_request
def require_signed_api_token():
    if request.method == "OPTIONS":
        return None
    if not request.path.startswith("/api/"):
        return None
    if is_public_api_path(request.path):
        return None

    token = extract_bearer_token()
    if not token:
        return jsonify({"success": False, "error": "请先登录。"}), 401

    try:
        user, payload = verify_auth_token(token)
    except PermissionError as exc:
        return jsonify({"success": False, "error": str(exc)}), 401

    g.current_user = user
    g.auth_payload = payload
    touch_online_user_presence(user)
    current_user_id = normalize_identity_for_compare(user["id"])
    for key, value in iter_request_identity_values():
        requested_user_id = normalize_identity_for_compare(value)
        if requested_user_id and requested_user_id != current_user_id:
            return (
                jsonify(
                    {
                        "success": False,
                        "error": f"请求中的{key}与当前登录账号不一致，已拒绝操作。",
                    }
                ),
                403,
            )

    return None


@app.before_request
def ensure_backup_scheduler_started():
    start_backup_scheduler_once()


def get_checklist_field_meta(cur, inspection_table_id):
    rows = get_management_checklist_fields(cur, inspection_table_id, include_public=True)
    return [(row["field_key"], row["field_label"]) for row in rows]


def build_standard_detail_text(field_meta, row):
    lines = []
    for field_key, field_label in field_meta:
        value = row.get(field_key)
        if value is None:
            continue
        value_text = str(value).strip()
        if not value_text:
            continue
        lines.append(f"{field_label}：{value_text}")
    return "\n".join(lines)


def normalize_boolean_flag(value, default=True):
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().lower() not in {"false", "0", "no", "否"}
    return bool(value)


def build_register_display_field_meta(fields):
    return [
        (field["field_key"], field["field_label"])
        for field in fields
        if normalize_boolean_flag(field.get("is_register_visible"), True)
    ]


def normalize_checklist_code(value):
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9_]+", "_", text).strip("_")
    text = re.sub(r"_+", "_", text)
    if not text:
        raise ValueError("请填写检查表编码。")
    if not re.match(r"^[a-z][a-z0-9_]{1,41}$", text):
        raise ValueError("检查表编码只能使用小写英文字母、数字和下划线，且必须以字母开头，长度 2-42 位。")
    return text


def normalize_checklist_mode(value):
    text = str(value or "online").strip().lower()
    if text not in {"online", "offline"}:
        raise ValueError("检查表模式只能选择线上或线下。")
    return text


def normalize_checklist_field_key(value):
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9_]+", "_", text).strip("_")
    text = re.sub(r"_+", "_", text)
    if not text:
        raise ValueError("字段系统标识缺失，请重新保存。")
    if not re.match(rf"^[a-z][a-z0-9_]{{1,{CHECKLIST_FIELD_KEY_MAX_LENGTH - 1}}}$", text):
        raise ValueError(
            f"字段系统标识只能使用小写英文字母、数字和下划线，且必须以字母开头，长度 2-{CHECKLIST_FIELD_KEY_MAX_LENGTH} 位。"
        )
    if text in RESERVED_CHECKLIST_FIELD_KEYS:
        raise ValueError(f"字段系统标识 {text} 为系统保留字段，不能使用。")
    return text


def generate_checklist_field_key(table_code, sort_order, existing_keys=None):
    normalized_code = normalize_checklist_code(table_code)
    occupied_keys = set(existing_keys or [])
    order_text = str(max(int(sort_order or 1), 1)).zfill(3)
    for _ in range(20):
        candidate = f"f_{normalized_code}_{order_text}_{uuid.uuid4().hex[:6]}"
        if len(candidate) > CHECKLIST_FIELD_KEY_MAX_LENGTH:
            raise ValueError("检查表编码过长，无法生成字段系统标识。")
        if candidate not in occupied_keys:
            return candidate
    raise ValueError("字段系统标识自动生成失败，请重试。")


def normalize_checklist_field_rows(fields, table_code=None, allow_empty=False):
    if not isinstance(fields, list) or not fields:
        if allow_empty:
            return []
        raise ValueError("请至少配置一个检查表字段。")

    normalized = []
    seen_keys = set()
    seen_labels = set()
    for index, field in enumerate(fields, start=1):
        raw_field_key = field.get("field_key") if isinstance(field, dict) else ""
        if raw_field_key:
            field_key = normalize_checklist_field_key(raw_field_key)
        elif table_code:
            field_key = generate_checklist_field_key(table_code, index, seen_keys)
        else:
            raise ValueError("字段系统标识自动生成失败，请重新进入页面后再试。")
        field_label = normalize_text(field.get("field_label") if isinstance(field, dict) else "", 80)
        if not field_label:
            raise ValueError(f"第 {index} 个字段缺少字段名称。")
        if field_key in seen_keys:
            raise ValueError(f"字段系统标识 {field_key} 重复。")
        if field_label in seen_labels:
            raise ValueError(f"字段名称【{field_label}】重复。")
        seen_keys.add(field_key)
        seen_labels.add(field_label)
        normalized.append(
            {
                "field_key": field_key,
                "field_label": field_label,
                "is_filterable": normalize_boolean_flag(field.get("is_filterable"), True) if isinstance(field, dict) else True,
                "is_register_visible": normalize_boolean_flag(field.get("is_register_visible"), True) if isinstance(field, dict) else True,
                "is_long_text": normalize_boolean_flag(field.get("is_long_text"), False) if isinstance(field, dict) else False,
                "is_scorable": normalize_boolean_flag(field.get("is_scorable"), False) if isinstance(field, dict) else False,
                "sort_order": index,
            }
        )

    return normalized


def get_physical_table_name_by_code(table_code):
    try:
        table_code = normalize_checklist_code(table_code)
    except ValueError:
        return None
    return f"{CHECKLIST_PHYSICAL_TABLE_PREFIX}{table_code}"


def get_default_checklist_standard_id_base(inspection_table_id):
    table_id = int(inspection_table_id or 0)
    if table_id <= 0:
        raise ValueError("检查表编号异常，无法生成规范ID号段。")
    return table_id * CHECKLIST_STANDARD_ID_BLOCK_SIZE


def normalize_checklist_standard_id_base(value, fallback_table_id=None):
    if value in (None, ""):
        if fallback_table_id is None:
            raise ValueError("检查表规范ID号段缺失。")
        value = get_default_checklist_standard_id_base(fallback_table_id)
    try:
        standard_id_base = int(value)
    except (TypeError, ValueError):
        raise ValueError("检查表规范ID号段必须是整数。")
    if (
        standard_id_base < CHECKLIST_STANDARD_ID_BLOCK_SIZE
        or standard_id_base % CHECKLIST_STANDARD_ID_BLOCK_SIZE != 0
    ):
        raise ValueError("检查表规范ID号段必须从 1000、2000、3000 这样的整千数开始。")
    return standard_id_base


def get_checklist_standard_id_bounds(standard_id_base):
    start = normalize_checklist_standard_id_base(standard_id_base)
    return start, start + CHECKLIST_STANDARD_ID_BLOCK_SIZE - 1


def is_standard_id_in_checklist_range(standard_id, standard_id_base):
    start, end = get_checklist_standard_id_bounds(standard_id_base)
    try:
        value = int(standard_id)
    except (TypeError, ValueError):
        return False
    return start <= value <= end


def migrate_checklist_standard_ids_to_base(cur, checklist_row):
    physical_table_name = get_physical_table_name_by_code(checklist_row["table_code"])
    if not physical_table_name or not checklist_physical_table_exists(cur, physical_table_name):
        return

    standard_id_base = normalize_checklist_standard_id_base(
        checklist_row.get("standard_id_base"),
        checklist_row["id"],
    )
    range_start, range_end = get_checklist_standard_id_bounds(standard_id_base)

    cur.execute(
        sql.SQL("SELECT id, standard_id FROM {} ORDER BY standard_id ASC, id ASC;").format(
            sql.Identifier(physical_table_name)
        )
    )
    rows = cur.fetchall()
    if not rows:
        return
    if len(rows) > CHECKLIST_STANDARD_ID_BLOCK_SIZE:
        raise ValueError(
            f"检查表【{checklist_row['table_name']}】规范数量超过 {CHECKLIST_STANDARD_ID_BLOCK_SIZE} 条，"
            "无法按当前规范ID号段规则迁移。"
        )

    if all(range_start <= int(row["standard_id"]) <= range_end for row in rows):
        return

    mappings = []
    for index, row in enumerate(rows):
        new_standard_id = range_start + index
        mappings.append(
            {
                "row_id": row["id"],
                "old_standard_id": int(row["standard_id"]),
                "new_standard_id": new_standard_id,
                "temp_standard_id": -(range_start + index + 1000000000),
            }
        )

    for item in mappings:
        cur.execute(
            sql.SQL("UPDATE {} SET standard_id = %s WHERE id = %s;").format(
                sql.Identifier(physical_table_name)
            ),
            (item["temp_standard_id"], item["row_id"]),
        )

    cur.execute("SELECT to_regclass('public.issues') AS table_name;")
    issues_table_exists = bool(cur.fetchone().get("table_name"))
    if issues_table_exists:
        for item in mappings:
            cur.execute(
                """
                UPDATE issues
                SET standard_id = %s
                WHERE inspection_table_id = %s
                  AND standard_id = %s;
                """,
                (item["temp_standard_id"], checklist_row["id"], item["old_standard_id"]),
            )

    for item in mappings:
        cur.execute(
            sql.SQL("UPDATE {} SET standard_id = %s WHERE id = %s;").format(
                sql.Identifier(physical_table_name)
            ),
            (item["new_standard_id"], item["row_id"]),
        )

    if issues_table_exists:
        for item in mappings:
            cur.execute(
                """
                UPDATE issues
                SET standard_id = %s
                WHERE inspection_table_id = %s
                  AND standard_id = %s;
                """,
                (item["new_standard_id"], checklist_row["id"], item["temp_standard_id"]),
            )


def ensure_checklist_standard_id_bases(cur):
    cur.execute(
        """
        ALTER TABLE inspection_tables
        ADD COLUMN IF NOT EXISTS standard_id_base INTEGER;
        """
    )
    cur.execute(
        """
        UPDATE inspection_tables
        SET standard_id_base = id * %s
        WHERE standard_id_base IS NULL;
        """,
        (CHECKLIST_STANDARD_ID_BLOCK_SIZE,),
    )
    cur.execute(
        """
        CREATE UNIQUE INDEX IF NOT EXISTS idx_inspection_tables_standard_id_base_unique
        ON inspection_tables (standard_id_base);
        """
    )
    cur.execute(
        """
        DO $$
        BEGIN
            IF NOT EXISTS (
                SELECT 1
                FROM pg_constraint
                WHERE conname = 'inspection_tables_standard_id_base_check'
            ) THEN
                ALTER TABLE inspection_tables
                ADD CONSTRAINT inspection_tables_standard_id_base_check
                CHECK (
                    standard_id_base >= 1000
                    AND standard_id_base % 1000 = 0
                );
            END IF;
        END $$;
        """
    )
    cur.execute(
        """
        SELECT id, table_code, table_name, standard_id_base
        FROM inspection_tables
        ORDER BY id ASC;
        """
    )
    for row in cur.fetchall():
        migrate_checklist_standard_ids_to_base(cur, row)


def ensure_inspection_checklist_management_schema(cur):
    global INSPECTION_CHECKLIST_MANAGEMENT_SCHEMA_READY
    if INSPECTION_CHECKLIST_MANAGEMENT_SCHEMA_READY:
        return

    acquire_schema_migration_lock(cur)
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS inspection_tables (
            id SERIAL PRIMARY KEY,
            table_code TEXT UNIQUE NOT NULL,
            table_name TEXT UNIQUE NOT NULL,
            checklist_mode TEXT NOT NULL DEFAULT 'online',
            standard_id_base INTEGER UNIQUE,
            description TEXT,
            is_active BOOLEAN DEFAULT TRUE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """
    )
    cur.execute(
        """
        ALTER TABLE inspection_tables
        ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP;
        """
    )
    cur.execute(
        """
        ALTER TABLE inspection_tables
        ADD COLUMN IF NOT EXISTS checklist_mode TEXT NOT NULL DEFAULT 'online';
        """
    )
    ensure_checklist_standard_id_bases(cur)
    cur.execute(
        """
        UPDATE inspection_tables
        SET checklist_mode = 'online'
        WHERE checklist_mode IS NULL OR checklist_mode NOT IN ('online', 'offline');
        """
    )
    cur.execute(
        """
        DO $$
        BEGIN
            IF NOT EXISTS (
                SELECT 1
                FROM pg_constraint
                WHERE conname = 'inspection_tables_checklist_mode_check'
            ) THEN
                ALTER TABLE inspection_tables
                ADD CONSTRAINT inspection_tables_checklist_mode_check
                CHECK (checklist_mode IN ('online', 'offline'));
            END IF;
        END $$;
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS inspection_table_fields (
            id SERIAL PRIMARY KEY,
            inspection_table_id INTEGER NOT NULL REFERENCES inspection_tables(id) ON DELETE CASCADE,
            field_key TEXT NOT NULL,
            field_label TEXT NOT NULL,
            is_filterable BOOLEAN DEFAULT TRUE,
            is_register_visible BOOLEAN DEFAULT TRUE,
            is_long_text BOOLEAN DEFAULT FALSE,
            is_scorable BOOLEAN DEFAULT FALSE,
            sort_order INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS inspection_standard_export_templates (
            id INTEGER PRIMARY KEY DEFAULT 1,
            template_config JSONB NOT NULL DEFAULT '{}'::jsonb,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            CONSTRAINT inspection_standard_export_templates_singleton CHECK (id = 1)
        );
        """
    )
    cur.execute(
        """
        ALTER TABLE inspection_table_fields
        ADD COLUMN IF NOT EXISTS is_filterable BOOLEAN DEFAULT TRUE;
        """
    )
    cur.execute(
        """
        ALTER TABLE inspection_table_fields
        ADD COLUMN IF NOT EXISTS is_register_visible BOOLEAN DEFAULT TRUE;
        """
    )
    cur.execute(
        """
        ALTER TABLE inspection_table_fields
        ADD COLUMN IF NOT EXISTS is_long_text BOOLEAN DEFAULT FALSE;
        """
    )
    cur.execute(
        """
        ALTER TABLE inspection_table_fields
        ADD COLUMN IF NOT EXISTS is_scorable BOOLEAN DEFAULT FALSE;
        """
    )
    cur.execute(
        """
        CREATE UNIQUE INDEX IF NOT EXISTS idx_inspection_table_fields_table_key_unique
        ON inspection_table_fields (inspection_table_id, field_key);
        """
    )
    cur.execute(
        """
        DO $$
        BEGIN
            IF NOT EXISTS (
                SELECT 1
                FROM pg_indexes
                WHERE schemaname = 'public'
                  AND indexname = 'idx_inspection_table_fields_key_unique'
            )
            AND NOT EXISTS (
                SELECT 1
                FROM inspection_table_fields
                GROUP BY field_key
                HAVING COUNT(*) > 1
            ) THEN
                CREATE UNIQUE INDEX idx_inspection_table_fields_key_unique
                ON inspection_table_fields (field_key);
            END IF;
        END $$;
        """
    )
    INSPECTION_CHECKLIST_MANAGEMENT_SCHEMA_READY = True


def ensure_checklist_physical_table(cur, physical_table_name):
    cur.execute(
        sql.SQL(
            """
            CREATE TABLE IF NOT EXISTS {} (
                id SERIAL PRIMARY KEY,
                standard_id BIGINT UNIQUE NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            """
        ).format(sql.Identifier(physical_table_name))
    )


def ensure_checklist_field_columns(cur, physical_table_name, fields):
    ensure_checklist_physical_table(cur, physical_table_name)
    for field in fields:
        cur.execute(
            sql.SQL("ALTER TABLE {} ADD COLUMN IF NOT EXISTS {} TEXT;").format(
                sql.Identifier(physical_table_name),
                sql.Identifier(field["field_key"]),
            )
        )


def checklist_physical_table_exists(cur, physical_table_name):
    cur.execute("SELECT to_regclass(%s) AS table_name;", (f"public.{physical_table_name}",))
    row = cur.fetchone()
    return bool(row and row.get("table_name"))


def get_checklist_row_count(cur, physical_table_name):
    if not physical_table_name or not checklist_physical_table_exists(cur, physical_table_name):
        return 0
    cur.execute(
        sql.SQL("SELECT COUNT(*) AS total FROM {};").format(sql.Identifier(physical_table_name))
    )
    row = cur.fetchone()
    return int(row["total"] or 0)


def get_management_checklist_fields(cur, inspection_table_id, include_public=False):
    def column_exists(column_name):
        cur.execute(
            """
            SELECT EXISTS (
                SELECT 1
                FROM information_schema.columns
                WHERE table_schema = 'public'
                  AND table_name = 'inspection_table_fields'
                  AND column_name = %s
            ) AS has_column;
            """,
            (column_name,),
        )
        return bool(cur.fetchone().get("has_column"))

    has_register_visible_column = column_exists("is_register_visible")
    has_long_text_column = column_exists("is_long_text")
    has_scorable_column = column_exists("is_scorable")
    register_visible_select = "is_register_visible" if has_register_visible_column else "TRUE AS is_register_visible"
    long_text_select = "is_long_text" if has_long_text_column else "FALSE AS is_long_text"
    scorable_select = "is_scorable" if has_scorable_column else "FALSE AS is_scorable"
    cur.execute(
        """
        SELECT
            id,
            inspection_table_id,
            field_key,
            field_label,
            is_filterable,
            {register_visible_select},
            {long_text_select},
            {scorable_select},
            sort_order,
            FALSE AS is_public
        FROM inspection_table_fields
        WHERE inspection_table_id = %s
        ORDER BY sort_order ASC, id ASC;
        """.format(
            register_visible_select=register_visible_select,
            long_text_select=long_text_select,
            scorable_select=scorable_select,
        ),
        (inspection_table_id,),
    )
    return cur.fetchall()


def normalize_standard_export_template_config(cur, raw_config):
    config = raw_config if isinstance(raw_config, dict) else {}
    cur.execute(
        """
        SELECT id
        FROM inspection_tables
        WHERE is_active = TRUE
        ORDER BY id ASC;
        """
    )
    normalized = {}
    for table in cur.fetchall():
        table_key = str(table["id"])
        allowed_keys = {
            str(field["field_key"])
            for field in get_management_checklist_fields(cur, table["id"], include_public=True)
        }
        incoming_keys = config.get(table_key, config.get(table["id"], []))
        if not isinstance(incoming_keys, list):
            incoming_keys = []
        normalized[table_key] = []
        seen = set()
        for field_key in incoming_keys:
            text = str(field_key or "").strip()
            if not text or text in seen or text not in allowed_keys:
                continue
            normalized[table_key].append(text)
            seen.add(text)
    return normalized


def get_standard_export_template_config(cur):
    cur.execute(
        """
        SELECT template_config
        FROM inspection_standard_export_templates
        WHERE id = 1;
        """
    )
    row = cur.fetchone()
    if not row:
        return {"has_saved": False, "tables": {}}
    return {
        "has_saved": True,
        "tables": normalize_standard_export_template_config(
            cur, row.get("template_config") or {}
        ),
    }


def save_standard_export_template_config(cur, template_config):
    normalized = normalize_standard_export_template_config(cur, template_config)
    cur.execute(
        """
        INSERT INTO inspection_standard_export_templates (
            id,
            template_config,
            updated_at
        )
        VALUES (1, %s::jsonb, CURRENT_TIMESTAMP)
        ON CONFLICT (id) DO UPDATE
        SET template_config = EXCLUDED.template_config,
            updated_at = CURRENT_TIMESTAMP;
        """,
        (json.dumps(normalized, ensure_ascii=False),),
    )
    return normalized


def upsert_checklist_fields(cur, inspection_table_id, fields):
    incoming_keys = [field["field_key"] for field in fields]
    cur.execute(
        """
        DELETE FROM inspection_table_fields
        WHERE inspection_table_id = %s
          AND field_key <> ALL(%s::text[]);
        """,
        (inspection_table_id, incoming_keys),
    )
    for field in fields:
        cur.execute(
            """
            INSERT INTO inspection_table_fields (
                inspection_table_id,
                field_key,
                field_label,
                is_filterable,
                is_register_visible,
                is_long_text,
                is_scorable,
                sort_order
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (inspection_table_id, field_key)
            DO UPDATE SET
                field_label = EXCLUDED.field_label,
                is_filterable = EXCLUDED.is_filterable,
                is_register_visible = EXCLUDED.is_register_visible,
                is_long_text = EXCLUDED.is_long_text,
                is_scorable = EXCLUDED.is_scorable,
                sort_order = EXCLUDED.sort_order;
            """,
            (
                inspection_table_id,
                field["field_key"],
                field["field_label"],
                field["is_filterable"],
                field["is_register_visible"],
                field.get("is_long_text", False),
                field.get("is_scorable", False),
                field["sort_order"],
            ),
        )


def ensure_unique_checklist_field_keys(cur, fields, inspection_table_id=None):
    incoming_keys = [field["field_key"] for field in fields]
    if not incoming_keys:
        return
    cur.execute(
        """
        SELECT field_key, inspection_table_id
        FROM inspection_table_fields
        WHERE field_key = ANY(%s::text[])
          AND (%s::integer IS NULL OR inspection_table_id <> %s::integer)
        LIMIT 1;
        """,
        (incoming_keys, inspection_table_id, inspection_table_id),
    )
    conflict = cur.fetchone()
    if conflict:
        raise ValueError("字段系统标识已被其他检查表使用，请重新保存生成新的字段键。")


def serialize_management_checklist(cur, row):
    physical_table_name = get_physical_table_name_by_code(row["table_code"])
    local_fields = [dict(field) for field in get_management_checklist_fields(cur, row["id"])]
    return {
        "id": row["id"],
        "table_code": row["table_code"],
        "table_name": row["table_name"],
        "checklist_mode": normalize_checklist_mode(row.get("checklist_mode")),
        "standard_id_base": normalize_checklist_standard_id_base(
            row.get("standard_id_base"),
            row["id"],
        ),
        "description": row["description"],
        "is_active": row["is_active"],
        "created_at": row.get("created_at"),
        "updated_at": row.get("updated_at"),
        "physical_table_name": physical_table_name,
        "fields": local_fields,
        "local_fields": local_fields,
        "standard_count": get_checklist_row_count(cur, physical_table_name),
    }


def fetch_management_checklist(cur, inspection_table_id):
    cur.execute(
        """
        SELECT
            id,
            table_code,
            table_name,
            checklist_mode,
            standard_id_base,
            description,
            is_active,
            TO_CHAR(created_at, 'YYYY-MM-DD HH24:MI') AS created_at,
            TO_CHAR(updated_at, 'YYYY-MM-DD HH24:MI') AS updated_at
        FROM inspection_tables
        WHERE id = %s
        LIMIT 1;
        """,
        (inspection_table_id,),
    )
    return cur.fetchone()


def normalize_import_standard_id(value):
    text = str(value or "").strip()
    if not text:
        raise ValueError("请填写外部规范ID。")
    if re.match(r"^\d+\.0$", text):
        text = text[:-2]
    if not re.match(r"^-?\d+$", text):
        raise ValueError(f"外部规范ID【{text}】不是有效整数。")
    return int(text)


def get_import_value(row, field):
    for key in (field["field_key"], field["field_label"]):
        if key in row:
            return row.get(key)
    return ""


def normalize_checklist_import_rows(raw_rows, fields, standard_id_base=None, allow_reassign=False):
    if not raw_rows:
        raise ValueError("导入文件没有可导入的数据行。")
    range_start = range_end = None
    if standard_id_base is not None:
        range_start, range_end = get_checklist_standard_id_bounds(standard_id_base)
    if len(raw_rows) > CHECKLIST_STANDARD_ID_BLOCK_SIZE:
        raise ValueError(f"单张检查表最多导入 {CHECKLIST_STANDARD_ID_BLOCK_SIZE} 条规范。")

    normalized = []
    seen_ids = set()
    for index, row in enumerate(raw_rows):
        raw_standard_id = (
            row.get("standard_id")
            or row.get("规范ID")
            or row.get("标准ID")
            or row.get("id")
        )
        if raw_standard_id in (None, "") and standard_id_base is not None:
            standard_id = range_start + index
        else:
            standard_id = normalize_import_standard_id(raw_standard_id)
        if standard_id_base is not None and not (range_start <= standard_id <= range_end):
            if not allow_reassign:
                raise ValueError(
                    f"外部规范ID【{standard_id}】不在当前检查表号段 {range_start}-{range_end} 内。"
                )
            standard_id = range_start + index
        if standard_id in seen_ids:
            raise ValueError(f"导入文件内外部规范ID【{standard_id}】重复。")
        seen_ids.add(standard_id)
        item = {"standard_id": standard_id}
        for field in fields:
            item[field["field_key"]] = str(get_import_value(row, field) or "").strip()
        normalized.append(item)
    return normalized


def rebase_checklist_standard_rows(rows, source_base, target_base):
    source_start, source_end = get_checklist_standard_id_bounds(source_base)
    target_start, target_end = get_checklist_standard_id_bounds(target_base)
    if source_start == target_start:
        return [dict(row) for row in rows]

    rebased_rows = []
    for index, row in enumerate(sorted(rows, key=lambda item: int(item.get("standard_id") or 0))):
        old_standard_id = int(row.get("standard_id") or 0)
        if source_start <= old_standard_id <= source_end:
            offset = old_standard_id - source_start
        else:
            offset = index
        new_standard_id = target_start + offset
        if new_standard_id > target_end:
            raise ValueError(
                f"检查表规范数量超过号段 {target_start}-{target_end} 可容纳范围，无法导入。"
            )
        item = dict(row)
        item["standard_id"] = new_standard_id
        rebased_rows.append(item)
    return rebased_rows


def fetch_checklist_standard_rows(cur, physical_table_name, fields):
    if not checklist_physical_table_exists(cur, physical_table_name):
        return []
    columns = ["standard_id", *[field["field_key"] for field in fields]]
    cur.execute(
        sql.SQL("SELECT {} FROM {} ORDER BY standard_id ASC;").format(
            sql.SQL(", ").join(sql.Identifier(column) for column in columns),
            sql.Identifier(physical_table_name),
        )
    )
    return [dict(row) for row in cur.fetchall()]


def insert_checklist_standard_rows(cur, physical_table_name, fields, rows):
    columns = ["standard_id", *[field["field_key"] for field in fields]]
    update_columns = [field["field_key"] for field in fields]
    if not update_columns:
        return
    insert_sql = sql.SQL(
        """
        INSERT INTO {} ({})
        VALUES ({})
        ON CONFLICT (standard_id)
        DO UPDATE SET {};
        """
    ).format(
        sql.Identifier(physical_table_name),
        sql.SQL(", ").join(sql.Identifier(column) for column in columns),
        sql.SQL(", ").join(sql.Placeholder() for _ in columns),
        sql.SQL(", ").join(
            sql.SQL("{} = EXCLUDED.{}").format(
                sql.Identifier(column),
                sql.Identifier(column),
            )
            for column in update_columns
        ),
    )
    for row in rows:
        cur.execute(insert_sql, [row.get(column) for column in columns])


def normalize_checklist_standard_payload(data, fields):
    if not isinstance(data, dict):
        raise ValueError("规范数据格式不正确。")

    values = data.get("values") if isinstance(data.get("values"), dict) else data
    row = {}
    for field in fields:
        row[field["field_key"]] = normalize_text(values.get(field["field_key"], ""), 3000)
    if not any(value for value in row.values()):
        raise ValueError("请至少填写一项规范内容。")
    return row


def get_next_checklist_standard_id(cur, checklist, physical_table_name):
    cur.execute(
        """
        SELECT id, standard_id_base
        FROM inspection_tables
        WHERE id = %s
        FOR UPDATE;
        """,
        (checklist["id"],),
    )
    locked_checklist = cur.fetchone()
    if not locked_checklist:
        raise LookupError("检查表不存在。")

    standard_id_base = normalize_checklist_standard_id_base(
        locked_checklist.get("standard_id_base"),
        locked_checklist["id"],
    )
    range_start, range_end = get_checklist_standard_id_bounds(standard_id_base)

    if not checklist_physical_table_exists(cur, physical_table_name):
        return range_start

    cur.execute(
        sql.SQL("LOCK TABLE {} IN SHARE ROW EXCLUSIVE MODE;").format(
            sql.Identifier(physical_table_name)
        )
    )
    cur.execute(
        sql.SQL(
            """
            SELECT COALESCE(MAX(standard_id), %s - 1) AS max_standard_id
            FROM {}
            WHERE standard_id BETWEEN %s AND %s;
            """
        ).format(sql.Identifier(physical_table_name)),
        (range_start, range_start, range_end),
    )
    row = cur.fetchone()
    next_standard_id = int(row["max_standard_id"] or (range_start - 1)) + 1
    if next_standard_id > range_end:
        raise ValueError(
            f"该检查表外部规范ID号段 {range_start}-{range_end} 已用完，"
            f"单张检查表最多维护 {CHECKLIST_STANDARD_ID_BLOCK_SIZE} 条规范。"
        )
    return next_standard_id


def get_checklist_standard_reference_counts(cur, inspection_table_id, standard_id):
    cur.execute(
        """
        SELECT COUNT(*) AS total
        FROM issues
        WHERE inspection_table_id = %s
          AND standard_id = %s;
        """,
        (inspection_table_id, standard_id),
    )
    issue_count = int(cur.fetchone()["total"] or 0)
    return {"issue_count": issue_count}


def format_checklist_standard_reference_message(counts):
    parts = []
    if counts.get("issue_count", 0) > 0:
        parts.append(f"巡检问题 {counts['issue_count']} 条")
    return "、".join(parts)


def sync_referenced_standard_detail_text(cur, inspection_table_id, standard_id, detail_text):
    cur.execute(
        """
        UPDATE issues
        SET standard_detail_text = %s
        WHERE inspection_table_id = %s
          AND standard_id = %s;
        """,
        (detail_text, inspection_table_id, standard_id),
    )
    return cur.rowcount


def sync_referenced_internal_standard_detail_text(cur, internal_standard_id, detail_text):
    cur.execute(
        """
        UPDATE issues
        SET internal_standard_detail_text = %s
        WHERE UPPER(COALESCE(internal_standard_id, '')) = %s;
        """,
        (detail_text, str(internal_standard_id or "").upper()),
    )
    return cur.rowcount


def get_management_checklist_standard_context(cur, inspection_table_id, require_fields=True):
    checklist = fetch_management_checklist(cur, inspection_table_id)
    if not checklist:
        raise LookupError("检查表不存在。")

    fields = [dict(field) for field in get_management_checklist_fields(cur, inspection_table_id, include_public=True)]
    if require_fields and not fields:
        raise ValueError("请先配置检查表字段，再维护规范数据。")

    physical_table_name = get_physical_table_name_by_code(checklist["table_code"])
    if not physical_table_name:
        raise ValueError("检查表物理表名生成失败。")
    ensure_checklist_field_columns(cur, physical_table_name, fields)
    return checklist, fields, physical_table_name


def build_standard_search_sql(fields, keyword):
    keyword = normalize_text(keyword, 120)
    if not keyword:
        return sql.SQL(""), []

    search_columns = [
        sql.SQL("standard_id::text"),
        *[
            sql.SQL("COALESCE({}, '')").format(sql.Identifier(field["field_key"]))
            for field in fields
        ],
    ]
    return (
        sql.SQL(" WHERE CONCAT_WS(' ', {}) ILIKE %s").format(
            sql.SQL(", ").join(search_columns)
        ),
        [f"%{keyword}%"],
    )


def normalize_page_args(page_value, page_size_value):
    try:
        page = max(int(page_value or 1), 1)
    except (TypeError, ValueError):
        page = 1
    try:
        page_size = int(page_size_value or 10)
    except (TypeError, ValueError):
        page_size = 10
    page_size = min(max(page_size, 5), 100)
    return page, page_size


def normalize_checklist_backup_payload(payload):
    if not isinstance(payload, dict):
        raise ValueError("备份文件格式不正确：根内容必须是对象。")
    if payload.get("backup_type") != "ywddzx_checklists":
        raise ValueError("备份文件类型不匹配，请选择巡检表数据备份文件。")
    checklists = payload.get("checklists")
    if not isinstance(checklists, list):
        raise ValueError("备份文件缺少 checklists 数组。")

    normalized = []
    seen_codes = set()
    seen_names = set()
    seen_field_keys = set()
    seen_standard_id_bases = set()
    for index, item in enumerate(checklists, start=1):
        if not isinstance(item, dict):
            raise ValueError(f"第 {index} 张检查表数据格式不正确。")
        table_code = normalize_checklist_code(item.get("table_code"))
        table_name = sanitize_display_string(normalize_text(item.get("table_name"), 120))
        checklist_mode = normalize_checklist_mode(item.get("checklist_mode"))
        standard_id_base = normalize_checklist_standard_id_base(
            item.get("standard_id_base"),
            index,
        )
        if not table_name:
            raise ValueError(f"第 {index} 张检查表缺少名称。")
        if table_code in seen_codes:
            raise ValueError(f"备份文件内检查表编码 {table_code} 重复。")
        if table_name in seen_names:
            raise ValueError(f"备份文件内检查表名称【{table_name}】重复。")
        if standard_id_base in seen_standard_id_bases:
            raise ValueError(f"备份文件内外部规范ID号段【{standard_id_base}】重复。")
        seen_codes.add(table_code)
        seen_names.add(table_name)
        seen_standard_id_bases.add(standard_id_base)
        fields = normalize_checklist_field_rows(item.get("fields"), table_code)
        for field in fields:
            field_key = field["field_key"]
            if field_key in seen_field_keys:
                raise ValueError(f"备份文件内字段系统标识 {field_key} 重复。")
            seen_field_keys.add(field_key)
        standards = (
            normalize_checklist_import_rows(
                item.get("standards") or [],
                fields,
                standard_id_base,
                allow_reassign=True,
            )
            if item.get("standards")
            else []
        )
        normalized.append(
            {
                "table_code": table_code,
                "table_name": table_name,
                "checklist_mode": checklist_mode,
                "standard_id_base": standard_id_base,
                "description": normalize_text(item.get("description"), 300),
                "is_active": bool(item.get("is_active", True)),
                "fields": fields,
                "standards": standards,
            }
        )
    return {
        "checklists": normalized,
    }


def parse_checklist_backup_file(file_storage):
    if not file_storage or not file_storage.filename:
        raise ValueError("请选择需要导入的巡检表备份文件。")
    if not file_storage.filename.lower().endswith(".json"):
        raise ValueError("巡检表备份文件只能导入 JSON 格式。")
    file_bytes = file_storage.read()
    if not file_bytes:
        raise ValueError("备份文件内容为空。")
    try:
        payload = json.loads(file_bytes.decode("utf-8-sig"))
    except Exception as exc:
        raise ValueError("备份文件 JSON 解析失败，请检查文件是否损坏。") from exc
    return normalize_checklist_backup_payload(payload)


def get_blocking_checklist_references(cur, inspection_table_ids):
    ids = [int(value) for value in inspection_table_ids if value]
    if not ids:
        return []
    reference_tables = [
        ("inspections", "巡检记录"),
        ("issues", "巡检问题"),
        ("inspection_plan_configs", "巡检计划"),
        ("inspection_table_original_files", "检查表原件"),
    ]
    blockers = []
    for table_name, label in reference_tables:
        cur.execute("SELECT to_regclass(%s) AS table_name;", (f"public.{table_name}",))
        if not cur.fetchone().get("table_name"):
            continue
        cur.execute(
            sql.SQL("SELECT COUNT(*) AS total FROM {} WHERE inspection_table_id = ANY(%s);").format(
                sql.Identifier(table_name)
            ),
            (ids,),
        )
        total = int(cur.fetchone()["total"] or 0)
        if total:
            blockers.append(f"{label} {total} 条")
    return blockers


def get_inspection_table_record(cur, inspection_table_id):
    cur.execute(
        """
        SELECT
            id,
            table_code,
            table_name,
            standard_id_base,
            description,
            is_active
        FROM inspection_tables
        WHERE id = %s
        LIMIT 1;
        """,
        (inspection_table_id,),
    )
    return cur.fetchone()


def fetch_standard_from_table(cur, physical_table_name, standard_id):
    cur.execute(
        sql.SQL("SELECT * FROM {} WHERE standard_id = %s LIMIT 1;").format(
            sql.Identifier(physical_table_name)
        ),
        (standard_id,),
    )
    return cur.fetchone()


def ensure_issue_inspector_schema(cur):
    global ISSUE_INSPECTOR_SCHEMA_READY
    if ISSUE_INSPECTOR_SCHEMA_READY:
        return
    acquire_schema_migration_lock(cur)
    cur.execute("SELECT to_regclass('public.issues') AS table_name;")
    if not cur.fetchone().get("table_name"):
        return

    cur.execute(
        """
        ALTER TABLE issues
        ADD COLUMN IF NOT EXISTS inspector_id INTEGER REFERENCES users(id) ON DELETE RESTRICT;
        """
    )
    cur.execute(
        """
        ALTER TABLE issues
        ADD COLUMN IF NOT EXISTS internal_standard_id TEXT;
        """
    )
    cur.execute(
        """
        ALTER TABLE issues
        ADD COLUMN IF NOT EXISTS internal_standard_detail_text TEXT;
        """
    )
    cur.execute(
        """
        ALTER TABLE issues
        ADD COLUMN IF NOT EXISTS audit_status TEXT NOT NULL DEFAULT 'pending';
        """
    )
    cur.execute(
        """
        ALTER TABLE issues
        ADD COLUMN IF NOT EXISTS audited_by INTEGER REFERENCES users(id) ON DELETE SET NULL;
        """
    )
    cur.execute(
        """
        ALTER TABLE issues
        ADD COLUMN IF NOT EXISTS audited_at TIMESTAMP;
        """
    )
    cur.execute(
        """
        ALTER TABLE issues
        ADD COLUMN IF NOT EXISTS is_excellent BOOLEAN NOT NULL DEFAULT FALSE;
        """
    )
    cur.execute(
        """
        ALTER TABLE issues
        ADD COLUMN IF NOT EXISTS rectification_at TIMESTAMP;
        """
    )
    cur.execute(
        """
        ALTER TABLE issues
        ADD COLUMN IF NOT EXISTS review_at TIMESTAMP;
        """
    )
    cur.execute(
        """
        UPDATE issues
        SET is_excellent = FALSE
        WHERE is_excellent IS NULL
           OR COALESCE(audit_status, 'pending') = 'rejected';
        """
    )
    cur.execute("SELECT to_regclass('public.inspections') AS table_name;")
    if cur.fetchone().get("table_name"):
        cur.execute(
            """
            UPDATE issues i
            SET inspector_id = ins.inspector_id
            FROM inspections ins
            WHERE i.inspection_id = ins.id
              AND i.inspector_id IS NULL;
            """
        )
        cur.execute(
            """
            UPDATE issues i
            SET audit_status = 'approved',
                audited_at = COALESCE(ins.station_manager_signed_at, i.audited_at, i.created_at, CURRENT_TIMESTAMP)
            FROM inspections ins
            WHERE i.inspection_id = ins.id
              AND (
                  COALESCE(ins.sign_status, '') = '已签名确认'
                  OR ins.station_manager_signed_at IS NOT NULL
                  OR NULLIF(ins.station_manager_signature_path, '') IS NOT NULL
                  OR NULLIF(ins.station_manager_signed_name, '') IS NOT NULL
              )
              AND COALESCE(i.audit_status, 'pending') = 'pending';
            """
        )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_issues_inspector_id
        ON issues (inspector_id);
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_issues_internal_standard_id
        ON issues (internal_standard_id);
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_issues_audit_status
        ON issues (audit_status);
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_issues_is_excellent
        ON issues (is_excellent);
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_issues_station_status_audit
        ON issues (station_id, status, audit_status);
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_issues_created_at
        ON issues (created_at DESC);
        """
    )
    ISSUE_INSPECTOR_SCHEMA_READY = True


def create_inspection_record(
    cur, station_id, inspector_id, inspection_table_id, batch_id, inspection_date=None
):
    target_date = inspection_date or beijing_today()
    cur.execute(
        """
        INSERT INTO inspections (
            station_id,
            inspector_id,
            inspection_table_id,
            inspection_date,
            batch_id
        )
        VALUES (%s, %s, %s, %s, %s)
        RETURNING id;
        """,
        (station_id, inspector_id, inspection_table_id, target_date, batch_id),
    )
    row = cur.fetchone()
    return row["id"]


def ensure_inspection_completion_schema(cur):
    global INSPECTION_COMPLETION_SCHEMA_READY
    if INSPECTION_COMPLETION_SCHEMA_READY:
        return
    acquire_schema_migration_lock(cur)
    cur.execute(
        """
        ALTER TABLE inspections
        ADD COLUMN IF NOT EXISTS inspector_completion_status TEXT NOT NULL DEFAULT '待检查人确认';
        """
    )
    cur.execute(
        """
        ALTER TABLE inspections
        ADD COLUMN IF NOT EXISTS inspector_completed_by INTEGER REFERENCES users(id) ON DELETE SET NULL;
        """
    )
    cur.execute(
        """
        ALTER TABLE inspections
        ADD COLUMN IF NOT EXISTS inspector_completed_at TIMESTAMP;
        """
    )
    cur.execute(
        """
        ALTER TABLE inspections
        ADD COLUMN IF NOT EXISTS inspector_completion_source TEXT;
        """
    )
    cur.execute(
        """
        ALTER TABLE inspections
        ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP;
        """
    )
    cur.execute(
        """
        UPDATE inspections
        SET inspector_completion_status = '待检查人确认'
        WHERE inspector_completion_status IS NULL
           OR inspector_completion_status NOT IN ('待检查人确认', '已确认完成');
        """
    )
    cur.execute(
        """
        UPDATE inspections
        SET updated_at = COALESCE(updated_at, created_at, CURRENT_TIMESTAMP)
        WHERE updated_at IS NULL;
        """
    )
    cur.execute(
        """
        UPDATE inspections
        SET sign_status = '已签名确认',
            station_manager_signed_at = COALESCE(
                station_manager_signed_at,
                updated_at,
                created_at,
                CURRENT_TIMESTAMP
            ),
            updated_at = COALESCE(updated_at, CURRENT_TIMESTAMP)
        WHERE COALESCE(sign_status, '') <> '已签名确认'
          AND (
              station_manager_signed_at IS NOT NULL
              OR NULLIF(station_manager_signature_path, '') IS NOT NULL
              OR NULLIF(station_manager_signed_name, '') IS NOT NULL
          );
        """
    )
    cur.execute(
        """
        UPDATE inspections
        SET inspector_completion_status = '已确认完成',
            inspector_completed_at = COALESCE(
                inspector_completed_at,
                station_manager_signed_at,
                updated_at,
                created_at,
                CURRENT_TIMESTAMP
            ),
            inspector_completion_source = COALESCE(NULLIF(inspector_completion_source, ''), 'signature'),
            updated_at = COALESCE(updated_at, CURRENT_TIMESTAMP)
        WHERE sign_status = '已签名确认'
          AND inspector_completion_status <> '已确认完成';
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_inspections_station_table_month
        ON inspections (station_id, inspection_table_id, inspection_date);
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_inspections_station_date
        ON inspections (station_id, inspection_date DESC);
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_inspections_completion_status
        ON inspections (inspector_completion_status, inspection_date);
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS inspection_inspector_confirmations (
            id SERIAL PRIMARY KEY,
            inspection_id INTEGER NOT NULL REFERENCES inspections(id) ON DELETE CASCADE,
            inspector_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            confirmed_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
            source TEXT NOT NULL DEFAULT 'manual',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            CONSTRAINT uq_inspection_inspector_confirmations UNIQUE (inspection_id, inspector_id)
        );
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_inspection_inspector_confirmations_inspection
        ON inspection_inspector_confirmations (inspection_id);
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_inspection_inspector_confirmations_inspector
        ON inspection_inspector_confirmations (inspector_id);
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS inspection_completion_settings (
            singleton BOOLEAN PRIMARY KEY DEFAULT TRUE,
            auto_complete_enabled BOOLEAN NOT NULL DEFAULT TRUE,
            auto_complete_days INTEGER NOT NULL DEFAULT 7,
            record_uniqueness_period TEXT NOT NULL DEFAULT 'month',
            updated_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            CONSTRAINT chk_inspection_completion_singleton CHECK (singleton = TRUE),
            CONSTRAINT chk_inspection_completion_days CHECK (auto_complete_days BETWEEN 1 AND 31)
        );
        """
    )
    cur.execute(
        """
        ALTER TABLE inspection_completion_settings
        ADD COLUMN IF NOT EXISTS record_uniqueness_period TEXT NOT NULL DEFAULT 'month';
        """
    )
    cur.execute(
        """
        UPDATE inspection_completion_settings
        SET record_uniqueness_period = 'month'
        WHERE record_uniqueness_period IS NULL
           OR record_uniqueness_period NOT IN ('week', 'month', 'quarter', 'year');
        """
    )
    cur.execute(
        """
        INSERT INTO inspection_completion_settings (
            singleton,
            auto_complete_enabled,
            auto_complete_days,
            record_uniqueness_period
        )
        VALUES (TRUE, TRUE, %s, %s)
        ON CONFLICT (singleton) DO NOTHING;
        """,
        (DEFAULT_INSPECTION_AUTO_COMPLETE_DAYS, DEFAULT_INSPECTION_RECORD_UNIQUENESS_PERIOD),
    )
    connection = getattr(cur, "connection", None)
    if connection:
        connection.commit()
    INSPECTION_COMPLETION_SCHEMA_READY = True


def ensure_inspection_plan_assignment_schema(cur):
    global INSPECTION_PLAN_ASSIGNMENT_SCHEMA_READY
    if INSPECTION_PLAN_ASSIGNMENT_SCHEMA_READY:
        return

    acquire_schema_migration_lock(cur)
    cur.execute(
        """
        ALTER TABLE stations
        ADD COLUMN IF NOT EXISTS monitoring_status TEXT NOT NULL DEFAULT '运行中';
        """
    )
    cur.execute(
        """
        UPDATE stations
        SET monitoring_status = '运行中'
        WHERE monitoring_status IS NULL
           OR monitoring_status NOT IN ('运行中', '未运行');
        """
    )
    cur.execute(
        """
        DO $$
        BEGIN
            IF NOT EXISTS (
                SELECT 1
                FROM pg_constraint
                WHERE conname = 'stations_monitoring_status_check'
            ) THEN
                ALTER TABLE stations
                ADD CONSTRAINT stations_monitoring_status_check
                CHECK (monitoring_status IN ('运行中', '未运行'));
            END IF;
        END
        $$;
        """
    )
    cur.execute(
        """
        ALTER TABLE inspection_plan_station_items
        ADD COLUMN IF NOT EXISTS assigned_inspector_id INTEGER;
        """
    )
    cur.execute(
        """
        ALTER TABLE inspection_plan_station_items
        ADD COLUMN IF NOT EXISTS assigned_inspector_ids JSONB NOT NULL DEFAULT '[]'::jsonb;
        """
    )
    cur.execute(
        """
        UPDATE inspection_plan_station_items
        SET assigned_inspector_ids = CASE
                WHEN assigned_inspector_id IS NULL THEN '[]'::jsonb
                ELSE jsonb_build_array(assigned_inspector_id)
            END
        WHERE assigned_inspector_ids IS NULL
           OR assigned_inspector_ids = '[]'::jsonb
           OR jsonb_typeof(assigned_inspector_ids) <> 'array';
        """
    )
    cur.execute(
        """
        ALTER TABLE inspection_plan_station_items
        ADD COLUMN IF NOT EXISTS assigned_by INTEGER;
        """
    )
    cur.execute(
        """
        ALTER TABLE inspection_plan_station_items
        ADD COLUMN IF NOT EXISTS assigned_at TIMESTAMP;
        """
    )
    cur.execute(
        """
        DO $$
        BEGIN
            IF NOT EXISTS (
                SELECT 1
                FROM pg_constraint c
                JOIN pg_attribute a
                  ON a.attrelid = c.conrelid
                 AND a.attnum = ANY(c.conkey)
                WHERE c.contype = 'f'
                  AND c.conrelid = 'inspection_plan_station_items'::regclass
                  AND a.attname = 'assigned_inspector_id'
            ) THEN
                ALTER TABLE inspection_plan_station_items
                ADD CONSTRAINT fk_inspection_plan_items_assigned_inspector
                FOREIGN KEY (assigned_inspector_id) REFERENCES users(id) ON DELETE SET NULL;
            END IF;
        END
        $$;
        """
    )
    cur.execute(
        """
        DO $$
        BEGIN
            IF NOT EXISTS (
                SELECT 1
                FROM pg_constraint c
                JOIN pg_attribute a
                  ON a.attrelid = c.conrelid
                 AND a.attnum = ANY(c.conkey)
                WHERE c.contype = 'f'
                  AND c.conrelid = 'inspection_plan_station_items'::regclass
                  AND a.attname = 'assigned_by'
            ) THEN
                ALTER TABLE inspection_plan_station_items
                ADD CONSTRAINT fk_inspection_plan_items_assigned_by
                FOREIGN KEY (assigned_by) REFERENCES users(id) ON DELETE SET NULL;
            END IF;
        END
        $$;
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_inspection_plan_station_items_assigned_inspector_id
        ON inspection_plan_station_items(assigned_inspector_id);
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_inspection_plan_station_items_assigned_inspector_ids
        ON inspection_plan_station_items USING GIN (assigned_inspector_ids);
        """
    )
    connection = getattr(cur, "connection", None)
    if connection:
        connection.commit()
    INSPECTION_PLAN_ASSIGNMENT_SCHEMA_READY = True


def normalize_plan_assigned_inspector_ids(item):
    raw_values = item.get("assigned_inspector_ids")
    if raw_values in (None, ""):
        legacy_value = item.get("assigned_inspector_id")
        raw_values = [] if legacy_value in (None, "") else [legacy_value]
    elif not isinstance(raw_values, list):
        raw_values = [raw_values]

    result = []
    seen = set()
    for raw_value in raw_values:
        if raw_value in (None, ""):
            continue
        try:
            inspector_id = int(raw_value)
        except (TypeError, ValueError) as exc:
            raise ValueError("检查人参数不正确。") from exc
        if inspector_id <= 0 or inspector_id in seen:
            continue
        seen.add(inspector_id)
        result.append(inspector_id)
    return result


def serialize_inspection_completion_config(row=None):
    period = str(row.get("record_uniqueness_period") if row else "").strip() or DEFAULT_INSPECTION_RECORD_UNIQUENESS_PERIOD
    if period not in INSPECTION_RECORD_UNIQUENESS_PERIODS:
        period = DEFAULT_INSPECTION_RECORD_UNIQUENESS_PERIOD
    return {
        "auto_complete_enabled": bool(row.get("auto_complete_enabled")) if row else True,
        "auto_complete_days": int(row.get("auto_complete_days") or DEFAULT_INSPECTION_AUTO_COMPLETE_DAYS)
        if row
        else DEFAULT_INSPECTION_AUTO_COMPLETE_DAYS,
        "record_uniqueness_period": period,
        "record_uniqueness_period_label": inspection_period_label(period),
        "updated_by": row.get("updated_by") if row else None,
        "updated_by_username": row.get("updated_by_username") if row else "",
        "updated_by_name": row.get("updated_by_name") if row else "",
        "updated_at": row.get("updated_at") if row else "",
    }


def get_inspection_completion_config(cur):
    ensure_inspection_completion_schema(cur)
    cur.execute(
        """
        SELECT
            s.auto_complete_enabled,
            s.auto_complete_days,
            s.record_uniqueness_period,
            s.updated_by,
            COALESCE(u.username, '') AS updated_by_username,
            COALESCE(u.real_name, '') AS updated_by_name,
            TO_CHAR(s.updated_at, 'YYYY-MM-DD HH24:MI') AS updated_at
        FROM inspection_completion_settings s
        LEFT JOIN users u ON u.id = s.updated_by
        WHERE s.singleton = TRUE
        LIMIT 1;
        """
    )
    return serialize_inspection_completion_config(cur.fetchone())


def save_inspection_completion_config(cur, data, user_id):
    enabled = bool(data.get("auto_complete_enabled", True))
    period = str(
        data.get("record_uniqueness_period") or DEFAULT_INSPECTION_RECORD_UNIQUENESS_PERIOD
    ).strip()
    if period not in INSPECTION_RECORD_UNIQUENESS_PERIODS:
        raise ValueError("巡检记录唯一周期参数不合法。")
    try:
        days = int(data.get("auto_complete_days", DEFAULT_INSPECTION_AUTO_COMPLETE_DAYS))
    except (TypeError, ValueError):
        raise ValueError("自动确认天数必须是数字。")
    if days < 1 or days > 31:
        raise ValueError("自动确认天数需设置为 1-31 天。")

    ensure_inspection_completion_schema(cur)
    cur.execute(
        """
        INSERT INTO inspection_completion_settings (
            singleton,
            auto_complete_enabled,
            auto_complete_days,
            record_uniqueness_period,
            updated_by,
            updated_at
        )
        VALUES (TRUE, %s, %s, %s, %s, CURRENT_TIMESTAMP)
        ON CONFLICT (singleton)
        DO UPDATE SET
            auto_complete_enabled = EXCLUDED.auto_complete_enabled,
            auto_complete_days = EXCLUDED.auto_complete_days,
            record_uniqueness_period = EXCLUDED.record_uniqueness_period,
            updated_by = EXCLUDED.updated_by,
            updated_at = CURRENT_TIMESTAMP;
        """,
        (enabled, days, period, user_id or None),
    )
    return get_inspection_completion_config(cur)


def auto_complete_overdue_inspections(cur):
    config = get_inspection_completion_config(cur)
    if not config["auto_complete_enabled"]:
        return 0

    cutoff_date = beijing_today() - timedelta(days=config["auto_complete_days"])
    cur.execute(
        """
        UPDATE inspections
        SET inspector_completion_status = %s,
            inspector_completed_by = NULL,
            inspector_completed_at = COALESCE(inspector_completed_at, CURRENT_TIMESTAMP),
            inspector_completion_source = 'auto',
            updated_at = CURRENT_TIMESTAMP
        WHERE inspector_completion_status <> %s
          AND inspection_date <= %s
          AND COALESCE(sign_status, '') <> '已签名确认'
          AND COALESCE(inspector_completion_source, '') <> 'admin_reopen';
        """,
        (INSPECTION_COMPLETION_DONE, INSPECTION_COMPLETION_DONE, cutoff_date),
    )
    return cur.rowcount


def normalize_inspection_period(value):
    period = str(value or DEFAULT_INSPECTION_RECORD_UNIQUENESS_PERIOD).strip()
    if period not in INSPECTION_RECORD_UNIQUENESS_PERIODS:
        return DEFAULT_INSPECTION_RECORD_UNIQUENESS_PERIOD
    return period


def inspection_period_label(period):
    return {
        "week": "自然周",
        "month": "自然月",
        "quarter": "自然季度",
        "year": "自然年",
    }.get(normalize_inspection_period(period), "自然月")


def inspection_period_scope_text(period):
    return {
        "week": "本自然周",
        "month": "本自然月",
        "quarter": "本自然季度",
        "year": "本自然年",
    }.get(normalize_inspection_period(period), "本自然月")


def get_inspection_period_range(target_date, period=None):
    period = normalize_inspection_period(period)
    if isinstance(target_date, str):
        target_date = datetime.strptime(target_date, "%Y-%m-%d").date()
    if period == "week":
        period_start = target_date - timedelta(days=target_date.weekday())
        period_end = period_start + timedelta(days=7)
    elif period == "quarter":
        start_month = ((target_date.month - 1) // 3) * 3 + 1
        period_start = target_date.replace(month=start_month, day=1)
        if start_month == 10:
            period_end = period_start.replace(year=period_start.year + 1, month=1)
        else:
            period_end = period_start.replace(month=start_month + 3)
    elif period == "year":
        period_start = target_date.replace(month=1, day=1)
        period_end = period_start.replace(year=period_start.year + 1)
    else:
        period_start = target_date.replace(day=1)
        if period_start.month == 12:
            period_end = period_start.replace(year=period_start.year + 1, month=1)
        else:
            period_end = period_start.replace(month=period_start.month + 1)
    return period_start, period_end


def format_inspection_period_key(target_date, period=None):
    period = normalize_inspection_period(period)
    if isinstance(target_date, str):
        target_date = datetime.strptime(target_date, "%Y-%m-%d").date()
    period_start, _ = get_inspection_period_range(target_date, period)
    if period == "week":
        iso_year, iso_week, _ = target_date.isocalendar()
        return f"{iso_year}-W{iso_week:02d}"
    if period == "quarter":
        quarter = ((target_date.month - 1) // 3) + 1
        return f"{target_date.year}-Q{quarter}"
    if period == "year":
        return str(target_date.year)
    return period_start.strftime("%Y-%m")


def format_inspection_period_label(target_date, period=None):
    period = normalize_inspection_period(period)
    if isinstance(target_date, str):
        target_date = datetime.strptime(target_date, "%Y-%m-%d").date()
    if period == "week":
        period_start, period_end = get_inspection_period_range(target_date, period)
        return f"{period_start.strftime('%Y-%m-%d')}至{(period_end - timedelta(days=1)).strftime('%Y-%m-%d')}"
    if period == "quarter":
        quarter = ((target_date.month - 1) // 3) + 1
        return f"{target_date.year}年第{quarter}季度"
    if period == "year":
        return f"{target_date.year}年"
    return f"{target_date.year}年{target_date.month}月"


def lock_inspection_period_scope(cur, station_id, inspection_table_id, target_date, period=None):
    period = normalize_inspection_period(period)
    period_start, _ = get_inspection_period_range(target_date, period)
    lock_seed = f"inspection-period:{period}:{station_id}:{inspection_table_id}:{period_start.isoformat()}"
    lock_key = int.from_bytes(
        hashlib.blake2b(lock_seed.encode("utf-8"), digest_size=8).digest(),
        byteorder="big",
        signed=True,
    )
    cur.execute("SELECT pg_advisory_xact_lock(%s);", (lock_key,))


def find_period_inspection(cur, station_id, inspection_table_id, target_date, period=None):
    period_start, period_end = get_inspection_period_range(target_date, period)
    cur.execute(
        """
        SELECT
            id,
            batch_id,
            inspection_date,
            inspector_completion_status,
            inspector_completed_at
        FROM inspections
        WHERE station_id = %s
          AND inspection_table_id = %s
          AND inspection_date >= %s
          AND inspection_date < %s
        ORDER BY
            CASE WHEN inspector_completion_status = %s THEN 1 ELSE 0 END ASC,
            inspection_date ASC,
            id ASC
        LIMIT 1;
        """,
        (
            station_id,
            inspection_table_id,
            period_start,
            period_end,
            INSPECTION_COMPLETION_DONE,
        ),
    )
    return cur.fetchone()


def get_or_create_period_inspection(
    cur,
    station_id,
    inspector_id,
    inspection_table_id,
    target_date,
):
    period = get_inspection_completion_config(cur)["record_uniqueness_period"]
    lock_inspection_period_scope(cur, station_id, inspection_table_id, target_date, period)
    inspection = find_period_inspection(cur, station_id, inspection_table_id, target_date, period)
    if inspection:
        if inspection.get("inspector_completion_status") == INSPECTION_COMPLETION_DONE:
            raise ValueError(f"该站点{inspection_period_scope_text(period)}已完成确认该检查表，不能继续登记。")
        return inspection["id"]

    batch_id = get_or_create_inspection_batch(cur, station_id, inspector_id, target_date)
    return create_inspection_record(
        cur,
        station_id,
        inspector_id,
        inspection_table_id,
        batch_id,
        target_date,
    )


def inspection_completion_source_label(value):
    return {
        "manual": "检查人手动确认",
        "manual_all": "全部检查人确认",
        "auto": "系统自动确认",
        "admin": "后台管理确认",
        "signature": "站经理签字确认",
        "admin_reopen": "后台恢复未完成",
    }.get(str(value or ""), "")


def sync_signed_inspections_completion(cur):
    cur.execute(
        """
        UPDATE inspections
        SET inspector_completion_status = %s,
            inspector_completed_at = COALESCE(
                inspector_completed_at,
                station_manager_signed_at,
                updated_at,
                created_at,
                CURRENT_TIMESTAMP
            ),
            inspector_completion_source = COALESCE(NULLIF(inspector_completion_source, ''), 'signature'),
            updated_at = COALESCE(updated_at, CURRENT_TIMESTAMP)
        WHERE sign_status = '已签名确认'
          AND inspector_completion_status <> %s;
        """,
        (INSPECTION_COMPLETION_DONE, INSPECTION_COMPLETION_DONE),
    )
    return cur.rowcount


def complete_inspection_record(cur, inspection_id, user_id=None, source="manual"):
    normalized_source = source if source in INSPECTION_COMPLETION_SOURCES else "manual"
    cur.execute(
        """
        UPDATE inspections
        SET inspector_completion_status = %s,
            inspector_completed_by = %s,
            inspector_completed_at = CURRENT_TIMESTAMP,
            inspector_completion_source = %s,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = %s
          AND inspector_completion_status <> %s;
        """,
        (
            INSPECTION_COMPLETION_DONE,
            user_id,
            normalized_source,
            inspection_id,
            INSPECTION_COMPLETION_DONE,
        ),
    )
    return cur.rowcount


def reopen_inspection_record(cur, inspection_id):
    cur.execute(
        """
        DELETE FROM inspection_inspector_confirmations
        WHERE inspection_id = %s;
        """,
        (inspection_id,),
    )
    cur.execute(
        """
        UPDATE inspections
        SET inspector_completion_status = %s,
            inspector_completed_by = NULL,
            inspector_completed_at = NULL,
            inspector_completion_source = 'admin_reopen',
            updated_at = CURRENT_TIMESTAMP
        WHERE id = %s
          AND COALESCE(sign_status, '') <> '已签名确认';
        """,
        (INSPECTION_COMPLETION_PENDING, inspection_id),
    )
    return cur.rowcount


def user_participated_in_inspection(cur, inspection_id, user_id):
    cur.execute(
        """
        SELECT 1
        WHERE EXISTS (
            SELECT 1
            FROM inspections ins
            WHERE ins.id = %s
              AND ins.inspector_id = %s
        )
        OR EXISTS (
            SELECT 1
            FROM issues i
            WHERE i.inspection_id = %s
              AND i.inspector_id = %s
        )
        LIMIT 1;
        """,
        (inspection_id, user_id, inspection_id, user_id),
    )
    return bool(cur.fetchone())


def fetch_inspection_completion_progress(cur, inspection_ids):
    normalized_ids = []
    for value in inspection_ids or []:
        try:
            normalized_ids.append(int(value))
        except (TypeError, ValueError):
            continue
    normalized_ids = sorted(set(normalized_ids))
    if not normalized_ids:
        return {}

    cur.execute(
        """
        WITH participant_rows AS (
            SELECT ins.id AS inspection_id, ins.inspector_id AS inspector_id
            FROM inspections ins
            WHERE ins.id = ANY(%s)
              AND ins.inspector_id IS NOT NULL
            UNION
            SELECT i.inspection_id, i.inspector_id
            FROM issues i
            WHERE i.inspection_id = ANY(%s)
              AND i.inspector_id IS NOT NULL
              AND COALESCE(i.audit_status, 'pending') <> 'rejected'
        )
        SELECT
            pr.inspection_id,
            participant.id AS inspector_id,
            participant.username,
            participant.real_name,
            participant.phone,
            ins.inspector_completion_status,
            ins.inspector_completion_source,
            TO_CHAR(ins.inspector_completed_at, 'YYYY-MM-DD HH24:MI') AS inspector_completed_at,
            TO_CHAR(conf.confirmed_at, 'YYYY-MM-DD HH24:MI') AS confirmed_at,
            conf.source AS confirmation_source
        FROM participant_rows pr
        JOIN inspections ins ON ins.id = pr.inspection_id
        JOIN users participant ON participant.id = pr.inspector_id
        LEFT JOIN inspection_inspector_confirmations conf
          ON conf.inspection_id = pr.inspection_id
         AND conf.inspector_id = pr.inspector_id
        ORDER BY
            pr.inspection_id ASC,
            COALESCE(NULLIF(TRIM(participant.real_name), ''), participant.username, participant.id::text) ASC,
            participant.id ASC;
        """,
        (normalized_ids, normalized_ids),
    )

    progress_map = {
        inspection_id: {
            "total": 0,
            "confirmed": 0,
            "pending": 0,
            "is_complete": False,
            "participants": [],
        }
        for inspection_id in normalized_ids
    }
    for row in cur.fetchall():
        inspection_id = int(row["inspection_id"])
        progress = progress_map.setdefault(
            inspection_id,
            {
                "total": 0,
                "confirmed": 0,
                "pending": 0,
                "is_complete": False,
                "participants": [],
            },
        )
        is_overall_done = row.get("inspector_completion_status") == INSPECTION_COMPLETION_DONE
        confirmed_at = row.get("confirmed_at")
        confirmation_source = row.get("confirmation_source") or "manual"
        if is_overall_done and not confirmed_at:
            confirmed_at = row.get("inspector_completed_at")
            confirmation_source = row.get("inspector_completion_source") or "manual_all"

        is_confirmed = bool(confirmed_at) or is_overall_done
        display_name = (
            row.get("real_name")
            or row.get("username")
            or row.get("phone")
            or str(row.get("inspector_id"))
        )
        progress["participants"].append(
            {
                "id": row.get("inspector_id"),
                "username": row.get("username") or "",
                "real_name": row.get("real_name") or "",
                "phone": row.get("phone") or "",
                "display_name": display_name,
                "confirmed": is_confirmed,
                "confirmed_at": confirmed_at or "",
                "source": confirmation_source,
                "source_label": inspection_completion_source_label(confirmation_source),
            }
        )

    for inspection_id, progress in progress_map.items():
        total = len(progress["participants"])
        confirmed = sum(1 for item in progress["participants"] if item.get("confirmed"))
        progress["total"] = total
        progress["confirmed"] = confirmed
        progress["pending"] = max(total - confirmed, 0)
        progress["is_complete"] = bool(total and confirmed >= total)
        progress["pending_names"] = [
            item.get("display_name") or "未命名检查人"
            for item in progress["participants"]
            if not item.get("confirmed")
        ]
        progress["confirmed_names"] = [
            item.get("display_name") or "未命名检查人"
            for item in progress["participants"]
            if item.get("confirmed")
        ]
    return progress_map


def apply_inspection_completion_progress(cur, items, hide_inspector_contact=False):
    if not items:
        return items
    progress_map = fetch_inspection_completion_progress(
        cur,
        [item.get("id") for item in items if item.get("id")],
    )
    for item in items:
        progress = progress_map.get(int(item.get("id") or 0), {
            "total": 0,
            "confirmed": 0,
            "pending": 0,
            "is_complete": False,
            "participants": [],
            "pending_names": [],
            "confirmed_names": [],
        })
        if hide_inspector_contact:
            masked_participants = []
            for index, participant in enumerate(progress.get("participants") or [], start=1):
                masked_participants.append(
                    {
                        "id": participant.get("id"),
                        "display_name": f"检查人{index}",
                        "confirmed": bool(participant.get("confirmed")),
                        "confirmed_at": participant.get("confirmed_at") or "",
                        "source": participant.get("source") or "",
                        "source_label": participant.get("source_label") or "",
                    }
                )
            progress = {
                **progress,
                "participants": masked_participants,
                "pending_names": [
                    item.get("display_name") or "检查人"
                    for item in masked_participants
                    if not item.get("confirmed")
                ],
                "confirmed_names": [
                    item.get("display_name") or "检查人"
                    for item in masked_participants
                    if item.get("confirmed")
                ],
            }
        item["inspector_completion_progress"] = progress
    return items


def get_inspection_participant_ids(cur, inspection_id):
    progress = fetch_inspection_completion_progress(cur, [inspection_id]).get(int(inspection_id), {})
    return {
        int(item["id"])
        for item in progress.get("participants") or []
        if item.get("id") is not None
    }


def confirm_inspection_participant(cur, inspection_id, inspector_id):
    cur.execute(
        """
        INSERT INTO inspection_inspector_confirmations (
            inspection_id,
            inspector_id,
            confirmed_at,
            source,
            updated_at
        )
        VALUES (%s, %s, CURRENT_TIMESTAMP, 'manual', CURRENT_TIMESTAMP)
        ON CONFLICT (inspection_id, inspector_id)
        DO UPDATE SET
            confirmed_at = EXCLUDED.confirmed_at,
            source = EXCLUDED.source,
            updated_at = CURRENT_TIMESTAMP;
        """,
        (inspection_id, inspector_id),
    )
    return fetch_inspection_completion_progress(cur, [inspection_id]).get(int(inspection_id), {})


# === Permission helpers ===
def get_user_by_id(cur, user_id):
    cur.execute(
        """
        SELECT id, username, role, real_name, phone, station_id
        FROM users
        WHERE id = %s
        LIMIT 1;
        """,
        (user_id,),
    )
    return cur.fetchone()


def ensure_user_birthday_schema(cur):
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS user_birthdays (
            id SERIAL PRIMARY KEY,
            real_name TEXT NOT NULL UNIQUE,
            birthday_month INTEGER NOT NULL CHECK (birthday_month BETWEEN 1 AND 12),
            birthday_day INTEGER NOT NULL CHECK (birthday_day BETWEEN 1 AND 31),
            updated_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_user_birthdays_real_name
        ON user_birthdays(real_name);
        """
    )
    for real_name, month, day in DEFAULT_USER_BIRTHDAYS:
        cur.execute(
            """
            INSERT INTO user_birthdays (real_name, birthday_month, birthday_day, created_at, updated_at)
            VALUES (%s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            ON CONFLICT (real_name) DO NOTHING;
            """,
            (real_name, month, day),
        )


def validate_birthday_month_day(month, day):
    try:
        birthday_month = int(month)
        birthday_day = int(day)
        datetime(2026, birthday_month, birthday_day)
    except (TypeError, ValueError):
        raise ValueError("生日日期不正确，请选择有效的月份和日期。")
    return birthday_month, birthday_day


def format_birthday_label(month, day):
    return f"{int(month)}月{int(day)}日"


def build_work_duration_payload(today=None):
    current_date = today or beijing_today()
    elapsed_days = max(0, (current_date - WORK_ANNIVERSARY_START_DATE).days + 1)
    return {
        "work_start_date": WORK_ANNIVERSARY_START_DATE.isoformat(),
        "work_days": elapsed_days,
        "work_duration_text": f"从2026年4月15日起，已并肩工作 {elapsed_days} 天",
    }


def build_birthday_event_payload(user, birthday_row, *, force=False):
    today = beijing_today()
    birthday_month = int(birthday_row["birthday_month"])
    birthday_day = int(birthday_row["birthday_day"])
    is_today = birthday_month == today.month and birthday_day == today.day
    if not force and not is_today:
        return None

    real_name = (
        birthday_row.get("real_name")
        or (user or {}).get("real_name")
        or (user or {}).get("username")
        or "伙伴"
    )
    work_payload = build_work_duration_payload(today)
    return {
        "is_today": is_today,
        "is_test": bool(force),
        "event_key": f"{today.isoformat()}:{real_name}:{birthday_month}-{birthday_day}",
        "real_name": real_name,
        "birthday_month": birthday_month,
        "birthday_day": birthday_day,
        "birthday_label": format_birthday_label(birthday_month, birthday_day),
        **work_payload,
        "message": f"感谢你在业务督导中心的每一天付出，愿今天有光、有花，也有被认真看见的快乐。",
    }


def get_user_birthday_event(cur, user, *, force=False):
    real_name = normalize_text((user or {}).get("real_name"), 80)
    if not real_name:
        return None
    cur.execute(
        """
        SELECT id, real_name, birthday_month, birthday_day
        FROM user_birthdays
        WHERE TRIM(real_name) = TRIM(%s)
        LIMIT 1;
        """,
        (real_name,),
    )
    birthday_row = cur.fetchone()
    if not birthday_row:
        return None
    return build_birthday_event_payload(user, birthday_row, force=force)


def ensure_user_security_schema(cur):
    global USER_SECURITY_SCHEMA_READY
    if USER_SECURITY_SCHEMA_READY:
        return

    cur.execute(
        """
        INSERT INTO users (username, password, role, real_name, phone, station_id)
        VALUES ('root', %s, 'root', '系统管理员', '18801800773', NULL)
        ON CONFLICT (username) DO NOTHING;
        """,
        (DEFAULT_INITIAL_PASSWORD,),
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS user_permissions (
            user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            permission_key TEXT NOT NULL,
            is_allowed BOOLEAN NOT NULL,
            updated_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (user_id, permission_key)
        );
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS role_permissions (
            role TEXT NOT NULL,
            permission_key TEXT NOT NULL,
            is_allowed BOOLEAN NOT NULL,
            updated_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (role, permission_key)
        );
        """
    )
    ensure_user_birthday_schema(cur)
    cur.execute("SELECT to_regclass('inspection_tables') AS table_ref;")
    if cur.fetchone()["table_ref"]:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS user_inspection_table_scopes (
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                scope_key TEXT NOT NULL,
                inspection_table_id INTEGER NOT NULL REFERENCES inspection_tables(id) ON DELETE CASCADE,
                updated_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (user_id, scope_key, inspection_table_id)
            );
            """
        )
        cur.execute(
            """
            ALTER TABLE user_inspection_table_scopes
            ADD COLUMN IF NOT EXISTS scope_key TEXT;
            """
        )
        cur.execute(
            """
            ALTER TABLE user_inspection_table_scopes
            DROP CONSTRAINT IF EXISTS user_inspection_table_scopes_pkey;
            """
        )
        cur.execute(
            """
            UPDATE user_inspection_table_scopes
            SET scope_key = 'limit_issue_inspection_table_scope'
            WHERE NULLIF(TRIM(COALESCE(scope_key, '')), '') IS NULL;
            """
        )
        for scope_key in (
            "limit_record_inspection_table_scope",
            "limit_plan_inspection_table_scope",
        ):
            cur.execute(
                """
                INSERT INTO user_inspection_table_scopes (
                    user_id,
                    scope_key,
                    inspection_table_id,
                    updated_by,
                    created_at,
                    updated_at
                )
                SELECT
                    source.user_id,
                    %s,
                    source.inspection_table_id,
                    source.updated_by,
                    source.created_at,
                    source.updated_at
                FROM user_inspection_table_scopes source
                WHERE source.scope_key = 'limit_issue_inspection_table_scope'
                  AND NOT EXISTS (
                      SELECT 1
                      FROM user_inspection_table_scopes target
                      WHERE target.user_id = source.user_id
                        AND target.scope_key = %s
                        AND target.inspection_table_id = source.inspection_table_id
                  );
                """,
                (scope_key, scope_key),
            )
        cur.execute(
            """
            DELETE FROM user_inspection_table_scopes a
            USING user_inspection_table_scopes b
            WHERE a.ctid < b.ctid
              AND a.user_id = b.user_id
              AND a.scope_key = b.scope_key
              AND a.inspection_table_id = b.inspection_table_id;
            """
        )
        cur.execute(
            """
            ALTER TABLE user_inspection_table_scopes
            ALTER COLUMN scope_key SET NOT NULL;
            """
        )
        cur.execute(
            """
            ALTER TABLE user_inspection_table_scopes
            ADD CONSTRAINT user_inspection_table_scopes_pkey
            PRIMARY KEY (user_id, scope_key, inspection_table_id);
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS role_inspection_table_scopes (
                role TEXT NOT NULL,
                scope_key TEXT NOT NULL,
                inspection_table_id INTEGER NOT NULL REFERENCES inspection_tables(id) ON DELETE CASCADE,
                updated_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (role, scope_key, inspection_table_id)
            );
            """
        )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS user_station_region_scopes (
            user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            scope_key TEXT NOT NULL,
            station_region TEXT NOT NULL,
            updated_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (user_id, scope_key, station_region)
        );
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS role_station_region_scopes (
            role TEXT NOT NULL,
            scope_key TEXT NOT NULL,
            station_region TEXT NOT NULL,
            updated_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (role, scope_key, station_region)
        );
        """
    )
    USER_SECURITY_SCHEMA_READY = True


def ensure_system_page_visibility_schema(cur):
    global SYSTEM_PAGE_VISIBILITY_SCHEMA_READY
    if SYSTEM_PAGE_VISIBILITY_SCHEMA_READY:
        return

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS system_page_visibility (
            page_key TEXT PRIMARY KEY,
            is_visible BOOLEAN NOT NULL DEFAULT TRUE,
            updated_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_system_page_visibility_visible
        ON system_page_visibility(is_visible);
        """
    )
    SYSTEM_PAGE_VISIBILITY_SCHEMA_READY = True


def fetch_page_visibility_settings(cur):
    ensure_system_page_visibility_schema(cur)
    cur.execute(
        """
        SELECT page_key, is_visible
        FROM system_page_visibility
        ORDER BY page_key ASC;
        """
    )
    return {row["page_key"]: bool(row["is_visible"]) for row in cur.fetchall()}


def normalize_page_visibility_items(raw_items):
    if not isinstance(raw_items, list):
        raise ValueError("页面显示配置格式不正确。")

    normalized = []
    seen_keys = set()
    for item in raw_items:
        if not isinstance(item, dict):
            raise ValueError("页面显示配置中存在无效项目。")
        page_key = normalize_text(item.get("page_key"), 120)
        if not page_key:
            raise ValueError("页面标识不能为空。")
        if page_key in seen_keys:
            raise ValueError(f"页面显示配置中存在重复项目：{page_key}")
        seen_keys.add(page_key)
        normalized.append(
            {
                "page_key": page_key,
                "is_visible": bool(item.get("is_visible")),
            }
        )
    return normalized


def ensure_ai_usage_schema(cur):
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS ai_usage_logs (
            id SERIAL PRIMARY KEY,
            user_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
            username TEXT,
            real_name TEXT,
            role TEXT,
            usage_module TEXT NOT NULL,
            usage_action TEXT NOT NULL,
            model TEXT NOT NULL,
            base_url TEXT DEFAULT 'https://api.deepseek.com',
            ai_called BOOLEAN NOT NULL DEFAULT FALSE,
            ai_generated BOOLEAN NOT NULL DEFAULT FALSE,
            success BOOLEAN NOT NULL DEFAULT FALSE,
            fallback_used BOOLEAN NOT NULL DEFAULT FALSE,
            status_code INTEGER,
            prompt_chars INTEGER NOT NULL DEFAULT 0,
            prompt_chinese_chars INTEGER NOT NULL DEFAULT 0,
            prompt_other_chars INTEGER NOT NULL DEFAULT 0,
            completion_chars INTEGER NOT NULL DEFAULT 0,
            completion_chinese_chars INTEGER NOT NULL DEFAULT 0,
            completion_other_chars INTEGER NOT NULL DEFAULT 0,
            total_chars INTEGER NOT NULL DEFAULT 0,
            input_tokens_est NUMERIC(14, 2) NOT NULL DEFAULT 0,
            output_tokens_est NUMERIC(14, 2) NOT NULL DEFAULT 0,
            total_tokens_est NUMERIC(14, 2) NOT NULL DEFAULT 0,
            input_cost_est NUMERIC(14, 6) NOT NULL DEFAULT 0,
            output_cost_est NUMERIC(14, 6) NOT NULL DEFAULT 0,
            total_cost_est NUMERIC(14, 6) NOT NULL DEFAULT 0,
            message TEXT,
            request_summary TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_ai_usage_logs_created_at
        ON ai_usage_logs(created_at DESC);
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_ai_usage_logs_user_id
        ON ai_usage_logs(user_id);
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_ai_usage_logs_module_action
        ON ai_usage_logs(usage_module, usage_action);
        """
    )


def get_number(value, default=0):
    try:
        if value is None:
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def get_integer(value, default=0):
    try:
        if value is None:
            return default
        return int(value)
    except (TypeError, ValueError):
        return default


def record_ai_usage_log(
    cur,
    user,
    ai_result,
    usage_module,
    usage_action,
    request_summary="",
):
    ensure_ai_usage_schema(cur)
    usage = (ai_result or {}).get("usage") or {}
    model = str(usage.get("model") or "deepseek-v4-pro").strip() or "deepseek-v4-pro"
    base_url = str(usage.get("base_url") or "https://api.deepseek.com").strip()
    message = str((ai_result or {}).get("message") or "")[:500]
    user = user or {}
    cur.execute(
        """
        INSERT INTO ai_usage_logs (
            user_id,
            username,
            real_name,
            role,
            usage_module,
            usage_action,
            model,
            base_url,
            ai_called,
            ai_generated,
            success,
            fallback_used,
            status_code,
            prompt_chars,
            prompt_chinese_chars,
            prompt_other_chars,
            completion_chars,
            completion_chinese_chars,
            completion_other_chars,
            total_chars,
            input_tokens_est,
            output_tokens_est,
            total_tokens_est,
            input_cost_est,
            output_cost_est,
            total_cost_est,
            message,
            request_summary
        )
        VALUES (
            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
        );
        """,
        (
            user.get("id"),
            user.get("username"),
            user.get("real_name"),
            user.get("role"),
            str(usage_module or "AI功能")[:80],
            str(usage_action or "AI调用")[:80],
            model[:80],
            base_url[:160],
            bool(usage.get("ai_called")),
            bool((ai_result or {}).get("generated")),
            bool(usage.get("success")),
            bool(usage.get("fallback_used")),
            usage.get("status_code"),
            get_integer(usage.get("prompt_chars")),
            get_integer(usage.get("prompt_chinese_chars")),
            get_integer(usage.get("prompt_other_chars")),
            get_integer(usage.get("completion_chars")),
            get_integer(usage.get("completion_chinese_chars")),
            get_integer(usage.get("completion_other_chars")),
            get_integer(usage.get("total_chars")),
            get_number(usage.get("input_tokens_est")),
            get_number(usage.get("output_tokens_est")),
            get_number(usage.get("total_tokens_est")),
            get_number(usage.get("input_cost_est")),
            get_number(usage.get("output_cost_est")),
            get_number(usage.get("total_cost_est")),
            message,
            str(request_summary or "")[:500],
        ),
    )


def serialize_ai_usage_row(row):
    return {
        "id": row["id"],
        "user_id": row["user_id"],
        "username": row["username"],
        "real_name": row["real_name"],
        "role": row["role"],
        "usage_module": row["usage_module"],
        "usage_action": row["usage_action"],
        "model": row["model"],
        "base_url": row["base_url"],
        "ai_called": bool(row["ai_called"]),
        "ai_generated": bool(row["ai_generated"]),
        "success": bool(row["success"]),
        "fallback_used": bool(row["fallback_used"]),
        "status_code": row["status_code"],
        "prompt_chars": get_integer(row["prompt_chars"]),
        "prompt_chinese_chars": get_integer(row["prompt_chinese_chars"]),
        "prompt_other_chars": get_integer(row["prompt_other_chars"]),
        "completion_chars": get_integer(row["completion_chars"]),
        "completion_chinese_chars": get_integer(row["completion_chinese_chars"]),
        "completion_other_chars": get_integer(row["completion_other_chars"]),
        "total_chars": get_integer(row["total_chars"]),
        "input_tokens_est": get_number(row["input_tokens_est"]),
        "output_tokens_est": get_number(row["output_tokens_est"]),
        "total_tokens_est": get_number(row["total_tokens_est"]),
        "input_cost_est": get_number(row["input_cost_est"]),
        "output_cost_est": get_number(row["output_cost_est"]),
        "total_cost_est": get_number(row["total_cost_est"]),
        "message": row["message"] or "",
        "request_summary": row["request_summary"] or "",
        "created_at": row["created_at"],
    }


def build_ai_usage_aggregate(rows, group_keys, label_builder):
    grouped = {}
    for row in rows:
        key = tuple(str(row.get(item) or "") for item in group_keys)
        if key not in grouped:
            grouped[key] = {
                "key": "||".join(key),
                "label": label_builder(row),
                "calls": 0,
                "ai_called": 0,
                "success": 0,
                "fallback": 0,
                "total_chars": 0,
                "total_tokens_est": 0.0,
                "total_cost_est": 0.0,
            }
        item = grouped[key]
        item["calls"] += 1
        item["ai_called"] += 1 if row.get("ai_called") else 0
        item["success"] += 1 if row.get("success") else 0
        item["fallback"] += 1 if row.get("fallback_used") else 0
        item["total_chars"] += get_integer(row.get("total_chars"))
        item["total_tokens_est"] += get_number(row.get("total_tokens_est"))
        item["total_cost_est"] += get_number(row.get("total_cost_est"))

    return sorted(
        (
            {
                **item,
                "total_tokens_est": round(item["total_tokens_est"], 2),
                "total_cost_est": round(item["total_cost_est"], 6),
            }
            for item in grouped.values()
        ),
        key=lambda item: (-item["calls"], -item["total_cost_est"], item["label"]),
    )


def build_ai_usage_summary(rows):
    return {
        "total_calls": len(rows),
        "ai_called": sum(1 for row in rows if row.get("ai_called")),
        "success": sum(1 for row in rows if row.get("success")),
        "fallback": sum(1 for row in rows if row.get("fallback_used")),
        "prompt_chars": sum(get_integer(row.get("prompt_chars")) for row in rows),
        "completion_chars": sum(get_integer(row.get("completion_chars")) for row in rows),
        "total_chars": sum(get_integer(row.get("total_chars")) for row in rows),
        "input_tokens_est": round(
            sum(get_number(row.get("input_tokens_est")) for row in rows),
            2,
        ),
        "output_tokens_est": round(
            sum(get_number(row.get("output_tokens_est")) for row in rows),
            2,
        ),
        "total_tokens_est": round(
            sum(get_number(row.get("total_tokens_est")) for row in rows),
            2,
        ),
        "input_cost_est": round(
            sum(get_number(row.get("input_cost_est")) for row in rows),
            6,
        ),
        "output_cost_est": round(
            sum(get_number(row.get("output_cost_est")) for row in rows),
            6,
        ),
        "total_cost_est": round(
            sum(get_number(row.get("total_cost_est")) for row in rows),
            6,
        ),
    }


def is_root_user(user):
    return bool(user and user.get("role") == "root")


def is_supervisor_like(user):
    return bool(user and user.get("role") in ("root", "supervisor"))


def is_station_manager(user):
    return bool(user and user.get("role") == "station_manager")


def is_quality_safety_user(user):
    return bool(user and user.get("role") == "quality_safety")


def has_role_default_checklist_scope(user):
    return bool(user and user.get("role") in ROLE_DEFAULT_CHECKLIST_SCOPES)


def role_default_permission(role, permission_key):
    for item in PERMISSION_CATALOG:
        if item["key"] == permission_key:
            return bool(item.get("defaults", {}).get(role, False))
    return False


def enforce_exclusive_permissions(permission_map, role=None):
    normalized = dict(permission_map or {})
    for exclusive_keys in PERMISSION_EXCLUSIVE_GROUPS:
        active_keys = [key for key in exclusive_keys if normalized.get(key)]
        if len(active_keys) <= 1:
            continue

        default_active_keys = [key for key in active_keys if role_default_permission(role, key)]
        keep_key = default_active_keys[0] if len(default_active_keys) == 1 else active_keys[0]
        for key in active_keys:
            normalized[key] = key == keep_key

    for child_key, parent_key in PERMISSION_DEPENDENCIES.items():
        if normalized.get(child_key) and not normalized.get(parent_key):
            normalized[child_key] = False

    for child_key, parent_keys in PERMISSION_ANY_DEPENDENCIES.items():
        if normalized.get(child_key) and not any(normalized.get(parent_key) for parent_key in parent_keys):
            normalized[child_key] = False

    return normalized


def get_permission_overrides(cur, user_id):
    ensure_user_security_schema(cur)
    cur.execute(
        """
        SELECT permission_key, is_allowed
        FROM user_permissions
        WHERE user_id = %s;
        """,
        (user_id,),
    )
    overrides = {row["permission_key"]: bool(row["is_allowed"]) for row in cur.fetchall()}
    legacy_scope_value = overrides.pop("limit_inspection_table_scope", None)
    if legacy_scope_value is not None:
        for scope_key in INSPECTION_TABLE_SCOPE_PERMISSION_KEYS:
            overrides.setdefault(scope_key, bool(legacy_scope_value))
    return overrides


def get_role_permission_overrides(cur, role):
    ensure_user_security_schema(cur)
    normalized_role = normalize_text(role)
    if normalized_role not in ROLE_OPTIONS or normalized_role == "root":
        return {}
    cur.execute(
        """
        SELECT permission_key, is_allowed
        FROM role_permissions
        WHERE role = %s;
        """,
        (normalized_role,),
    )
    return {row["permission_key"]: bool(row["is_allowed"]) for row in cur.fetchall()}


def get_role_permission_overrides_map(cur):
    ensure_user_security_schema(cur)
    cur.execute(
        """
        SELECT role, permission_key, is_allowed
        FROM role_permissions
        ORDER BY role ASC, permission_key ASC;
        """
    )
    result = {role: {} for role in ROLE_OPTIONS if role != "root"}
    for row in cur.fetchall():
        role = row["role"]
        if role in result:
            result[role][row["permission_key"]] = bool(row["is_allowed"])
    return result


def build_role_effective_permissions(cur, role):
    if role == "root":
        return {item["key"]: True for item in PERMISSION_CATALOG}
    role_overrides = get_role_permission_overrides(cur, role)
    permissions = {
        item["key"]: role_overrides.get(item["key"], role_default_permission(role, item["key"]))
        for item in PERMISSION_CATALOG
    }
    return enforce_exclusive_permissions(permissions, role)


def apply_role_permission_updates(cur, role, permissions, actor_user_id):
    normalized_role = normalize_text(role)
    if normalized_role not in ROLE_OPTIONS:
        raise ValueError("角色类型不正确。")
    if normalized_role == "root":
        raise ValueError("root 固定拥有全部权限，不需要配置角色通用权限。")
    normalized_permissions = enforce_exclusive_permissions(
        normalize_permission_updates(permissions),
        normalized_role,
    )
    cur.execute("DELETE FROM role_permissions WHERE role = %s;", (normalized_role,))
    for permission_key, is_allowed in normalized_permissions.items():
        cur.execute(
            """
            INSERT INTO role_permissions (
                role,
                permission_key,
                is_allowed,
                updated_by,
                created_at,
                updated_at
            )
            VALUES (%s, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP);
            """,
            (normalized_role, permission_key, is_allowed, actor_user_id),
        )
    return normalized_permissions


def normalize_checklist_scope_name(value):
    return (
        sanitize_display_string(str(value or ""))
        .replace("（视频）", "")
        .replace("(视频)", "")
        .replace("（现场）", "")
        .replace("(现场)", "")
        .strip()
    )


def get_role_default_inspection_table_ids(cur, role):
    checklist_scope = ROLE_DEFAULT_CHECKLIST_SCOPES.get(role)
    if not checklist_scope:
        return []
    cur.execute("SELECT to_regclass('inspection_tables') AS table_ref;")
    if not cur.fetchone()["table_ref"]:
        return []
    cur.execute(
        """
        SELECT id, table_name, checklist_mode
        FROM inspection_tables
        WHERE is_active = TRUE;
        """
    )
    target_pairs = {
        (normalize_checklist_scope_name(table_name), checklist_mode)
        for table_name, checklist_mode in checklist_scope
    }
    result = []
    for row in cur.fetchall():
        row_pair = (
            normalize_checklist_scope_name(row["table_name"]),
            normalize_checklist_mode(row.get("checklist_mode")),
        )
        if row_pair in target_pairs:
            result.append(row["id"])
    return result


def get_quality_safety_default_inspection_table_ids(cur):
    return get_role_default_inspection_table_ids(cur, "quality_safety")


def normalize_inspection_table_scope_key(scope_key):
    value = str(scope_key or "").strip()
    if value not in INSPECTION_TABLE_SCOPE_PERMISSION_KEYS:
        raise ValueError("检查表范围类型不正确。")
    return value


def get_user_inspection_table_scope_overrides(cur, user_id, scope_key=None):
    ensure_user_security_schema(cur)
    cur.execute("SELECT to_regclass('user_inspection_table_scopes') AS table_ref;")
    if not cur.fetchone()["table_ref"]:
        return {} if scope_key is None else []

    if scope_key is None:
        cur.execute(
            """
            SELECT scope_key, inspection_table_id
            FROM user_inspection_table_scopes
            WHERE user_id = %s
            ORDER BY scope_key ASC, inspection_table_id ASC;
            """,
            (user_id,),
        )
        result = {key: [] for key in INSPECTION_TABLE_SCOPE_PERMISSION_KEYS}
        for row in cur.fetchall():
            key = row["scope_key"]
            if key in result:
                result[key].append(row["inspection_table_id"])
        return result

    normalized_scope_key = normalize_inspection_table_scope_key(scope_key)
    cur.execute(
        """
        SELECT inspection_table_id
        FROM user_inspection_table_scopes
        WHERE user_id = %s
          AND scope_key = %s
        ORDER BY inspection_table_id ASC;
        """,
        (user_id, normalized_scope_key),
    )
    return [row["inspection_table_id"] for row in cur.fetchall()]


def get_role_inspection_table_scope_overrides(cur, role, scope_key=None):
    ensure_user_security_schema(cur)
    normalized_role = normalize_text(role)
    if normalized_role not in ROLE_OPTIONS or normalized_role == "root":
        return {key: [] for key in INSPECTION_TABLE_SCOPE_PERMISSION_KEYS} if scope_key is None else []
    cur.execute("SELECT to_regclass('role_inspection_table_scopes') AS table_ref;")
    if not cur.fetchone()["table_ref"]:
        return {key: [] for key in INSPECTION_TABLE_SCOPE_PERMISSION_KEYS} if scope_key is None else []

    if scope_key is None:
        cur.execute(
            """
            SELECT scope_key, inspection_table_id
            FROM role_inspection_table_scopes
            WHERE role = %s
            ORDER BY scope_key ASC, inspection_table_id ASC;
            """,
            (normalized_role,),
        )
        result = {key: [] for key in INSPECTION_TABLE_SCOPE_PERMISSION_KEYS}
        for row in cur.fetchall():
            key = row["scope_key"]
            if key in result:
                result[key].append(row["inspection_table_id"])
        return result

    normalized_scope_key = normalize_inspection_table_scope_key(scope_key)
    cur.execute(
        """
        SELECT inspection_table_id
        FROM role_inspection_table_scopes
        WHERE role = %s
          AND scope_key = %s
        ORDER BY inspection_table_id ASC;
        """,
        (normalized_role, normalized_scope_key),
    )
    return [row["inspection_table_id"] for row in cur.fetchall()]


def get_role_inspection_table_scope_overrides_map(cur):
    ensure_user_security_schema(cur)
    result = {
        role: {key: [] for key in INSPECTION_TABLE_SCOPE_PERMISSION_KEYS}
        for role in ROLE_OPTIONS
        if role != "root"
    }
    cur.execute("SELECT to_regclass('role_inspection_table_scopes') AS table_ref;")
    if not cur.fetchone()["table_ref"]:
        return result
    cur.execute(
        """
        SELECT role, scope_key, inspection_table_id
        FROM role_inspection_table_scopes
        ORDER BY role ASC, scope_key ASC, inspection_table_id ASC;
        """
    )
    for row in cur.fetchall():
        role = row["role"]
        scope_key = row["scope_key"]
        if role in result and scope_key in result[role]:
            result[role][scope_key].append(row["inspection_table_id"])
    return result


def build_role_effective_inspection_table_scope_map(cur, role, permissions=None):
    normalized_role = normalize_text(role)
    if normalized_role not in ROLE_OPTIONS or normalized_role == "root":
        return {key: [] for key in INSPECTION_TABLE_SCOPE_PERMISSION_KEYS}
    effective_permissions = permissions or build_role_effective_permissions(cur, normalized_role)
    role_scope_map = get_role_inspection_table_scope_overrides(cur, normalized_role)
    default_scope_ids = get_role_default_inspection_table_ids(cur, normalized_role)
    result = {}
    for scope_key in INSPECTION_TABLE_SCOPE_PERMISSION_KEYS:
        if not effective_permissions.get(scope_key):
            result[scope_key] = []
        elif role_scope_map.get(scope_key):
            result[scope_key] = list(role_scope_map[scope_key])
        elif default_scope_ids:
            result[scope_key] = list(default_scope_ids)
        else:
            result[scope_key] = []
    return result


def build_role_effective_inspection_table_scope_map_all(cur):
    return {
        role: build_role_effective_inspection_table_scope_map(cur, role)
        for role in ROLE_OPTIONS
        if role != "root"
    }


def get_effective_inspection_table_scope_ids(cur, user, scope_key, permissions=None):
    if not user or is_root_user(user):
        return None

    normalized_scope_key = normalize_inspection_table_scope_key(scope_key)
    effective_permissions = permissions or get_effective_permissions(cur, user)
    if not effective_permissions.get(normalized_scope_key):
        return None

    scope_ids = get_user_inspection_table_scope_overrides(cur, user["id"], normalized_scope_key)
    if scope_ids:
        return set(scope_ids)
    role_scope_ids = get_role_inspection_table_scope_overrides(cur, user.get("role"), normalized_scope_key)
    if role_scope_ids:
        return set(role_scope_ids)
    if has_role_default_checklist_scope(user):
        return set(get_role_default_inspection_table_ids(cur, user.get("role")))
    return set()


def append_inspection_table_scope_filter(cur, user, where_clauses, params, column_sql, scope_key, permissions=None):
    scope_ids = get_effective_inspection_table_scope_ids(cur, user, scope_key, permissions)
    if scope_ids is None:
        return True
    if not scope_ids:
        return False
    where_clauses.append(f"{column_sql} = ANY(%s)")
    params.append(list(scope_ids))
    return True


def is_inspection_table_allowed_for_user(cur, user, inspection_table_id, scope_key, permissions=None):
    scope_ids = get_effective_inspection_table_scope_ids(cur, user, scope_key, permissions)
    if scope_ids is None:
        return True
    try:
        return int(inspection_table_id) in scope_ids
    except (TypeError, ValueError):
        return False


def normalize_station_region_scope_key(scope_key):
    value = str(scope_key or "").strip()
    if value not in STATION_REGION_SCOPE_PERMISSION_KEYS:
        raise ValueError("片区范围类型不正确。")
    return value


def normalize_station_region_value(value):
    text = normalize_text(value, 120)
    return text or "未填写片区"


def get_user_station_region_scope_overrides(cur, user_id, scope_key=None):
    ensure_user_security_schema(cur)
    cur.execute("SELECT to_regclass('user_station_region_scopes') AS table_ref;")
    if not cur.fetchone()["table_ref"]:
        return {} if scope_key is None else []

    if scope_key is None:
        cur.execute(
            """
            SELECT scope_key, station_region
            FROM user_station_region_scopes
            WHERE user_id = %s
            ORDER BY scope_key ASC, station_region ASC;
            """,
            (user_id,),
        )
        result = {key: [] for key in STATION_REGION_SCOPE_PERMISSION_KEYS}
        for row in cur.fetchall():
            key = row["scope_key"]
            if key in result:
                result[key].append(normalize_station_region_value(row["station_region"]))
        return result

    normalized_scope_key = normalize_station_region_scope_key(scope_key)
    cur.execute(
        """
        SELECT station_region
        FROM user_station_region_scopes
        WHERE user_id = %s
          AND scope_key = %s
        ORDER BY station_region ASC;
        """,
        (user_id, normalized_scope_key),
    )
    return [normalize_station_region_value(row["station_region"]) for row in cur.fetchall()]


def get_role_station_region_scope_overrides(cur, role, scope_key=None):
    ensure_user_security_schema(cur)
    normalized_role = normalize_text(role)
    if normalized_role not in ROLE_OPTIONS or normalized_role == "root":
        return {key: [] for key in STATION_REGION_SCOPE_PERMISSION_KEYS} if scope_key is None else []
    cur.execute("SELECT to_regclass('role_station_region_scopes') AS table_ref;")
    if not cur.fetchone()["table_ref"]:
        return {key: [] for key in STATION_REGION_SCOPE_PERMISSION_KEYS} if scope_key is None else []

    if scope_key is None:
        cur.execute(
            """
            SELECT scope_key, station_region
            FROM role_station_region_scopes
            WHERE role = %s
            ORDER BY scope_key ASC, station_region ASC;
            """,
            (normalized_role,),
        )
        result = {key: [] for key in STATION_REGION_SCOPE_PERMISSION_KEYS}
        for row in cur.fetchall():
            key = row["scope_key"]
            if key in result:
                result[key].append(normalize_station_region_value(row["station_region"]))
        return result

    normalized_scope_key = normalize_station_region_scope_key(scope_key)
    cur.execute(
        """
        SELECT station_region
        FROM role_station_region_scopes
        WHERE role = %s
          AND scope_key = %s
        ORDER BY station_region ASC;
        """,
        (normalized_role, normalized_scope_key),
    )
    return [normalize_station_region_value(row["station_region"]) for row in cur.fetchall()]


def get_role_station_region_scope_overrides_map(cur):
    ensure_user_security_schema(cur)
    result = {
        role: {key: [] for key in STATION_REGION_SCOPE_PERMISSION_KEYS}
        for role in ROLE_OPTIONS
        if role != "root"
    }
    cur.execute("SELECT to_regclass('role_station_region_scopes') AS table_ref;")
    if not cur.fetchone()["table_ref"]:
        return result
    cur.execute(
        """
        SELECT role, scope_key, station_region
        FROM role_station_region_scopes
        ORDER BY role ASC, scope_key ASC, station_region ASC;
        """
    )
    for row in cur.fetchall():
        role = row["role"]
        scope_key = row["scope_key"]
        if role in result and scope_key in result[role]:
            result[role][scope_key].append(normalize_station_region_value(row["station_region"]))
    return result


def build_role_effective_station_region_scope_map(cur, role, permissions=None):
    normalized_role = normalize_text(role)
    if normalized_role not in ROLE_OPTIONS or normalized_role == "root":
        return {key: [] for key in STATION_REGION_SCOPE_PERMISSION_KEYS}
    effective_permissions = permissions or build_role_effective_permissions(cur, normalized_role)
    role_scope_map = get_role_station_region_scope_overrides(cur, normalized_role)
    result = {}
    for scope_key in STATION_REGION_SCOPE_PERMISSION_KEYS:
        result[scope_key] = sorted(role_scope_map.get(scope_key, [])) if effective_permissions.get(scope_key) else []
    return result


def build_role_effective_station_region_scope_map_all(cur):
    return {
        role: build_role_effective_station_region_scope_map(cur, role)
        for role in ROLE_OPTIONS
        if role != "root"
    }


def get_effective_station_region_scope_values(cur, user, scope_key, permissions=None):
    if not user or is_root_user(user):
        return None

    normalized_scope_key = normalize_station_region_scope_key(scope_key)
    effective_permissions = permissions or get_effective_permissions(cur, user)
    if not effective_permissions.get(normalized_scope_key):
        return None

    scope_values = get_user_station_region_scope_overrides(cur, user["id"], normalized_scope_key)
    if scope_values:
        return set(scope_values)
    role_scope_values = get_role_station_region_scope_overrides(cur, user.get("role"), normalized_scope_key)
    if role_scope_values:
        return set(role_scope_values)
    return set(scope_values)


def append_station_region_scope_filter(cur, user, where_clauses, params, column_sql, scope_key, permissions=None):
    scope_values = get_effective_station_region_scope_values(cur, user, scope_key, permissions)
    if scope_values is None:
        return True
    if not scope_values:
        return False
    where_clauses.append(f"COALESCE(NULLIF(TRIM({column_sql}), ''), '未填写片区') = ANY(%s)")
    params.append(list(scope_values))
    return True


def is_station_region_allowed_for_user(cur, user, station_region, scope_key, permissions=None):
    scope_values = get_effective_station_region_scope_values(cur, user, scope_key, permissions)
    if scope_values is None:
        return True
    return normalize_station_region_value(station_region) in scope_values


def get_effective_permissions(cur, user):
    if not user:
        return {}

    if is_root_user(user):
        return {item["key"]: True for item in PERMISSION_CATALOG}

    role = user.get("role")
    role_permissions = build_role_effective_permissions(cur, role)
    overrides = get_permission_overrides(cur, user["id"])
    permissions = {
        item["key"]: overrides.get(item["key"], role_permissions.get(item["key"], False))
        for item in PERMISSION_CATALOG
    }
    return enforce_exclusive_permissions(permissions, role)


def has_permission(cur, user, permission_key):
    if permission_key not in PERMISSION_KEYS:
        return False
    if is_root_user(user):
        return True
    if not user:
        return False
    return bool(get_effective_permissions(cur, user).get(permission_key))


def should_hide_inspector_contact_info(cur, user):
    return bool(user and not is_root_user(user) and has_permission(cur, user, "hide_inspector_contact_info"))


def should_hide_pending_audit_flow_for_user(user):
    return bool(user and user.get("role") in ("station_manager", "area_account"))


def append_pending_audit_issue_visibility_filter(user, where_clauses):
    if should_hide_pending_audit_flow_for_user(user):
        where_clauses.append("COALESCE(i.audit_status, 'pending') <> 'pending'")


def append_pending_audit_inspection_visibility_filter(user, where_clauses):
    if should_hide_pending_audit_flow_for_user(user):
        where_clauses.append(
            """
            NOT EXISTS (
                SELECT 1
                FROM issues pending_issue
                WHERE pending_issue.inspection_id = ins.id
                  AND COALESCE(pending_issue.audit_status, 'pending') = 'pending'
            )
            """
        )


def can_manage_plan(cur, user):
    return has_permission(cur, user, "manage_inspection_plans")


def can_manage_system(cur, user, permission_key):
    return is_root_user(user) or has_permission(cur, user, permission_key)


def can_view_inspection_standards(cur, user):
    return has_permission(cur, user, "view_inspection_standards")


def can_view_checklist_originals(cur, user):
    return has_permission(cur, user, "view_checklist_originals")


def can_view_all_inspection_issues(cur, user):
    return bool(get_effective_permissions(cur, user).get("view_all_inspection_issues"))


def can_view_region_inspection_issues(cur, user):
    return bool(get_effective_permissions(cur, user).get("limit_issue_station_region_scope"))


def can_view_own_inspection_issues(cur, user):
    return bool(get_effective_permissions(cur, user).get("view_own_inspection_issues"))


def can_edit_inspection_issues(cur, user):
    return has_permission(cur, user, "edit_inspection_issues")


def can_delete_inspection_issues(cur, user):
    return has_permission(cur, user, "delete_inspection_issues")


def can_audit_inspection_issues(cur, user):
    return has_permission(cur, user, "audit_inspection_issues")


def can_change_issue_inspector(cur, user):
    return has_permission(cur, user, "change_issue_inspector")


def can_export_issue_photos(cur, user):
    return has_permission(cur, user, "export_issue_photos")


def normalize_issue_audit_status(value):
    status = str(value or "pending").strip().lower()
    return status if status in ISSUE_AUDIT_STATUS_OPTIONS else "pending"


def is_issue_audit_rejected(issue):
    return normalize_issue_audit_status((issue or {}).get("audit_status")) == "rejected"


def is_issue_audit_pending(issue):
    return normalize_issue_audit_status((issue or {}).get("audit_status")) == "pending"


def canonical_issue_status(value):
    normalized = str(value or "").strip()
    return ISSUE_STATUS_ALIASES.get(normalized, normalized)


def is_closed_issue_status(value):
    return canonical_issue_status(value) == "已闭环"


def canonical_issue_result(value):
    normalized = str(value or "").strip()
    return ISSUE_RESULT_ALIASES.get(normalized, normalized)


def issue_created_by_user(user, issue):
    if not user or not issue:
        return False
    return str(issue.get("inspector_id") or "") == str(user.get("id") or "")


def issue_station_rectification_started(issue):
    if not issue:
        return False
    if is_issue_audit_rejected(issue):
        return True
    return (
        canonical_issue_status(issue.get("status")) != "待整改"
        or normalize_issue_result_for_response(issue.get("rectification_result"))
        is not None
    )


def can_user_use_creator_issue_controls(user, issue):
    return (
        issue_created_by_user(user, issue)
        and is_issue_audit_pending(issue)
        and not issue_station_rectification_started(issue)
    )


def can_user_update_rectification_photo(user, issue, can_explicit_edit=False):
    if not user or not issue:
        return False
    if is_issue_audit_rejected(issue):
        return False
    if not normalize_issue_result_for_response(issue.get("rectification_result")):
        return False
    if is_root_user(user):
        return True
    if is_closed_issue_status(issue.get("status")):
        return False
    if can_explicit_edit:
        return True
    return (
        is_station_manager(user)
        and str(user.get("station_id") or "") == str(issue.get("station_id") or "")
        and canonical_issue_status(issue.get("status")) == "待复核"
    )


def normalize_issue_result_for_response(value):
    normalized = canonical_issue_result(value)
    if normalized == "未整改":
        return None
    return normalized or None


def is_issue_inspection_signed(issue):
    if not issue:
        return False
    return (
        issue.get("inspection_sign_status") == "已签名确认"
        or bool(issue.get("station_manager_signed_at"))
        or bool(issue.get("station_manager_signature_path"))
        or bool(issue.get("station_manager_signed_name"))
    )


def is_issue_inspection_completion_done(issue):
    return (issue or {}).get("inspection_completion_status") == INSPECTION_COMPLETION_DONE


def display_issue_status(issue):
    status = canonical_issue_status((issue or {}).get("status"))
    if is_issue_audit_pending(issue):
        return "待审核"
    if status == "待整改" and not is_issue_inspection_signed(issue):
        return "待签名"
    return status


def normalize_issue_row_for_response(
    row,
    user=None,
    can_explicit_edit=False,
    can_explicit_delete=False,
    can_explicit_audit=False,
    can_explicit_change_inspector=None,
    hide_inspector_contact_info=False,
):
    data = dict(row)
    data["inspector_user_id"] = data.get("inspector_id")
    data["audit_status"] = normalize_issue_audit_status(data.get("audit_status"))
    data["audit_status_label"] = ISSUE_AUDIT_STATUS_LABELS.get(
        data["audit_status"],
        "待审核",
    )
    data["is_excellent"] = bool(data.get("is_excellent")) and data["audit_status"] != "rejected"
    if "status" in data:
        data["status"] = canonical_issue_status(data.get("status"))
        data["raw_status"] = data["status"]
        data["workflow_status"] = data["status"]
    for key in ("rectification_result", "review_result"):
        if key in data:
            data[key] = normalize_issue_result_for_response(data.get(key))
    creator_can_modify = can_user_use_creator_issue_controls(user, data)
    closed = is_closed_issue_status(data.get("status"))
    inspection_signed = is_issue_inspection_signed(data)
    inspection_completion_done = is_issue_inspection_completion_done(data)
    issue_mutation_locked = inspection_signed
    creator_audit_locked = (
        issue_created_by_user(user, data)
        and not is_issue_audit_pending(data)
        and not is_root_user(user)
        and not can_explicit_edit
        and not can_explicit_delete
        and not can_explicit_change_inspector
    )
    data["inspection_signed"] = bool(inspection_signed)
    data["inspection_completion_done"] = bool(inspection_completion_done)
    data["inspection_locked"] = bool(issue_mutation_locked)
    if inspection_signed:
        data["operation_lock_reason"] = "已签字不可操作"
    elif creator_audit_locked:
        data["operation_lock_reason"] = "已审核，需重新判定后才能编辑删除"
    else:
        data["operation_lock_reason"] = ""
    data["can_edit_issue_workflow"] = bool(
        not issue_mutation_locked and (is_root_user(user) or (can_explicit_edit and not closed))
    )
    data["can_edit_issue"] = bool(
        not issue_mutation_locked and (is_root_user(user) or ((can_explicit_edit or creator_can_modify) and not closed))
    )
    data["can_delete_issue"] = bool(
        not issue_mutation_locked
        and (
            is_root_user(user)
            or ((can_explicit_delete or creator_can_modify) and not closed)
        )
    )
    data["can_update_rectification_photo"] = can_user_update_rectification_photo(
        user, data, can_explicit_edit
    )
    if can_explicit_change_inspector is None:
        can_explicit_change_inspector = is_root_user(user)
    data["can_change_issue_inspector"] = bool(
        not issue_mutation_locked
        and can_explicit_change_inspector
        and (is_root_user(user) or not closed)
    )
    audit_locked_by_completion = not inspection_completion_done
    data["audit_lock_reason"] = (
        "等待检查人确认"
        if can_explicit_audit and audit_locked_by_completion and not inspection_signed
        else ""
    )
    data["can_audit_issue"] = bool(
        can_explicit_audit
        and not inspection_signed
        and inspection_completion_done
    )
    data["can_mark_excellent_issue"] = bool(can_explicit_audit and data["audit_status"] != "rejected")
    if "status" in data:
        data["status"] = display_issue_status(data)
    if hide_inspector_contact_info:
        for key in ("inspector", "inspector_name", "inspector_username", "inspector_phone"):
            if key in data:
                data[key] = ""
    data.pop("inspector_id", None)
    return data


def fetch_issue_row_for_response(
    cur,
    issue_id,
    user,
    can_explicit_edit=False,
    can_explicit_delete=False,
    can_explicit_audit=False,
    can_explicit_change_inspector=None,
    hide_inspector_contact_info=False,
):
    cur.execute(
        """
        SELECT
            i.id,
            i.station_id,
            COALESCE(i.inspector_id, ins.inspector_id) AS inspector_id,
            TO_CHAR(i.created_at, 'YYYY-MM') AS month,
            TO_CHAR(i.created_at, 'YYYY-MM-DD HH24:MI') AS time,
            s.region,
            s.station_name AS station,
            s.station_manager_name AS station_manager,
            s.station_manager_phone AS station_manager_phone,
            issue_inspector.real_name AS inspector,
            issue_inspector.phone AS inspector_phone,
            t.table_name AS inspection_table_name,
            i.standard_id,
            i.standard_detail_text,
            i.internal_standard_id,
            i.internal_standard_detail_text,
            i.description,
            i.photo_path AS issue_photo,
            i.rectification_result,
            i.rectification_note,
            TO_CHAR(i.rectification_at, 'YYYY-MM-DD HH24:MI') AS rectification_at,
            i.rectification_photo_path AS rectification_photo,
            i.review_result,
            i.review_note,
            TO_CHAR(i.review_at, 'YYYY-MM-DD HH24:MI') AS review_at,
            i.review_photo_path AS review_photo,
            i.status,
            COALESCE(i.audit_status, 'pending') AS audit_status,
            COALESCE(i.is_excellent, FALSE) AS is_excellent,
            i.audited_by,
            audit_user.real_name AS audited_by_name,
            TO_CHAR(i.audited_at, 'YYYY-MM-DD HH24:MI') AS audited_at,
            ins.sign_status AS inspection_sign_status,
            ins.station_manager_signed_name,
            ins.station_manager_signature_path,
            TO_CHAR(ins.station_manager_signed_at, 'YYYY-MM-DD HH24:MI') AS station_manager_signed_at,
            ins.inspector_completion_status AS inspection_completion_status
        FROM issues i
        JOIN inspections ins ON i.inspection_id = ins.id
        JOIN stations s ON i.station_id = s.id
        JOIN inspection_tables t ON i.inspection_table_id = t.id
        JOIN users issue_inspector ON COALESCE(i.inspector_id, ins.inspector_id) = issue_inspector.id
        LEFT JOIN users audit_user ON audit_user.id = i.audited_by
        WHERE i.id = %s
        LIMIT 1;
        """,
        (issue_id,),
    )
    row = cur.fetchone()
    if not row:
        return None
    attach_internal_standard_tags_to_issue_rows(cur, [row])
    return normalize_issue_row_for_response(
        row,
        user,
        can_explicit_edit,
        can_explicit_delete,
        can_explicit_audit,
        can_explicit_change_inspector,
        hide_inspector_contact_info,
    )


def ensure_issue_export_schema(cur):
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS issue_export_tasks (
            task_id TEXT PRIMARY KEY,
            created_by INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            status TEXT NOT NULL DEFAULT 'pending',
            selected_count INTEGER NOT NULL DEFAULT 0,
            exported_count INTEGER NOT NULL DEFAULT 0,
            filter_summary JSONB NOT NULL DEFAULT '{}'::jsonb,
            export_options JSONB NOT NULL DEFAULT '{}'::jsonb,
            file_path TEXT,
            download_filename TEXT,
            error_message TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            completed_at TIMESTAMP,
            expires_at TIMESTAMP NOT NULL
        );
        """
    )
    cur.execute(
        """
        ALTER TABLE issue_export_tasks
        ADD COLUMN IF NOT EXISTS export_options JSONB NOT NULL DEFAULT '{}'::jsonb;
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_issue_export_tasks_created_by
        ON issue_export_tasks (created_by, created_at DESC);
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_issue_export_tasks_expires_at
        ON issue_export_tasks (expires_at);
        """
    )


def cleanup_expired_issue_exports(cur):
    ensure_issue_export_schema(cur)
    cur.execute(
        """
        SELECT file_path
        FROM issue_export_tasks
        WHERE expires_at < CURRENT_TIMESTAMP;
        """
    )
    for row in cur.fetchall():
        if row.get("file_path"):
            remove_storage_file(row["file_path"])
    cur.execute("DELETE FROM issue_export_tasks WHERE expires_at < CURRENT_TIMESTAMP;")

    os.makedirs(ISSUE_EXPORTS_STORAGE_DIR, exist_ok=True)
    cutoff = time.time() - ISSUE_EXPORT_RETENTION_DAYS * 24 * 60 * 60
    for name in os.listdir(ISSUE_EXPORTS_STORAGE_DIR):
        path = os.path.abspath(os.path.join(ISSUE_EXPORTS_STORAGE_DIR, name))
        if os.path.commonpath([os.path.abspath(ISSUE_EXPORTS_STORAGE_DIR), path]) != os.path.abspath(ISSUE_EXPORTS_STORAGE_DIR):
            continue
        if os.path.isfile(path) and path.lower().endswith(".xlsx") and os.path.getmtime(path) < cutoff:
            try:
                os.remove(path)
            except OSError:
                pass


def maybe_cleanup_expired_issue_exports():
    global issue_export_cleanup_last_run
    now = time.time()
    if now - issue_export_cleanup_last_run < ISSUE_EXPORT_CLEANUP_INTERVAL_SECONDS:
        return
    if not issue_export_cleanup_lock.acquire(blocking=False):
        return

    conn = None
    cur = None
    try:
        if now - issue_export_cleanup_last_run < ISSUE_EXPORT_CLEANUP_INTERVAL_SECONDS:
            return
        conn = get_db_connection()
        cur = conn.cursor()
        cleanup_expired_issue_exports(cur)
        conn.commit()
        issue_export_cleanup_last_run = now
    except Exception:
        if conn:
            conn.rollback()
    finally:
        close_db_resources(cur, conn)
        issue_export_cleanup_lock.release()


def normalize_issue_export_ids(raw_ids):
    if not isinstance(raw_ids, list):
        raise ValueError("导出任务缺少筛选后的问题ID列表。")
    normalized = []
    seen = set()
    for raw_id in raw_ids:
        try:
            issue_id = int(raw_id)
        except (TypeError, ValueError):
            continue
        if issue_id <= 0 or issue_id in seen:
            continue
        seen.add(issue_id)
        normalized.append(issue_id)
    if not normalized:
        raise ValueError("当前筛选结果为空，不能导出。")
    return normalized


def normalize_issue_export_filter_summary(raw_summary):
    if not isinstance(raw_summary, dict):
        return {}
    allowed_keys = {
        "issueId",
        "month",
        "date",
        "region",
        "station",
        "stationManager",
        "inspector",
        "inspectionTableName",
        "standardId",
        "standardDetail",
        "standardTags",
        "rectificationResult",
        "reviewResult",
        "status",
        "excellent",
        "auditStatus",
        "auditState",
    }
    return {
        key: str(value or "").strip()
        for key, value in raw_summary.items()
        if key in allowed_keys and str(value or "").strip()
    }


ISSUE_EXPORT_PHOTO_KEYS = {
    "issue_photo",
    "rectification_photo",
    "review_photo",
}

ISSUE_EXPORT_FIELD_KEYS = {
    "id",
    "month",
    "time",
    "region",
    "station",
    "station_manager",
    "station_manager_phone",
    "inspector",
    "inspector_phone",
    "inspection_table_name",
    "internal_standard",
    "external_standard",
    "standard_tags",
    "description",
    "is_excellent",
    "audit_result",
    "issue_photo",
    "rectification_result",
    "rectification_note",
    "rectification_at",
    "rectification_photo",
    "review_result",
    "review_note",
    "review_at",
    "review_photo",
    "status",
}


def normalize_issue_export_options(raw_options):
    if not isinstance(raw_options, dict):
        raw_options = {}
    raw_include_fields = raw_options.get("include_fields")
    has_explicit_fields = isinstance(raw_include_fields, dict)
    include_fields = {
        key: bool(raw_include_fields.get(key))
        for key in ISSUE_EXPORT_FIELD_KEYS
    } if has_explicit_fields else {
        key: True
        for key in ISSUE_EXPORT_FIELD_KEYS
    }

    raw_include_photos = raw_options.get("include_photos")
    if not isinstance(raw_include_photos, dict):
        raw_include_photos = {}
    include_photos = {
        key: bool(raw_include_photos.get(key)) and bool(include_fields.get(key))
        for key in ISSUE_EXPORT_PHOTO_KEYS
    }
    return {
        "include_fields": include_fields,
        "include_photos": include_photos,
    }


def issue_export_includes_photos(export_options):
    include_photos = (export_options or {}).get("include_photos") or {}
    return any(bool(include_photos.get(key)) for key in ISSUE_EXPORT_PHOTO_KEYS)


def selected_issue_export_field_keys(export_options):
    include_fields = (export_options or {}).get("include_fields") or {}
    return {
        key
        for key in ISSUE_EXPORT_FIELD_KEYS
        if bool(include_fields.get(key))
    }


def get_issue_export_task_for_user(cur, task_id, user):
    cur.execute(
        """
        SELECT
            task_id,
            created_by,
            status,
            selected_count,
            exported_count,
            filter_summary,
            export_options,
            file_path,
            download_filename,
            error_message,
            TO_CHAR(created_at, 'YYYY-MM-DD HH24:MI') AS created_at,
            TO_CHAR(updated_at, 'YYYY-MM-DD HH24:MI') AS updated_at,
            TO_CHAR(completed_at, 'YYYY-MM-DD HH24:MI') AS completed_at,
            TO_CHAR(expires_at, 'YYYY-MM-DD HH24:MI') AS expires_at
        FROM issue_export_tasks
        WHERE task_id = %s
        LIMIT 1;
        """,
        (task_id,),
    )
    task = cur.fetchone()
    if not task:
        raise LookupError("导出任务不存在或已过期。")
    if not is_root_user(user) and str(task["created_by"]) != str(user["id"]):
        raise PermissionError("当前账号无权查看该导出任务。")
    return task


def serialize_issue_export_task(task):
    status = task.get("status")
    task_id = task.get("task_id")
    file_size_bytes = get_storage_file_size(task.get("file_path")) if status == "completed" else 0
    return {
        "task_id": task_id,
        "status": status,
        "selected_count": int(task.get("selected_count") or 0),
        "exported_count": int(task.get("exported_count") or 0),
        "filter_summary": task.get("filter_summary") or {},
        "export_options": task.get("export_options") or {},
        "download_filename": task.get("download_filename") or "",
        "download_url": f"/api/issues/export-tasks/{task_id}/download" if status == "completed" else "",
        "file_size_bytes": file_size_bytes,
        "file_size_label": format_file_size(file_size_bytes),
        "error_message": task.get("error_message") or "",
        "created_at": task.get("created_at") or "",
        "updated_at": task.get("updated_at") or "",
        "completed_at": task.get("completed_at") or "",
        "expires_at": task.get("expires_at") or "",
    }


def fetch_issue_export_rows(cur, user, issue_ids):
    where_parts = ["i.id = ANY(%s)"]
    params = [issue_ids]

    can_view_all = can_view_all_inspection_issues(cur, user) or can_view_region_inspection_issues(cur, user)
    can_view_own = can_view_own_inspection_issues(cur, user)
    if can_view_all:
        pass
    elif can_view_own:
        if not user.get("station_id"):
            where_parts.append("COALESCE(i.inspector_id, ins.inspector_id) = %s")
            params.append(user["id"])
        else:
            where_parts.append("(i.station_id = %s OR COALESCE(i.inspector_id, ins.inspector_id) = %s)")
            params.extend([user["station_id"], user["id"]])
    else:
        where_parts.append("COALESCE(i.inspector_id, ins.inspector_id) = %s")
        params.append(user["id"])

    if not append_inspection_table_scope_filter(
        cur,
        user,
        where_parts,
        params,
        "i.inspection_table_id",
        "limit_issue_inspection_table_scope",
    ):
        return []
    if not append_station_region_scope_filter(
        cur,
        user,
        where_parts,
        params,
        "s.region",
        "limit_issue_station_region_scope",
    ):
        return []
    append_pending_audit_issue_visibility_filter(user, where_parts)
    hide_inspector_contact = should_hide_inspector_contact_info(cur, user)

    cur.execute(
        sql.SQL(
            """
            SELECT
                i.id,
                TO_CHAR(i.created_at, 'YYYY-MM') AS month,
                TO_CHAR(i.created_at, 'YYYY-MM-DD HH24:MI') AS time,
                s.region,
                s.station_name AS station,
                s.station_manager_name AS station_manager,
                s.station_manager_phone AS station_manager_phone,
                issue_inspector.real_name AS inspector,
                issue_inspector.phone AS inspector_phone,
                t.table_name AS inspection_table_name,
                i.inspection_table_id,
                i.internal_standard_id,
                i.standard_id,
                i.internal_standard_detail_text,
                i.standard_detail_text,
                i.description,
                i.photo_path AS issue_photo,
                CASE
                    WHEN COALESCE(i.is_excellent, FALSE)
                         AND COALESCE(i.audit_status, 'pending') <> 'rejected'
                    THEN '★'
                    ELSE ''
                END AS is_excellent,
                CASE
                    WHEN COALESCE(i.audit_status, 'pending') = 'approved' THEN '通过'
                    WHEN COALESCE(i.audit_status, 'pending') = 'rejected' THEN '否决'
                    ELSE ''
                END AS audit_result,
                i.rectification_result,
                i.rectification_note,
                TO_CHAR(i.rectification_at, 'YYYY-MM-DD HH24:MI') AS rectification_at,
                i.rectification_photo_path AS rectification_photo,
                i.review_result,
                i.review_note,
                TO_CHAR(i.review_at, 'YYYY-MM-DD HH24:MI') AS review_at,
                i.review_photo_path AS review_photo,
                CASE
                    WHEN COALESCE(i.audit_status, 'pending') = 'pending'
                    THEN '待审核'
                    WHEN i.status = '待整改'
                         AND NOT (
                             ins.sign_status = '已签名确认'
                             OR ins.station_manager_signed_at IS NOT NULL
                             OR COALESCE(ins.station_manager_signature_path, '') <> ''
                             OR COALESCE(ins.station_manager_signed_name, '') <> ''
                         )
                    THEN '待签名'
                    ELSE i.status
                END AS status
            FROM issues i
            JOIN inspections ins ON i.inspection_id = ins.id
            JOIN stations s ON i.station_id = s.id
            JOIN inspection_tables t ON i.inspection_table_id = t.id
            JOIN users issue_inspector ON COALESCE(i.inspector_id, ins.inspector_id) = issue_inspector.id
            WHERE {where_clause}
            ORDER BY i.id DESC;
            """
        ).format(where_clause=sql.SQL(" AND ").join(sql.SQL(part) for part in where_parts)),
        params,
    )
    rows = cur.fetchall()
    attach_internal_standard_tags_to_issue_rows(cur, rows)
    for row in rows:
        row["standard_tags"] = "；".join(
            f"{tag.get('group_name') or ''}：{tag.get('tag_name') or ''}".strip("：")
            for tag in row.get("standard_tags", [])
            if tag.get("tag_name")
        )
    if hide_inspector_contact:
        sanitized_rows = []
        for row in rows:
            data = dict(row)
            data["inspector"] = ""
            data["inspector_phone"] = ""
            sanitized_rows.append(data)
        return sanitized_rows
    return rows


ISSUE_EXPORT_BASE_COLUMNS_BEFORE_STANDARD = [
    ("ID", "id"),
    ("检查月度", "month"),
    ("检查时间", "time"),
    ("站点所属地", "region"),
    ("站点名称", "station"),
    ("站点负责人", "station_manager"),
    ("站点负责人手机号", "station_manager_phone"),
    ("检查人员", "inspector"),
    ("检查人员手机号", "inspector_phone"),
    ("检查表", "inspection_table_name"),
]

ISSUE_EXPORT_INTERNAL_STANDARD_COLUMNS = [
    ("内部规范ID", "internal_standard_id"),
    ("内部规范详情", "internal_standard_detail_text"),
]

ISSUE_EXPORT_EXTERNAL_STANDARD_ID_COLUMNS = [
    ("外部规范ID", "standard_id"),
]

ISSUE_EXPORT_BASE_COLUMNS_AFTER_STANDARD = [
    ("规范标签", "standard_tags"),
    ("问题描述", "description"),
    ("是否优秀", "is_excellent"),
    ("审核结果", "audit_result"),
    ("问题照片", "issue_photo"),
    ("站经理整改结果", "rectification_result"),
    ("站点反馈整改说明", "rectification_note"),
    ("整改时间", "rectification_at"),
    ("站点反馈整改照片", "rectification_photo"),
    ("督导组复核结果", "review_result"),
    ("督导组复核说明", "review_note"),
    ("复核时间", "review_at"),
    ("督导组复核照片", "review_photo"),
    ("问题状态", "status"),
]

ISSUE_EXPORT_LONG_TEXT_KEYS = {
    "internal_standard_detail_text",
    "standard_detail_text",
    "description",
    "rectification_note",
    "review_note",
}


def excel_column_name(index):
    name = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name


def selected_issue_export_photo_keys(export_options):
    include_photos = (export_options or {}).get("include_photos") or {}
    return {
        key
        for key in ISSUE_EXPORT_PHOTO_KEYS
        if bool(include_photos.get(key))
    }


def build_issue_export_excel_image(image_path, max_width=150, max_height=110):
    abs_path = resolve_storage_abs_path(image_path)
    if not abs_path or not os.path.isfile(abs_path):
        return None
    try:
        from openpyxl.drawing.image import Image as ExcelImage
    except ImportError as exc:
        raise RuntimeError("服务器缺少 openpyxl 组件，暂时无法导出照片。") from exc

    try:
        with Image.open(abs_path) as source_image:
            width, height = source_image.size
        if width <= 0 or height <= 0:
            return None
        scale = min(max_width / float(width), max_height / float(height), 1)
        excel_image = ExcelImage(abs_path)
        excel_image.width = max(1, int(width * scale))
        excel_image.height = max(1, int(height * scale))
        return excel_image
    except Exception:
        return None


def parse_standard_detail_entries(detail_text):
    entries = []
    normalized_text = str(detail_text or "").replace("\\n", "\n")
    for raw_line in normalized_text.split("\n"):
        line = str(raw_line or "").strip()
        if not line:
            continue
        full_width_index = line.find("：")
        raw_half_width_index = line.find(":")
        half_width_index = (
            raw_half_width_index
            if raw_half_width_index > -1
            and line[raw_half_width_index + 1: raw_half_width_index + 2] != "/"
            else -1
        )
        separator_index = full_width_index if full_width_index > -1 else half_width_index
        if separator_index >= 0:
            label = line[:separator_index].strip()
            if label and len(label) <= 48:
                entries.append(
                    {
                        "label": label,
                        "value": line[separator_index + 1:].strip() or "-",
                    }
                )
                continue
        if entries:
            entries[-1]["value"] = f"{entries[-1]['value']}\n{line}".strip()
        else:
            entries.append({"label": "规范内容", "value": line})
    return entries


def parse_standard_detail_value_map(detail_text):
    values = {}
    for entry in parse_standard_detail_entries(detail_text):
        label = entry["label"]
        value = entry["value"] or "-"
        if label in values and values[label] != value:
            values[label] = f"{values[label]}\n{value}".strip()
        else:
            values[label] = value
    return values


def build_issue_export_table_field_map(cur, rows):
    table_ids = []
    seen = set()
    for row in rows:
        table_id = row.get("inspection_table_id")
        if table_id is None or table_id in seen:
            continue
        seen.add(table_id)
        table_ids.append(table_id)

    field_map = {}
    for table_id in table_ids:
        field_map[table_id] = [
            {
                "field_key": field["field_key"],
                "field_label": field["field_label"],
                "sort_order": field.get("sort_order") or 0,
            }
            for field in get_management_checklist_fields(cur, table_id, include_public=True)
        ]
    return field_map


def safe_excel_sheet_title(raw_title, used_titles):
    base = re.sub(r"[\[\]\:\*\?\/\\]", "_", str(raw_title or "检查表").strip())
    base = re.sub(r"\s+", " ", base).strip() or "检查表"
    base = base[:31]
    title = base
    index = 2
    while title in used_titles:
        suffix = f"({index})"
        title = f"{base[:31 - len(suffix)]}{suffix}"
        index += 1
    used_titles.add(title)
    return title


def group_issue_export_rows_by_table(rows):
    grouped = OrderedDict()
    for row in rows:
        table_id = row.get("inspection_table_id")
        table_name = row.get("inspection_table_name") or "未命名检查表"
        group_key = table_id if table_id is not None else table_name
        if group_key not in grouped:
            grouped[group_key] = {
                "table_id": table_id,
                "table_name": table_name,
                "rows": [],
            }
        grouped[group_key]["rows"].append(row)
    return list(grouped.values())


def build_issue_export_standard_field_columns(group, table_field_map):
    table_id = group.get("table_id")
    configured_fields = table_field_map.get(table_id, []) if table_id is not None else []
    columns = []
    seen_labels = set()
    for field in configured_fields:
        label = str(field.get("field_label") or "").strip()
        if not label or label in seen_labels:
            continue
        columns.append({"label": label})
        seen_labels.add(label)

    # Keep old snapshot fields too. This protects exports when a checklist field
    # was later renamed or removed after issues had already been recorded.
    for row in group.get("rows", []):
        for entry in parse_standard_detail_entries(row.get("standard_detail_text")):
            label = entry["label"]
            if label and label not in seen_labels:
                columns.append({"label": label})
                seen_labels.add(label)
    return columns


def write_issue_export_xlsx(file_path, rows, export_options=None, table_field_map=None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise RuntimeError("服务器缺少 openpyxl 组件，暂时无法导出 Excel。") from exc

    table_field_map = table_field_map or {}
    selected_field_keys = selected_issue_export_field_keys(export_options)
    if not selected_field_keys:
        raise ValueError("请至少选择一个导出字段。")
    selected_photo_keys = selected_issue_export_photo_keys(export_options)
    os.makedirs(os.path.dirname(file_path), exist_ok=True)

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    header_fill = PatternFill("solid", fgColor="DCEBFF")
    header_font = Font(bold=True, color="0F172A")
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    wrap_alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    long_text_alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)

    used_sheet_titles = set()
    groups = group_issue_export_rows_by_table(rows)
    for group in groups:
        standard_field_columns = (
            build_issue_export_standard_field_columns(group, table_field_map)
            if "external_standard" in selected_field_keys
            else []
        )
        columns = []
        columns.extend(
            (header, key)
            for header, key in ISSUE_EXPORT_BASE_COLUMNS_BEFORE_STANDARD
            if key in selected_field_keys
        )
        if "internal_standard" in selected_field_keys:
            columns.extend(ISSUE_EXPORT_INTERNAL_STANDARD_COLUMNS)
        if "external_standard" in selected_field_keys:
            columns.extend(ISSUE_EXPORT_EXTERNAL_STANDARD_ID_COLUMNS)
            columns.extend(
                (field["label"], f"external_field::{field['label']}")
                for field in standard_field_columns
            )
        columns.extend(
            (header, key)
            for header, key in ISSUE_EXPORT_BASE_COLUMNS_AFTER_STANDARD
            if key in selected_field_keys
        )
        if not columns:
            raise ValueError("请至少选择一个导出字段。")
        worksheet = workbook.create_sheet(
            title=safe_excel_sheet_title(group.get("table_name"), used_sheet_titles)
        )
        worksheet.freeze_panes = "A2"

        for column_index, (header, key) in enumerate(columns, start=1):
            cell = worksheet.cell(row=1, column=column_index, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = wrap_alignment

            width = 18
            if key.startswith("external_field::"):
                width = 22
            if key in ISSUE_EXPORT_LONG_TEXT_KEYS or key in {"description", "rectification_note", "review_note"}:
                width = 34
            if key in ISSUE_EXPORT_PHOTO_KEYS:
                width = 22 if key in selected_photo_keys else 16
            worksheet.column_dimensions[excel_column_name(column_index)].width = width

        for row_index, row in enumerate(group.get("rows", []), start=2):
            row_has_image = False
            external_values = parse_standard_detail_value_map(row.get("standard_detail_text"))
            for column_index, (_header, key) in enumerate(columns, start=1):
                cell = worksheet.cell(row=row_index, column=column_index)
                cell.border = thin_border
                cell.alignment = long_text_alignment if (
                    key in ISSUE_EXPORT_LONG_TEXT_KEYS or key.startswith("external_field::")
                ) else wrap_alignment

                if key in ISSUE_EXPORT_PHOTO_KEYS:
                    cell.value = ""
                    if key in selected_photo_keys:
                        excel_image = build_issue_export_excel_image(row.get(key))
                        if excel_image:
                            worksheet.add_image(excel_image, cell.coordinate)
                            row_has_image = True
                    continue

                if key.startswith("external_field::"):
                    label = key.split("::", 1)[1]
                    cell.value = external_values.get(label, "-")
                    continue

                cell.value = "" if row.get(key) is None else str(row.get(key))

            if row_has_image:
                worksheet.row_dimensions[row_index].height = 92

        last_column = excel_column_name(len(columns))
        worksheet.auto_filter.ref = f"A1:{last_column}1"

    workbook.properties.title = "巡检问题列表导出"
    workbook.properties.creator = "业务督导中心数智管理平台"
    workbook.save(file_path)


def mark_issue_export_task_failed(task_id, message):
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_issue_export_schema(cur)
        cur.execute(
            """
            UPDATE issue_export_tasks
            SET status = 'failed',
                error_message = %s,
                updated_at = CURRENT_TIMESTAMP,
                completed_at = CURRENT_TIMESTAMP
            WHERE task_id = %s;
            """,
            (message, task_id),
        )
        conn.commit()
    finally:
        close_db_resources(cur, conn)


def run_issue_export_task(task_id, issue_ids, user_id, export_options=None):
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_issue_export_schema(cur)
        cur.execute(
            """
            UPDATE issue_export_tasks
            SET status = 'running',
                updated_at = CURRENT_TIMESTAMP
            WHERE task_id = %s;
            """,
            (task_id,),
        )
        conn.commit()
        user = get_user_by_id(cur, user_id)
        if not user:
            raise ValueError("导出用户不存在。")
        rows = fetch_issue_export_rows(cur, user, issue_ids)
        if not rows:
            raise ValueError("当前筛选结果中没有可导出的数据。")
        table_field_map = build_issue_export_table_field_map(cur, rows)

        now = beijing_now()
        filename = f"巡检问题列表_{now.strftime('%Y%m%d_%H%M%S')}_{task_id[:8]}.xlsx"
        safe_filename = secure_filename(filename) or f"inspection_issues_{task_id[:8]}.xlsx"
        if not safe_filename.lower().endswith(".xlsx"):
            safe_filename = f"inspection_issues_{task_id[:8]}.xlsx"
        abs_path = os.path.join(ISSUE_EXPORTS_STORAGE_DIR, safe_filename)
        relative_path = f"issue_exports/{safe_filename}"
        write_issue_export_xlsx(abs_path, rows, export_options, table_field_map)
        cur.execute(
            """
            UPDATE issue_export_tasks
            SET status = 'completed',
                exported_count = %s,
                file_path = %s,
                download_filename = %s,
                updated_at = CURRENT_TIMESTAMP,
                completed_at = CURRENT_TIMESTAMP
            WHERE task_id = %s;
            """,
            (len(rows), relative_path, filename, task_id),
        )
        conn.commit()
    except Exception as exc:
        if conn:
            conn.rollback()
        mark_issue_export_task_failed(task_id, str(exc))
    finally:
        close_db_resources(cur, conn)


def start_issue_export_task(task_id, issue_ids, user_id, export_options=None):
    thread = threading.Thread(
        target=run_issue_export_task,
        args=(task_id, issue_ids, user_id, export_options),
        daemon=True,
    )
    thread.start()


ATTENDANCE_MODE_META = {
    "online": {
        "label": "视频检查",
        "short_label": "视频",
        "description": "统计检查表模式为线上的巡检记录。",
    },
    "offline": {
        "label": "现场检查",
        "short_label": "现场",
        "description": "统计检查表模式为线下的巡检记录。",
    },
}


def parse_attendance_month(month_value):
    month_text = str(month_value or "").strip()
    if not month_text:
        today = beijing_today()
        month_text = today.strftime("%Y-%m")
    try:
        month_start = datetime.strptime(month_text, "%Y-%m").date()
    except ValueError as exc:
        raise ValueError("月份格式必须为 YYYY-MM。") from exc
    if month_start.month == 12:
        next_month = month_start.replace(year=month_start.year + 1, month=1)
    else:
        next_month = month_start.replace(month=month_start.month + 1)
    return month_text, month_start, next_month


def parse_reporting_date_range(month_value=None, date_from_value=None, date_to_value=None):
    date_from_text = str(date_from_value or "").strip()
    date_to_text = str(date_to_value or "").strip()

    if date_from_text or date_to_text:
        try:
            start_date = datetime.strptime(date_from_text, "%Y-%m-%d").date()
            end_date = datetime.strptime(date_to_text, "%Y-%m-%d").date()
        except ValueError as exc:
            raise ValueError("日期范围格式必须为 YYYY-MM-DD。") from exc
        if start_date > end_date:
            raise ValueError("开始日期不能晚于结束日期。")
        period_key = (
            start_date.strftime("%Y-%m")
            if start_date.day == 1
            and (end_date + timedelta(days=1)).day == 1
            and start_date.year == end_date.year
            and start_date.month == end_date.month
            else f"{start_date.isoformat()}_{end_date.isoformat()}"
        )
        return period_key, start_date, end_date + timedelta(days=1)

    return parse_attendance_month(month_value)


def normalize_attendance_mode(mode_value):
    mode = str(mode_value or "all").strip().lower()
    if mode not in {"all", "online", "offline"}:
        raise ValueError("检查方式参数不合法。")
    return mode


def append_unique_item(target_map, key, item):
    if key is None or key == "":
        return
    target_map.setdefault(str(key), item)


def sorted_dict_items(item_map, name_key="name"):
    return sorted(
        item_map.values(),
        key=lambda item: str(item.get(name_key) or item.get("date") or item.get("id") or ""),
    )


def build_attendance_payload(rows, month, mode_filter):
    modes = ["online", "offline"] if mode_filter == "all" else [mode_filter]
    buckets = {
        mode: {
            "mode": mode,
            **ATTENDANCE_MODE_META[mode],
            "people_map": {},
            "groups_map": {},
            "station_ids": set(),
            "inspection_ids": set(),
            "checklist_ids": set(),
        }
        for mode in modes
    }

    for row in rows:
        mode = row.get("checklist_mode") or "online"
        if mode not in buckets:
            continue

        bucket = buckets[mode]
        inspector_id = row.get("inspector_id")
        inspection_date = row.get("inspection_date")
        station_id = row.get("station_id")
        table_id = row.get("inspection_table_id")
        inspection_id = row.get("inspection_id")

        if not inspector_id or not inspection_date:
            continue

        inspector_name = (
            row.get("inspector_name")
            or row.get("inspector_username")
            or row.get("inspector_phone")
            or str(inspector_id)
        )
        person_key = str(inspector_id)
        person = bucket["people_map"].setdefault(
            person_key,
            {
                "inspector_id": inspector_id,
                "inspector_name": inspector_name,
                "username": row.get("inspector_username") or "",
                "phone": row.get("inspector_phone") or "",
                "issue_count": 0,
                "approved_issue_count": 0,
                "attendance_dates": set(),
                "inspection_ids": set(),
                "stations_map": {},
                "checklists_map": {},
                "activity_days": {},
            },
        )
        person["issue_count"] = max(
            int(person.get("issue_count") or 0),
            int(row.get("issue_count") or 0),
        )
        person["approved_issue_count"] = max(
            int(person.get("approved_issue_count") or 0),
            int(row.get("approved_issue_count") or 0),
        )
        person["attendance_dates"].add(inspection_date)
        person["inspection_ids"].add(inspection_id)
        append_unique_item(
            person["stations_map"],
            station_id,
            {
                "id": station_id,
                "name": row.get("station_name") or "未命名站点",
                "region": row.get("station_region") or "",
            },
        )
        append_unique_item(
            person["checklists_map"],
            table_id,
            {
                "id": table_id,
                "name": row.get("inspection_table_name") or "未命名检查表",
            },
        )
        activity = person["activity_days"].setdefault(
            inspection_date,
            {
                "date": inspection_date,
                "stations_map": {},
                "checklists_map": {},
            },
        )
        append_unique_item(
            activity["stations_map"],
            station_id,
            {
                "id": station_id,
                "name": row.get("station_name") or "未命名站点",
                "region": row.get("station_region") or "",
            },
        )
        append_unique_item(
            activity["checklists_map"],
            table_id,
            {
                "id": table_id,
                "name": row.get("inspection_table_name") or "未命名检查表",
            },
        )

        group_key = f"{inspection_date}:{station_id}"
        group = bucket["groups_map"].setdefault(
            group_key,
            {
                "date": inspection_date,
                "station_id": station_id,
                "station_name": row.get("station_name") or "未命名站点",
                "station_region": row.get("station_region") or "",
                "inspectors_map": {},
                "checklists_map": {},
                "inspection_ids": set(),
            },
        )
        append_unique_item(
            group["inspectors_map"],
            inspector_id,
            {
                "id": inspector_id,
                "name": inspector_name,
                "username": row.get("inspector_username") or "",
                "phone": row.get("inspector_phone") or "",
            },
        )
        append_unique_item(
            group["checklists_map"],
            table_id,
            {
                "id": table_id,
                "name": row.get("inspection_table_name") or "未命名检查表",
            },
        )
        group["inspection_ids"].add(inspection_id)

        bucket["station_ids"].add(station_id)
        bucket["inspection_ids"].add(inspection_id)
        bucket["checklist_ids"].add(table_id)

    mode_payloads = []
    for mode in modes:
        bucket = buckets[mode]
        people = []
        for person in bucket["people_map"].values():
            activity_days = []
            for activity in sorted(person["activity_days"].values(), key=lambda item: item["date"]):
                activity_days.append(
                    {
                        "date": activity["date"],
                        "stations": sorted_dict_items(activity["stations_map"]),
                        "checklists": sorted_dict_items(activity["checklists_map"]),
                    }
                )
            people.append(
                {
                    "inspector_id": person["inspector_id"],
                    "inspector_name": person["inspector_name"],
                    "username": person["username"],
                    "phone": person["phone"],
                    "attendance_days": len(person["attendance_dates"]),
                    "inspection_count": len(person["inspection_ids"]),
                    "issue_count": int(person.get("issue_count") or 0),
                    "approved_issue_count": int(person.get("approved_issue_count") or 0),
                    "station_count": len(person["stations_map"]),
                    "checklist_count": len(person["checklists_map"]),
                    "attendance_dates": sorted(person["attendance_dates"]),
                    "stations": sorted_dict_items(person["stations_map"]),
                    "checklists": sorted_dict_items(person["checklists_map"]),
                    "activity_days": activity_days,
                }
            )
        people.sort(
            key=lambda item: (
                -item["attendance_days"],
                -item["inspection_count"],
                item["inspector_name"],
            )
        )

        groups = []
        for group in bucket["groups_map"].values():
            inspectors = sorted_dict_items(group["inspectors_map"])
            checklists = sorted_dict_items(group["checklists_map"])
            groups.append(
                {
                    "date": group["date"],
                    "station_id": group["station_id"],
                    "station_name": group["station_name"],
                    "station_region": group["station_region"],
                    "inspectors": inspectors,
                    "checklists": checklists,
                    "inspector_count": len(inspectors),
                    "checklist_count": len(checklists),
                    "inspection_count": len(group["inspection_ids"]),
                }
            )
        groups.sort(key=lambda item: (item["date"], item["station_region"], item["station_name"]))

        mode_payloads.append(
            {
                "mode": bucket["mode"],
                "label": bucket["label"],
                "short_label": bucket["short_label"],
                "description": bucket["description"],
                "summary": {
                    "inspector_count": len(people),
                    "attendance_person_days": sum(item["attendance_days"] for item in people),
                    "group_count": len(groups),
                    "station_count": len([item for item in bucket["station_ids"] if item]),
                    "inspection_count": len([item for item in bucket["inspection_ids"] if item]),
                    "checklist_count": len([item for item in bucket["checklist_ids"] if item]),
                },
                "people": people,
                "groups": groups,
            }
        )

    return {
        "month": month,
        "mode": mode_filter,
        "modes": mode_payloads,
    }


def ensure_station_score_schema(cur):
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS station_score_adjustments (
            id SERIAL PRIMARY KEY,
            station_id INTEGER NOT NULL REFERENCES stations(id) ON DELETE CASCADE,
            inspection_table_id INTEGER NOT NULL REFERENCES inspection_tables(id) ON DELETE CASCADE,
            standard_id BIGINT NOT NULL,
            score_month TEXT NOT NULL,
            manual_score NUMERIC(8, 2) NOT NULL,
            note TEXT,
            adjusted_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
            adjusted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE (station_id, inspection_table_id, standard_id, score_month)
        );
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_station_score_adjustments_month
        ON station_score_adjustments (score_month);
        """
    )


def round_score(value):
    try:
        return round(float(value or 0), 2)
    except (TypeError, ValueError):
        return 0.0


def fetch_station_for_score(cur, station_id):
    cur.execute(
        """
        SELECT id, station_name, region, hos_station_code
        FROM stations
        WHERE id = %s;
        """,
        (station_id,),
    )
    return cur.fetchone()


def fetch_scorable_checklist_rows(cur):
    cur.execute(
        """
        SELECT DISTINCT
            t.id,
            t.table_code,
            t.table_name,
            t.checklist_mode,
            t.standard_id_base,
            t.description,
            t.is_active
        FROM inspection_tables t
        JOIN inspection_table_fields f ON f.inspection_table_id = t.id
        WHERE t.is_active = TRUE
          AND COALESCE(f.is_scorable, FALSE) = TRUE
        ORDER BY t.id ASC;
        """
    )
    return [dict(row) for row in cur.fetchall()]


def normalize_station_score_table_ids(raw_table_ids):
    if raw_table_ids in (None, "", []):
        return []
    if not isinstance(raw_table_ids, list):
        raw_table_ids = [raw_table_ids]
    normalized = []
    seen = set()
    for raw_id in raw_table_ids:
        try:
            table_id = int(raw_id)
        except (TypeError, ValueError):
            continue
        if table_id <= 0 or table_id in seen:
            continue
        seen.add(table_id)
        normalized.append(table_id)
    return normalized


def fetch_station_score_issue_map(cur, station_id, month_start, next_month, table_ids):
    if not table_ids:
        return {}
    cur.execute(
        """
        SELECT
            picked.id,
            picked.inspection_table_id,
            picked.standard_id,
            picked.internal_standard_id,
            picked.internal_standard_detail_text,
            picked.description,
            picked.issue_photo,
            picked.status,
            picked.audit_status,
            picked.created_at,
            picked.inspection_date,
            picked.inspector_name
        FROM (
            SELECT
                i.id,
                i.inspection_table_id,
                i.standard_id,
                i.internal_standard_id,
                i.internal_standard_detail_text,
                i.description,
                i.photo_path AS issue_photo,
                i.status,
                COALESCE(i.audit_status, 'pending') AS audit_status,
                TO_CHAR(i.created_at, 'YYYY-MM-DD HH24:MI') AS created_at,
                TO_CHAR(ins.inspection_date, 'YYYY-MM-DD') AS inspection_date,
                COALESCE(issue_inspector.real_name, issue_inspector.username, '') AS inspector_name,
                ROW_NUMBER() OVER (
                    PARTITION BY i.inspection_table_id, i.standard_id
                    ORDER BY RANDOM()
                ) AS row_number
            FROM issues i
            JOIN inspections ins ON ins.id = i.inspection_id
            LEFT JOIN users issue_inspector ON issue_inspector.id = COALESCE(i.inspector_id, ins.inspector_id)
            WHERE i.station_id = %s
              AND ins.inspection_date >= %s
              AND ins.inspection_date < %s
              AND i.inspection_table_id = ANY(%s::int[])
              AND COALESCE(i.audit_status, 'pending') = 'approved'
        ) picked
        WHERE picked.row_number = 1
        ORDER BY picked.inspection_table_id ASC, picked.standard_id ASC, picked.id ASC;
        """,
        (station_id, month_start, next_month, table_ids),
    )
    issue_map = {}
    for row in cur.fetchall():
        key = (int(row["inspection_table_id"]), int(row["standard_id"]))
        issue_map.setdefault(key, []).append(dict(row))
    return issue_map


def fetch_station_score_adjustment_map(cur, station_id, month):
    cur.execute(
        """
        SELECT
            a.id,
            a.station_id,
            a.inspection_table_id,
            a.standard_id,
            a.score_month,
            a.manual_score,
            a.note,
            TO_CHAR(a.adjusted_at, 'YYYY-MM-DD HH24:MI') AS adjusted_at,
            COALESCE(u.real_name, u.username, '') AS adjusted_by_name
        FROM station_score_adjustments a
        LEFT JOIN users u ON u.id = a.adjusted_by
        WHERE a.station_id = %s
          AND a.score_month = %s;
        """,
        (station_id, month),
    )
    result = {}
    for row in cur.fetchall():
        result[(int(row["inspection_table_id"]), int(row["standard_id"]))] = dict(row)
    return result


def build_station_score_payload(
    cur,
    station_id,
    month,
    month_start,
    next_month,
    user,
    inspection_table_id=None,
    inspection_table_ids=None,
):
    station = fetch_station_for_score(cur, station_id)
    if not station:
        raise LookupError("站点不存在。")

    selected_table_ids = set(normalize_station_score_table_ids(inspection_table_ids))
    if inspection_table_id:
        selected_table_ids.add(int(inspection_table_id))

    checklists = fetch_scorable_checklist_rows(cur)
    if selected_table_ids:
        checklists = [
            checklist
            for checklist in checklists
            if int(checklist.get("id") or 0) in selected_table_ids
        ]
    table_ids = [int(item["id"]) for item in checklists]
    issue_map = fetch_station_score_issue_map(cur, station_id, month_start, next_month, table_ids)
    adjustment_map = fetch_station_score_adjustment_map(cur, station_id, month)

    table_payloads = []
    summary = {
        "table_count": 0,
        "item_count": 0,
        "issue_count": 0,
        "deducted_item_count": 0,
        "adjusted_item_count": 0,
        "max_score": 0.0,
        "auto_score": 0.0,
        "final_score": 0.0,
    }

    for checklist in checklists:
        fields = [dict(field) for field in get_management_checklist_fields(cur, checklist["id"], include_public=True)]
        scorable_fields = [field for field in fields if normalize_boolean_flag(field.get("is_scorable"), False)]
        if not scorable_fields:
            continue
        physical_table_name = get_physical_table_name_by_code(checklist["table_code"])
        if not physical_table_name:
            continue
        ensure_checklist_field_columns(cur, physical_table_name, fields)
        standard_rows = fetch_checklist_standard_rows(cur, physical_table_name, fields)
        item_count = len(standard_rows)
        max_score_per_item = 100 / item_count if item_count else 0

        standards = []
        table_issue_count = 0
        table_deducted_count = 0
        table_adjusted_count = 0
        table_auto_score = 0.0
        table_final_score = 0.0
        field_meta = [(field["field_key"], field["field_label"]) for field in fields]

        for row in standard_rows:
            standard_id = int(row["standard_id"])
            key = (int(checklist["id"]), standard_id)
            issues = issue_map.get(key, [])
            adjustment = adjustment_map.get(key)
            auto_score = 0.0 if issues else max_score_per_item
            final_score = float(adjustment["manual_score"]) if adjustment else auto_score
            score_fields = [
                {
                    "field_key": field["field_key"],
                    "field_label": field["field_label"],
                    "value": str(row.get(field["field_key"]) or "").strip() or "-",
                }
                for field in scorable_fields
            ]
            all_fields = [
                {
                    "field_key": field["field_key"],
                    "field_label": field["field_label"],
                    "value": str(row.get(field["field_key"]) or "").strip() or "-",
                }
                for field in fields
            ]
            standards.append(
                {
                    "standard_id": standard_id,
                    "detail_text": build_standard_detail_text(field_meta, row),
                    "score_fields": score_fields,
                    "all_fields": all_fields,
                    "issues": issues,
                    "issue_count": len(issues),
                    "is_deducted": bool(issues),
                    "max_score": round_score(max_score_per_item),
                    "auto_score": round_score(auto_score),
                    "final_score": round_score(final_score),
                    "has_manual_adjustment": bool(adjustment),
                    "adjustment_note": adjustment.get("note") if adjustment else "",
                    "adjusted_by_name": adjustment.get("adjusted_by_name") if adjustment else "",
                    "adjusted_at": adjustment.get("adjusted_at") if adjustment else "",
                }
            )
            table_issue_count += len(issues)
            table_deducted_count += 1 if issues else 0
            table_adjusted_count += 1 if adjustment else 0
            table_auto_score += auto_score
            table_final_score += final_score

        table_payloads.append(
            {
                "id": checklist["id"],
                "table_code": checklist["table_code"],
                "table_name": checklist["table_name"],
                "checklist_mode": normalize_checklist_mode(checklist.get("checklist_mode")),
                "checklist_mode_label": "视频检查" if normalize_checklist_mode(checklist.get("checklist_mode")) == "online" else "现场检查",
                "scorable_fields": [
                    {
                        "field_key": field["field_key"],
                        "field_label": field["field_label"],
                    }
                    for field in scorable_fields
                ],
                "item_count": item_count,
                "issue_count": table_issue_count,
                "deducted_item_count": table_deducted_count,
                "adjusted_item_count": table_adjusted_count,
                "max_score": 100.0 if item_count else 0.0,
                "auto_score": round_score(table_auto_score),
                "final_score": round_score(table_final_score),
                "standards": standards,
            }
        )

        summary["table_count"] += 1
        summary["item_count"] += item_count
        summary["issue_count"] += table_issue_count
        summary["deducted_item_count"] += table_deducted_count
        summary["adjusted_item_count"] += table_adjusted_count
        summary["max_score"] += 100.0 if item_count else 0.0
        summary["auto_score"] += table_auto_score
        summary["final_score"] += table_final_score

    summary["auto_score"] = round_score(summary["auto_score"])
    summary["final_score"] = round_score(summary["final_score"])
    summary["max_score"] = round_score(summary["max_score"])

    return {
        "station": dict(station),
        "month": month,
        "can_adjust": can_adjust_station_scores(cur, user),
        "summary": summary,
        "tables": table_payloads,
    }


def safe_excel_sheet_title(value, fallback="站点评分"):
    title = re.sub(r"[\[\]\*:/\\?]", "", str(value or "").strip())[:31]
    return title or fallback


def normalize_station_score_export_options(raw_options):
    if not isinstance(raw_options, dict):
        raw_options = {}
    include_photos = raw_options.get("include_photos")
    if not isinstance(include_photos, dict):
        include_photos = {}
    return {
        "include_photos": {
            "issue_photo": bool(include_photos.get("issue_photo")),
        }
    }


def station_score_export_includes_issue_photos(export_options):
    include_photos = (export_options or {}).get("include_photos") or {}
    return bool(include_photos.get("issue_photo"))


def write_station_score_export_xlsx(score_payload, export_options=None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise RuntimeError("服务器缺少 openpyxl 组件，暂时无法导出站点评分。") from exc

    output = BytesIO()
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    include_issue_photos = station_score_export_includes_issue_photos(export_options)

    header_fill = PatternFill("solid", fgColor="E7F0E7")
    score_fill = PatternFill("solid", fgColor="F4F8EF")
    deducted_fill = PatternFill("solid", fgColor="FFF0EA")
    adjusted_fill = PatternFill("solid", fgColor="EAF6EF")
    header_font = Font(bold=True, color="172033")
    title_font = Font(bold=True, size=14, color="173D2C")
    thin_border = Border(
        left=Side(style="thin", color="CAD6CF"),
        right=Side(style="thin", color="CAD6CF"),
        top=Side(style="thin", color="CAD6CF"),
        bottom=Side(style="thin", color="CAD6CF"),
    )
    center = Alignment(vertical="center", horizontal="center", wrap_text=True)
    left = Alignment(vertical="center", horizontal="left", wrap_text=True)

    with tempfile.TemporaryDirectory() as temp_dir:
        for table in score_payload.get("tables") or []:
            sheet = workbook.create_sheet(safe_excel_sheet_title(table.get("table_name")))
            fields = table.get("standards", [{}])[0].get("all_fields", []) if table.get("standards") else []
            headers = ["外部规范ID", *[field.get("field_label") for field in fields], "问题描述", "问题照片", "评分"]
            last_column = excel_column_name(len(headers))

            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            title_cell = sheet.cell(row=1, column=1)
            title_cell.value = (
                f"{score_payload.get('station', {}).get('station_name', '')}"
                f"｜{score_payload.get('month', '')}"
                f"｜{table.get('table_name', '')}"
                f"｜得分 {round_score(table.get('final_score'))}/{round_score(table.get('max_score'))}"
            )
            title_cell.font = title_font
            title_cell.alignment = left

            for column_index, header in enumerate(headers, start=1):
                cell = sheet.cell(row=2, column=column_index, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = center
                if header == "问题照片":
                    sheet.column_dimensions[excel_column_name(column_index)].width = 26
                elif header in {"问题描述"}:
                    sheet.column_dimensions[excel_column_name(column_index)].width = 36
                elif header == "评分":
                    sheet.column_dimensions[excel_column_name(column_index)].width = 22
                else:
                    sheet.column_dimensions[excel_column_name(column_index)].width = 18

            photo_column_index = len(headers) - 1
            score_column_index = len(headers)
            for row_index, standard in enumerate(table.get("standards") or [], start=3):
                row_fill = deducted_fill if standard.get("is_deducted") else None
                if standard.get("has_manual_adjustment"):
                    row_fill = adjusted_fill

                values = [standard.get("standard_id")]
                values.extend(field.get("value") for field in standard.get("all_fields") or [])
                issue_text = "\n".join(
                    f"#{issue.get('id')} {issue.get('description') or '未填写问题描述'}"
                    for issue in standard.get("issues") or []
                )
                score_text = (
                    f"{round_score(standard.get('final_score'))}/{round_score(standard.get('max_score'))}"
                    f"\n自动：{round_score(standard.get('auto_score'))}"
                )
                if standard.get("has_manual_adjustment"):
                    score_text += (
                        f"\n人工调整：{standard.get('adjusted_by_name') or '-'}"
                        f"\n{standard.get('adjusted_at') or ''}"
                    )
                values.extend([issue_text, "", score_text])

                for column_index, value in enumerate(values, start=1):
                    cell = sheet.cell(row=row_index, column=column_index, value=value)
                    cell.border = thin_border
                    cell.alignment = left if column_index not in {1, score_column_index} else center
                    if row_fill:
                        cell.fill = row_fill
                    if column_index == score_column_index:
                        cell.fill = score_fill

                excel_image = None
                if include_issue_photos:
                    photo_path = next(
                        (
                            issue.get("issue_photo")
                            for issue in standard.get("issues") or []
                            if issue.get("issue_photo")
                        ),
                        None,
                    )
                    excel_image = build_issue_export_excel_image(
                        photo_path,
                        max_width=190,
                        max_height=120,
                    )
                if excel_image:
                    sheet.add_image(excel_image, sheet.cell(row=row_index, column=photo_column_index).coordinate)
                    sheet.row_dimensions[row_index].height = 92
                elif len(issue_text) > 80:
                    sheet.row_dimensions[row_index].height = 60

            sheet.freeze_panes = "A3"
            sheet.auto_filter.ref = f"A2:{last_column}2"

        if not workbook.sheetnames:
            sheet = workbook.create_sheet("站点评分")
            sheet["A1"] = "暂无可导出的站点评分数据"

        workbook.properties.title = "站点评分导出"
        workbook.properties.creator = "业务督导中心数智管理平台"
        workbook.save(output)

    output.seek(0)
    return output


def safe_zip_path_part(value, fallback):
    text = str(value or "").strip() or fallback
    text = re.sub(r"[\x00-\x1f<>:\"/\\|?*]+", "_", text)
    text = re.sub(r"\s+", " ", text).strip(" .")
    return text or fallback


def write_station_score_all_export_zip(score_payloads, month, export_options=None):
    output = BytesIO()
    used_paths = set()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        for payload in score_payloads:
            station = payload.get("station") or {}
            region_name = safe_zip_path_part(station.get("region"), "未设置片区")
            station_name = safe_zip_path_part(station.get("station_name"), f"站点{station.get('id') or ''}")
            hos_code = safe_zip_path_part(station.get("hos_station_code"), "")
            filename_parts = [station_name]
            if hos_code:
                filename_parts.append(hos_code)
            filename_parts.append(month)
            filename = "_".join(filename_parts) + ".xlsx"
            zip_path = f"{region_name}/{filename}"
            if zip_path in used_paths:
                zip_path = f"{region_name}/{station_name}_{station.get('id') or uuid.uuid4().hex[:6]}_{month}.xlsx"
            used_paths.add(zip_path)
            workbook_stream = write_station_score_export_xlsx(payload, export_options)
            archive.writestr(zip_path, workbook_stream.getvalue())
    output.seek(0)
    return output


def ensure_station_score_export_schema(cur):
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS station_score_export_tasks (
            task_id TEXT PRIMARY KEY,
            created_by INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            status TEXT NOT NULL DEFAULT 'pending',
            export_mode TEXT NOT NULL DEFAULT 'single',
            selected_count INTEGER NOT NULL DEFAULT 0,
            exported_count INTEGER NOT NULL DEFAULT 0,
            filter_summary JSONB NOT NULL DEFAULT '{}'::jsonb,
            export_options JSONB NOT NULL DEFAULT '{}'::jsonb,
            file_path TEXT,
            download_filename TEXT,
            error_message TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            completed_at TIMESTAMP,
            expires_at TIMESTAMP NOT NULL
        );
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_station_score_export_tasks_created_by
        ON station_score_export_tasks (created_by, created_at DESC);
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_station_score_export_tasks_expires_at
        ON station_score_export_tasks (expires_at);
        """
    )


def cleanup_expired_station_score_exports(cur):
    ensure_station_score_export_schema(cur)
    cur.execute(
        """
        SELECT file_path
        FROM station_score_export_tasks
        WHERE expires_at < CURRENT_TIMESTAMP;
        """
    )
    for row in cur.fetchall():
        if row.get("file_path"):
            remove_storage_file(row["file_path"])
    cur.execute("DELETE FROM station_score_export_tasks WHERE expires_at < CURRENT_TIMESTAMP;")

    os.makedirs(STATION_SCORE_EXPORTS_STORAGE_DIR, exist_ok=True)
    cutoff = time.time() - ISSUE_EXPORT_RETENTION_DAYS * 24 * 60 * 60
    storage_dir_abs = os.path.abspath(STATION_SCORE_EXPORTS_STORAGE_DIR)
    for name in os.listdir(STATION_SCORE_EXPORTS_STORAGE_DIR):
        path = os.path.abspath(os.path.join(STATION_SCORE_EXPORTS_STORAGE_DIR, name))
        if os.path.commonpath([storage_dir_abs, path]) != storage_dir_abs:
            continue
        if os.path.isfile(path) and path.lower().endswith((".xlsx", ".zip")) and os.path.getmtime(path) < cutoff:
            try:
                os.remove(path)
            except OSError:
                pass


def maybe_cleanup_expired_station_score_exports():
    global station_score_export_cleanup_last_run
    now = time.time()
    if now - station_score_export_cleanup_last_run < ISSUE_EXPORT_CLEANUP_INTERVAL_SECONDS:
        return
    if not station_score_export_cleanup_lock.acquire(blocking=False):
        return

    conn = None
    cur = None
    try:
        if now - station_score_export_cleanup_last_run < ISSUE_EXPORT_CLEANUP_INTERVAL_SECONDS:
            return
        conn = get_db_connection()
        cur = conn.cursor()
        cleanup_expired_station_score_exports(cur)
        conn.commit()
        station_score_export_cleanup_last_run = now
    except Exception:
        if conn:
            conn.rollback()
    finally:
        close_db_resources(cur, conn)
        station_score_export_cleanup_lock.release()


def station_score_export_download_url(task_id, status):
    return (
        f"/api/assessment/station-scores/export-tasks/{task_id}/download"
        if status == "completed"
        else ""
    )


def get_station_score_export_task_for_user(cur, task_id, user):
    cur.execute(
        """
        SELECT
            task_id,
            created_by,
            status,
            export_mode,
            selected_count,
            exported_count,
            filter_summary,
            export_options,
            file_path,
            download_filename,
            error_message,
            TO_CHAR(created_at, 'YYYY-MM-DD HH24:MI') AS created_at,
            TO_CHAR(updated_at, 'YYYY-MM-DD HH24:MI') AS updated_at,
            TO_CHAR(completed_at, 'YYYY-MM-DD HH24:MI') AS completed_at,
            TO_CHAR(expires_at, 'YYYY-MM-DD HH24:MI') AS expires_at
        FROM station_score_export_tasks
        WHERE task_id = %s
        LIMIT 1;
        """,
        (task_id,),
    )
    task = cur.fetchone()
    if not task:
        raise LookupError("导出任务不存在或已过期。")
    if not is_root_user(user) and str(task["created_by"]) != str(user["id"]):
        raise PermissionError("当前账号无权查看该导出任务。")
    return task


def serialize_station_score_export_task(task):
    status = task.get("status")
    task_id = task.get("task_id")
    file_size_bytes = get_storage_file_size(task.get("file_path")) if status == "completed" else 0
    return {
        "task_id": task_id,
        "status": status,
        "export_mode": task.get("export_mode") or "single",
        "selected_count": int(task.get("selected_count") or 0),
        "exported_count": int(task.get("exported_count") or 0),
        "filter_summary": task.get("filter_summary") or {},
        "export_options": task.get("export_options") or {},
        "download_filename": task.get("download_filename") or "",
        "download_url": station_score_export_download_url(task_id, status),
        "file_size_bytes": file_size_bytes,
        "file_size_label": format_file_size(file_size_bytes),
        "error_message": task.get("error_message") or "",
        "created_at": task.get("created_at") or "",
        "updated_at": task.get("updated_at") or "",
        "completed_at": task.get("completed_at") or "",
        "expires_at": task.get("expires_at") or "",
    }


def build_station_score_export_filter_summary(
    mode,
    month,
    station=None,
    table=None,
    station_count=0,
    selected_tables=None,
):
    summary = {
        "mode": mode,
        "month": month,
    }
    if station:
        summary["station"] = station.get("station_name") or ""
        summary["region"] = station.get("region") or ""
        summary["hos_station_code"] = station.get("hos_station_code") or ""
    if table:
        summary["inspection_table"] = table.get("table_name") or ""
        summary["checklist_mode"] = table.get("checklist_mode_label") or ""
    if selected_tables is not None:
        summary["inspection_tables"] = [
            table.get("table_name") or ""
            for table in selected_tables
            if table.get("table_name")
        ]
        summary["inspection_table_count"] = len(selected_tables)
    if station_count:
        summary["station_count"] = station_count
    return summary


def write_station_score_export_stream_to_file(stream, abs_path):
    os.makedirs(os.path.dirname(abs_path), exist_ok=True)
    with open(abs_path, "wb") as target:
        target.write(stream.getvalue())


def mark_station_score_export_task_failed(task_id, message):
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_station_score_export_schema(cur)
        cur.execute(
            """
            UPDATE station_score_export_tasks
            SET status = 'failed',
                error_message = %s,
                updated_at = CURRENT_TIMESTAMP,
                completed_at = CURRENT_TIMESTAMP
            WHERE task_id = %s;
            """,
            (message, task_id),
        )
        conn.commit()
    finally:
        close_db_resources(cur, conn)


def run_station_score_export_task(
    task_id,
    user_id,
    mode,
    month,
    month_start,
    next_month,
    export_options=None,
    station_id=None,
    inspection_table_id=None,
    inspection_table_ids=None,
):
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_inspection_checklist_management_schema(cur)
        ensure_issue_inspector_schema(cur)
        ensure_station_score_schema(cur)
        ensure_station_score_export_schema(cur)
        cur.execute(
            """
            UPDATE station_score_export_tasks
            SET status = 'running',
                updated_at = CURRENT_TIMESTAMP
            WHERE task_id = %s;
            """,
            (task_id,),
        )
        conn.commit()

        user = get_user_by_id(cur, user_id)
        if not user:
            raise ValueError("导出用户不存在。")
        if not can_adjust_station_scores(cur, user):
            raise PermissionError("当前账号无权导出站点评分。")

        now = beijing_now()
        if mode == "all":
            cur.execute(
                """
                SELECT id
                FROM stations
                ORDER BY region NULLS LAST, station_name ASC, id ASC;
                """,
            )
            station_ids = [int(row["id"]) for row in cur.fetchall()]
            if not station_ids:
                raise ValueError("当前没有可导出的站点。")
            payloads = [
                build_station_score_payload(
                    cur,
                    item_id,
                    month,
                    month_start,
                    next_month,
                    user,
                    inspection_table_ids=inspection_table_ids,
                )
                for item_id in station_ids
            ]
            stream = write_station_score_all_export_zip(payloads, month, export_options)
            download_filename = f"站点评分全部站点_{month}_{now.strftime('%Y%m%d_%H%M%S')}.zip"
            safe_filename = secure_filename(download_filename) or f"station_scores_all_{task_id[:8]}.zip"
            if not safe_filename.lower().endswith(".zip"):
                safe_filename = f"{safe_filename}.zip"
            exported_count = len(station_ids)
        else:
            payload = build_station_score_payload(
                cur,
                station_id,
                month,
                month_start,
                next_month,
                user,
                inspection_table_id=inspection_table_id,
            )
            if not payload.get("tables"):
                raise ValueError("当前站点和检查表暂无可导出的评分数据。")
            table = payload["tables"][0]
            stream = write_station_score_export_xlsx(payload, export_options)
            station_name = payload.get("station", {}).get("station_name") or "站点"
            table_name = table.get("table_name") or "检查表"
            download_filename = (
                f"站点评分_{station_name}_{table_name}_{month}_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            safe_filename = secure_filename(download_filename) or f"station_score_{task_id[:8]}.xlsx"
            if not safe_filename.lower().endswith(".xlsx"):
                safe_filename = f"{safe_filename}.xlsx"
            exported_count = int(table.get("item_count") or 0)

        abs_path = os.path.join(STATION_SCORE_EXPORTS_STORAGE_DIR, safe_filename)
        relative_path = f"station_score_exports/{safe_filename}"
        write_station_score_export_stream_to_file(stream, abs_path)
        cur.execute(
            """
            UPDATE station_score_export_tasks
            SET status = 'completed',
                exported_count = %s,
                file_path = %s,
                download_filename = %s,
                updated_at = CURRENT_TIMESTAMP,
                completed_at = CURRENT_TIMESTAMP
            WHERE task_id = %s;
            """,
            (exported_count, relative_path, download_filename, task_id),
        )
        conn.commit()
    except Exception as exc:
        if conn:
            conn.rollback()
        mark_station_score_export_task_failed(task_id, str(exc))
    finally:
        close_db_resources(cur, conn)


def start_station_score_export_task(
    task_id,
    user_id,
    mode,
    month,
    month_start,
    next_month,
    export_options=None,
    station_id=None,
    inspection_table_id=None,
    inspection_table_ids=None,
):
    thread = threading.Thread(
        target=run_station_score_export_task,
        args=(
            task_id,
            user_id,
            mode,
            month,
            month_start,
            next_month,
            export_options,
            station_id,
            inspection_table_id,
            inspection_table_ids,
        ),
        daemon=True,
    )
    thread.start()


def fetch_station_score_standard_context(
    cur,
    station_id,
    month,
    inspection_table_id,
    standard_id,
    date_from=None,
    date_to=None,
):
    month_text, month_start, next_month = parse_reporting_date_range(month, date_from, date_to)
    station = fetch_station_for_score(cur, station_id)
    if not station:
        raise LookupError("站点不存在。")
    checklist = fetch_management_checklist(cur, inspection_table_id)
    if not checklist or not checklist.get("is_active"):
        raise LookupError("检查表不存在或未启用。")
    fields = [dict(field) for field in get_management_checklist_fields(cur, inspection_table_id, include_public=True)]
    if not any(normalize_boolean_flag(field.get("is_scorable"), False) for field in fields):
        raise ValueError("该检查表未配置可评分字段。")
    physical_table_name = get_physical_table_name_by_code(checklist["table_code"])
    if not physical_table_name or not checklist_physical_table_exists(cur, physical_table_name):
        raise ValueError("检查表规范数据不存在。")
    ensure_checklist_field_columns(cur, physical_table_name, fields)
    standard = fetch_standard_from_table(cur, physical_table_name, standard_id)
    if not standard:
        raise LookupError("外部规范不存在。")
    item_count = get_checklist_row_count(cur, physical_table_name)
    if item_count <= 0:
        raise ValueError("该检查表暂无规范数据，不能调整评分。")
    return {
        "month": month_text,
        "month_start": month_start,
        "next_month": next_month,
        "station": station,
        "checklist": checklist,
        "standard": standard,
        "item_count": item_count,
        "max_score": 100 / item_count,
    }


def normalize_optional_issue_result(value, field_label):
    normalized = canonical_issue_result(value)
    if not normalized:
        return None
    if normalized not in ISSUE_RESULT_OPTIONS:
        raise ValueError(f"{field_label}参数不合法。")
    return normalized


def can_view_all_inspection_records(cur, user):
    return bool(get_effective_permissions(cur, user).get("view_all_inspection_records"))


def can_view_region_inspection_records(cur, user):
    return bool(get_effective_permissions(cur, user).get("limit_record_station_region_scope"))


def can_view_own_inspection_records(cur, user):
    return bool(get_effective_permissions(cur, user).get("view_own_inspection_records"))


def can_delete_inspection_records(cur, user):
    return has_permission(cur, user, "delete_inspection_records")


def can_reset_inspection_signature(cur, user):
    return has_permission(cur, user, "reset_inspection_signature")


def can_sign_inspection_records(_cur, user):
    return is_station_manager(user) and bool(user.get("station_id"))


def can_view_assessment(cur, user):
    return has_permission(cur, user, "view_assessment")


def can_view_attendance(cur, user):
    return has_permission(cur, user, "view_attendance")


def can_view_station_scores(cur, user):
    return has_permission(cur, user, "view_station_scores")


def can_adjust_station_scores(cur, user):
    return has_permission(cur, user, "adjust_station_scores")


def can_view_peer_reviews(cur, user):
    return has_permission(cur, user, "view_peer_reviews")


def can_manage_peer_review_tasks(cur, user):
    return has_permission(cur, user, "manage_peer_review_tasks")


def ensure_peer_review_schema(cur):
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS peer_review_templates (
            id SERIAL PRIMARY KEY,
            title TEXT NOT NULL,
            description TEXT,
            default_deadline_at TIMESTAMP,
            show_participation BOOLEAN NOT NULL DEFAULT TRUE,
            show_reviewer BOOLEAN NOT NULL DEFAULT TRUE,
            is_active BOOLEAN NOT NULL DEFAULT TRUE,
            created_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS peer_review_template_items (
            id SERIAL PRIMARY KEY,
            template_id INTEGER NOT NULL REFERENCES peer_review_templates(id) ON DELETE CASCADE,
            item_type TEXT NOT NULL DEFAULT 'score',
            title TEXT NOT NULL,
            description TEXT,
            max_score INTEGER NOT NULL DEFAULT 5,
            sort_order INTEGER NOT NULL DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS peer_review_template_participants (
            template_id INTEGER NOT NULL REFERENCES peer_review_templates(id) ON DELETE CASCADE,
            user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            PRIMARY KEY (template_id, user_id)
        );
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS peer_review_template_reviewees (
            template_id INTEGER NOT NULL REFERENCES peer_review_templates(id) ON DELETE CASCADE,
            user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            PRIMARY KEY (template_id, user_id)
        );
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS peer_review_tasks (
            id SERIAL PRIMARY KEY,
            template_id INTEGER REFERENCES peer_review_templates(id) ON DELETE SET NULL,
            title TEXT NOT NULL,
            description TEXT,
            deadline_at TIMESTAMP,
            show_participation BOOLEAN NOT NULL DEFAULT TRUE,
            show_reviewer BOOLEAN NOT NULL DEFAULT TRUE,
            status TEXT NOT NULL DEFAULT 'active',
            created_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS peer_review_task_items (
            id SERIAL PRIMARY KEY,
            task_id INTEGER NOT NULL REFERENCES peer_review_tasks(id) ON DELETE CASCADE,
            source_template_item_id INTEGER REFERENCES peer_review_template_items(id) ON DELETE SET NULL,
            item_type TEXT NOT NULL DEFAULT 'score',
            title TEXT NOT NULL,
            description TEXT,
            max_score INTEGER NOT NULL DEFAULT 5,
            sort_order INTEGER NOT NULL DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS peer_review_task_participants (
            task_id INTEGER NOT NULL REFERENCES peer_review_tasks(id) ON DELETE CASCADE,
            user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            PRIMARY KEY (task_id, user_id)
        );
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS peer_review_task_reviewees (
            task_id INTEGER NOT NULL REFERENCES peer_review_tasks(id) ON DELETE CASCADE,
            user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            PRIMARY KEY (task_id, user_id)
        );
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS peer_review_responses (
            id SERIAL PRIMARY KEY,
            task_id INTEGER NOT NULL REFERENCES peer_review_tasks(id) ON DELETE CASCADE,
            reviewer_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            reviewee_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            submitted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE (task_id, reviewer_id, reviewee_id)
        );
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS peer_review_response_items (
            id SERIAL PRIMARY KEY,
            response_id INTEGER NOT NULL REFERENCES peer_review_responses(id) ON DELETE CASCADE,
            task_item_id INTEGER NOT NULL REFERENCES peer_review_task_items(id) ON DELETE CASCADE,
            item_type TEXT NOT NULL DEFAULT 'score',
            score_value INTEGER,
            text_value TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE (response_id, task_item_id)
        );
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_peer_review_tasks_status_deadline
        ON peer_review_tasks (status, deadline_at);
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_peer_review_responses_task_reviewer
        ON peer_review_responses (task_id, reviewer_id);
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_peer_review_responses_task_reviewee
        ON peer_review_responses (task_id, reviewee_id);
        """
    )


def normalize_peer_review_item_type(value):
    item_type = str(value or "score").strip().lower()
    return item_type if item_type in {"score", "text"} else "score"


def normalize_peer_review_status(value):
    status = str(value or "active").strip().lower()
    return status if status in {"active", "closed"} else "active"


def normalize_peer_review_id_list(raw_values):
    if raw_values in (None, ""):
        return []
    if not isinstance(raw_values, list):
        raw_values = [raw_values]
    result = []
    seen = set()
    for raw_value in raw_values:
        try:
            item_id = int(raw_value)
        except (TypeError, ValueError):
            continue
        if item_id <= 0 or item_id in seen:
            continue
        seen.add(item_id)
        result.append(item_id)
    return result


def normalize_peer_review_deadline(value):
    text = normalize_text(value, 40)
    if not text:
        return None
    for fmt in ("%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            parsed = datetime.strptime(text, fmt)
            if fmt == "%Y-%m-%d":
                parsed = parsed.replace(hour=23, minute=59)
            return parsed
        except ValueError:
            continue
    raise ValueError("截止时间格式不正确。")


def normalize_peer_review_items(raw_items):
    if not isinstance(raw_items, list) or not raw_items:
        raise ValueError("请至少配置一个评价项目。")
    items = []
    for index, raw_item in enumerate(raw_items, start=1):
        if not isinstance(raw_item, dict):
            continue
        title = normalize_text(raw_item.get("title"), 120)
        if not title:
            continue
        item_type = normalize_peer_review_item_type(raw_item.get("item_type"))
        try:
            max_score = int(raw_item.get("max_score") or 5)
        except (TypeError, ValueError):
            max_score = 5
        max_score = min(max(max_score, 1), 10)
        items.append(
            {
                "id": raw_item.get("id"),
                "item_type": item_type,
                "title": title,
                "description": normalize_text(raw_item.get("description"), 300),
                "max_score": max_score if item_type == "score" else 0,
                "sort_order": index,
            }
        )
    if not items:
        raise ValueError("请至少配置一个有效评价项目。")
    return items


def fetch_peer_review_people(cur):
    cur.execute(
        """
        SELECT id, username, real_name, role, phone
        FROM users
        WHERE role <> 'root'
        ORDER BY
            CASE role
                WHEN 'supervisor' THEN 1
                WHEN 'quality_safety' THEN 2
                WHEN 'development_plan' THEN 3
                WHEN 'oil_gas' THEN 4
                WHEN 'non_oil' THEN 5
                WHEN 'finance' THEN 6
                WHEN 'area_account' THEN 7
                WHEN 'station_manager' THEN 8
                ELSE 9
            END,
            real_name ASC,
            username ASC;
        """
    )
    return [
        {
            "id": row["id"],
            "username": row["username"],
            "real_name": row["real_name"],
            "display_name": row.get("real_name") or row.get("username") or f"用户{row['id']}",
            "role": row["role"],
            "phone": row["phone"],
        }
        for row in cur.fetchall()
    ]


def get_peer_review_user_map(cur):
    return {int(user["id"]): user for user in fetch_peer_review_people(cur)}


def validate_peer_review_user_ids(cur, user_ids, field_label):
    normalized_ids = normalize_peer_review_id_list(user_ids)
    if not normalized_ids:
        raise ValueError(f"请至少选择{field_label}。")
    user_map = get_peer_review_user_map(cur)
    invalid_ids = [user_id for user_id in normalized_ids if user_id not in user_map]
    if invalid_ids:
        raise ValueError(f"{field_label}中包含不存在或不可参与的用户。")
    return normalized_ids


def serialize_peer_review_template(cur, template_id):
    cur.execute(
        """
        SELECT
            t.id,
            t.title,
            t.description,
            TO_CHAR(t.default_deadline_at, 'YYYY-MM-DD"T"HH24:MI') AS default_deadline_at,
            t.show_participation,
            t.show_reviewer,
            t.is_active,
            COALESCE(u.real_name, u.username, '') AS created_by_name,
            TO_CHAR(t.created_at, 'YYYY-MM-DD HH24:MI') AS created_at,
            TO_CHAR(t.updated_at, 'YYYY-MM-DD HH24:MI') AS updated_at
        FROM peer_review_templates t
        LEFT JOIN users u ON u.id = t.created_by
        WHERE t.id = %s;
        """,
        (template_id,),
    )
    template = cur.fetchone()
    if not template:
        raise LookupError("互评模板不存在。")
    payload = dict(template)
    cur.execute(
        """
        SELECT id, item_type, title, description, max_score, sort_order
        FROM peer_review_template_items
        WHERE template_id = %s
        ORDER BY sort_order ASC, id ASC;
        """,
        (template_id,),
    )
    payload["items"] = [dict(row) for row in cur.fetchall()]
    cur.execute(
        "SELECT user_id FROM peer_review_template_participants WHERE template_id = %s ORDER BY user_id ASC;",
        (template_id,),
    )
    payload["participant_ids"] = [row["user_id"] for row in cur.fetchall()]
    cur.execute(
        "SELECT user_id FROM peer_review_template_reviewees WHERE template_id = %s ORDER BY user_id ASC;",
        (template_id,),
    )
    payload["reviewee_ids"] = [row["user_id"] for row in cur.fetchall()]
    return payload


def upsert_peer_review_template(cur, data, actor, template_id=None):
    title = normalize_text(data.get("title"), 120)
    if not title:
        raise ValueError("请填写互评任务标题。")
    description = normalize_text(data.get("description"), 1000)
    default_deadline_at = normalize_peer_review_deadline(data.get("default_deadline_at") or data.get("deadline_at"))
    show_participation = normalize_boolean_flag(data.get("show_participation"), True)
    show_reviewer = normalize_boolean_flag(data.get("show_reviewer"), True)
    items = normalize_peer_review_items(data.get("items"))
    participant_ids = validate_peer_review_user_ids(cur, data.get("participant_ids"), "填写人员")
    reviewee_ids = validate_peer_review_user_ids(cur, data.get("reviewee_ids"), "被评人")

    if template_id:
        cur.execute(
            """
            UPDATE peer_review_templates
            SET title = %s,
                description = %s,
                default_deadline_at = %s,
                show_participation = %s,
                show_reviewer = %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = %s
            RETURNING id;
            """,
            (title, description, default_deadline_at, show_participation, show_reviewer, template_id),
        )
        row = cur.fetchone()
        if not row:
            raise LookupError("互评模板不存在。")
        target_template_id = row["id"]
        cur.execute("DELETE FROM peer_review_template_items WHERE template_id = %s;", (target_template_id,))
        cur.execute("DELETE FROM peer_review_template_participants WHERE template_id = %s;", (target_template_id,))
        cur.execute("DELETE FROM peer_review_template_reviewees WHERE template_id = %s;", (target_template_id,))
    else:
        cur.execute(
            """
            INSERT INTO peer_review_templates (
                title,
                description,
                default_deadline_at,
                show_participation,
                show_reviewer,
                created_by,
                created_at,
                updated_at
            )
            VALUES (%s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            RETURNING id;
            """,
            (title, description, default_deadline_at, show_participation, show_reviewer, actor["id"]),
        )
        target_template_id = cur.fetchone()["id"]

    for item in items:
        cur.execute(
            """
            INSERT INTO peer_review_template_items (
                template_id,
                item_type,
                title,
                description,
                max_score,
                sort_order,
                created_at,
                updated_at
            )
            VALUES (%s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP);
            """,
            (
                target_template_id,
                item["item_type"],
                item["title"],
                item["description"],
                item["max_score"],
                item["sort_order"],
            ),
        )
    for participant_id in participant_ids:
        cur.execute(
            "INSERT INTO peer_review_template_participants (template_id, user_id) VALUES (%s, %s);",
            (target_template_id, participant_id),
        )
    for reviewee_id in reviewee_ids:
        cur.execute(
            "INSERT INTO peer_review_template_reviewees (template_id, user_id) VALUES (%s, %s);",
            (target_template_id, reviewee_id),
        )
    return serialize_peer_review_template(cur, target_template_id)


def create_peer_review_task_from_template(cur, data, actor):
    template_id = int(data.get("template_id") or 0)
    if template_id <= 0:
        raise ValueError("请选择需要发起的互评模板。")
    template = serialize_peer_review_template(cur, template_id)
    if not template.get("is_active"):
        raise ValueError("该互评模板已停用，不能继续发起任务。")
    title = normalize_text(data.get("title"), 120) or template["title"]
    description = normalize_text(data.get("description"), 1000) or template.get("description") or ""
    deadline_at = normalize_peer_review_deadline(data.get("deadline_at") or template.get("default_deadline_at"))
    show_participation = normalize_boolean_flag(data.get("show_participation"), bool(template.get("show_participation")))
    show_reviewer = normalize_boolean_flag(data.get("show_reviewer"), bool(template.get("show_reviewer")))
    participant_ids = validate_peer_review_user_ids(
        cur,
        data.get("participant_ids") or template.get("participant_ids"),
        "填写人员",
    )
    reviewee_ids = validate_peer_review_user_ids(
        cur,
        data.get("reviewee_ids") or template.get("reviewee_ids"),
        "被评人",
    )
    items = normalize_peer_review_items(data.get("items") or template.get("items"))

    cur.execute(
        """
        INSERT INTO peer_review_tasks (
            template_id,
            title,
            description,
            deadline_at,
            show_participation,
            show_reviewer,
            status,
            created_by,
            created_at,
            updated_at
        )
        VALUES (%s, %s, %s, %s, %s, %s, 'active', %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
        RETURNING id;
        """,
        (template_id, title, description, deadline_at, show_participation, show_reviewer, actor["id"]),
    )
    task_id = cur.fetchone()["id"]
    source_items = template.get("items") or []
    for index, item in enumerate(items):
        source_id = item.get("id") or (source_items[index].get("id") if index < len(source_items) else None)
        cur.execute(
            """
            INSERT INTO peer_review_task_items (
                task_id,
                source_template_item_id,
                item_type,
                title,
                description,
                max_score,
                sort_order,
                created_at
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP);
            """,
            (
                task_id,
                source_id,
                item["item_type"],
                item["title"],
                item["description"],
                item["max_score"],
                item["sort_order"],
            ),
        )
    for participant_id in participant_ids:
        cur.execute("INSERT INTO peer_review_task_participants (task_id, user_id) VALUES (%s, %s);", (task_id, participant_id))
    for reviewee_id in reviewee_ids:
        cur.execute("INSERT INTO peer_review_task_reviewees (task_id, user_id) VALUES (%s, %s);", (task_id, reviewee_id))
    return task_id


def get_peer_review_task_rows(cur, user, can_manage):
    if can_manage:
        cur.execute(
            """
            SELECT
                t.*,
                TO_CHAR(t.deadline_at, 'YYYY-MM-DD HH24:MI') AS deadline_label,
                TO_CHAR(t.created_at, 'YYYY-MM-DD HH24:MI') AS created_at_label,
                COALESCE(u.real_name, u.username, '') AS created_by_name
            FROM peer_review_tasks t
            LEFT JOIN users u ON u.id = t.created_by
            ORDER BY t.created_at DESC, t.id DESC;
            """
        )
    else:
        cur.execute(
            """
            SELECT
                t.*,
                TO_CHAR(t.deadline_at, 'YYYY-MM-DD HH24:MI') AS deadline_label,
                TO_CHAR(t.created_at, 'YYYY-MM-DD HH24:MI') AS created_at_label,
                COALESCE(u.real_name, u.username, '') AS created_by_name
            FROM peer_review_tasks t
            JOIN peer_review_task_participants p ON p.task_id = t.id
            LEFT JOIN users u ON u.id = t.created_by
            WHERE p.user_id = %s
            ORDER BY t.created_at DESC, t.id DESC;
            """,
            (user["id"],),
        )
    return [dict(row) for row in cur.fetchall()]


def fetch_peer_review_task_people_maps(cur, task_ids, table_name):
    if not task_ids:
        return {}
    if table_name not in {"peer_review_task_participants", "peer_review_task_reviewees"}:
        return {task_id: [] for task_id in task_ids}
    column_name = "task_id"
    cur.execute(
        f"""
        SELECT
            r.{column_name},
            u.id,
            u.username,
            u.real_name,
            u.role
        FROM {table_name} r
        JOIN users u ON u.id = r.user_id
        WHERE r.{column_name} = ANY(%s)
        ORDER BY u.real_name ASC, u.username ASC;
        """,
        (task_ids,),
    )
    result = {task_id: [] for task_id in task_ids}
    for row in cur.fetchall():
        task_id = row[column_name]
        result.setdefault(task_id, []).append(
            {
                "id": row["id"],
                "username": row["username"],
                "real_name": row["real_name"],
                "display_name": row.get("real_name") or row.get("username") or f"用户{row['id']}",
                "role": row["role"],
            }
        )
    return result


def fetch_peer_review_task_items_map(cur, task_ids):
    if not task_ids:
        return {}
    cur.execute(
        """
        SELECT id, task_id, item_type, title, description, max_score, sort_order
        FROM peer_review_task_items
        WHERE task_id = ANY(%s)
        ORDER BY task_id ASC, sort_order ASC, id ASC;
        """,
        (task_ids,),
    )
    result = {task_id: [] for task_id in task_ids}
    for row in cur.fetchall():
        result.setdefault(row["task_id"], []).append(dict(row))
    return result


def fetch_peer_review_responses_map(cur, task_ids, user, can_manage):
    if not task_ids:
        return {}
    if can_manage:
        cur.execute(
            """
            SELECT
                r.id,
                r.task_id,
                r.reviewer_id,
                r.reviewee_id,
                TO_CHAR(r.submitted_at, 'YYYY-MM-DD HH24:MI') AS submitted_at,
                TO_CHAR(r.updated_at, 'YYYY-MM-DD HH24:MI') AS updated_at,
                COALESCE(reviewer.real_name, reviewer.username, '') AS reviewer_name,
                COALESCE(reviewee.real_name, reviewee.username, '') AS reviewee_name
            FROM peer_review_responses r
            JOIN users reviewer ON reviewer.id = r.reviewer_id
            JOIN users reviewee ON reviewee.id = r.reviewee_id
            WHERE r.task_id = ANY(%s)
            ORDER BY r.task_id ASC, r.submitted_at DESC, r.id DESC;
            """,
            (task_ids,),
        )
    else:
        cur.execute(
            """
            SELECT
                r.id,
                r.task_id,
                r.reviewer_id,
                r.reviewee_id,
                TO_CHAR(r.submitted_at, 'YYYY-MM-DD HH24:MI') AS submitted_at,
                TO_CHAR(r.updated_at, 'YYYY-MM-DD HH24:MI') AS updated_at,
                COALESCE(reviewer.real_name, reviewer.username, '') AS reviewer_name,
                COALESCE(reviewee.real_name, reviewee.username, '') AS reviewee_name
            FROM peer_review_responses r
            JOIN users reviewer ON reviewer.id = r.reviewer_id
            JOIN users reviewee ON reviewee.id = r.reviewee_id
            WHERE r.task_id = ANY(%s)
              AND r.reviewer_id = %s
            ORDER BY r.task_id ASC, r.submitted_at DESC, r.id DESC;
            """,
            (task_ids, user["id"]),
        )
    responses = [dict(row) for row in cur.fetchall()]
    response_ids = [row["id"] for row in responses]
    item_map = {response_id: [] for response_id in response_ids}
    if response_ids:
        cur.execute(
            """
            SELECT
                ri.response_id,
                ri.task_item_id,
                ri.item_type,
                ri.score_value,
                ri.text_value,
                ti.title,
                ti.max_score,
                ti.sort_order
            FROM peer_review_response_items ri
            JOIN peer_review_task_items ti ON ti.id = ri.task_item_id
            WHERE ri.response_id = ANY(%s)
            ORDER BY ti.sort_order ASC, ti.id ASC;
            """,
            (response_ids,),
        )
        for row in cur.fetchall():
            item_map.setdefault(row["response_id"], []).append(dict(row))
    result = {task_id: [] for task_id in task_ids}
    for response in responses:
        response["items"] = item_map.get(response["id"], [])
        result.setdefault(response["task_id"], []).append(response)
    return result


def build_peer_review_dashboard(cur, user, can_manage):
    task_rows = get_peer_review_task_rows(cur, user, can_manage)
    task_ids = [row["id"] for row in task_rows]
    participants_map = fetch_peer_review_task_people_maps(cur, task_ids, "peer_review_task_participants")
    reviewees_map = fetch_peer_review_task_people_maps(cur, task_ids, "peer_review_task_reviewees")
    items_map = fetch_peer_review_task_items_map(cur, task_ids)
    all_response_map = fetch_peer_review_responses_map(cur, task_ids, user, True)
    visible_response_map = all_response_map if can_manage else fetch_peer_review_responses_map(cur, task_ids, user, False)

    tasks = []
    for row in task_rows:
        task_id = row["id"]
        participants = participants_map.get(task_id, [])
        reviewees = reviewees_map.get(task_id, [])
        items = items_map.get(task_id, [])
        all_responses = all_response_map.get(task_id, [])
        visible_responses = visible_response_map.get(task_id, [])
        participant_ids = [person["id"] for person in participants]
        reviewee_ids = [person["id"] for person in reviewees]
        response_pairs = {(response["reviewer_id"], response["reviewee_id"]) for response in all_responses}
        completed_participants = []
        pending_participants = []
        for participant in participants:
            expected_reviewees = [reviewee_id for reviewee_id in reviewee_ids if reviewee_id != participant["id"]]
            completed_count = len([reviewee_id for reviewee_id in expected_reviewees if (participant["id"], reviewee_id) in response_pairs])
            is_completed = bool(expected_reviewees) and completed_count >= len(expected_reviewees)
            if is_completed:
                completed_participants.append(participant)
            else:
                pending_participants.append(participant)
        current_expected_reviewees = [reviewee for reviewee in reviewees if reviewee["id"] != user["id"]]
        my_completed_reviewee_ids = {
            response["reviewee_id"]
            for response in all_responses
            if response["reviewer_id"] == user["id"]
        }
        my_pending_reviewees = [
            reviewee
            for reviewee in current_expected_reviewees
            if reviewee["id"] not in my_completed_reviewee_ids
        ] if user["id"] in participant_ids else []
        task_status = normalize_peer_review_status(row.get("status"))
        deadline_at = row.get("deadline_at")
        is_expired = bool(deadline_at and deadline_at < datetime.now())

        tasks.append(
            {
                "id": task_id,
                "template_id": row.get("template_id"),
                "title": row.get("title"),
                "description": row.get("description") or "",
                "deadline_at": row.get("deadline_label") or "",
                "show_participation": bool(row.get("show_participation")),
                "show_reviewer": bool(row.get("show_reviewer")),
                "status": task_status,
                "is_expired": is_expired,
                "created_by_name": row.get("created_by_name") or "",
                "created_at": row.get("created_at_label") or "",
                "items": items,
                "participants": participants,
                "reviewees": reviewees,
                "progress": {
                    "participant_count": len(participants),
                    "completed_count": len(completed_participants),
                    "completed_participants": completed_participants if bool(row.get("show_participation")) or can_manage else [],
                    "pending_participants": pending_participants if bool(row.get("show_participation")) or can_manage else [],
                },
                "my_pending_reviewees": my_pending_reviewees,
                "my_completed_count": len(my_completed_reviewee_ids),
                "responses": visible_responses,
                "can_submit": bool(user["id"] in participant_ids and task_status == "active" and not is_expired),
            }
        )
    return tasks


def can_view_own_certificates(cur, user):
    return bool(get_effective_permissions(cur, user).get("view_own_certificates"))


def can_edit_own_certificates(cur, user):
    return bool(get_effective_permissions(cur, user).get("edit_own_certificates"))


def can_view_all_certificates(cur, user):
    return bool(get_effective_permissions(cur, user).get("view_all_certificates"))


def can_view_region_certificates(cur, user):
    return bool(get_effective_permissions(cur, user).get("limit_certificate_station_region_scope"))


def can_edit_all_certificates(user):
    return is_root_user(user)


def can_view_certificates(cur, user):
    return (
        can_edit_all_certificates(user)
        or can_view_all_certificates(cur, user)
        or can_view_region_certificates(cur, user)
        or can_view_own_certificates(cur, user)
        or can_edit_own_certificates(cur, user)
    )


def normalize_text(value, max_length=None):
    text = str(value or "").strip()
    if max_length and len(text) > max_length:
        return text[:max_length]
    return text


def ensure_feedback_schema(cur):
    global FEEDBACK_SCHEMA_READY
    if FEEDBACK_SCHEMA_READY:
        return
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS system_feedbacks (
            id SERIAL PRIMARY KEY,
            feedback_type TEXT NOT NULL,
            module TEXT NOT NULL,
            title TEXT NOT NULL,
            description TEXT NOT NULL,
            created_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
            author_name TEXT,
            author_phone TEXT,
            author_role TEXT,
            accepted_at TIMESTAMP,
            accepted_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """
    )
    cur.execute(
        """
        ALTER TABLE system_feedbacks
            ADD COLUMN IF NOT EXISTS accepted_at TIMESTAMP,
            ADD COLUMN IF NOT EXISTS accepted_by INTEGER REFERENCES users(id) ON DELETE SET NULL;
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS system_feedback_screenshots (
            id SERIAL PRIMARY KEY,
            feedback_id INTEGER NOT NULL REFERENCES system_feedbacks(id) ON DELETE CASCADE,
            file_path TEXT NOT NULL,
            sort_order INTEGER NOT NULL DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS system_feedback_comments (
            id SERIAL PRIMARY KEY,
            feedback_id INTEGER NOT NULL REFERENCES system_feedbacks(id) ON DELETE CASCADE,
            comment_text TEXT NOT NULL,
            created_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
            author_name TEXT,
            author_phone TEXT,
            author_role TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS system_feedback_read_states (
            user_id INTEGER PRIMARY KEY REFERENCES users(id) ON DELETE CASCADE,
            last_read_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_system_feedbacks_created_at
        ON system_feedbacks(created_at DESC);
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_system_feedbacks_accepted_at
        ON system_feedbacks(accepted_at DESC);
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_system_feedback_screenshots_feedback_id
        ON system_feedback_screenshots(feedback_id);
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_system_feedback_comments_feedback_id
        ON system_feedback_comments(feedback_id);
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_system_feedback_comments_created_at
        ON system_feedback_comments(created_at DESC);
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_system_feedback_read_states_last_read
        ON system_feedback_read_states(last_read_at DESC);
        """
    )
    FEEDBACK_SCHEMA_READY = True


def get_or_create_feedback_last_read_at(cur, user_id):
    cur.execute(
        """
        SELECT last_read_at
        FROM system_feedback_read_states
        WHERE user_id = %s
        LIMIT 1;
        """,
        (user_id,),
    )
    row = cur.fetchone()
    if row:
        return row["last_read_at"]

    cur.execute(
        """
        INSERT INTO system_feedback_read_states (user_id, last_read_at, updated_at)
        VALUES (%s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
        ON CONFLICT (user_id) DO NOTHING
        RETURNING last_read_at;
        """,
        (user_id,),
    )
    created = cur.fetchone()
    return created["last_read_at"] if created else datetime.now()


def get_feedback_unread_count(cur, user_id):
    last_read_at = get_or_create_feedback_last_read_at(cur, user_id)
    cur.execute(
        """
        SELECT
            (
                SELECT COUNT(*)
                FROM system_feedbacks
                WHERE created_at > %s
            ) + (
                SELECT COUNT(*)
                FROM system_feedback_comments
                WHERE created_at > %s
            ) AS unread_count;
        """,
        (last_read_at, last_read_at),
    )
    row = cur.fetchone()
    return int(row["unread_count"] or 0)


def mark_feedbacks_read(cur, user_id):
    cur.execute(
        """
        INSERT INTO system_feedback_read_states (user_id, last_read_at, updated_at)
        VALUES (%s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
        ON CONFLICT (user_id) DO UPDATE
        SET last_read_at = EXCLUDED.last_read_at,
            updated_at = CURRENT_TIMESTAMP;
        """,
        (user_id,),
    )


def get_current_request_user():
    user = getattr(g, "current_user", None)
    if not user:
        raise PermissionError("请先登录。")
    return user


def build_feedback_author_snapshot(user):
    return {
        "author_name": normalize_text(
            user.get("real_name") or user.get("username") or "未命名用户", 80
        ),
        "author_phone": normalize_text(user.get("phone"), 40),
        "author_role": normalize_text(user.get("role"), 40),
    }


def normalize_feedback_type(value):
    feedback_type = normalize_text(value, 40)
    if feedback_type not in FEEDBACK_TYPE_OPTIONS:
        raise ValueError("反馈类型不正确。")
    return feedback_type


def normalize_feedback_module(value):
    module = normalize_text(value, 40)
    if module not in FEEDBACK_MODULE_OPTIONS:
        raise ValueError("问题模块不正确。")
    return module


def collect_feedback_files(request_files):
    files = []
    for key in ("screenshots", "screenshots[]", "screenshot"):
        files.extend(request_files.getlist(key))
    return [file for file in files if file and file.filename]


def serialize_feedback_comment(row, can_delete=False):
    item = dict(row)
    item["can_delete"] = bool(can_delete)
    return item


def serialize_feedback_row(row, screenshots=None, comments=None, can_delete=False):
    item = dict(row)
    item["screenshots"] = screenshots or []
    item["comments"] = comments or []
    item["can_delete"] = bool(can_delete)
    item["can_accept"] = bool(can_delete)
    item["is_accepted"] = bool(item.get("accepted_at"))
    return item


def validate_new_login_password(username, new_password, confirm_password):
    password = normalize_text(new_password)
    confirm = normalize_text(confirm_password)
    username_text = normalize_text(username)

    if not password:
        raise ValueError("请填写新密码。")
    if len(password) < PASSWORD_MIN_LENGTH or len(password) > PASSWORD_MAX_LENGTH:
        raise ValueError(
            f"新密码长度需为 {PASSWORD_MIN_LENGTH}-{PASSWORD_MAX_LENGTH} 位。"
        )
    if re.search(r"\s", password):
        raise ValueError("新密码不能包含空格。")
    if not re.search(r"[A-Za-z]", password) or not re.search(r"\d", password):
        raise ValueError("新密码需同时包含字母和数字。")
    if password == DEFAULT_INITIAL_PASSWORD:
        raise ValueError("新密码不能继续使用初始密码 123456。")
    if username_text and password.lower() == username_text.lower():
        raise ValueError("新密码不能与用户名相同。")
    if password != confirm:
        raise ValueError("两次输入的新密码不一致。")

    return password


def normalize_decimal_text(value):
    text = normalize_text(value)
    if text == "":
        return None
    return text


def validate_option(value, options, field_label, default_value=None):
    text = normalize_text(value)
    if not text and default_value is not None:
        text = default_value
    if text not in options:
        raise ValueError(f"{field_label}只能选择：{'、'.join(options)}。")
    return text


def normalize_asset_type_option(value):
    text = normalize_text(value)
    if any(keyword in text for keyword in ("股权", "控股", "参股")):
        return "股权"
    if "全资" in text:
        return "全资"
    return text


def normalize_station_type_option(value):
    text = normalize_text(value)
    if text == DISPLAY_REMOVED_STATION_PHRASE:
        return DISPLAY_OIL_STATION_TYPE
    return text


def normalize_hos_station_code(value):
    text = normalize_text(value, 40).upper()
    if not text:
        raise ValueError("请填写 HOS编码。")
    return text


def normalize_operating_hours(value):
    text = normalize_text(value, 40)
    text = re.sub(r"\s+", "", text)
    text = (
        text.replace("－", "-")
        .replace("—", "-")
        .replace("–", "-")
        .replace("至", "-")
        .replace("到", "-")
    )
    if not text or text in {"24小时", "全天", "全天营业"}:
        return "24小时"

    matched = re.fullmatch(r"(\d{1,2}):([0-5]\d)-(\d{1,2}):([0-5]\d)", text)
    if not matched:
        raise ValueError("营运时间只能选择 24小时 或 HH:MM-HH:MM 格式。")

    start_hour = int(matched.group(1))
    start_minute = matched.group(2)
    end_hour = int(matched.group(3))
    end_minute = matched.group(4)
    if start_hour > 23 or end_hour > 23:
        raise ValueError("营运时间小时必须在 00-23 之间。")

    start_time = f"{start_hour:02d}:{start_minute}"
    end_time = f"{end_hour:02d}:{end_minute}"
    if start_time == end_time:
        raise ValueError("营运时间的开始时间和结束时间不能相同。")

    return f"{start_time}-{end_time}"


def ensure_station_management_columns(cur):
    global STATION_MANAGEMENT_SCHEMA_READY
    if STATION_MANAGEMENT_SCHEMA_READY:
        return

    cur.execute(
        """
        ALTER TABLE stations
        ADD COLUMN IF NOT EXISTS is_consolidated TEXT DEFAULT '否';
        """
    )
    cur.execute(
        """
        ALTER TABLE stations
        ADD COLUMN IF NOT EXISTS online_3_status TEXT DEFAULT '未上线';
        """
    )
    cur.execute(
        """
        ALTER TABLE stations
        ADD COLUMN IF NOT EXISTS monitoring_status TEXT NOT NULL DEFAULT '运行中';
        """
    )
    cur.execute(
        """
        UPDATE stations
        SET monitoring_status = '运行中'
        WHERE monitoring_status IS NULL
           OR monitoring_status NOT IN ('运行中', '未运行');
        """
    )
    cur.execute(
        """
        ALTER TABLE stations
        DROP CONSTRAINT IF EXISTS stations_station_type_check;
        """
    )
    cur.execute(
        """
        UPDATE stations
        SET station_type = %s
        WHERE station_type = %s;
        """,
        (DISPLAY_OIL_STATION_TYPE, DISPLAY_REMOVED_STATION_PHRASE),
    )
    cur.execute(
        """
        UPDATE stations
        SET station_type = %s
        WHERE station_type IS NULL
           OR station_type NOT IN (%s, '充电站');
        """,
        (DISPLAY_OIL_STATION_TYPE, DISPLAY_OIL_STATION_TYPE),
    )
    cur.execute(
        """
        ALTER TABLE stations
        ALTER COLUMN station_type SET DEFAULT '油站';
        """
    )
    cur.execute(
        """
        ALTER TABLE stations
        ADD CONSTRAINT stations_station_type_check
        CHECK (station_type IN ('油站', '充电站'));
        """
    )
    cur.execute(
        """
        ALTER TABLE stations
        ADD COLUMN IF NOT EXISTS hos_station_code TEXT;
        """
    )
    cur.execute(
        """
        UPDATE stations
        SET hos_station_code = NULLIF(UPPER(BTRIM(hos_station_code)), '')
        WHERE hos_station_code IS NOT NULL;
        """
    )
    cur.execute(
        """
        CREATE UNIQUE INDEX IF NOT EXISTS idx_stations_hos_station_code_unique
        ON stations (hos_station_code)
        WHERE hos_station_code IS NOT NULL;
        """
    )
    cur.execute(
        """
        ALTER TABLE stations
        ADD COLUMN IF NOT EXISTS landline_phone TEXT;
        """
    )
    cur.execute(
        """
        UPDATE stations
        SET asset_type = CASE
            WHEN asset_type LIKE '%股权%' OR asset_type LIKE '%控股%' OR asset_type LIKE '%参股%' THEN '股权'
            ELSE '全资'
        END
        WHERE asset_type IS NULL OR asset_type NOT IN ('全资', '股权');
        """
    )
    STATION_MANAGEMENT_SCHEMA_READY = True


def require_management_user(cur, user_id, permission_key):
    user_id = get_authenticated_request_user_id(user_id)
    if not user_id:
        raise PermissionError("缺少用户信息。")

    user = get_user_by_id(cur, user_id)
    if not user:
        raise LookupError("用户不存在。")

    if not can_manage_system(cur, user, permission_key):
        raise PermissionError("当前账号无权操作管理系统。")

    return user


def build_station_payload(data):
    station_name = sanitize_display_string(normalize_text(data.get("station_name"), 120))
    if not station_name:
        raise ValueError("请填写站点名称。")

    station_type = validate_option(
        normalize_station_type_option(data.get("station_type")),
        STATION_TYPE_OPTIONS,
        "站点类型",
        DISPLAY_OIL_STATION_TYPE,
    )
    asset_type = validate_option(
        normalize_asset_type_option(data.get("asset_type")),
        STATION_ASSET_TYPE_OPTIONS,
        "资产类型",
        "全资",
    )
    is_consolidated = validate_option(
        data.get("is_consolidated"),
        STATION_CONSOLIDATED_OPTIONS,
        "是否并表",
        "否",
    )
    online_3_status = validate_option(
        data.get("online_3_status"),
        STATION_ONLINE_3_STATUS_OPTIONS,
        "是否上线3.0",
        "未上线",
    )
    monitoring_status = validate_option(
        data.get("monitoring_status"),
        STATION_MONITORING_STATUS_OPTIONS,
        "监控状态",
        "运行中",
    )
    status = validate_option(
        data.get("status"),
        STATION_STATUS_OPTIONS,
        "站点状态",
        "营业中",
    )

    return {
        "station_name": station_name,
        "region": sanitize_display_string(normalize_text(data.get("region"), 80)) or None,
        "address": sanitize_display_string(normalize_text(data.get("address"), 220)) or None,
        "longitude": normalize_decimal_text(data.get("longitude")),
        "latitude": normalize_decimal_text(data.get("latitude")),
        "station_manager_name": normalize_text(data.get("station_manager_name"), 80) or None,
        "station_manager_phone": normalize_text(data.get("station_manager_phone"), 40) or None,
        "station_type": station_type,
        "asset_type": asset_type,
        "is_consolidated": is_consolidated,
        "online_3_status": online_3_status,
        "monitoring_status": monitoring_status,
        "hos_station_code": normalize_hos_station_code(data.get("hos_station_code")),
        "landline_phone": normalize_text(data.get("landline_phone"), 40) or None,
        "status": status,
        "operating_hours": normalize_operating_hours(data.get("operating_hours")),
    }


STATION_DATA_EXPORT_FIELDS = [
    {"key": "station_name", "label": "站点名称", "width": 24},
    {"key": "station_usernames", "label": "站点登录用户名", "width": 22},
    {"key": "region", "label": "所属片区/归属地", "width": 22},
    {"key": "address", "label": "站点地址", "width": 36},
    {"key": "longitude", "label": "经度", "width": 16},
    {"key": "latitude", "label": "纬度", "width": 16},
    {"key": "station_manager_name", "label": "站点负责人姓名", "width": 18},
    {"key": "station_manager_phone", "label": "站点负责人手机号", "width": 20},
    {"key": "station_type", "label": "站点类型", "width": 14},
    {"key": "asset_type", "label": "资产类型", "width": 14},
    {"key": "is_consolidated", "label": "是否并表", "width": 14},
    {"key": "online_3_status", "label": "是否上线3.0", "width": 18},
    {"key": "monitoring_status", "label": "监控状态", "width": 14},
    {"key": "hos_station_code", "label": "HOS编码", "width": 18},
    {"key": "landline_phone", "label": "固定电话", "width": 18},
    {"key": "status", "label": "站点状态", "width": 14},
    {"key": "operating_hours", "label": "营运时间", "width": 16},
    {"key": "created_at", "label": "创建时间", "width": 20},
    {"key": "updated_at", "label": "更新时间", "width": 20},
]
STATION_DATA_EXPORT_FIELD_MAP = {field["key"]: field for field in STATION_DATA_EXPORT_FIELDS}
DEFAULT_STATION_DATA_EXPORT_FIELD_KEYS = ["station_name", "station_usernames", "region"]


def normalize_station_export_field_keys(raw_field_keys):
    if not isinstance(raw_field_keys, list):
        raw_field_keys = DEFAULT_STATION_DATA_EXPORT_FIELD_KEYS
    selected_keys = []
    for raw_key in raw_field_keys:
        key = str(raw_key or "").strip()
        if key in STATION_DATA_EXPORT_FIELD_MAP and key not in selected_keys:
            selected_keys.append(key)
    if not selected_keys:
        raise ValueError("请至少选择一个导出字段。")
    return selected_keys


def normalize_station_export_ids(raw_ids):
    if not isinstance(raw_ids, list):
        return []
    station_ids = []
    seen = set()
    for raw_id in raw_ids:
        try:
            station_id = int(raw_id)
        except (TypeError, ValueError):
            continue
        if station_id <= 0 or station_id in seen:
            continue
        station_ids.append(station_id)
        seen.add(station_id)
    return station_ids


def fetch_station_export_rows(cur, station_ids=None):
    where_clause = ""
    params = []
    if station_ids:
        where_clause = "WHERE s.id = ANY(%s)"
        params.append(station_ids)

    cur.execute(
        f"""
        SELECT
            s.id,
            s.station_name,
            s.region,
            s.address,
            s.longitude::TEXT AS longitude,
            s.latitude::TEXT AS latitude,
            s.station_manager_name,
            s.station_manager_phone,
            s.station_type,
            CASE
                WHEN s.asset_type LIKE '%%股权%%' OR s.asset_type LIKE '%%控股%%' OR s.asset_type LIKE '%%参股%%' THEN '股权'
                ELSE '全资'
            END AS asset_type,
            COALESCE(s.is_consolidated, '否') AS is_consolidated,
            COALESCE(s.online_3_status, '未上线') AS online_3_status,
            COALESCE(s.monitoring_status, '运行中') AS monitoring_status,
            s.hos_station_code,
            s.landline_phone,
            COALESCE(s.status, '营业中') AS status,
            COALESCE(s.operating_hours, '24小时') AS operating_hours,
            COALESCE(station_accounts.station_usernames, '') AS station_usernames,
            TO_CHAR(s.created_at, 'YYYY-MM-DD HH24:MI') AS created_at,
            TO_CHAR(s.updated_at, 'YYYY-MM-DD HH24:MI') AS updated_at
        FROM stations s
        LEFT JOIN (
            SELECT
                station_id,
                STRING_AGG(username, ' ' ORDER BY username) AS station_usernames
            FROM users
            WHERE role = 'station_manager'
              AND station_id IS NOT NULL
            GROUP BY station_id
        ) station_accounts ON station_accounts.station_id = s.id
        {where_clause}
        ORDER BY s.id ASC;
        """,
        params,
    )
    return cur.fetchall()


def build_station_data_export_workbook(rows, selected_field_keys):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise RuntimeError("服务器缺少 openpyxl 组件，暂时无法导出 Excel。") from exc

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "站点数据"
    worksheet.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="DCEBFF")
    header_font = Font(bold=True, color="0F172A")
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    center_alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    text_alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)

    for column_index, field_key in enumerate(selected_field_keys, start=1):
        field = STATION_DATA_EXPORT_FIELD_MAP[field_key]
        cell = worksheet.cell(row=1, column=column_index, value=field["label"])
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_alignment
        worksheet.column_dimensions[excel_column_name(column_index)].width = field.get("width") or 18

    for row_index, row in enumerate(rows, start=2):
        for column_index, field_key in enumerate(selected_field_keys, start=1):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.border = thin_border
            cell.alignment = text_alignment if field_key in {"address", "station_usernames"} else center_alignment
            value = row.get(field_key)
            cell.value = "" if value is None else str(value)

    if selected_field_keys:
        last_column = excel_column_name(len(selected_field_keys))
        worksheet.auto_filter.ref = f"A1:{last_column}1"

    workbook.properties.title = "站点数据导出"
    workbook.properties.creator = "业务督导中心数智管理平台"
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def parse_station_backup_json(file_storage):
    if not file_storage or not file_storage.filename:
        raise ValueError("请选择需要导入的站点备份文件。")

    filename = str(file_storage.filename or "").lower()
    if not filename.endswith(".json"):
        raise ValueError("仅支持导入 JSON 格式的站点备份文件。")

    file_bytes = file_storage.read()
    if not file_bytes:
        raise ValueError("站点备份文件内容为空。")
    if len(file_bytes) > 5 * 1024 * 1024:
        raise ValueError("站点备份文件不能超过 5MB。")

    try:
        payload = json.loads(file_bytes.decode("utf-8-sig"))
    except Exception as exc:
        raise ValueError("站点备份文件不是有效的 JSON。") from exc

    stations = payload.get("stations") if isinstance(payload, dict) else payload
    if not isinstance(stations, list):
        raise ValueError("站点备份文件缺少 stations 数据。")
    if not stations:
        raise ValueError("站点备份文件中没有可导入的站点。")

    return stations


def parse_user_backup_json(file_storage):
    if not file_storage or not file_storage.filename:
        raise ValueError("请选择需要导入的用户备份文件。")

    filename = str(file_storage.filename or "").lower()
    if not filename.endswith(".json"):
        raise ValueError("仅支持导入 JSON 格式的用户备份文件。")

    file_bytes = file_storage.read()
    if not file_bytes:
        raise ValueError("用户备份文件内容为空。")
    if len(file_bytes) > 5 * 1024 * 1024:
        raise ValueError("用户备份文件不能超过 5MB。")

    try:
        payload = json.loads(file_bytes.decode("utf-8-sig"))
    except Exception as exc:
        raise ValueError("用户备份文件不是有效的 JSON。") from exc

    users = payload.get("users") if isinstance(payload, dict) else payload
    if not isinstance(users, list):
        raise ValueError("用户备份文件缺少 users 数据。")
    if not users:
        raise ValueError("用户备份文件中没有可导入的用户。")

    if isinstance(payload, dict):
        return payload
    return {"users": users}


def normalize_station_backup_id(value):
    if value in (None, ""):
        return None
    try:
        station_id = int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError("站点备份文件中存在无效的站点ID。") from exc
    if station_id <= 0:
        raise ValueError("站点备份文件中存在无效的站点ID。")
    return station_id


def normalize_user_backup_id(value):
    if value in (None, ""):
        return None
    try:
        user_id = int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError("用户备份文件中存在无效的用户ID。") from exc
    if user_id <= 0:
        raise ValueError("用户备份文件中存在无效的用户ID。")
    return user_id


def normalize_user_role(value):
    role = normalize_text(value)
    if role not in ROLE_OPTIONS:
        raise ValueError("用户角色只能选择：root、supervisor、station_manager、quality_safety、development_plan、oil_gas、non_oil、finance、area_account。")
    return role


def build_management_user_payload(data, is_create=False):
    username = normalize_text(data.get("username"), 80)
    if not username:
        raise ValueError("请填写用户名。")

    password = normalize_text(data.get("password"), 120)
    if is_create and not password:
        raise ValueError("请填写初始密码。")

    role = normalize_user_role(data.get("role"))
    real_name = normalize_text(data.get("real_name"), 80)
    if role == "area_account" and not real_name:
        real_name = username
    if not real_name:
        raise ValueError("请填写用户姓名。")

    station_id = data.get("station_id")
    if station_id in (None, "", "null"):
        station_id = None
    elif role == "station_manager":
        try:
            station_id = int(station_id)
        except (TypeError, ValueError) as exc:
            raise ValueError("请选择有效的所属站点。") from exc
    else:
        station_id = None

    if role == "station_manager" and not station_id:
        raise ValueError("站点账号必须选择所属站点。")

    return {
        "username": username,
        "password": password,
        "role": role,
        "real_name": real_name,
        "phone": None if role == "area_account" else normalize_text(data.get("phone"), 40) or None,
        "station_id": station_id,
    }


def normalize_permission_updates(raw_permissions):
    if raw_permissions in (None, ""):
        return {}
    if not isinstance(raw_permissions, dict):
        raise ValueError("权限配置格式不正确。")

    updates = {}
    for key, value in raw_permissions.items():
        if key not in PERMISSION_KEYS:
            continue
        updates[key] = bool(value)
    return updates


def normalize_single_inspection_table_scope_ids(raw_scope_ids):
    if raw_scope_ids in (None, ""):
        return []
    if not isinstance(raw_scope_ids, list):
        raise ValueError("检查表范围配置格式不正确。")

    scope_ids = []
    seen = set()
    for raw_id in raw_scope_ids:
        try:
            table_id = int(raw_id)
        except (TypeError, ValueError) as exc:
            raise ValueError("检查表范围中存在无效检查表。") from exc
        if table_id <= 0 or table_id in seen:
            continue
        seen.add(table_id)
        scope_ids.append(table_id)
    return scope_ids


def normalize_inspection_table_scope_updates(raw_scope_ids):
    if raw_scope_ids in (None, ""):
        return {key: [] for key in INSPECTION_TABLE_SCOPE_PERMISSION_KEYS}

    if isinstance(raw_scope_ids, list):
        # 兼容旧备份：原来一套范围同时作用于问题、记录和计划。
        legacy_ids = normalize_single_inspection_table_scope_ids(raw_scope_ids)
        return {key: list(legacy_ids) for key in INSPECTION_TABLE_SCOPE_PERMISSION_KEYS}

    if not isinstance(raw_scope_ids, dict):
        raise ValueError("检查表范围配置格式不正确。")

    scope_map = {key: [] for key in INSPECTION_TABLE_SCOPE_PERMISSION_KEYS}
    for key, value in raw_scope_ids.items():
        if key not in scope_map:
            continue
        scope_map[key] = normalize_single_inspection_table_scope_ids(value)
    return scope_map


def normalize_single_station_region_scope_values(raw_scope_values):
    if raw_scope_values in (None, ""):
        return []
    if not isinstance(raw_scope_values, list):
        raise ValueError("片区范围配置格式不正确。")

    scope_values = []
    seen = set()
    for raw_value in raw_scope_values:
        region = normalize_station_region_value(raw_value)
        if region in seen:
            continue
        seen.add(region)
        scope_values.append(region)
    return scope_values


def normalize_station_region_scope_updates(raw_scope_values):
    if raw_scope_values in (None, ""):
        return {key: [] for key in STATION_REGION_SCOPE_PERMISSION_KEYS}

    if isinstance(raw_scope_values, list):
        legacy_values = normalize_single_station_region_scope_values(raw_scope_values)
        return {key: list(legacy_values) for key in STATION_REGION_SCOPE_PERMISSION_KEYS}

    if not isinstance(raw_scope_values, dict):
        raise ValueError("片区范围配置格式不正确。")

    scope_map = {key: [] for key in STATION_REGION_SCOPE_PERMISSION_KEYS}
    for key, value in raw_scope_values.items():
        if key not in scope_map:
            continue
        scope_map[key] = normalize_single_station_region_scope_values(value)
    return scope_map


def apply_user_permission_updates(cur, target_user, permissions, actor_user_id):
    if is_root_user(target_user):
        cur.execute("DELETE FROM user_permissions WHERE user_id = %s;", (target_user["id"],))
        cur.execute("DELETE FROM user_inspection_table_scopes WHERE user_id = %s;", (target_user["id"],))
        cur.execute("DELETE FROM user_station_region_scopes WHERE user_id = %s;", (target_user["id"],))
        return

    permissions = enforce_exclusive_permissions(permissions, target_user.get("role"))
    for permission_key, is_allowed in permissions.items():
        cur.execute(
            """
            INSERT INTO user_permissions (
                user_id,
                permission_key,
                is_allowed,
                updated_by,
                created_at,
                updated_at
            )
            VALUES (%s, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            ON CONFLICT (user_id, permission_key)
            DO UPDATE SET
                is_allowed = EXCLUDED.is_allowed,
                updated_by = EXCLUDED.updated_by,
                updated_at = CURRENT_TIMESTAMP;
            """,
            (target_user["id"], permission_key, is_allowed, actor_user_id),
        )


def apply_user_inspection_table_scope_updates(cur, target_user, scope_map, actor_user_id):
    if is_root_user(target_user):
        cur.execute("DELETE FROM user_inspection_table_scopes WHERE user_id = %s;", (target_user["id"],))
        return

    normalized_scope_map = normalize_inspection_table_scope_updates(scope_map)
    all_scope_ids = sorted(
        {
            table_id
            for scope_ids in normalized_scope_map.values()
            for table_id in scope_ids
        }
    )
    cur.execute("DELETE FROM user_inspection_table_scopes WHERE user_id = %s;", (target_user["id"],))
    if not all_scope_ids:
        return

    cur.execute(
        """
        SELECT id
        FROM inspection_tables
        WHERE id = ANY(%s)
          AND is_active = TRUE;
        """,
        (all_scope_ids,),
    )
    valid_ids = {row["id"] for row in cur.fetchall()}
    invalid_ids = [table_id for table_id in all_scope_ids if table_id not in valid_ids]
    if invalid_ids:
        raise ValueError("检查表范围中包含不存在或未启用的检查表。")

    for scope_key, scope_ids in normalized_scope_map.items():
        for table_id in scope_ids:
            cur.execute(
                """
                INSERT INTO user_inspection_table_scopes (
                    user_id,
                    scope_key,
                    inspection_table_id,
                    updated_by,
                    created_at,
                    updated_at
                )
                VALUES (%s, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                ON CONFLICT (user_id, scope_key, inspection_table_id)
                DO UPDATE SET
                    updated_by = EXCLUDED.updated_by,
                    updated_at = CURRENT_TIMESTAMP;
                """,
                (target_user["id"], scope_key, table_id, actor_user_id),
            )


def apply_role_inspection_table_scope_updates(cur, role, scope_map, actor_user_id):
    normalized_role = normalize_text(role)
    if normalized_role not in ROLE_OPTIONS:
        raise ValueError("角色类型不正确。")
    if normalized_role == "root":
        cur.execute("DELETE FROM role_inspection_table_scopes WHERE role = %s;", (normalized_role,))
        return

    normalized_scope_map = normalize_inspection_table_scope_updates(scope_map)
    all_scope_ids = sorted(
        {
            table_id
            for scope_ids in normalized_scope_map.values()
            for table_id in scope_ids
        }
    )
    cur.execute("DELETE FROM role_inspection_table_scopes WHERE role = %s;", (normalized_role,))
    if not all_scope_ids:
        return

    cur.execute("SELECT to_regclass('inspection_tables') AS table_ref;")
    if not cur.fetchone()["table_ref"]:
        raise ValueError("检查表范围需要先完成检查表数据初始化。")

    cur.execute(
        """
        SELECT id
        FROM inspection_tables
        WHERE id = ANY(%s)
          AND is_active = TRUE;
        """,
        (all_scope_ids,),
    )
    valid_ids = {row["id"] for row in cur.fetchall()}
    invalid_ids = [table_id for table_id in all_scope_ids if table_id not in valid_ids]
    if invalid_ids:
        raise ValueError("角色检查表范围中包含不存在或未启用的检查表。")

    for scope_key, scope_ids in normalized_scope_map.items():
        for table_id in scope_ids:
            cur.execute(
                """
                INSERT INTO role_inspection_table_scopes (
                    role,
                    scope_key,
                    inspection_table_id,
                    updated_by,
                    created_at,
                    updated_at
                )
                VALUES (%s, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                ON CONFLICT (role, scope_key, inspection_table_id)
                DO UPDATE SET
                    updated_by = EXCLUDED.updated_by,
                    updated_at = CURRENT_TIMESTAMP;
                """,
                (normalized_role, scope_key, table_id, actor_user_id),
            )


def apply_user_station_region_scope_updates(cur, target_user, scope_map, actor_user_id):
    if is_root_user(target_user):
        cur.execute("DELETE FROM user_station_region_scopes WHERE user_id = %s;", (target_user["id"],))
        return

    normalized_scope_map = normalize_station_region_scope_updates(scope_map)
    all_scope_values = sorted(
        {
            region
            for scope_values in normalized_scope_map.values()
            for region in scope_values
        }
    )
    cur.execute("DELETE FROM user_station_region_scopes WHERE user_id = %s;", (target_user["id"],))
    if not all_scope_values:
        return

    cur.execute(
        """
        SELECT DISTINCT COALESCE(NULLIF(TRIM(region), ''), '未填写片区') AS station_region
        FROM stations;
        """
    )
    valid_regions = {normalize_station_region_value(row["station_region"]) for row in cur.fetchall()}
    invalid_regions = [region for region in all_scope_values if region not in valid_regions]
    if invalid_regions:
        raise ValueError("片区范围中包含不存在的所属片区/归属地。")

    for scope_key, scope_values in normalized_scope_map.items():
        for region in scope_values:
            cur.execute(
                """
                INSERT INTO user_station_region_scopes (
                    user_id,
                    scope_key,
                    station_region,
                    updated_by,
                    created_at,
                    updated_at
                )
                VALUES (%s, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                ON CONFLICT (user_id, scope_key, station_region)
                DO UPDATE SET
                    updated_by = EXCLUDED.updated_by,
                    updated_at = CURRENT_TIMESTAMP;
                """,
                (target_user["id"], scope_key, region, actor_user_id),
            )


def apply_role_station_region_scope_updates(cur, role, scope_map, actor_user_id):
    normalized_role = normalize_text(role)
    if normalized_role not in ROLE_OPTIONS:
        raise ValueError("角色类型不正确。")
    if normalized_role == "root":
        cur.execute("DELETE FROM role_station_region_scopes WHERE role = %s;", (normalized_role,))
        return

    normalized_scope_map = normalize_station_region_scope_updates(scope_map)
    all_scope_values = sorted(
        {
            region
            for scope_values in normalized_scope_map.values()
            for region in scope_values
        }
    )
    cur.execute("DELETE FROM role_station_region_scopes WHERE role = %s;", (normalized_role,))
    if not all_scope_values:
        return

    cur.execute(
        """
        SELECT DISTINCT COALESCE(NULLIF(TRIM(region), ''), '未填写片区') AS station_region
        FROM stations;
        """
    )
    valid_regions = {normalize_station_region_value(row["station_region"]) for row in cur.fetchall()}
    invalid_regions = [region for region in all_scope_values if region not in valid_regions]
    if invalid_regions:
        raise ValueError("角色片区范围中包含不存在的所属片区/归属地。")

    for scope_key, scope_values in normalized_scope_map.items():
        for region in scope_values:
            cur.execute(
                """
                INSERT INTO role_station_region_scopes (
                    role,
                    scope_key,
                    station_region,
                    updated_by,
                    created_at,
                    updated_at
                )
                VALUES (%s, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                ON CONFLICT (role, scope_key, station_region)
                DO UPDATE SET
                    updated_by = EXCLUDED.updated_by,
                    updated_at = CURRENT_TIMESTAMP;
                """,
                (normalized_role, scope_key, region, actor_user_id),
            )


def can_manage_checklist_originals(cur, user):
    return has_permission(cur, user, "manage_checklist_originals")


def can_upload_training_materials(cur, user):
    return has_permission(cur, user, "upload_training_materials")


def can_delete_any_training_material(cur, user):
    return is_root_user(user)


def can_edit_training_material(cur, user, material_row):
    if not user or not material_row:
        return False
    if is_root_user(user):
        return True
    return str(material_row.get("uploaded_by") or "") == str(user.get("id") or "")


def can_delete_training_material(cur, user, material_row):
    return can_edit_training_material(cur, user, material_row) or can_delete_any_training_material(cur, user)


def ensure_training_materials_table(cur):
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS training_materials (
            id SERIAL PRIMARY KEY,
            title TEXT NOT NULL,
            file_type TEXT NOT NULL,
            file_path TEXT NOT NULL,
            original_filename TEXT,
            file_size BIGINT,
            uploaded_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """
    )
    cur.execute(
        """
        ALTER TABLE training_materials
        ADD COLUMN IF NOT EXISTS title TEXT;
        """
    )
    cur.execute(
        """
        ALTER TABLE training_materials
        ADD COLUMN IF NOT EXISTS file_type TEXT;
        """
    )
    cur.execute(
        """
        ALTER TABLE training_materials
        ADD COLUMN IF NOT EXISTS original_filename TEXT;
        """
    )
    cur.execute(
        """
        ALTER TABLE training_materials
        ADD COLUMN IF NOT EXISTS file_size BIGINT;
        """
    )
    cur.execute(
        """
        ALTER TABLE training_materials
        ADD COLUMN IF NOT EXISTS uploaded_by INTEGER REFERENCES users(id) ON DELETE SET NULL;
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_training_materials_updated_at
        ON training_materials (updated_at DESC);
        """
    )


def ensure_inspection_table_original_files_table(cur):
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS inspection_table_original_files (
            id SERIAL PRIMARY KEY,
            inspection_table_id INTEGER NOT NULL REFERENCES inspection_tables(id) ON DELETE CASCADE,
            file_path TEXT NOT NULL,
            original_filename TEXT,
            file_size BIGINT,
            uploaded_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE (inspection_table_id)
        );
        """
    )
    cur.execute(
        """
        ALTER TABLE inspection_table_original_files
        ADD COLUMN IF NOT EXISTS original_filename TEXT;
        """
    )
    cur.execute(
        """
        ALTER TABLE inspection_table_original_files
        ADD COLUMN IF NOT EXISTS file_size BIGINT;
        """
    )
    cur.execute(
        """
        ALTER TABLE inspection_table_original_files
        ADD COLUMN IF NOT EXISTS uploaded_by INTEGER REFERENCES users(id) ON DELETE SET NULL;
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_inspection_table_original_files_table_id
        ON inspection_table_original_files (inspection_table_id);
        """
    )


def ensure_station_certificates_table(cur):
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS station_certificates (
            id SERIAL PRIMARY KEY,
            station_id INTEGER NOT NULL REFERENCES stations(id) ON DELETE CASCADE,
            certificate_type TEXT NOT NULL,
            certificate_name TEXT NOT NULL,
            start_date DATE,
            expiry_date DATE NOT NULL,
            remark TEXT,
            created_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
            updated_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE (station_id, certificate_type)
        );
        """
    )
    cur.execute(
        """
        ALTER TABLE station_certificates
        ADD COLUMN IF NOT EXISTS remark TEXT;
        """
    )
    cur.execute(
        """
        ALTER TABLE station_certificates
        ADD COLUMN IF NOT EXISTS created_by INTEGER REFERENCES users(id) ON DELETE SET NULL;
        """
    )
    cur.execute(
        """
        ALTER TABLE station_certificates
        ADD COLUMN IF NOT EXISTS updated_by INTEGER REFERENCES users(id) ON DELETE SET NULL;
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_station_certificates_expiry_date
        ON station_certificates (expiry_date);
        """
    )


def normalize_optional_date(value):
    value_text = str(value or "").strip()
    if not value_text:
        return None
    try:
        return datetime.strptime(value_text, "%Y-%m-%d").date()
    except ValueError as exc:
        raise ValueError("日期格式必须为 YYYY-MM-DD。") from exc


# === Helper functions for automatic inspection-plan completion writeback ===


def get_plan_period_key_for_date(coverage_type, target_date):
    if coverage_type == "monthly":
        return f"{target_date.year}年{target_date.month}月"

    if coverage_type == "quarterly":
        quarter = ((target_date.month - 1) // 3) + 1
        quarter_label_map = {
            1: "第一季度",
            2: "第二季度",
            3: "第三季度",
            4: "第四季度",
        }
        return (
            f"{target_date.year}年{quarter_label_map.get(quarter, f'第{quarter}季度')}"
        )

    if coverage_type == "yearly":
        return f"{target_date.year}年"

    return None


def get_period_date_range(coverage_type, period_key):
    if coverage_type == "monthly":
        try:
            year_part, month_part = str(period_key).split("年", 1)
            month_text = month_part.replace("月", "").strip()
            year = int(year_part)
            month = int(month_text)
            start_date = datetime(year, month, 1).date()
            if month == 12:
                end_date = datetime(year + 1, 1, 1).date()
            else:
                end_date = datetime(year, month + 1, 1).date()
            return start_date, end_date
        except Exception:
            return None, None

    if coverage_type == "quarterly":
        try:
            year_part, quarter_part = str(period_key).split("年", 1)
            year = int(year_part)
            normalized_quarter_text = (
                quarter_part.replace("季度", "").replace("第", "").strip()
            )
            quarter_map = {
                "1": 1,
                "2": 2,
                "3": 3,
                "4": 4,
                "一": 1,
                "二": 2,
                "三": 3,
                "四": 4,
                "第一": 1,
                "第二": 2,
                "第三": 3,
                "第四": 4,
            }
            quarter = quarter_map.get(normalized_quarter_text)
            if quarter not in {1, 2, 3, 4}:
                return None, None
            start_month = (quarter - 1) * 3 + 1
            start_date = datetime(year, start_month, 1).date()
            if quarter == 4:
                end_date = datetime(year + 1, 1, 1).date()
            else:
                end_date = datetime(year, start_month + 3, 1).date()
            return start_date, end_date
        except Exception:
            return None, None

    if coverage_type == "yearly":
        try:
            year_text = str(period_key).replace("年", "").strip()
            year = int(year_text)
            start_date = datetime(year, 1, 1).date()
            end_date = datetime(year + 1, 1, 1).date()
            return start_date, end_date
        except Exception:
            return None, None

    return None, None


def sync_plan_station_items_completion_by_history(cur, plan_config_id, wait_for_lock=True):
    ensure_inspection_completion_schema(cur)
    try:
        normalized_plan_config_id = int(plan_config_id)
    except (TypeError, ValueError):
        return 0

    if wait_for_lock:
        cur.execute(
            """
            SELECT pg_advisory_xact_lock(%s, %s);
            """,
            (PLAN_COMPLETION_SYNC_LOCK_NAMESPACE, normalized_plan_config_id),
        )
    else:
        cur.execute(
            """
            SELECT pg_try_advisory_xact_lock(%s, %s) AS lock_acquired;
            """,
            (PLAN_COMPLETION_SYNC_LOCK_NAMESPACE, normalized_plan_config_id),
        )
        lock_row = cur.fetchone()
        if not lock_row or not lock_row.get("lock_acquired"):
            return 0

    cur.execute(
        """
        SELECT id, inspection_table_id, coverage_type, period_key
        FROM inspection_plan_configs
        WHERE id = %s
        LIMIT 1;
        """,
        (normalized_plan_config_id,),
    )
    plan_config = cur.fetchone()
    if not plan_config:
        return 0

    start_date, end_date = get_period_date_range(
        plan_config["coverage_type"], plan_config["period_key"]
    )
    if not start_date or not end_date:
        return 0

    cur.execute(
        """
        WITH matched AS (
            SELECT
                psi_inner.id AS psi_id,
                matched_ins.id AS inspection_id,
                matched_ins.inspector_completed_at AS inspection_completed_at,
                matched_ins.updated_at AS inspection_updated_at,
                matched_ins.created_at AS inspection_created_at
            FROM inspection_plan_station_items psi_inner
            LEFT JOIN LATERAL (
                SELECT
                    ins.id,
                    ins.inspector_completed_at,
                    ins.updated_at,
                    ins.created_at
                FROM inspections ins
                WHERE ins.station_id = psi_inner.station_id
                  AND ins.inspection_table_id = %s
                  AND ins.inspection_date >= %s
                  AND ins.inspection_date < %s
                  AND ins.inspector_completion_status = %s
                ORDER BY
                    COALESCE(ins.inspector_completed_at, ins.updated_at, ins.created_at) DESC,
                    ins.inspection_date DESC,
                    ins.id DESC
                LIMIT 1
            ) AS matched_ins ON TRUE
            WHERE psi_inner.plan_config_id = %s
              AND psi_inner.is_included = TRUE
        ),
        desired AS (
            SELECT
                psi_id,
                CASE
                    WHEN inspection_id IS NOT NULL THEN 'completed'
                    ELSE 'pending'
                END AS completion_status,
                inspection_id AS completed_inspection_id,
                CASE
                    WHEN inspection_id IS NOT NULL
                        THEN COALESCE(inspection_completed_at, inspection_updated_at, inspection_created_at, CURRENT_TIMESTAMP)
                    ELSE NULL
                END AS completed_at
            FROM matched
        )
        UPDATE inspection_plan_station_items psi
        SET completion_status = desired.completion_status,
            completed_inspection_id = desired.completed_inspection_id,
            completed_at = desired.completed_at,
            updated_at = CURRENT_TIMESTAMP
        FROM desired
        WHERE psi.id = desired.psi_id
          AND (
              psi.completion_status IS DISTINCT FROM desired.completion_status
              OR psi.completed_inspection_id IS DISTINCT FROM desired.completed_inspection_id
              OR psi.completed_at IS DISTINCT FROM desired.completed_at
          );
        """,
        (
            plan_config["inspection_table_id"],
            start_date,
            end_date,
            INSPECTION_COMPLETION_DONE,
            normalized_plan_config_id,
        ),
    )
    updated_count = max(cur.rowcount or 0, 0)

    cur.execute(
        """
        UPDATE inspection_plan_station_items
        SET completion_status = 'pending',
            completed_inspection_id = NULL,
            completed_at = NULL,
            updated_at = CURRENT_TIMESTAMP
        WHERE plan_config_id = %s
          AND is_included = FALSE
          AND (
              completion_status IS DISTINCT FROM 'pending'
              OR completed_inspection_id IS NOT NULL
              OR completed_at IS NOT NULL
          );
        """,
        (normalized_plan_config_id,),
    )
    updated_count += max(cur.rowcount or 0, 0)
    return updated_count


def mark_related_plan_items_completed(
    cur,
    station_id,
    inspection_table_id,
    inspection_id,
    inspection_date,
):
    coverage_types = ["monthly", "quarterly", "yearly"]

    for coverage_type in coverage_types:
        period_key = get_plan_period_key_for_date(coverage_type, inspection_date)
        if not period_key:
            continue

        cur.execute(
            """
            SELECT pc.id
            FROM inspection_plan_configs pc
            JOIN inspection_plan_station_items psi ON psi.plan_config_id = pc.id
            WHERE pc.inspection_table_id = %s
              AND pc.coverage_type = %s
              AND pc.period_key = %s
              AND psi.station_id = %s
              AND psi.is_included = TRUE;
            """,
            (
                inspection_table_id,
                coverage_type,
                period_key,
                station_id,
            ),
        )
        for row in cur.fetchall():
            sync_plan_station_items_completion_by_history(cur, row["id"])


def normalize_plan_assignment_match_text(value):
    text = sanitize_display_string(str(value or "")).strip().lower()
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"[（）()【】\[\]《》<>·,，。:：;；/\\|_\-]+", "", text)
    return text


PLAN_ASSIGNMENT_TEMPLATE_STATION_HEADER = "站点"
PLAN_ASSIGNMENT_TEMPLATE_LEGACY_STATION_HEADER = DISPLAY_REMOVED_STATION_PHRASE
PLAN_ASSIGNMENT_TEMPLATE_REQUIRED_HEADERS = ("序号", "片区", PLAN_ASSIGNMENT_TEMPLATE_STATION_HEADER)
PLAN_ASSIGNMENT_TEMPLATE_MAX_DATA_ROWS = 1000
PLAN_ASSIGNMENT_TEMPLATE_MAX_ASSIGNMENTS = 3000


def normalize_plan_assignment_template_header(value):
    return re.sub(r"\s+", "", str(value or "").strip())


def plan_assignment_template_cell_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def split_plan_assignment_inspector_text(value):
    text = plan_assignment_template_cell_text(value)
    if not text:
        return []
    parts = re.split(r"[、,，;；/\\|｜\n\r\t ]+", text)
    result = []
    seen = set()
    for part in parts:
        name = str(part or "").strip()
        if not name:
            continue
        key = normalize_plan_assignment_match_text(name)
        if not key or key in seen:
            continue
        seen.add(key)
        result.append(name)
    return result


def parse_plan_assignment_template_file(file_storage):
    if not file_storage or not getattr(file_storage, "filename", ""):
        raise ValueError("请先上传固定派工模板文件。")

    original_filename = str(getattr(file_storage, "filename", "") or "").strip()
    display_filename = original_filename.replace("\\", "/").split("/")[-1] or "未命名文件"
    safe_filename = secure_filename(display_filename) or display_filename
    ext = os.path.splitext(display_filename)[1].lower() or os.path.splitext(safe_filename)[1].lower()
    if not ext:
        mimetype = str(getattr(file_storage, "mimetype", "") or "").lower()
        if "spreadsheetml.sheet" in mimetype:
            ext = ".xlsx"
    if ext == ".xls":
        raise ValueError("暂不支持旧版 .xls 文件，请另存为 .xlsx 后上传。")
    if ext not in {".xlsx", ".xlsm"}:
        raise ValueError("请上传固定模板 .xlsx 或 .xlsm 文件。")

    raw_bytes = file_storage.read()
    if not raw_bytes:
        raise ValueError("上传文件为空，请重新选择文件。")

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("服务器缺少 openpyxl 组件，暂时无法解析 Excel。") from exc

    workbook = None
    try:
        workbook = load_workbook(BytesIO(raw_bytes), read_only=True, data_only=True)
        worksheet = workbook.worksheets[0]
        header_row_number = None
        header_values = []
        normalized_headers = []

        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=10, values_only=True),
            start=1,
        ):
            values = [plan_assignment_template_cell_text(value) for value in row]
            normalized = [normalize_plan_assignment_template_header(value) for value in values]
            has_required_headers = (
                "序号" in normalized
                and "片区" in normalized
                and (
                    PLAN_ASSIGNMENT_TEMPLATE_STATION_HEADER in normalized
                    or PLAN_ASSIGNMENT_TEMPLATE_LEGACY_STATION_HEADER in normalized
                )
            )
            if has_required_headers:
                header_row_number = row_number
                header_values = values
                normalized_headers = normalized
                break

        if header_row_number is None:
            raise ValueError("未找到模板表头，请确认前 3 列包含“序号、片区、站点”。")

        header_positions = {}
        for index, header in enumerate(normalized_headers):
            if header and header not in header_positions:
                header_positions[header] = index

        region_col = header_positions.get("片区")
        station_col = header_positions.get(PLAN_ASSIGNMENT_TEMPLATE_STATION_HEADER)
        if station_col is None:
            station_col = header_positions.get(PLAN_ASSIGNMENT_TEMPLATE_LEGACY_STATION_HEADER)
        table_columns = []
        required_set = set(PLAN_ASSIGNMENT_TEMPLATE_REQUIRED_HEADERS) | {PLAN_ASSIGNMENT_TEMPLATE_LEGACY_STATION_HEADER}
        for column_index, header in enumerate(header_values):
            normalized_header = normalized_headers[column_index] if column_index < len(normalized_headers) else ""
            table_name = sanitize_display_string(str(header or "").strip())
            if column_index <= station_col or not table_name or normalized_header in required_set:
                continue
            table_columns.append({"index": column_index, "table_name": table_name})

        if not table_columns:
            raise ValueError("模板中没有检查表列，请在“站点”后方保留检查表名称列。")

        parsed_rows = []
        max_row = header_row_number + PLAN_ASSIGNMENT_TEMPLATE_MAX_DATA_ROWS
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=header_row_number + 1, max_row=max_row, values_only=True),
            start=header_row_number + 1,
        ):
            region_name = plan_assignment_template_cell_text(row[region_col]) if region_col is not None and region_col < len(row) else ""
            station_name = (
                sanitize_display_string(plan_assignment_template_cell_text(row[station_col]))
                if station_col is not None and station_col < len(row)
                else ""
            )
            has_any_assignment = False
            for table_column in table_columns:
                cell_value = (
                    row[table_column["index"]]
                    if table_column["index"] < len(row)
                    else None
                )
                inspectors = split_plan_assignment_inspector_text(cell_value)
                if not inspectors:
                    continue
                has_any_assignment = True
                if not station_name:
                    parsed_rows.append(
                        {
                            "table_name": table_column["table_name"],
                            "station_name": "",
                            "station_region": region_name,
                            "inspectors": inspectors,
                            "raw_text": f"第{row_number}行缺少站点名称",
                            "source_row_number": row_number,
                        }
                    )
                else:
                    parsed_rows.append(
                        {
                            "table_name": table_column["table_name"],
                            "station_name": station_name,
                            "station_region": region_name,
                            "inspectors": inspectors,
                            "raw_text": f"第{row_number}行：{region_name or '-'} / {station_name} / {table_column['table_name']} / {'、'.join(inspectors)}",
                            "source_row_number": row_number,
                        }
                    )
                if len(parsed_rows) > PLAN_ASSIGNMENT_TEMPLATE_MAX_ASSIGNMENTS:
                    raise ValueError("单次模板最多解析 3000 条派工项，请拆分文件后上传。")

            if not has_any_assignment and not station_name and not region_name:
                continue

        if not parsed_rows:
            raise ValueError("模板里没有填写检查人，请在检查表列下填写督导组成员姓名。")

        return parsed_rows, {
            "filename": display_filename,
            "sheet_name": worksheet.title,
            "header_row": header_row_number,
            "table_columns": [item["table_name"] for item in table_columns],
        }
    finally:
        if workbook is not None:
            workbook.close()


def build_plan_assignment_catalog_context(cur):
    cur.execute(
        """
        SELECT
            id,
            table_name,
            COALESCE(NULLIF(checklist_mode, ''), 'online') AS checklist_mode
        FROM inspection_tables
        WHERE is_active = TRUE
        ORDER BY id ASC;
        """
    )
    table_rows = cur.fetchall()

    cur.execute(
        """
        SELECT
            id,
            station_name,
            COALESCE(region, '') AS region,
            COALESCE(hos_station_code, '') AS hos_station_code,
            COALESCE(monitoring_status, '运行中') AS monitoring_status
        FROM stations
        ORDER BY station_name ASC, id ASC;
        """
    )
    station_rows = cur.fetchall()

    cur.execute(
        """
        SELECT id, username, real_name, phone
        FROM users
        WHERE role = 'supervisor'
        ORDER BY COALESCE(NULLIF(real_name, ''), username), id ASC;
        """
    )
    inspector_rows = cur.fetchall()

    context = {
        "inspection_tables": [
            {
                "id": row["id"],
                "table_name": row["table_name"],
                "mode": normalize_checklist_mode(row.get("checklist_mode")),
                "mode_label": "视频检查"
                if normalize_checklist_mode(row.get("checklist_mode")) == "online"
                else "现场检查",
            }
            for row in table_rows
        ],
        "stations": [
            {
                "id": row["id"],
                "station_name": row["station_name"],
                "region": row["region"],
                "hos_station_code": row["hos_station_code"],
                "monitoring_status": row["monitoring_status"],
            }
            for row in station_rows
        ],
        "inspectors": [
            {
                "id": row["id"],
                "username": row["username"],
                "real_name": row["real_name"],
                "phone": row["phone"],
            }
            for row in inspector_rows
        ],
    }
    return context, table_rows, station_rows, inspector_rows


def build_plan_assignment_alias_map(rows, alias_builder):
    aliases = []
    for row in rows:
        for alias in alias_builder(row):
            key = normalize_plan_assignment_match_text(alias)
            if key:
                aliases.append((key, row))
    aliases.sort(key=lambda item: -len(item[0]))
    return aliases


def find_plan_assignment_catalog_candidates(value, aliases):
    key = normalize_plan_assignment_match_text(value)
    if not key:
        return []
    exact_matches = []
    for alias_key, row in aliases:
        if key == alias_key:
            exact_matches.append(row)
    if exact_matches:
        return unique_plan_assignment_rows(exact_matches)

    partial_matches = []
    for alias_key, row in aliases:
        if len(key) >= 2 and (key in alias_key or alias_key in key):
            partial_matches.append(row)
    return unique_plan_assignment_rows(partial_matches)


def unique_plan_assignment_rows(rows):
    result = []
    seen = set()
    for row in rows:
        row_id = row.get("id") if isinstance(row, dict) else id(row)
        if row_id in seen:
            continue
        seen.add(row_id)
        result.append(row)
    return result


def match_plan_assignment_catalog_row(value, aliases):
    candidates = find_plan_assignment_catalog_candidates(value, aliases)
    return candidates[0] if candidates else None


def match_plan_assignment_station_row(station_name, station_region, aliases):
    candidates = find_plan_assignment_catalog_candidates(station_name, aliases)
    if not candidates:
        return None
    region_key = normalize_plan_assignment_match_text(station_region).replace("片区", "")
    if not region_key:
        return candidates[0]
    region_matches = []
    for row in candidates:
        row_region_key = normalize_plan_assignment_match_text(row.get("region")).replace("片区", "")
        if row_region_key and (region_key in row_region_key or row_region_key in region_key):
            region_matches.append(row)
    return region_matches[0] if region_matches else candidates[0]


def select_current_plan_config(plan_configs, inspection_table_id, target_date):
    candidates = []
    for config in plan_configs:
        if int(config["inspection_table_id"]) != int(inspection_table_id):
            continue
        start_date, end_date = get_period_date_range(
            config["coverage_type"],
            config["period_key"],
        )
        if start_date and end_date and start_date <= target_date < end_date:
            candidates.append(config)
    if not candidates:
        return None, "当前日期没有匹配到该检查表的有效计划，请先建立当前周期计划。"

    priority_map = {"monthly": 1, "quarterly": 2, "yearly": 3}
    candidates.sort(
        key=lambda item: (
            priority_map.get(item["coverage_type"], 9),
            -int(item["id"]),
        )
    )
    selected = candidates[0]
    if len(candidates) > 1:
        return selected, "匹配到多个当前周期计划，已优先选择更细周期计划，请人工复核。"
    return selected, ""


def format_plan_assignment_inspector_names(inspectors):
    return "、".join(
        [
            str(item.get("real_name") or item.get("username") or item.get("id") or "").strip()
            for item in inspectors
            if item
        ]
    )


def fetch_plan_assignment_template_tables(cur, user):
    cur.execute(
        """
        SELECT
            id,
            table_name,
            COALESCE(NULLIF(checklist_mode, ''), 'online') AS checklist_mode
        FROM inspection_tables
        WHERE is_active = TRUE
        ORDER BY id ASC;
        """
    )
    table_rows = cur.fetchall()
    return [
        row
        for row in table_rows
        if is_inspection_table_allowed_for_user(
            cur,
            user,
            row["id"],
            "limit_plan_inspection_table_scope",
        )
    ]


def build_plan_assignment_template_workbook(cur, user):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise RuntimeError("服务器缺少 openpyxl 组件，暂时无法生成模板。") from exc

    table_rows = fetch_plan_assignment_template_tables(cur, user)
    if not table_rows:
        raise ValueError("当前账号没有可维护的检查表，无法生成派工模板。")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "派工模板"

    headers = list(PLAN_ASSIGNMENT_TEMPLATE_REQUIRED_HEADERS) + [
        row["table_name"] for row in table_rows
    ]
    worksheet.append(headers)
    for row_index in range(1, 201):
        worksheet.append([row_index, "", ""] + [""] * len(table_rows))

    header_fill = PatternFill("solid", fgColor="DBEAFE")
    required_fill = PatternFill("solid", fgColor="E0F2FE")
    header_font = Font(bold=True, color="0F172A")
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for column_index, cell in enumerate(worksheet[1], start=1):
        cell.fill = required_fill if column_index <= 3 else header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border

    for row in worksheet.iter_rows(min_row=2, max_row=201, max_col=len(headers)):
        for cell in row:
            cell.alignment = center_alignment
            cell.border = thin_border

    worksheet.freeze_panes = "D2"
    worksheet.auto_filter.ref = f"A1:{excel_column_name(len(headers))}201"
    worksheet.column_dimensions["A"].width = 10
    worksheet.column_dimensions["B"].width = 14
    worksheet.column_dimensions["C"].width = 20
    for column_index in range(4, len(headers) + 1):
        worksheet.column_dimensions[excel_column_name(column_index)].width = 32
    worksheet.row_dimensions[1].height = 38

    guide = workbook.create_sheet("填写说明")
    guide_rows = [
        ["填写规则", "说明"],
        ["片区", "填写站点所属片区/归属地，可帮助系统更准确匹配站点。"],
        ["站点", "填写站点名称，可填写简称，系统会结合片区匹配真实站点。"],
        ["检查表列", "每个检查表列下填写负责检查人姓名；多人可用“、”分隔。"],
        ["写入规则", "上传后只生成预览清单，必须人工确认后才会写入当前日期所在周期的计划。"],
        ["注意事项", "请不要修改表头名称；旧版 .xls 文件请另存为 .xlsx 后上传。"],
    ]
    for row in guide_rows:
        guide.append(row)
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 86
    for row in guide.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_border
    for cell in guide[1]:
        cell.fill = header_fill
        cell.font = header_font

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)
    return output


def resolve_plan_assignment_preview_rows(cur, user, parsed_rows):
    context, table_rows, station_rows, inspector_rows = build_plan_assignment_catalog_context(cur)

    table_aliases = build_plan_assignment_alias_map(
        table_rows,
        lambda row: [
            row["table_name"],
            f"{row['table_name']}{'视频检查' if normalize_checklist_mode(row.get('checklist_mode')) == 'online' else '现场检查'}",
            f"{row['table_name']}（{'视频' if normalize_checklist_mode(row.get('checklist_mode')) == 'online' else '现场'}）",
        ],
    )
    station_aliases = build_plan_assignment_alias_map(
        station_rows,
        lambda row: [row["station_name"], row.get("hos_station_code")],
    )
    inspector_aliases = build_plan_assignment_alias_map(
        inspector_rows,
        lambda row: [row.get("real_name"), row.get("username"), row.get("phone")],
    )

    cur.execute(
        """
        SELECT
            pc.id,
            pc.inspection_table_id,
            it.table_name AS inspection_table_name,
            COALESCE(NULLIF(it.checklist_mode, ''), 'online') AS checklist_mode,
            pc.coverage_type,
            pc.period_key,
            pc.status
        FROM inspection_plan_configs pc
        JOIN inspection_tables it ON it.id = pc.inspection_table_id
        WHERE pc.status = 'active'
        ORDER BY pc.id DESC;
        """
    )
    plan_configs = cur.fetchall()
    today = datetime.now(BEIJING_TZ).date()

    initial_rows = []
    plan_station_pairs = []
    for index, raw_row in enumerate(parsed_rows or [], start=1):
        row = raw_row if isinstance(raw_row, dict) else {}
        messages = []
        status = "ready"

        table_row = match_plan_assignment_catalog_row(row.get("table_name"), table_aliases)
        station_row = match_plan_assignment_station_row(
            row.get("station_name"),
            row.get("station_region"),
            station_aliases,
        )

        raw_inspectors = row.get("inspectors") if isinstance(row.get("inspectors"), list) else []
        matched_inspectors = []
        seen_inspector_ids = set()
        for raw_inspector in raw_inspectors:
            inspector_row = match_plan_assignment_catalog_row(raw_inspector, inspector_aliases)
            if not inspector_row:
                continue
            inspector_id = int(inspector_row["id"])
            if inspector_id in seen_inspector_ids:
                continue
            seen_inspector_ids.add(inspector_id)
            matched_inspectors.append(inspector_row)

        selected_plan = None
        plan_warning = ""
        if not table_row:
            status = "error"
            messages.append("未匹配到检查表。")
        elif not is_inspection_table_allowed_for_user(
            cur,
            user,
            table_row["id"],
            "limit_plan_inspection_table_scope",
        ):
            status = "error"
            messages.append("当前账号无权维护该检查表计划。")
        else:
            selected_plan, plan_warning = select_current_plan_config(
                plan_configs,
                table_row["id"],
                today,
            )
            if not selected_plan:
                status = "error"
                messages.append(plan_warning)
            elif plan_warning:
                status = "warning"
                messages.append(plan_warning)

        if not station_row:
            status = "error"
            messages.append("未匹配到站点。")
        elif not is_station_region_allowed_for_user(
            cur,
            user,
            station_row.get("region"),
            "limit_plan_station_region_scope",
        ):
            status = "error"
            messages.append("当前账号无权维护该片区站点计划。")

        if not matched_inspectors:
            status = "error"
            messages.append("未匹配到督导组检查人。")

        if selected_plan and station_row:
            mode = normalize_checklist_mode(selected_plan.get("checklist_mode"))
            if mode == "online" and station_row.get("monitoring_status") == "未运行":
                status = "error"
                messages.append("该站点监控未运行，视频检查计划不可分配。")
            plan_station_pairs.append((selected_plan["id"], station_row["id"]))

        initial_rows.append(
            {
                "source_index": index,
                "raw_text": row.get("raw_text") or "",
                "source_table_name": row.get("table_name") or "",
                "source_station_name": row.get("station_name") or "",
                "source_station_region": row.get("station_region") or "",
                "source_inspector_text": "、".join(raw_inspectors),
                "inspection_table_id": table_row["id"] if table_row else None,
                "inspection_table_name": table_row["table_name"] if table_row else "",
                "checklist_mode": normalize_checklist_mode(table_row.get("checklist_mode")) if table_row else "",
                "checklist_mode_label": (
                    "视频检查"
                    if table_row and normalize_checklist_mode(table_row.get("checklist_mode")) == "online"
                    else "现场检查"
                    if table_row
                    else ""
                ),
                "plan_config_id": selected_plan["id"] if selected_plan else None,
                "coverage_type": selected_plan["coverage_type"] if selected_plan else "",
                "coverage_type_label": COVERAGE_TYPE_LABELS.get(selected_plan["coverage_type"], selected_plan["coverage_type"]) if selected_plan else "",
                "period_key": selected_plan["period_key"] if selected_plan else "",
                "station_id": station_row["id"] if station_row else None,
                "station_name": station_row["station_name"] if station_row else "",
                "station_region": station_row["region"] if station_row else "",
                "monitoring_status": station_row["monitoring_status"] if station_row else "",
                "inspector_ids": [item["id"] for item in matched_inspectors],
                "inspectors": [
                    {
                        "id": item["id"],
                        "username": item["username"],
                        "real_name": item["real_name"],
                        "phone": item["phone"],
                        "display_name": item["real_name"] or item["username"],
                    }
                    for item in matched_inspectors
                ],
                "status": status,
                "message": " ".join(messages) or "已匹配到当前周期计划，待人工确认。",
                "can_apply": status in {"ready", "warning"},
            }
        )

    item_map = {}
    if plan_station_pairs:
        plan_ids = sorted({int(plan_id) for plan_id, _station_id in plan_station_pairs})
        station_ids = sorted({int(station_id) for _plan_id, station_id in plan_station_pairs})
        cur.execute(
            """
            SELECT
                plan_config_id,
                station_id,
                is_included,
                assigned_inspector_id,
                COALESCE(assigned_inspector_ids, '[]'::jsonb) AS assigned_inspector_ids,
                completion_status,
                TO_CHAR(completed_at, 'YYYY-MM-DD HH24:MI') AS completed_at
            FROM inspection_plan_station_items
            WHERE plan_config_id = ANY(%s)
              AND station_id = ANY(%s);
            """,
            (plan_ids, station_ids),
        )
        item_map = {
            (int(row["plan_config_id"]), int(row["station_id"])): row
            for row in cur.fetchall()
        }

    merged_rows = []
    merged_key_map = {}
    for row in initial_rows:
        key = (row.get("plan_config_id"), row.get("station_id"))
        existing_item = item_map.get(key)
        if existing_item:
            row["previous_is_included"] = bool(existing_item["is_included"])
            row["previous_inspector_ids"] = normalize_plan_assigned_inspector_ids(existing_item)
            row["previous_completion_status"] = existing_item["completion_status"]
            row["completed_at"] = existing_item["completed_at"]
            if existing_item["completion_status"] == "completed":
                row["status"] = "error"
                row["can_apply"] = False
                row["message"] = "该站点当前周期任务已完成，不可再调整派工。"
        else:
            row["previous_is_included"] = False
            row["previous_inspector_ids"] = []
            row["previous_completion_status"] = "pending"
            row["completed_at"] = ""

        if not row.get("can_apply"):
            merged_rows.append(row)
            continue

        merge_key = (row["plan_config_id"], row["station_id"])
        if merge_key not in merged_key_map:
            merged_key_map[merge_key] = row
            merged_rows.append(row)
            continue

        target = merged_key_map[merge_key]
        existing_ids = {int(value) for value in target.get("inspector_ids") or []}
        for inspector in row.get("inspectors") or []:
            inspector_id = int(inspector["id"])
            if inspector_id in existing_ids:
                continue
            target["inspector_ids"].append(inspector_id)
            target["inspectors"].append(inspector)
            existing_ids.add(inspector_id)
        if row.get("raw_text"):
            target["raw_text"] = "；".join([item for item in [target.get("raw_text"), row.get("raw_text")] if item])[:120]
        target["source_inspector_text"] = format_plan_assignment_inspector_names(target["inspectors"])
        target["message"] = "已合并同一站点同一检查表的多名检查人，请人工复核。"
        target["status"] = "warning"

    for row in merged_rows:
        previous_names = []
        previous_ids = row.get("previous_inspector_ids") or []
        for previous_id in previous_ids:
            inspector = next((item for item in inspector_rows if int(item["id"]) == int(previous_id)), None)
            if inspector:
                previous_names.append(inspector["real_name"] or inspector["username"])
        row["previous_inspector_names"] = "、".join(previous_names)
        row["inspector_names"] = format_plan_assignment_inspector_names(row.get("inspectors") or [])
    return merged_rows, context, today


def validate_plan_assignment_apply_row(cur, user, row, target_date):
    plan_config_id = int(row.get("plan_config_id") or 0)
    station_id = int(row.get("station_id") or 0)
    inspector_ids = normalize_plan_assigned_inspector_ids(
        {"assigned_inspector_ids": row.get("inspector_ids") or []}
    )
    if plan_config_id <= 0 or station_id <= 0 or not inspector_ids:
        raise ValueError("存在缺少计划、站点或检查人的派工项。")

    cur.execute(
        """
        SELECT
            pc.id,
            pc.inspection_table_id,
            it.table_name AS inspection_table_name,
            COALESCE(NULLIF(it.checklist_mode, ''), 'online') AS checklist_mode,
            pc.coverage_type,
            pc.period_key,
            pc.status
        FROM inspection_plan_configs pc
        JOIN inspection_tables it ON it.id = pc.inspection_table_id
        WHERE pc.id = %s
        LIMIT 1;
        """,
        (plan_config_id,),
    )
    plan_config = cur.fetchone()
    if not plan_config or plan_config["status"] != "active":
        raise ValueError("存在无效或已停用的巡检计划。")
    if not is_inspection_table_allowed_for_user(
        cur,
        user,
        plan_config["inspection_table_id"],
        "limit_plan_inspection_table_scope",
    ):
        raise PermissionError("当前账号无权维护某张检查表计划。")

    start_date, end_date = get_period_date_range(
        plan_config["coverage_type"],
        plan_config["period_key"],
    )
    if not start_date or not end_date or not (start_date <= target_date < end_date):
        raise ValueError("存在不属于当前日期周期的计划，请重新生成预览后再写入。")

    cur.execute(
        """
        SELECT
            id,
            station_name,
            COALESCE(region, '') AS region,
            COALESCE(monitoring_status, '运行中') AS monitoring_status
        FROM stations
        WHERE id = %s
        LIMIT 1;
        """,
        (station_id,),
    )
    station = cur.fetchone()
    if not station:
        raise ValueError("存在无效站点。")
    if not is_station_region_allowed_for_user(
        cur,
        user,
        station.get("region"),
        "limit_plan_station_region_scope",
    ):
        raise PermissionError("当前账号无权维护某个片区站点计划。")
    if normalize_checklist_mode(plan_config.get("checklist_mode")) == "online" and station.get("monitoring_status") == "未运行":
        raise ValueError(f"【{station['station_name']}】监控未运行，不能分配视频检查任务。")

    cur.execute(
        """
        SELECT id
        FROM users
        WHERE id = ANY(%s)
          AND role = 'supervisor';
        """,
        (inspector_ids,),
    )
    existing_inspector_ids = {int(item["id"]) for item in cur.fetchall()}
    if any(inspector_id not in existing_inspector_ids for inspector_id in inspector_ids):
        raise ValueError("只能分配给督导组角色账号。")

    cur.execute(
        """
        SELECT id, completion_status
        FROM inspection_plan_station_items
        WHERE plan_config_id = %s
          AND station_id = %s
        FOR UPDATE;
        """,
        (plan_config_id, station_id),
    )
    existing_item = cur.fetchone()
    if existing_item and existing_item["completion_status"] == "completed":
        raise ValueError(f"【{station['station_name']}】当前周期任务已完成，不能调整派工。")

    return plan_config, station, inspector_ids


# === Inspection Plan Configs API ===


@app.route("/api/inspection-plan-configs")
def get_inspection_plan_configs():
    user_id = str(request.args.get("user_id", "")).strip()
    inspection_table_id = str(request.args.get("inspection_table_id", "")).strip()
    coverage_type = str(request.args.get("coverage_type", "")).strip()
    period_key = str(request.args.get("period_key", "")).strip()

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()

        user = get_user_by_id(cur, user_id) if user_id else None
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404
        if not has_permission(cur, user, "view_inspection_plans"):
            return jsonify({"success": False, "error": "当前账号无权查看巡检计划。"}), 403

        ensure_inspection_completion_schema(cur)
        ensure_inspection_plan_assignment_schema(cur)

        where_clauses = []
        params = []

        if inspection_table_id:
            where_clauses.append("pc.inspection_table_id = %s")
            params.append(inspection_table_id)

        if coverage_type:
            where_clauses.append("pc.coverage_type = %s")
            params.append(coverage_type)

        if period_key:
            where_clauses.append("pc.period_key = %s")
            params.append(period_key)

        if not append_inspection_table_scope_filter(
            cur,
            user,
            where_clauses,
            params,
            "pc.inspection_table_id",
            "limit_plan_inspection_table_scope",
        ):
            return jsonify({"success": True, "items": []})
        if not append_station_region_scope_filter(
            cur,
            user,
            where_clauses,
            params,
            "ps.region",
            "limit_plan_station_region_scope",
        ):
            return jsonify({"success": True, "items": []})

        where_sql = ""
        if where_clauses:
            where_sql = "WHERE " + " AND ".join(where_clauses)

        cur.execute(
            sql.SQL(
                """
                SELECT
                    pc.id,
                    pc.inspection_table_id,
                    it.table_name AS inspection_table_name,
                    it.checklist_mode,
                    pc.coverage_type,
                    pc.period_key,
                    pc.status,
                    pc.remark,
                    creator.username AS created_by_username,
                    updater.username AS updated_by_username,
                    COALESCE(SUM(CASE WHEN psi.is_included = TRUE THEN 1 ELSE 0 END), 0) AS included_station_count,
                    COALESCE(SUM(CASE WHEN psi.is_included = TRUE AND psi.completion_status = 'completed' THEN 1 ELSE 0 END), 0) AS completed_station_count,
                    COALESCE(SUM(CASE WHEN psi.is_included = TRUE AND psi.completion_status = 'pending' THEN 1 ELSE 0 END), 0) AS pending_station_count,
                    pc.created_at,
                    pc.updated_at
                FROM inspection_plan_configs pc
                JOIN inspection_tables it ON pc.inspection_table_id = it.id
                JOIN users creator ON pc.created_by = creator.id
                LEFT JOIN users updater ON pc.updated_by = updater.id
                LEFT JOIN inspection_plan_station_items psi ON psi.plan_config_id = pc.id
                LEFT JOIN stations ps ON ps.id = psi.station_id
                {where_sql}
                GROUP BY
                    pc.id,
                    pc.inspection_table_id,
                    it.table_name,
                    it.checklist_mode,
                    pc.coverage_type,
                    pc.period_key,
                    pc.status,
                    pc.remark,
                    creator.username,
                    updater.username,
                    pc.created_at,
                    pc.updated_at
                ORDER BY pc.id DESC;
                """
            ).format(where_sql=sql.SQL(where_sql)),
            params,
        )
        rows = cur.fetchall()

        result = []
        for row in rows:
            included_count = int(row["included_station_count"] or 0)
            completed_count = int(row["completed_station_count"] or 0)
            pending_count = int(row["pending_station_count"] or 0)
            completion_rate = (
                round((completed_count / included_count) * 100)
                if included_count > 0
                else 0
            )
            result.append(
                {
                    "id": row["id"],
                    "inspection_table_id": row["inspection_table_id"],
                    "inspection_table_name": row["inspection_table_name"],
                    "checklist_mode": normalize_checklist_mode(row.get("checklist_mode")),
                    "checklist_mode_label": "视频检查" if normalize_checklist_mode(row.get("checklist_mode")) == "online" else "现场检查",
                    "coverage_type": row["coverage_type"],
                    "coverage_type_label": COVERAGE_TYPE_LABELS.get(
                        row["coverage_type"], row["coverage_type"]
                    ),
                    "period_key": row["period_key"],
                    "status": row["status"],
                    "remark": row["remark"],
                    "created_by_username": row["created_by_username"],
                    "updated_by_username": row["updated_by_username"],
                    "included_station_count": included_count,
                    "completed_station_count": completed_count,
                    "pending_station_count": pending_count,
                    "completion_rate": completion_rate,
                    "created_at": row["created_at"],
                    "updated_at": row["updated_at"],
                }
            )

        return jsonify({"success": True, "items": result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/inspection-plan-configs/<int:plan_config_id>")
def get_inspection_plan_config_detail(plan_config_id):
    user_id = str(request.args.get("user_id", "")).strip()

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()

        user = get_user_by_id(cur, user_id) if user_id else None
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404
        if not has_permission(cur, user, "view_inspection_plans"):
            return jsonify({"success": False, "error": "当前账号无权查看巡检计划。"}), 403

        cur.execute(
            """
            SELECT
                pc.id,
                pc.inspection_table_id,
                it.table_name AS inspection_table_name,
                it.checklist_mode,
                pc.coverage_type,
                pc.period_key,
                pc.status,
                pc.remark,
                creator.username AS created_by_username,
                updater.username AS updated_by_username,
                pc.created_at,
                pc.updated_at
            FROM inspection_plan_configs pc
            JOIN inspection_tables it ON pc.inspection_table_id = it.id
            JOIN users creator ON pc.created_by = creator.id
            LEFT JOIN users updater ON pc.updated_by = updater.id
            WHERE pc.id = %s
            LIMIT 1;
            """,
            (plan_config_id,),
        )
        config_row = cur.fetchone()

        if not config_row:
            return jsonify({"success": False, "error": "巡检计划配置不存在。"}), 404
        if not is_inspection_table_allowed_for_user(
            cur,
            user,
            config_row["inspection_table_id"],
            "limit_plan_inspection_table_scope",
        ):
            return jsonify({"success": False, "error": "当前账号无权查看该检查表的巡检计划。"}), 403

        ensure_inspection_completion_schema(cur)
        ensure_inspection_plan_assignment_schema(cur)

        station_where_clauses = ["psi.plan_config_id = %s"]
        station_params = [plan_config_id]
        if not append_station_region_scope_filter(
            cur,
            user,
            station_where_clauses,
            station_params,
            "s.region",
            "limit_plan_station_region_scope",
        ):
            station_rows = []
        else:
            station_where_sql = " AND ".join(station_where_clauses)
            cur.execute(
                sql.SQL(
                    """
            SELECT
                psi.id,
                psi.station_id,
                s.station_name,
                s.region,
                s.address,
                COALESCE(s.monitoring_status, '运行中') AS monitoring_status,
                psi.is_included,
                psi.assigned_inspector_id,
                COALESCE(assigned_group.assigned_inspector_ids, '[]'::jsonb) AS assigned_inspector_ids,
                COALESCE(assigned_group.assigned_inspectors, '[]'::jsonb) AS assigned_inspectors,
                assigned_group.assigned_inspector_usernames AS assigned_inspector_username,
                assigned_group.assigned_inspector_names AS assigned_inspector_name,
                assigned_group.assigned_inspector_phones AS assigned_inspector_phone,
                TO_CHAR(psi.assigned_at, 'YYYY-MM-DD HH24:MI') AS assigned_at,
                psi.completion_status,
                psi.completed_inspection_id,
                TO_CHAR(psi.completed_at, 'YYYY-MM-DD HH24:MI') AS completed_at,
                psi.note
            FROM inspection_plan_station_items psi
            JOIN stations s ON psi.station_id = s.id
            LEFT JOIN LATERAL (
                SELECT
                    JSONB_AGG(assigned_user.id ORDER BY COALESCE(assigned_user.real_name, assigned_user.username, assigned_user.id::text)) AS assigned_inspector_ids,
                    JSONB_AGG(
                        JSONB_BUILD_OBJECT(
                            'id', assigned_user.id,
                            'username', assigned_user.username,
                            'real_name', assigned_user.real_name,
                            'phone', assigned_user.phone,
                            'display_name', COALESCE(assigned_user.real_name, assigned_user.username, assigned_user.id::text)
                        )
                        ORDER BY COALESCE(assigned_user.real_name, assigned_user.username, assigned_user.id::text)
                    ) AS assigned_inspectors,
                    STRING_AGG(assigned_user.username, '、' ORDER BY COALESCE(assigned_user.real_name, assigned_user.username, assigned_user.id::text)) AS assigned_inspector_usernames,
                    STRING_AGG(COALESCE(assigned_user.real_name, assigned_user.username, assigned_user.id::text), '、' ORDER BY COALESCE(assigned_user.real_name, assigned_user.username, assigned_user.id::text)) AS assigned_inspector_names,
                    STRING_AGG(assigned_user.phone, '、' ORDER BY COALESCE(assigned_user.real_name, assigned_user.username, assigned_user.id::text)) AS assigned_inspector_phones
                FROM (
                    SELECT DISTINCT inspector_id
                    FROM (
                        SELECT jsonb_array_elements_text(COALESCE(psi.assigned_inspector_ids, '[]'::jsonb))::integer AS inspector_id
                        UNION ALL
                        SELECT psi.assigned_inspector_id
                        WHERE psi.assigned_inspector_id IS NOT NULL
                    ) raw_assigned_ids
                ) assigned_ids
                JOIN users assigned_user ON assigned_user.id = assigned_ids.inspector_id
            ) assigned_group ON TRUE
            WHERE {station_where_sql}
            ORDER BY s.id ASC;
            """,
                ).format(station_where_sql=sql.SQL(station_where_sql)),
                station_params,
            )
            station_rows = cur.fetchall()

        included_count = sum(1 for row in station_rows if row["is_included"])
        completed_count = sum(
            1
            for row in station_rows
            if row["is_included"] and row["completion_status"] == "completed"
        )
        pending_count = sum(
            1
            for row in station_rows
            if row["is_included"] and row["completion_status"] == "pending"
        )
        completion_rate = (
            round((completed_count / included_count) * 100) if included_count > 0 else 0
        )

        return jsonify(
            {
                "success": True,
                "item": {
                    "id": config_row["id"],
                    "inspection_table_id": config_row["inspection_table_id"],
                    "inspection_table_name": config_row["inspection_table_name"],
                    "checklist_mode": normalize_checklist_mode(config_row.get("checklist_mode")),
                    "checklist_mode_label": "视频检查" if normalize_checklist_mode(config_row.get("checklist_mode")) == "online" else "现场检查",
                    "coverage_type": config_row["coverage_type"],
                    "coverage_type_label": COVERAGE_TYPE_LABELS.get(
                        config_row["coverage_type"], config_row["coverage_type"]
                    ),
                    "period_key": config_row["period_key"],
                    "status": config_row["status"],
                    "remark": config_row["remark"],
                    "created_by_username": config_row["created_by_username"],
                    "updated_by_username": config_row["updated_by_username"],
                    "included_station_count": included_count,
                    "completed_station_count": completed_count,
                    "pending_station_count": pending_count,
                    "completion_rate": completion_rate,
                    "created_at": config_row["created_at"],
                    "updated_at": config_row["updated_at"],
                    "stations": station_rows,
                },
            }
        )
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route(
    "/api/inspection-plan-configs/<int:plan_config_id>/stations", methods=["PUT"]
)
def save_inspection_plan_config_stations(plan_config_id):
    data = request.get_json(silent=True) or {}
    user_id = str(data.get("user_id", "")).strip()
    stations = data.get("stations") or []

    if not user_id:
        return jsonify({"success": False, "error": "缺少用户信息。"}), 400

    if not isinstance(stations, list):
        return jsonify({"success": False, "error": "stations 参数格式不正确。"}), 400

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()

        user = get_user_by_id(cur, user_id)
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404
        current_user_id = int(user["id"])

        if not can_manage_plan(cur, user):
            return (
                jsonify({"success": False, "error": "当前账号无权维护巡检计划。"}),
                403,
            )

        ensure_inspection_plan_assignment_schema(cur)

        cur.execute(
            """
            SELECT
                pc.id,
                pc.inspection_table_id,
                COALESCE(NULLIF(it.checklist_mode, ''), 'online') AS checklist_mode
            FROM inspection_plan_configs pc
            JOIN inspection_tables it ON it.id = pc.inspection_table_id
            WHERE pc.id = %s
            LIMIT 1;
            """,
            (plan_config_id,),
        )
        config_row = cur.fetchone()

        if not config_row:
            return jsonify({"success": False, "error": "巡检计划配置不存在。"}), 404
        if not is_inspection_table_allowed_for_user(
            cur,
            user,
            config_row["inspection_table_id"],
            "limit_plan_inspection_table_scope",
        ):
            return jsonify({"success": False, "error": "当前账号无权维护该检查表的巡检计划。"}), 403

        station_ids = []
        normalized_items = []
        seen_station_ids = set()
        inspector_ids = []
        for item in stations:
            if not isinstance(item, dict):
                return (
                    jsonify({"success": False, "error": "stations 中存在非法项。"}),
                    400,
                )

            station_id = item.get("station_id")
            is_included = bool(item.get("is_included", True))
            note = str(item.get("note", "")).strip() or None
            assigned_inspector_ids = normalize_plan_assigned_inspector_ids(item)
            assigned_inspector_id = assigned_inspector_ids[0] if assigned_inspector_ids else None

            if not station_id:
                return (
                    jsonify(
                        {
                            "success": False,
                            "error": "stations 中存在缺少 station_id 的项。",
                        }
                    ),
                    400,
                )
            try:
                station_id = int(station_id)
            except (TypeError, ValueError):
                return jsonify({"success": False, "error": "stations 中存在非法 station_id。"}), 400
            if station_id in seen_station_ids:
                return jsonify({"success": False, "error": "stations 中存在重复站点。"}), 400
            seen_station_ids.add(station_id)

            for inspector_id in assigned_inspector_ids:
                if inspector_id not in inspector_ids:
                    inspector_ids.append(inspector_id)

            station_ids.append(station_id)
            normalized_items.append(
                {
                    "station_id": station_id,
                    "is_included": is_included,
                    "assigned_inspector_id": assigned_inspector_id if is_included else None,
                    "assigned_inspector_ids": assigned_inspector_ids if is_included else [],
                    "note": note,
                }
            )

        station_map = {}
        if station_ids:
            cur.execute(
                """
                SELECT
                    id,
                    station_name,
                    COALESCE(monitoring_status, '运行中') AS monitoring_status
                FROM stations
                WHERE id = ANY(%s);
                """,
                (station_ids,),
            )
            existing_station_rows = cur.fetchall()
            station_map = {row["id"]: row for row in existing_station_rows}
            existing_station_ids = set(station_map)
            missing_station_ids = [
                sid for sid in station_ids if sid not in existing_station_ids
            ]
            if missing_station_ids:
                return (
                    jsonify(
                        {
                            "success": False,
                            "error": f"以下站点不存在：{', '.join(str(x) for x in missing_station_ids)}",
                        }
                    ),
                    400,
                )

        is_online_plan = normalize_checklist_mode(config_row.get("checklist_mode")) == "online"
        blocked_station_names = [
            station_map[item["station_id"]]["station_name"]
            for item in normalized_items
            if is_online_plan
            and item["is_included"]
            and station_map.get(item["station_id"])
            and station_map[item["station_id"]].get("monitoring_status") == "未运行"
        ]
        if blocked_station_names:
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "以下站点没有监控，目前是建立视频检查任务，无法勾选纳入计划："
                        + "、".join(blocked_station_names),
                    }
                ),
                400,
            )

        if inspector_ids:
            cur.execute(
                """
                SELECT id
                FROM users
                WHERE id = ANY(%s)
                  AND role = 'supervisor';
                """,
                (inspector_ids,),
            )
            existing_inspector_ids = {row["id"] for row in cur.fetchall()}
            missing_inspector_ids = [
                inspector_id
                for inspector_id in inspector_ids
                if inspector_id not in existing_inspector_ids
            ]
            if missing_inspector_ids:
                return jsonify({"success": False, "error": "只能分配给督导组角色账号。"}), 400

        if station_ids:
            cur.execute(
                """
                DELETE FROM inspection_plan_station_items
                WHERE plan_config_id = %s
                  AND NOT (station_id = ANY(%s));
                """,
                (plan_config_id, station_ids),
            )
        else:
            cur.execute(
                "DELETE FROM inspection_plan_station_items WHERE plan_config_id = %s;",
                (plan_config_id,),
            )

        for item in normalized_items:
            cur.execute(
                """
                INSERT INTO inspection_plan_station_items (
                    plan_config_id,
                    station_id,
                    is_included,
                    assigned_inspector_id,
                    assigned_inspector_ids,
                    assigned_by,
                    assigned_at,
                    completion_status,
                    completed_inspection_id,
                    completed_at,
                    note,
                    updated_at
                )
                VALUES (
                    %s,
                    %s,
                    %s,
                    %s,
                    %s::jsonb,
                    CASE WHEN %s::integer IS NULL THEN NULL ELSE %s::integer END,
                    CASE WHEN %s::integer IS NULL THEN NULL ELSE CURRENT_TIMESTAMP END,
                    'pending',
                    NULL,
                    NULL,
                    %s,
                    CURRENT_TIMESTAMP
                )
                ON CONFLICT (plan_config_id, station_id)
                DO UPDATE SET
                    is_included = CASE
                        WHEN inspection_plan_station_items.completion_status = 'completed'
                            THEN inspection_plan_station_items.is_included
                        ELSE EXCLUDED.is_included
                    END,
                    assigned_inspector_id = CASE
                        WHEN inspection_plan_station_items.completion_status = 'completed'
                            THEN inspection_plan_station_items.assigned_inspector_id
                        WHEN EXCLUDED.is_included = FALSE
                            THEN NULL
                        ELSE EXCLUDED.assigned_inspector_id
                    END,
                    assigned_inspector_ids = CASE
                        WHEN inspection_plan_station_items.completion_status = 'completed'
                            THEN inspection_plan_station_items.assigned_inspector_ids
                        WHEN EXCLUDED.is_included = FALSE
                            THEN '[]'::jsonb
                        ELSE EXCLUDED.assigned_inspector_ids
                    END,
                    assigned_by = CASE
                        WHEN inspection_plan_station_items.completion_status = 'completed'
                            THEN inspection_plan_station_items.assigned_by
                        WHEN EXCLUDED.is_included = FALSE
                            THEN NULL
                        WHEN EXCLUDED.assigned_inspector_id IS NULL
                            THEN NULL
                        WHEN EXCLUDED.assigned_inspector_ids IS DISTINCT FROM inspection_plan_station_items.assigned_inspector_ids
                            THEN EXCLUDED.assigned_by
                        ELSE inspection_plan_station_items.assigned_by
                    END,
                    assigned_at = CASE
                        WHEN inspection_plan_station_items.completion_status = 'completed'
                            THEN inspection_plan_station_items.assigned_at
                        WHEN EXCLUDED.is_included = FALSE
                            THEN NULL
                        WHEN EXCLUDED.assigned_inspector_id IS NULL
                            THEN NULL
                        WHEN EXCLUDED.assigned_inspector_id IS NOT NULL
                             AND (
                                EXCLUDED.assigned_inspector_ids IS DISTINCT FROM inspection_plan_station_items.assigned_inspector_ids
                                OR inspection_plan_station_items.assigned_at IS NULL
                             )
                            THEN CURRENT_TIMESTAMP
                        ELSE inspection_plan_station_items.assigned_at
                    END,
                    completion_status = CASE
                        WHEN inspection_plan_station_items.completion_status = 'completed'
                            THEN inspection_plan_station_items.completion_status
                        ELSE 'pending'
                    END,
                    completed_inspection_id = CASE
                        WHEN inspection_plan_station_items.completion_status = 'completed'
                            THEN inspection_plan_station_items.completed_inspection_id
                        ELSE NULL
                    END,
                    completed_at = CASE
                        WHEN inspection_plan_station_items.completion_status = 'completed'
                            THEN inspection_plan_station_items.completed_at
                        ELSE NULL
                    END,
                    note = EXCLUDED.note,
                    updated_at = CURRENT_TIMESTAMP;
                """,
                (
                    plan_config_id,
                    item["station_id"],
                    item["is_included"],
                    item["assigned_inspector_id"],
                    json.dumps(item["assigned_inspector_ids"], ensure_ascii=False),
                    item["assigned_inspector_id"],
                    current_user_id,
                    item["assigned_inspector_id"],
                    item["note"],
                ),
            )

        cur.execute(
            """
            UPDATE inspection_plan_configs
            SET updated_by = %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = %s;
            """,
            (current_user_id, plan_config_id),
        )

        sync_plan_station_items_completion_by_history(cur, plan_config_id)
        conn.commit()
        return jsonify({"success": True, "message": "巡检计划站点明细保存成功。"})
    except ValueError as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/inspection-plan-inspectors")
def get_inspection_plan_inspectors():
    conn = None
    cur = None

    try:
        current_user = get_current_request_user()
        conn = get_db_connection()
        cur = conn.cursor()

        if not can_manage_plan(cur, current_user):
            return jsonify({"success": False, "error": "当前账号无权维护巡检计划。"}), 403

        cur.execute(
            """
            SELECT
                id,
                username,
                real_name,
                phone
            FROM users
            WHERE role = 'supervisor'
            ORDER BY COALESCE(NULLIF(real_name, ''), username), id ASC;
            """
        )
        return jsonify(
            {
                "success": True,
                "items": [
                    {
                        "id": row["id"],
                        "username": row["username"],
                        "real_name": row["real_name"],
                        "phone": row["phone"],
                        "display_name": row["real_name"] or row["username"],
                    }
                    for row in cur.fetchall()
                ],
            }
        )
    except PermissionError as exc:
        return jsonify({"success": False, "error": str(exc)}), 401
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/inspection-plan-assignments/template", methods=["GET"])
def download_inspection_plan_assignment_template():
    conn = None
    cur = None
    try:
        current_user = get_current_request_user()
        conn = get_db_connection()
        cur = conn.cursor()

        if not can_manage_plan(cur, current_user):
            return jsonify({"success": False, "error": "当前账号无权维护巡检计划。"}), 403

        output = build_plan_assignment_template_workbook(cur, current_user)
        response = send_file(
            output,
            as_attachment=True,
            download_name="巡检计划派工模板.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response.headers["Cache-Control"] = "no-store"
        return response
    except PermissionError as exc:
        return jsonify({"success": False, "error": str(exc)}), 403
    except ValueError as exc:
        return jsonify({"success": False, "error": str(exc)}), 400
    except RuntimeError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500
    except Exception as exc:
        logging.exception("Failed to generate inspection plan assignment template.")
        return jsonify({"success": False, "error": f"生成派工模板失败：{str(exc)}"}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/inspection-plan-assignments/template-preview", methods=["POST"])
@app.route("/api/inspection-plan-assignments/ai-preview", methods=["POST"])
def preview_inspection_plan_template_assignments():
    upload_file = request.files.get("file") or request.files.get("assignment_file")

    try:
        parsed_rows, template_info = parse_plan_assignment_template_file(upload_file)
    except ValueError as exc:
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as exc:
        logging.exception("Failed to parse inspection plan assignment template.")
        return jsonify({"success": False, "error": f"读取派工模板失败：{str(exc)}"}), 500

    conn = None
    cur = None
    try:
        current_user = get_current_request_user()
        conn = get_db_connection()
        cur = conn.cursor()

        if not can_manage_plan(cur, current_user):
            return jsonify({"success": False, "error": "当前账号无权维护巡检计划。"}), 403

        ensure_inspection_plan_assignment_schema(cur)
        preview_rows, _context, target_date = resolve_plan_assignment_preview_rows(
            cur,
            current_user,
            parsed_rows,
        )
        conn.commit()

        ready_count = sum(1 for row in preview_rows if row.get("can_apply"))
        warning_count = sum(1 for row in preview_rows if row.get("status") == "warning")
        error_count = sum(1 for row in preview_rows if row.get("status") == "error")
        return jsonify(
            {
                "success": True,
                "message": "模板解析完成，请确认后再写入计划。",
                "ai_generated": False,
                "template_info": template_info,
                "target_date": target_date.isoformat(),
                "rows": preview_rows,
                "summary": {
                    "total": len(preview_rows),
                    "ready": ready_count,
                    "warning": warning_count,
                    "error": error_count,
                },
            }
        )
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except Exception as exc:
        if conn:
            conn.rollback()
        logging.exception("Inspection plan template assignment preview failed.")
        return jsonify({"success": False, "error": f"生成派工预览失败：{str(exc)}"}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/inspection-plan-assignments/template-apply", methods=["POST"])
@app.route("/api/inspection-plan-assignments/ai-apply", methods=["POST"])
def apply_inspection_plan_template_assignments():
    data = request.get_json(silent=True) or {}
    rows = data.get("rows") or []
    if not isinstance(rows, list):
        return jsonify({"success": False, "error": "rows 参数格式不正确。"}), 400
    if not rows:
        return jsonify({"success": False, "error": "请先选择需要写入的派工项。"}), 400
    if len(rows) > 300:
        return jsonify({"success": False, "error": "单次最多写入 300 条派工项。"}), 400

    grouped_rows = OrderedDict()
    for row in rows:
        if not isinstance(row, dict):
            return jsonify({"success": False, "error": "rows 中存在非法项。"}), 400
        if row.get("selected") is False:
            continue
        key = (str(row.get("plan_config_id") or ""), str(row.get("station_id") or ""))
        if not key[0] or not key[1]:
            return jsonify({"success": False, "error": "存在缺少计划或站点的派工项。"}), 400
        if key not in grouped_rows:
            grouped_rows[key] = {
                **row,
                "inspector_ids": [],
            }
        current_ids = grouped_rows[key]["inspector_ids"]
        for inspector_id in row.get("inspector_ids") or []:
            try:
                normalized_id = int(inspector_id)
            except (TypeError, ValueError):
                return jsonify({"success": False, "error": "检查人参数不正确。"}), 400
            if normalized_id not in current_ids:
                current_ids.append(normalized_id)

    if not grouped_rows:
        return jsonify({"success": False, "error": "请至少勾选一条可写入的派工项。"}), 400

    conn = None
    cur = None
    try:
        current_user = get_current_request_user()
        conn = get_db_connection()
        cur = conn.cursor()

        if not can_manage_plan(cur, current_user):
            return jsonify({"success": False, "error": "当前账号无权维护巡检计划。"}), 403

        ensure_inspection_plan_assignment_schema(cur)
        target_date = datetime.now(BEIJING_TZ).date()
        affected_plan_config_ids = set()
        applied_count = 0

        for row in grouped_rows.values():
            plan_config, station, inspector_ids = validate_plan_assignment_apply_row(
                cur,
                current_user,
                row,
                target_date,
            )
            assigned_inspector_id = inspector_ids[0] if inspector_ids else None
            cur.execute(
                """
                INSERT INTO inspection_plan_station_items (
                    plan_config_id,
                    station_id,
                    is_included,
                    assigned_inspector_id,
                    assigned_inspector_ids,
                    assigned_by,
                    assigned_at,
                    completion_status,
                    completed_inspection_id,
                    completed_at,
                    note,
                    updated_at
                )
                VALUES (
                    %s,
                    %s,
                    TRUE,
                    %s,
                    %s::jsonb,
                    %s,
                    CURRENT_TIMESTAMP,
                    'pending',
                    NULL,
                    NULL,
                    %s,
                    CURRENT_TIMESTAMP
                )
                ON CONFLICT (plan_config_id, station_id)
                DO UPDATE SET
                    is_included = TRUE,
                    assigned_inspector_id = EXCLUDED.assigned_inspector_id,
                    assigned_inspector_ids = EXCLUDED.assigned_inspector_ids,
                    assigned_by = EXCLUDED.assigned_by,
                    assigned_at = CASE
                        WHEN inspection_plan_station_items.assigned_inspector_ids IS DISTINCT FROM EXCLUDED.assigned_inspector_ids
                            OR inspection_plan_station_items.assigned_at IS NULL
                            THEN CURRENT_TIMESTAMP
                        ELSE inspection_plan_station_items.assigned_at
                    END,
                    completion_status = 'pending',
                    completed_inspection_id = NULL,
                    completed_at = NULL,
                    note = COALESCE(inspection_plan_station_items.note, EXCLUDED.note),
                    updated_at = CURRENT_TIMESTAMP
                WHERE inspection_plan_station_items.completion_status <> 'completed';
                """,
                (
                    int(plan_config["id"]),
                    int(station["id"]),
                    assigned_inspector_id,
                    json.dumps(inspector_ids, ensure_ascii=False),
                    int(current_user["id"]),
                    str(row.get("note") or "模板批量派工").strip()[:200],
                ),
            )
            if cur.rowcount <= 0:
                raise ValueError(f"【{station['station_name']}】当前周期任务已完成，不能调整派工。")
            affected_plan_config_ids.add(int(plan_config["id"]))
            applied_count += 1

        for plan_config_id in affected_plan_config_ids:
            cur.execute(
                """
                UPDATE inspection_plan_configs
                SET updated_by = %s,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = %s;
                """,
                (int(current_user["id"]), plan_config_id),
            )
            sync_plan_station_items_completion_by_history(cur, plan_config_id, wait_for_lock=False)

        conn.commit()
        return jsonify(
            {
                "success": True,
                "message": f"模板派工已写入 {applied_count} 个站点任务。",
                "applied_count": applied_count,
            }
        )
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except ValueError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/inspection-plan-assignments/board")
def get_inspection_plan_assignment_board():
    user_id = str(request.args.get("user_id", "")).strip()
    assigned_date = str(request.args.get("assigned_date", "")).strip()
    assigned_month = str(request.args.get("assigned_month", "")).strip()
    calendar_month = str(request.args.get("calendar_month", "")).strip()
    assigned_from = str(request.args.get("assigned_from", "")).strip()
    assigned_to = str(request.args.get("assigned_to", "")).strip()
    assigned_inspector_filter = str(request.args.get("assigned_inspector_filter", "all")).strip()
    completion_status = str(request.args.get("completion_status", "all")).strip()

    conn = None
    cur = None

    try:
        try:
            current_user = get_current_request_user()
        except PermissionError:
            current_user = None

        conn = get_db_connection()
        cur = conn.cursor()

        user = current_user or (get_user_by_id(cur, user_id) if user_id else None)
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404
        if not has_permission(cur, user, "view_inspection_plans"):
            return jsonify({"success": False, "error": "当前账号无权查看巡检计划。"}), 403

        ensure_inspection_completion_schema(cur)
        ensure_inspection_plan_assignment_schema(cur)

        where_clauses = [
            "psi.is_included = TRUE",
            "pc.status = 'active'",
        ]
        params = []

        if assigned_date:
            where_clauses.append("COALESCE(psi.assigned_at, psi.updated_at, pc.updated_at, pc.created_at)::date = %s::date")
            params.append(assigned_date)
        elif assigned_month:
            where_clauses.append("TO_CHAR(COALESCE(psi.assigned_at, psi.updated_at, pc.updated_at, pc.created_at), 'YYYY-MM') = %s")
            params.append(assigned_month)
        else:
            if assigned_from:
                where_clauses.append("COALESCE(psi.assigned_at, psi.updated_at, pc.updated_at, pc.created_at)::date >= %s::date")
                params.append(assigned_from)
            if assigned_to:
                where_clauses.append("COALESCE(psi.assigned_at, psi.updated_at, pc.updated_at, pc.created_at)::date <= %s::date")
                params.append(assigned_to)

        if assigned_inspector_filter and assigned_inspector_filter != "all":
            try:
                inspector_id_value = int(assigned_inspector_filter)
            except (TypeError, ValueError):
                return jsonify({"success": False, "error": "检查人参数不正确。"}), 400
            where_clauses.append("assigned.inspector_id = %s")
            params.append(inspector_id_value)

        if completion_status in {"pending", "completed"}:
            where_clauses.append("psi.completion_status = %s")
            params.append(completion_status)
        elif completion_status != "all":
            return jsonify({"success": False, "error": "完成状态参数不正确。"}), 400

        if not append_inspection_table_scope_filter(
            cur,
            user,
            where_clauses,
            params,
            "pc.inspection_table_id",
            "limit_plan_inspection_table_scope",
        ):
            return jsonify({"success": True, "items": [], "summary": {"total": 0, "completed": 0, "pending": 0}})
        if not append_station_region_scope_filter(
            cur,
            user,
            where_clauses,
            params,
            "s.region",
            "limit_plan_station_region_scope",
        ):
            return jsonify({"success": True, "items": [], "summary": {"total": 0, "completed": 0, "pending": 0}})

        calendar_where_clauses = [
            "psi.is_included = TRUE",
            "pc.status = 'active'",
        ]
        calendar_params = []
        if calendar_month:
            calendar_where_clauses.append("TO_CHAR(COALESCE(psi.assigned_at, psi.updated_at, pc.updated_at, pc.created_at), 'YYYY-MM') = %s")
            calendar_params.append(calendar_month)
        if assigned_inspector_filter and assigned_inspector_filter != "all":
            calendar_where_clauses.append("assigned.inspector_id = %s")
            calendar_params.append(inspector_id_value)
        if append_inspection_table_scope_filter(
            cur,
            user,
            calendar_where_clauses,
            calendar_params,
            "pc.inspection_table_id",
            "limit_plan_inspection_table_scope",
        ) and append_station_region_scope_filter(
            cur,
            user,
            calendar_where_clauses,
            calendar_params,
            "s.region",
            "limit_plan_station_region_scope",
        ):
            calendar_where_sql = " AND ".join(calendar_where_clauses)
            cur.execute(
                sql.SQL(
                    """
                    SELECT
                        TO_CHAR(COALESCE(psi.assigned_at, psi.updated_at, pc.updated_at, pc.created_at), 'YYYY-MM-DD') AS assigned_date,
                        COUNT(*) AS total_count,
                        COUNT(*) FILTER (WHERE psi.completion_status = 'completed') AS completed_count,
                        COUNT(*) FILTER (WHERE psi.completion_status <> 'completed') AS pending_count
                    FROM inspection_plan_station_items psi
                    JOIN inspection_plan_configs pc ON pc.id = psi.plan_config_id
                    JOIN stations s ON s.id = psi.station_id
                    JOIN LATERAL (
                        SELECT DISTINCT inspector_id
                        FROM (
                            SELECT jsonb_array_elements_text(COALESCE(psi.assigned_inspector_ids, '[]'::jsonb))::integer AS inspector_id
                            UNION ALL
                            SELECT psi.assigned_inspector_id
                            WHERE psi.assigned_inspector_id IS NOT NULL
                        ) raw_assigned_ids
                    ) assigned ON TRUE
                    WHERE {calendar_where_sql}
                    GROUP BY assigned_date
                    ORDER BY assigned_date DESC;
                    """
                ).format(calendar_where_sql=sql.SQL(calendar_where_sql)),
                calendar_params,
            )
            calendar_days = [
                {
                    "date": row["assigned_date"],
                    "total": int(row["total_count"] or 0),
                    "completed": int(row["completed_count"] or 0),
                    "pending": int(row["pending_count"] or 0),
                    "status": "completed" if int(row["pending_count"] or 0) == 0 else "pending",
                }
                for row in cur.fetchall()
            ]
        else:
            calendar_days = []

        where_sql = " AND ".join(where_clauses)
        cur.execute(
            sql.SQL(
                """
                SELECT
                    psi.id,
                    psi.plan_config_id,
                    psi.station_id,
                    s.station_name,
                    COALESCE(s.region, '') AS region,
                    pc.inspection_table_id,
                    it.table_name AS inspection_table_name,
                    COALESCE(NULLIF(it.checklist_mode, ''), 'online') AS checklist_mode,
                    pc.coverage_type,
                    pc.period_key,
                    assigned.inspector_id AS assigned_inspector_id,
                    assigned_user.username AS assigned_inspector_username,
                    assigned_user.real_name AS assigned_inspector_name,
                    assigned_user.phone AS assigned_inspector_phone,
                    TO_CHAR(COALESCE(psi.assigned_at, psi.updated_at, pc.updated_at, pc.created_at), 'YYYY-MM-DD') AS assigned_date,
                    TO_CHAR(COALESCE(psi.assigned_at, psi.updated_at, pc.updated_at, pc.created_at), 'YYYY-MM-DD HH24:MI') AS assigned_at,
                    assigner.username AS assigned_by_username,
                    assigner.real_name AS assigned_by_name,
                    psi.completion_status,
                    TO_CHAR(psi.completed_at, 'YYYY-MM-DD HH24:MI') AS completed_at,
                    psi.completed_inspection_id,
                    psi.note
                FROM inspection_plan_station_items psi
                JOIN inspection_plan_configs pc ON pc.id = psi.plan_config_id
                JOIN inspection_tables it ON it.id = pc.inspection_table_id
                JOIN stations s ON s.id = psi.station_id
                JOIN LATERAL (
                    SELECT DISTINCT inspector_id
                    FROM (
                        SELECT jsonb_array_elements_text(COALESCE(psi.assigned_inspector_ids, '[]'::jsonb))::integer AS inspector_id
                        UNION ALL
                        SELECT psi.assigned_inspector_id
                        WHERE psi.assigned_inspector_id IS NOT NULL
                    ) raw_assigned_ids
                ) assigned ON TRUE
                JOIN users assigned_user ON assigned_user.id = assigned.inspector_id
                LEFT JOIN users assigner ON assigner.id = psi.assigned_by
                WHERE {where_sql}
                ORDER BY
                    COALESCE(psi.assigned_at, psi.updated_at, pc.updated_at, pc.created_at) DESC,
                    assigned_user.real_name NULLS LAST,
                    assigned_user.username,
                    s.region,
                    s.station_name,
                    it.table_name;
                """
            ).format(where_sql=sql.SQL(where_sql)),
            params,
        )

        items = []
        completed_count = 0
        pending_count = 0
        inspector_options_map = {}
        for row in cur.fetchall():
            is_completed = row["completion_status"] == "completed"
            if is_completed:
                completed_count += 1
            else:
                pending_count += 1
            mode = normalize_checklist_mode(row.get("checklist_mode"))
            inspector_name = row["assigned_inspector_name"] or row["assigned_inspector_username"]
            inspector_options_map[row["assigned_inspector_id"]] = {
                "id": row["assigned_inspector_id"],
                "name": inspector_name,
            }
            items.append(
                {
                    "id": row["id"],
                    "plan_config_id": row["plan_config_id"],
                    "station_id": row["station_id"],
                    "station_name": row["station_name"],
                    "region": row["region"],
                    "inspection_table_id": row["inspection_table_id"],
                    "inspection_table_name": row["inspection_table_name"],
                    "checklist_mode": mode,
                    "checklist_mode_label": "视频检查" if mode == "online" else "现场检查",
                    "coverage_type": row["coverage_type"],
                    "coverage_type_label": COVERAGE_TYPE_LABELS.get(row["coverage_type"], row["coverage_type"]),
                    "period_key": row["period_key"],
                    "assigned_inspector_id": row["assigned_inspector_id"],
                    "assigned_inspector_name": inspector_name,
                    "assigned_inspector_phone": row["assigned_inspector_phone"],
                    "assigned_date": row["assigned_date"],
                    "assigned_at": row["assigned_at"],
                    "assigned_by_name": row["assigned_by_name"] or row["assigned_by_username"] or "",
                    "completion_status": row["completion_status"],
                    "completion_status_label": "已完成" if is_completed else "未完成",
                    "completed_at": row["completed_at"],
                    "completed_inspection_id": row["completed_inspection_id"],
                    "note": row["note"],
                }
            )

        inspector_options = sorted(
            inspector_options_map.values(),
            key=lambda item: str(item.get("name") or ""),
        )
        return jsonify(
            {
                "success": True,
                "items": items,
                "inspectors": inspector_options,
                "summary": {
                    "total": len(items),
                    "completed": completed_count,
                    "pending": pending_count,
                },
                "calendar": calendar_days,
            }
        )
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


def get_inspection_sign_pending_count_for_user(cur, user):
    if not is_station_manager(user) or not user.get("station_id"):
        return 0
    ensure_issue_inspector_schema(cur)
    ensure_inspection_completion_schema(cur)
    if not can_sign_inspection_records(cur, user):
        return 0
    cur.execute(
        """
        SELECT COUNT(*) AS pending_count
        FROM inspections ins
        WHERE ins.station_id = %s
          AND COALESCE(ins.inspector_completion_status, '待检查人确认') = %s
          AND COALESCE(ins.sign_status, '待签名确认') <> '已签名确认'
          AND NOT EXISTS (
              SELECT 1
              FROM issues i
              WHERE i.inspection_id = ins.id
                AND COALESCE(i.audit_status, 'pending') = 'pending'
          );
        """,
        (user["station_id"], INSPECTION_COMPLETION_DONE),
    )
    return int(cur.fetchone()["pending_count"] or 0)


def get_my_pending_rectification_count_for_user(cur, user):
    if (
        not is_station_manager(user)
        and not is_root_user(user)
        and user.get("role") != "supervisor"
    ):
        return 0

    ensure_issue_inspector_schema(cur)
    ensure_inspection_completion_schema(cur)
    if is_station_manager(user):
        if not user.get("station_id"):
            return 0
        cur.execute(
            """
            SELECT COUNT(*) AS pending_count
            FROM issues i
            JOIN inspections ins ON i.inspection_id = ins.id
            WHERE i.station_id = %s
              AND i.status = '待整改'
              AND ins.sign_status = '已签名确认'
              AND COALESCE(i.audit_status, 'pending') = 'approved';
            """,
            (user["station_id"],),
        )
    else:
        cur.execute(
            """
            SELECT COUNT(*) AS pending_count
            FROM issues i
            WHERE i.status = '待复核'
              AND COALESCE(i.audit_status, 'pending') <> 'rejected';
            """
        )
    return int(cur.fetchone()["pending_count"] or 0)


def get_peer_review_pending_count_for_user(cur, user):
    ensure_peer_review_schema(cur)
    if not user or not can_view_peer_reviews(cur, user):
        return 0
    cur.execute(
        """
        SELECT COUNT(DISTINCT t.id) AS pending_count
        FROM peer_review_tasks t
        JOIN peer_review_task_participants p ON p.task_id = t.id
        JOIN peer_review_task_reviewees r ON r.task_id = t.id
        LEFT JOIN peer_review_responses resp
          ON resp.task_id = t.id
         AND resp.reviewer_id = p.user_id
         AND resp.reviewee_id = r.user_id
        WHERE p.user_id = %s
          AND r.user_id <> %s
          AND resp.id IS NULL
          AND COALESCE(t.status, 'active') = 'active'
          AND (t.deadline_at IS NULL OR t.deadline_at >= CURRENT_TIMESTAMP);
        """,
        (user["id"], user["id"]),
    )
    return int(cur.fetchone()["pending_count"] or 0)


def get_plan_assignment_pending_summary_for_user(cur, user, include_items=False, item_limit=8):
    ensure_inspection_completion_schema(cur)
    ensure_inspection_plan_assignment_schema(cur)
    if not has_permission(cur, user, "submit_inspections"):
        return {"pending_count": 0, "items": []}

    cur.execute(
        """
        SELECT COUNT(*) AS pending_count
        FROM inspection_plan_station_items psi
        JOIN inspection_plan_configs pc ON pc.id = psi.plan_config_id
        JOIN LATERAL (
            SELECT DISTINCT inspector_id
            FROM (
                SELECT jsonb_array_elements_text(COALESCE(psi.assigned_inspector_ids, '[]'::jsonb))::integer AS inspector_id
                UNION ALL
                SELECT psi.assigned_inspector_id
                WHERE psi.assigned_inspector_id IS NOT NULL
            ) raw_assigned_ids
        ) assigned ON TRUE
        WHERE assigned.inspector_id = %s
          AND psi.is_included = TRUE
          AND psi.completion_status = 'pending'
          AND pc.status = 'active';
        """,
        (user["id"],),
    )
    pending_count = int(cur.fetchone()["pending_count"] or 0)
    if not include_items or pending_count <= 0:
        return {"pending_count": pending_count, "items": []}

    cur.execute(
        """
        SELECT
            psi.id,
            psi.plan_config_id,
            psi.station_id,
            s.station_name,
            COALESCE(s.region, '') AS region,
            pc.inspection_table_id,
            t.table_name AS inspection_table_name,
            COALESCE(NULLIF(t.checklist_mode, ''), 'online') AS checklist_mode,
            pc.coverage_type,
            pc.period_key,
            TO_CHAR(psi.assigned_at, 'YYYY-MM-DD HH24:MI') AS assigned_at,
            psi.note
        FROM inspection_plan_station_items psi
        JOIN inspection_plan_configs pc ON pc.id = psi.plan_config_id
        JOIN inspection_tables t ON t.id = pc.inspection_table_id
        JOIN stations s ON s.id = psi.station_id
        JOIN LATERAL (
            SELECT DISTINCT inspector_id
            FROM (
                SELECT jsonb_array_elements_text(COALESCE(psi.assigned_inspector_ids, '[]'::jsonb))::integer AS inspector_id
                UNION ALL
                SELECT psi.assigned_inspector_id
                WHERE psi.assigned_inspector_id IS NOT NULL
            ) raw_assigned_ids
        ) assigned ON TRUE
        WHERE assigned.inspector_id = %s
          AND psi.is_included = TRUE
          AND psi.completion_status = 'pending'
          AND pc.status = 'active'
        ORDER BY
            pc.updated_at DESC,
            pc.id DESC,
            COALESCE(s.region, ''),
            s.station_name
        LIMIT %s;
        """,
        (user["id"], max(1, int(item_limit or 8))),
    )
    items = []
    for row in cur.fetchall():
        mode = normalize_checklist_mode(row.get("checklist_mode"))
        items.append(
            {
                "id": row["id"],
                "plan_config_id": row["plan_config_id"],
                "station_id": row["station_id"],
                "station_name": row["station_name"],
                "region": row["region"],
                "inspection_table_id": row["inspection_table_id"],
                "inspection_table_name": row["inspection_table_name"],
                "checklist_mode": mode,
                "checklist_mode_label": "视频检查" if mode == "online" else "现场检查",
                "coverage_type": row["coverage_type"],
                "coverage_type_label": COVERAGE_TYPE_LABELS.get(row["coverage_type"], row["coverage_type"]),
                "period_key": row["period_key"],
                "assigned_at": row["assigned_at"],
                "note": row["note"],
            }
        )
    return {"pending_count": pending_count, "items": items}


@app.route("/api/notifications/summary", methods=["GET"])
def get_notification_summary():
    current_user = get_current_request_user()
    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_feedback_schema(cur)
        feedback_unread_count = get_feedback_unread_count(cur, current_user["id"])
        inspection_sign_pending_count = get_inspection_sign_pending_count_for_user(cur, current_user)
        my_pending_rectification_count = get_my_pending_rectification_count_for_user(cur, current_user)
        peer_review_pending_count = get_peer_review_pending_count_for_user(cur, current_user)
        plan_assignment_summary = get_plan_assignment_pending_summary_for_user(
            cur,
            current_user,
            include_items=True,
            item_limit=8,
        )
        conn.commit()
        return jsonify(
            {
                "success": True,
                "feedback_unread_count": feedback_unread_count,
                "inspection_sign_pending_count": inspection_sign_pending_count,
                "my_pending_rectification_count": my_pending_rectification_count,
                "peer_review_pending_count": peer_review_pending_count,
                "plan_assignment_pending_count": plan_assignment_summary["pending_count"],
                "plan_assignment_pending_items": plan_assignment_summary["items"],
            }
        )
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/inspection-plan-assignments/my-pending-count")
def get_my_pending_inspection_plan_assignment_count():
    conn = None
    cur = None

    try:
        current_user = get_current_request_user()
        conn = get_db_connection()
        cur = conn.cursor()
        summary = get_plan_assignment_pending_summary_for_user(
            cur,
            current_user,
            include_items=False,
        )
        conn.commit()
        return jsonify({"success": True, "pending_count": summary["pending_count"]})
    except PermissionError as exc:
        return jsonify({"success": False, "error": str(exc)}), 401
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/inspection-plan-assignments/my-pending")
def get_my_pending_inspection_plan_assignments():
    return frontend_version_expired_response()


# Inspection batch support
def get_or_create_inspection_batch(cur, station_id, inspector_id, batch_date):
    cur.execute(
        """
        SELECT id
        FROM inspection_batches
        WHERE station_id = %s
          AND batch_date = %s
        ORDER BY id ASC
        LIMIT 1;
        """,
        (station_id, batch_date),
    )
    existing_batch = cur.fetchone()
    if existing_batch:
        return existing_batch["id"]

    cur.execute(
        """
        INSERT INTO inspection_batches (
            station_id,
            inspector_id,
            batch_date
        )
        VALUES (%s, %s, %s)
        RETURNING id;
        """,
        (station_id, inspector_id, batch_date),
    )
    row = cur.fetchone()
    return row["id"]


def serialize_standard_row(field_meta, row, register_field_meta=None):
    item = {
        "id": row.get("id"),
        "standard_id": row.get("standard_id"),
        "created_at": row.get("created_at"),
    }
    for field_key, _field_label in field_meta:
        item[field_key] = row.get(field_key)
    item["standard_detail_text"] = build_standard_detail_text(field_meta, row)
    if register_field_meta is not None:
        item["register_display_text"] = build_standard_detail_text(register_field_meta, row)
    return item


INSPECTION_STANDARD_USAGE_MODES = {"internal", "external"}
DEFAULT_INSPECTION_STANDARD_USAGE_MODE = "internal"
INSPECTION_STANDARD_USAGE_MODE_LABELS = {
    "internal": "内部规范库",
    "external": "外部规范库",
}


def normalize_inspection_standard_usage_mode(value):
    mode = str(value or "").strip().lower()
    return mode if mode in INSPECTION_STANDARD_USAGE_MODES else DEFAULT_INSPECTION_STANDARD_USAGE_MODE


def serialize_inspection_standard_usage_mode(mode, row=None):
    normalized_mode = normalize_inspection_standard_usage_mode(mode)
    return {
        "mode": normalized_mode,
        "mode_label": INSPECTION_STANDARD_USAGE_MODE_LABELS.get(normalized_mode, "内部规范库"),
        "updated_by": row.get("updated_by") if row else None,
        "updated_by_username": row.get("updated_by_username") if row else "",
        "updated_by_name": row.get("updated_by_name") if row else "",
        "updated_at": row.get("updated_at") if row else "",
    }


def ensure_inspection_standard_usage_settings_schema(cur):
    acquire_schema_migration_lock(cur)
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS inspection_standard_usage_settings (
            singleton BOOLEAN PRIMARY KEY DEFAULT TRUE,
            register_standard_source TEXT NOT NULL DEFAULT 'internal',
            updated_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            CONSTRAINT chk_inspection_standard_usage_singleton CHECK (singleton = TRUE),
            CONSTRAINT chk_inspection_standard_usage_source
                CHECK (register_standard_source IN ('internal', 'external'))
        );
        """
    )
    cur.execute(
        """
        INSERT INTO inspection_standard_usage_settings (singleton, register_standard_source)
        VALUES (TRUE, %s)
        ON CONFLICT (singleton) DO NOTHING;
        """,
        (DEFAULT_INSPECTION_STANDARD_USAGE_MODE,),
    )


def get_inspection_standard_usage_mode(cur):
    cur.execute("SELECT to_regclass('public.inspection_standard_usage_settings') AS table_name;")
    row = cur.fetchone()
    if not row or not row.get("table_name"):
        return serialize_inspection_standard_usage_mode(DEFAULT_INSPECTION_STANDARD_USAGE_MODE)

    cur.execute(
        """
        SELECT
            s.register_standard_source,
            s.updated_by,
            COALESCE(u.username, '') AS updated_by_username,
            COALESCE(u.real_name, '') AS updated_by_name,
            TO_CHAR(s.updated_at, 'YYYY-MM-DD HH24:MI') AS updated_at
        FROM inspection_standard_usage_settings s
        LEFT JOIN users u ON u.id = s.updated_by
        WHERE s.singleton = TRUE
        LIMIT 1;
        """
    )
    setting = cur.fetchone()
    if not setting:
        return serialize_inspection_standard_usage_mode(DEFAULT_INSPECTION_STANDARD_USAGE_MODE)
    return serialize_inspection_standard_usage_mode(
        setting.get("register_standard_source"),
        setting,
    )


def save_inspection_standard_usage_mode(cur, mode, user_id):
    normalized_mode = normalize_inspection_standard_usage_mode(mode)
    cur.execute(
        """
        INSERT INTO inspection_standard_usage_settings (
            singleton,
            register_standard_source,
            updated_by,
            updated_at
        )
        VALUES (TRUE, %s, %s, CURRENT_TIMESTAMP)
        ON CONFLICT (singleton)
        DO UPDATE SET
            register_standard_source = EXCLUDED.register_standard_source,
            updated_by = EXCLUDED.updated_by,
            updated_at = CURRENT_TIMESTAMP;
        """,
        (normalized_mode, user_id or None),
    )
    return get_inspection_standard_usage_mode(cur)


def ensure_internal_standard_schema(cur):
    acquire_schema_migration_lock(cur)
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS inspection_internal_standards (
            id SERIAL PRIMARY KEY,
            internal_standard_id TEXT UNIQUE NOT NULL,
            path_values JSONB NOT NULL DEFAULT '[]'::jsonb,
            field_values JSONB NOT NULL DEFAULT '{}'::jsonb,
            content TEXT NOT NULL DEFAULT '',
            notes TEXT,
            is_active BOOLEAN DEFAULT TRUE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """
    )
    cur.execute(
        """
        ALTER TABLE inspection_internal_standards
        ADD COLUMN IF NOT EXISTS field_values JSONB NOT NULL DEFAULT '{}'::jsonb;
        """
    )
    cur.execute(
        """
        ALTER TABLE inspection_internal_standards
        ALTER COLUMN content SET DEFAULT '';
        """
    )
    cur.execute(
        """
        UPDATE inspection_internal_standards
        SET content = ''
        WHERE content IS NULL;
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS inspection_internal_standard_fields (
            id SERIAL PRIMARY KEY,
            field_key TEXT UNIQUE NOT NULL,
            field_label TEXT UNIQUE NOT NULL,
            is_filterable BOOLEAN DEFAULT TRUE,
            is_long_text BOOLEAN DEFAULT FALSE,
            is_register_visible BOOLEAN DEFAULT TRUE,
            sort_order INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """
    )
    cur.execute(
        """
        ALTER TABLE inspection_internal_standard_fields
        ADD COLUMN IF NOT EXISTS is_long_text BOOLEAN DEFAULT FALSE;
        """
    )
    cur.execute(
        """
        ALTER TABLE inspection_internal_standard_fields
        ADD COLUMN IF NOT EXISTS is_register_visible BOOLEAN DEFAULT TRUE;
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS inspection_internal_standard_links (
            id SERIAL PRIMARY KEY,
            internal_standard_id INTEGER NOT NULL REFERENCES inspection_internal_standards(id) ON DELETE CASCADE,
            external_standard_id BIGINT NOT NULL UNIQUE,
            external_inspection_table_id INTEGER REFERENCES inspection_tables(id) ON DELETE SET NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS inspection_internal_standard_tag_groups (
            id SERIAL PRIMARY KEY,
            group_name TEXT UNIQUE NOT NULL,
            group_type TEXT NOT NULL DEFAULT 'custom',
            color TEXT NOT NULL DEFAULT '#2563EB',
            is_system BOOLEAN NOT NULL DEFAULT FALSE,
            is_required BOOLEAN NOT NULL DEFAULT FALSE,
            is_filterable BOOLEAN NOT NULL DEFAULT TRUE,
            sort_order INTEGER NOT NULL DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            CONSTRAINT chk_internal_standard_tag_group_type
                CHECK (group_type IN ('custom', 'external_standard', 'inspection_table'))
        );
        """
    )
    cur.execute(
        """
        ALTER TABLE inspection_internal_standard_tag_groups
        ADD COLUMN IF NOT EXISTS color TEXT NOT NULL DEFAULT '#2563EB';
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS inspection_internal_standard_tags (
            id SERIAL PRIMARY KEY,
            group_id INTEGER NOT NULL REFERENCES inspection_internal_standard_tag_groups(id) ON DELETE CASCADE,
            tag_name TEXT NOT NULL,
            tag_key TEXT NOT NULL,
            color TEXT NOT NULL DEFAULT '#2563EB',
            sort_order INTEGER NOT NULL DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE (group_id, tag_key)
        );
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS inspection_internal_standard_tag_links (
            internal_standard_id INTEGER NOT NULL REFERENCES inspection_internal_standards(id) ON DELETE CASCADE,
            tag_id INTEGER NOT NULL REFERENCES inspection_internal_standard_tags(id) ON DELETE CASCADE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (internal_standard_id, tag_id)
        );
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_internal_standards_path_values
        ON inspection_internal_standards USING GIN (path_values);
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_internal_standards_field_values
        ON inspection_internal_standards USING GIN (field_values);
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_internal_standard_links_internal
        ON inspection_internal_standard_links (internal_standard_id);
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_internal_standard_tags_group
        ON inspection_internal_standard_tags (group_id, sort_order);
        """
    )
    cur.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_internal_standard_tag_links_tag
        ON inspection_internal_standard_tag_links (tag_id);
        """
    )
    cur.execute(
        """
        INSERT INTO inspection_internal_standard_tag_groups (
            group_name,
            group_type,
            color,
            is_system,
            is_required,
            is_filterable,
            sort_order
        )
        VALUES
            ('外部规范ID', 'external_standard', '#2563EB', TRUE, TRUE, TRUE, 0),
            ('检查表', 'inspection_table', '#0F766E', TRUE, FALSE, TRUE, 1)
        ON CONFLICT (group_name)
        DO UPDATE SET
            group_type = EXCLUDED.group_type,
            color = EXCLUDED.color,
            is_system = EXCLUDED.is_system,
            is_required = EXCLUDED.is_required,
            is_filterable = EXCLUDED.is_filterable,
            sort_order = EXCLUDED.sort_order,
            updated_at = CURRENT_TIMESTAMP;
        """
    )
    ensure_inspection_standard_usage_settings_schema(cur)


def get_internal_standard_fields(cur):
    cur.execute(
        """
        SELECT EXISTS (
            SELECT 1
            FROM information_schema.columns
            WHERE table_schema = 'public'
              AND table_name = 'inspection_internal_standard_fields'
              AND column_name = 'is_register_visible'
        ) AS has_column;
        """
    )
    has_register_visible_column = bool(cur.fetchone().get("has_column"))
    register_visible_select = "is_register_visible" if has_register_visible_column else "TRUE AS is_register_visible"
    cur.execute(
        f"""
        SELECT
            id,
            field_key,
            field_label,
            is_filterable,
            is_long_text,
            {register_visible_select},
            sort_order
        FROM inspection_internal_standard_fields
        ORDER BY sort_order ASC, id ASC;
        """
    )
    return cur.fetchall()


def normalize_internal_standard_field_rows(fields, allow_empty=True):
    return normalize_checklist_field_rows(
        fields,
        table_code="internal_standard",
        allow_empty=allow_empty,
    )


def upsert_internal_standard_fields(cur, fields):
    incoming_keys = [field["field_key"] for field in fields]
    current_rows = get_internal_standard_fields(cur)
    current_keys = {row["field_key"] for row in current_rows}
    removed_keys = [field_key for field_key in current_keys if field_key not in incoming_keys]

    if incoming_keys:
        cur.execute(
            """
            DELETE FROM inspection_internal_standard_fields
            WHERE field_key <> ALL(%s::text[]);
            """,
            (incoming_keys,),
        )
    else:
        cur.execute("DELETE FROM inspection_internal_standard_fields;")

    for field in fields:
        cur.execute(
            """
            INSERT INTO inspection_internal_standard_fields (
                field_key,
                field_label,
                is_filterable,
                is_long_text,
                is_register_visible,
                sort_order
            )
            VALUES (%s, %s, %s, %s, %s, %s)
            ON CONFLICT (field_key)
            DO UPDATE SET
                field_label = EXCLUDED.field_label,
                is_filterable = EXCLUDED.is_filterable,
                is_long_text = EXCLUDED.is_long_text,
                is_register_visible = EXCLUDED.is_register_visible,
                sort_order = EXCLUDED.sort_order;
            """,
            (
                field["field_key"],
                field["field_label"],
                field["is_filterable"],
                field.get("is_long_text", False),
                field["is_register_visible"],
                field["sort_order"],
            ),
        )

    for field_key in removed_keys:
        cur.execute(
            """
            UPDATE inspection_internal_standards
            SET field_values = field_values - %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE field_values ? %s;
            """,
            (field_key, field_key),
        )


INTERNAL_STANDARD_TAG_COLORS = [
    "#2563EB",
    "#0F766E",
    "#D97706",
    "#DC2626",
    "#7C3AED",
    "#0891B2",
    "#65A30D",
    "#DB2777",
    "#4F46E5",
    "#EA580C",
]


def normalize_internal_tag_key(value):
    return re.sub(r"\s+", "", str(value or "").strip().lower())


def next_internal_tag_color(seed=None):
    if seed:
        digest = hashlib.sha1(str(seed).encode("utf-8")).hexdigest()
        return INTERNAL_STANDARD_TAG_COLORS[int(digest[:2], 16) % len(INTERNAL_STANDARD_TAG_COLORS)]
    return INTERNAL_STANDARD_TAG_COLORS[uuid.uuid4().int % len(INTERNAL_STANDARD_TAG_COLORS)]


def serialize_internal_tag_group(row, tags=None):
    return {
        "id": row["id"],
        "group_name": row["group_name"],
        "group_type": row.get("group_type") or "custom",
        "color": row.get("color") or next_internal_tag_color(row.get("group_name")),
        "is_system": bool(row.get("is_system")),
        "is_required": bool(row.get("is_required")),
        "is_filterable": bool(row.get("is_filterable")),
        "sort_order": row.get("sort_order") or 0,
        "tags": tags or [],
    }


def serialize_internal_custom_tag(row):
    return {
        "id": row["id"],
        "tag_id": row["id"],
        "group_id": row["group_id"],
        "group_name": row.get("group_name") or "",
        "group_type": row.get("group_type") or "custom",
        "tag_name": row["tag_name"],
        "tag_key": row["tag_key"],
        "color": row.get("color") or next_internal_tag_color(row.get("group_name")),
        "sort_order": row.get("sort_order") or 0,
        "is_system": bool(row.get("is_system")),
    }


def get_internal_standard_tag_groups(cur, include_system=True):
    where_sql = "" if include_system else "WHERE is_system = FALSE"
    cur.execute(
        f"""
        SELECT
            id,
            group_name,
            group_type,
            color,
            is_system,
            is_required,
            is_filterable,
            sort_order
        FROM inspection_internal_standard_tag_groups
        {where_sql}
        ORDER BY sort_order ASC, id ASC;
        """
    )
    groups = [dict(row) for row in cur.fetchall()]
    if not groups:
        return []

    cur.execute(
        """
        SELECT
            t.id,
            t.group_id,
            g.group_name,
            g.group_type,
            g.is_system,
            t.tag_name,
            t.tag_key,
            g.color,
            t.sort_order
        FROM inspection_internal_standard_tags t
        JOIN inspection_internal_standard_tag_groups g ON g.id = t.group_id
        WHERE t.group_id = ANY(%s)
        ORDER BY g.sort_order ASC, t.sort_order ASC, t.id ASC;
        """,
        ([group["id"] for group in groups],),
    )
    tags_by_group = {}
    for tag in cur.fetchall():
        tags_by_group.setdefault(tag["group_id"], []).append(serialize_internal_custom_tag(tag))

    return [
        serialize_internal_tag_group(group, tags_by_group.get(group["id"], []))
        for group in groups
    ]


def normalize_internal_tag_group_rows(value):
    if not isinstance(value, list):
        raise ValueError("标签群组参数格式不正确。")
    normalized_groups = []
    seen_group_names = set()
    for group_index, raw_group in enumerate(value, start=1):
        if not isinstance(raw_group, dict):
            continue
        group_name = normalize_text(raw_group.get("group_name"), 50)
        if not group_name:
            raise ValueError(f"第 {group_index} 个标签群组缺少名称。")
        group_key = normalize_internal_tag_key(group_name)
        if group_key in {"外部规范id", "外部规范ID".lower(), "检查表"} or group_name in {"外部规范ID", "检查表"}:
            raise ValueError("外部规范ID和检查表是系统标签群组，不能手工维护。")
        if group_key in seen_group_names:
            raise ValueError(f"标签群组【{group_name}】重复。")
        seen_group_names.add(group_key)
        raw_color = str(raw_group.get("color") or "").strip()
        if not raw_color:
            for raw_tag in raw_group.get("tags") or []:
                if isinstance(raw_tag, dict) and str(raw_tag.get("color") or "").strip():
                    raw_color = str(raw_tag.get("color")).strip()
                    break
        group_color = raw_color if re.match(r"^#[0-9a-fA-F]{6}$", raw_color) else next_internal_tag_color(group_name)

        raw_tags = raw_group.get("tags") or []
        if not isinstance(raw_tags, list):
            raise ValueError(f"标签群组【{group_name}】的标签格式不正确。")
        tags = []
        seen_tag_names = set()
        for tag_index, raw_tag in enumerate(raw_tags, start=1):
            if not isinstance(raw_tag, dict):
                continue
            tag_name = normalize_text(raw_tag.get("tag_name"), 60)
            if not tag_name:
                raise ValueError(f"标签群组【{group_name}】第 {tag_index} 个标签缺少名称。")
            tag_key = normalize_internal_tag_key(tag_name)
            if tag_key in seen_tag_names:
                raise ValueError(f"标签群组【{group_name}】内标签【{tag_name}】重复。")
            seen_tag_names.add(tag_key)
            tags.append(
                {
                    "id": raw_tag.get("id"),
                    "tag_name": tag_name,
                    "tag_key": tag_key,
                    "color": group_color,
                    "sort_order": tag_index,
                }
            )
        normalized_groups.append(
            {
                "id": raw_group.get("id"),
                "group_name": group_name,
                "group_type": "custom",
                "color": group_color,
                "is_system": False,
                "is_required": False,
                "is_filterable": normalize_boolean_flag(raw_group.get("is_filterable"), True),
                "sort_order": group_index + 1,
                "tags": tags,
            }
        )
    return normalized_groups


def upsert_internal_standard_tag_groups(cur, groups):
    incoming_group_ids = []

    for group in groups:
        raw_group_id = group.get("id")
        group_id = None
        if raw_group_id:
            cur.execute(
                """
                UPDATE inspection_internal_standard_tag_groups
                SET group_name = %s,
                    color = %s,
                    is_filterable = %s,
                    sort_order = %s,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
                  AND is_system = FALSE
                RETURNING id;
                """,
                (
                    group["group_name"],
                    group["color"],
                    group["is_filterable"],
                    group["sort_order"],
                    raw_group_id,
                ),
            )
            row = cur.fetchone()
            group_id = row["id"] if row else None
        if not group_id:
            cur.execute(
                """
                INSERT INTO inspection_internal_standard_tag_groups (
                    group_name,
                    group_type,
                    color,
                    is_system,
                    is_required,
                    is_filterable,
                    sort_order,
                    created_at,
                    updated_at
                )
                VALUES (%s, 'custom', %s, FALSE, FALSE, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                ON CONFLICT (group_name)
                DO UPDATE SET
                    color = EXCLUDED.color,
                    is_filterable = EXCLUDED.is_filterable,
                    sort_order = EXCLUDED.sort_order,
                    updated_at = CURRENT_TIMESTAMP
                RETURNING id;
                """,
                (group["group_name"], group["color"], group["is_filterable"], group["sort_order"]),
            )
            group_id = cur.fetchone()["id"]
        incoming_group_ids.append(int(group_id))

        incoming_tag_ids = []
        for tag in group["tags"]:
            tag_id = None
            if tag.get("id"):
                cur.execute(
                    """
                    UPDATE inspection_internal_standard_tags
                    SET tag_name = %s,
                        tag_key = %s,
                        color = %s,
                        sort_order = %s,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = %s
                      AND group_id = %s
                    RETURNING id;
                    """,
                    (
                        tag["tag_name"],
                        tag["tag_key"],
                        group["color"],
                        tag["sort_order"],
                        tag["id"],
                        group_id,
                    ),
                )
                row = cur.fetchone()
                tag_id = row["id"] if row else None
            if not tag_id:
                cur.execute(
                    """
                    INSERT INTO inspection_internal_standard_tags (
                        group_id,
                        tag_name,
                        tag_key,
                        color,
                        sort_order,
                        created_at,
                        updated_at
                    )
                    VALUES (%s, %s, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                    ON CONFLICT (group_id, tag_key)
                    DO UPDATE SET
                        tag_name = EXCLUDED.tag_name,
                        color = EXCLUDED.color,
                        sort_order = EXCLUDED.sort_order,
                        updated_at = CURRENT_TIMESTAMP
                    RETURNING id;
                    """,
                    (group_id, tag["tag_name"], tag["tag_key"], group["color"], tag["sort_order"]),
                )
                tag_id = cur.fetchone()["id"]
            incoming_tag_ids.append(int(tag_id))

        if group["tags"]:
            cur.execute(
                """
                DELETE FROM inspection_internal_standard_tags
                WHERE group_id = %s
                  AND id <> ALL(%s::int[]);
                """,
                (group_id, [int(tag_id) for tag_id in incoming_tag_ids]),
            )
        else:
            cur.execute("DELETE FROM inspection_internal_standard_tags WHERE group_id = %s;", (group_id,))

    if incoming_group_ids:
        cur.execute(
            """
            DELETE FROM inspection_internal_standard_tag_groups
            WHERE is_system = FALSE
              AND id <> ALL(%s::int[]);
            """,
            (incoming_group_ids,),
        )
    else:
        cur.execute("DELETE FROM inspection_internal_standard_tag_groups WHERE is_system = FALSE;")


def normalize_internal_custom_tag_ids(value):
    if not isinstance(value, list):
        return []
    result = []
    seen = set()
    for raw_id in value:
        try:
            tag_id = int(raw_id)
        except (TypeError, ValueError):
            continue
        if tag_id <= 0 or tag_id in seen:
            continue
        seen.add(tag_id)
        result.append(tag_id)
    return result


def replace_internal_standard_custom_tags(cur, internal_id, tag_ids):
    normalized_ids = normalize_internal_custom_tag_ids(tag_ids)
    if normalized_ids:
        cur.execute(
            """
            SELECT t.id
            FROM inspection_internal_standard_tags t
            JOIN inspection_internal_standard_tag_groups g ON g.id = t.group_id
            WHERE t.id = ANY(%s)
              AND g.group_type = 'custom'
              AND g.is_system = FALSE;
            """,
            (normalized_ids,),
        )
        existing_ids = {int(row["id"]) for row in cur.fetchall()}
        missing_ids = [tag_id for tag_id in normalized_ids if tag_id not in existing_ids]
        if missing_ids:
            raise ValueError(f"标签ID【{missing_ids[0]}】不存在或不可用于内部规范。")

    cur.execute("DELETE FROM inspection_internal_standard_tag_links WHERE internal_standard_id = %s;", (internal_id,))
    for tag_id in normalized_ids:
        cur.execute(
            """
            INSERT INTO inspection_internal_standard_tag_links (
                internal_standard_id,
                tag_id,
                created_at
            )
            VALUES (%s, %s, CURRENT_TIMESTAMP)
            ON CONFLICT DO NOTHING;
            """,
            (internal_id, tag_id),
        )


def fetch_internal_standard_custom_tags(cur, internal_ids):
    ids = [int(item) for item in internal_ids if item]
    if not ids:
        return {}
    cur.execute(
        """
        SELECT
            l.internal_standard_id,
            t.id,
            t.group_id,
            g.group_name,
            g.group_type,
            g.is_system,
            t.tag_name,
            t.tag_key,
            g.color,
            t.sort_order
        FROM inspection_internal_standard_tag_links l
        JOIN inspection_internal_standard_tags t ON t.id = l.tag_id
        JOIN inspection_internal_standard_tag_groups g ON g.id = t.group_id
        WHERE l.internal_standard_id = ANY(%s)
        ORDER BY g.sort_order ASC, t.sort_order ASC, t.id ASC;
        """,
        (ids,),
    )
    result = {item: [] for item in ids}
    for row in cur.fetchall():
        result.setdefault(row["internal_standard_id"], []).append(serialize_internal_custom_tag(row))
    return result


def build_internal_system_tags_from_links(linked_externals):
    external_tags = []
    table_tags = []
    seen_tables = set()
    for link in linked_externals or []:
        external_id = str(link.get("external_standard_id") or "").strip()
        if external_id:
            external_tags.append(
                {
                    "id": f"external:{external_id}",
                    "tag_id": f"external:{external_id}",
                    "group_id": "external_standard",
                    "group_name": "外部规范ID",
                    "group_type": "external_standard",
                    "tag_name": external_id,
                    "tag_key": external_id.lower(),
                    "color": "#2563EB",
                    "is_system": True,
                }
            )
        table_id = str(link.get("external_inspection_table_id") or "").strip()
        table_name = str(link.get("inspection_table_name") or "").strip()
        table_key = table_id or table_name
        if table_key and table_key not in seen_tables:
            seen_tables.add(table_key)
            table_tags.append(
                {
                    "id": f"table:{table_key}",
                    "tag_id": f"table:{table_key}",
                    "group_id": "inspection_table",
                    "group_name": "检查表",
                    "group_type": "inspection_table",
                    "tag_name": table_name or f"检查表 {table_key}",
                    "tag_key": normalize_internal_tag_key(table_name or table_key),
                    "color": "#0F766E",
                    "is_system": True,
                }
            )
    return external_tags + table_tags


def fetch_internal_standard_tags_by_codes(cur, internal_standard_ids):
    codes = sorted({str(item or "").strip().upper() for item in internal_standard_ids if str(item or "").strip()})
    if not codes:
        return {}
    cur.execute(
        """
        SELECT
            id,
            internal_standard_id,
            path_values,
            field_values,
            content,
            notes,
            is_active,
            TO_CHAR(created_at, 'YYYY-MM-DD HH24:MI') AS created_at,
            TO_CHAR(updated_at, 'YYYY-MM-DD HH24:MI') AS updated_at
        FROM inspection_internal_standards
        WHERE UPPER(internal_standard_id) = ANY(%s);
        """,
        (codes,),
    )
    rows = cur.fetchall()
    if not rows:
        return {}
    link_map = fetch_internal_standard_links(cur, [row["id"] for row in rows])
    tag_map = fetch_internal_standard_custom_tags(cur, [row["id"] for row in rows])
    result = {}
    for row in rows:
        item = serialize_internal_standard(
            row,
            link_map.get(row["id"], []),
            [],
            tag_map.get(row["id"], []),
        )
        result[item["internal_standard_id"]] = item["tags"]
    return result


def attach_internal_standard_tags_to_issue_rows(cur, rows):
    if not rows:
        return rows
    tag_map = fetch_internal_standard_tags_by_codes(
        cur,
        [row.get("internal_standard_id") for row in rows],
    )
    for row in rows:
        code = str(row.get("internal_standard_id") or "").strip().upper()
        row["standard_tags"] = tag_map.get(code, [])
    return rows


def parse_json_field(value, fallback):
    if value in (None, ""):
        return fallback
    if isinstance(value, (dict, list)):
        return value
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
            return parsed
        except Exception:
            return fallback
    return fallback


def normalize_internal_field_values(raw_values, fields):
    if not fields:
        raise ValueError("请先配置内部规范字段。")
    if not isinstance(raw_values, dict):
        raise ValueError("内部规范字段内容格式不正确。")
    normalized = {}
    for field in fields:
        field_key = field["field_key"]
        normalized[field_key] = normalize_text(raw_values.get(field_key), 1000)
    first_field = fields[0]
    if not normalized.get(first_field["field_key"]):
        raise ValueError(f"请填写首个字段【{first_field['field_label']}】，系统需要用它生成内部规范ID。")
    return normalized


def build_internal_path_values_from_field_values(fields, field_values):
    return [
        str(field_values.get(field["field_key"]) or "").strip()
        for field in fields
        if str(field_values.get(field["field_key"]) or "").strip()
    ]


def build_internal_standard_summary(fields, field_values):
    parts = []
    for field in fields:
        value = str(field_values.get(field["field_key"]) or "").strip()
        if value:
            parts.append(f"{field['field_label']}：{value}")
    return "\n".join(parts)


def normalize_external_link_rows(value):
    if not isinstance(value, list):
        return []
    links = []
    seen = set()
    for item in value:
        raw_standard_id = item.get("external_standard_id") if isinstance(item, dict) else item
        raw_table_id = item.get("external_inspection_table_id") if isinstance(item, dict) else None
        standard_id = normalize_import_standard_id(raw_standard_id)
        if standard_id in seen:
            continue
        seen.add(standard_id)
        table_id = None
        if raw_table_id not in (None, ""):
            try:
                table_id = int(raw_table_id)
            except (TypeError, ValueError):
                table_id = None
        links.append(
            {
                "external_standard_id": standard_id,
                "external_inspection_table_id": table_id,
            }
        )
    return links


def get_pinyin_initial_prefix(value):
    text = str(value or "").strip()
    if not text:
        return "NB"
    try:
        from pypinyin import Style, lazy_pinyin

        prefix = "".join(lazy_pinyin(text, style=Style.FIRST_LETTER)).upper()
    except Exception:
        basic_initials = {
            "配": "P",
            "电": "D",
            "间": "J",
            "加": "J",
            "油": "Y",
            "站": "Z",
            "安": "A",
            "全": "Q",
            "服": "F",
            "务": "W",
            "卫": "W",
            "生": "S",
            "财": "C",
            "质": "Z",
            "量": "L",
        }
        prefix = "".join(
            basic_initials.get(char, char[0].upper() if char.isascii() and char.isalnum() else "")
            for char in text
        )
    prefix = re.sub(r"[^A-Z0-9]+", "", prefix)
    return prefix or "NB"


def generate_internal_standard_id(cur, first_field_value):
    prefix = get_pinyin_initial_prefix(first_field_value)
    cur.execute(
        """
        SELECT internal_standard_id
        FROM inspection_internal_standards
        WHERE internal_standard_id ~ %s;
        """,
        (f"^{prefix}[0-9]+$",),
    )
    max_number = 0
    for row in cur.fetchall():
        match = re.match(rf"^{re.escape(prefix)}(\d+)$", row["internal_standard_id"] or "")
        if match:
            max_number = max(max_number, int(match.group(1)))
    return f"{prefix}{max_number + 1}"


def fetch_external_standard_map(cur, external_standard_ids=None):
    cur.execute("SELECT to_regclass('public.inspection_tables') AS table_name;")
    row = cur.fetchone()
    if not row or not row.get("table_name"):
        return {}
    params = []
    filter_ids = {int(item) for item in external_standard_ids or [] if str(item).strip()}
    result = {}
    cur.execute(
        """
        SELECT id, table_code, table_name
        FROM inspection_tables
        WHERE is_active = TRUE
        ORDER BY id ASC;
        """
    )
    for table in cur.fetchall():
        physical_table_name = get_physical_table_name_by_code(table["table_code"])
        if not physical_table_name or not checklist_physical_table_exists(cur, physical_table_name):
            continue
        fields = [dict(field) for field in get_management_checklist_fields(cur, table["id"])]
        field_meta = [(field["field_key"], field["field_label"]) for field in fields]
        register_field_meta = build_register_display_field_meta(fields)
        where_sql = sql.SQL("")
        params = []
        if filter_ids:
            where_sql = sql.SQL(" WHERE standard_id = ANY(%s)")
            params = [list(filter_ids)]
        cur.execute(
            sql.SQL("SELECT * FROM {}{} ORDER BY standard_id ASC;").format(
                sql.Identifier(physical_table_name),
                where_sql,
            ),
            params,
        )
        for row in cur.fetchall():
            item = serialize_standard_row(field_meta, row, register_field_meta)
            standard_id = int(item["standard_id"])
            result[standard_id] = {
                **item,
                "external_standard_id": standard_id,
                "inspection_table_id": table["id"],
                "inspection_table_name": table["table_name"],
                "inspection_table_code": table["table_code"],
            }
    return result


def fetch_internal_links_by_external_ids(cur, external_standard_ids):
    ids = [int(item) for item in external_standard_ids if str(item).strip()]
    if not ids:
        return {}
    cur.execute("SELECT to_regclass('public.inspection_internal_standard_links') AS table_name;")
    row = cur.fetchone()
    if not row or not row.get("table_name"):
        return {}
    cur.execute(
        """
        SELECT
            l.external_standard_id,
            s.id,
            s.internal_standard_id,
            s.path_values,
            s.field_values,
            s.content
        FROM inspection_internal_standard_links l
        JOIN inspection_internal_standards s ON s.id = l.internal_standard_id
        WHERE l.external_standard_id = ANY(%s);
        """,
        (ids,),
    )
    return {int(row["external_standard_id"]): dict(row) for row in cur.fetchall()}


def serialize_internal_standard(row, linked_externals=None, fields=None, custom_tags=None):
    fields = fields or []
    linked_externals = linked_externals or []
    custom_tags = custom_tags or []
    path_values = parse_json_field(row.get("path_values"), [])
    field_values = parse_json_field(row.get("field_values"), {})
    if not isinstance(path_values, list):
        path_values = []
    if not isinstance(field_values, dict):
        field_values = {}
    if fields and not field_values and path_values:
        for index, field in enumerate(fields):
            if index < len(path_values):
                field_values[field["field_key"]] = path_values[index]
    content = row.get("content") or ""
    if fields and not content:
        content = build_internal_standard_summary(fields, field_values)
    register_display_text = build_internal_standard_summary(
        [field for field in fields if normalize_boolean_flag(field.get("is_register_visible"), True)],
        field_values,
    ) if fields else ""
    system_tags = build_internal_system_tags_from_links(linked_externals)
    all_tags = custom_tags + system_tags
    return {
        "id": row["id"],
        "internal_standard_id": str(row["internal_standard_id"] or "").upper(),
        "path_values": path_values,
        "field_values": field_values,
        "content": content,
        "register_display_text": register_display_text,
        "notes": row.get("notes") or "",
        "is_active": bool(row.get("is_active")),
        "created_at": row.get("created_at"),
        "updated_at": row.get("updated_at"),
        "linked_externals": linked_externals,
        "custom_tags": custom_tags,
        "custom_tag_ids": [tag["id"] for tag in custom_tags if isinstance(tag.get("id"), int)],
        "system_tags": system_tags,
        "tags": all_tags,
    }


def fetch_internal_standard_links(cur, internal_ids):
    ids = [int(item) for item in internal_ids if item]
    if not ids:
        return {}
    cur.execute(
        """
        SELECT
            l.internal_standard_id,
            l.external_standard_id,
            l.external_inspection_table_id
        FROM inspection_internal_standard_links l
        WHERE l.internal_standard_id = ANY(%s)
        ORDER BY l.external_standard_id ASC;
        """,
        (ids,),
    )
    link_rows = cur.fetchall()
    external_map = fetch_external_standard_map(cur, [row["external_standard_id"] for row in link_rows])
    result = {item: [] for item in ids}
    for row in link_rows:
        external = external_map.get(int(row["external_standard_id"]))
        result.setdefault(row["internal_standard_id"], []).append(
            {
                "external_standard_id": row["external_standard_id"],
                "external_inspection_table_id": row["external_inspection_table_id"],
                "inspection_table_name": external.get("inspection_table_name") if external else "",
                "standard_detail_text": external.get("standard_detail_text") if external else "",
            }
        )
    return result


def replace_internal_standard_links(cur, internal_id, links):
    external_ids = [item["external_standard_id"] for item in links]
    if external_ids:
        cur.execute(
            """
            SELECT
                l.external_standard_id,
                s.internal_standard_id
            FROM inspection_internal_standard_links l
            JOIN inspection_internal_standards s ON s.id = l.internal_standard_id
            WHERE l.external_standard_id = ANY(%s)
              AND l.internal_standard_id <> %s
            LIMIT 1;
            """,
            (external_ids, internal_id),
        )
        conflict = cur.fetchone()
        if conflict:
            raise ValueError(
                f"外部规范ID【{conflict['external_standard_id']}】已挂载到内部规范【{conflict['internal_standard_id']}】，不能重复挂载。"
            )

    external_map = fetch_external_standard_map(cur, external_ids)
    missing = [item for item in external_ids if int(item) not in external_map]
    if missing:
        raise ValueError(f"外部规范ID【{missing[0]}】不存在。")

    cur.execute("DELETE FROM inspection_internal_standard_links WHERE internal_standard_id = %s;", (internal_id,))
    for item in links:
        external = external_map[int(item["external_standard_id"])]
        table_id = item.get("external_inspection_table_id") or external["inspection_table_id"]
        cur.execute(
            """
            INSERT INTO inspection_internal_standard_links (
                internal_standard_id,
                external_standard_id,
                external_inspection_table_id
            )
            VALUES (%s, %s, %s);
            """,
            (internal_id, item["external_standard_id"], table_id),
        )


def fetch_internal_standard_by_code(cur, internal_standard_id):
    text = str(internal_standard_id or "").strip().upper()
    if not text:
        return None, [], []
    cur.execute("SELECT to_regclass('public.inspection_internal_standards') AS table_name;")
    row = cur.fetchone()
    if not row or not row.get("table_name"):
        return None, [], []
    fields = [dict(field) for field in get_internal_standard_fields(cur)]
    cur.execute(
        """
        SELECT
            id,
            internal_standard_id,
            path_values,
            field_values,
            content,
            notes,
            is_active,
            TO_CHAR(created_at, 'YYYY-MM-DD HH24:MI') AS created_at,
            TO_CHAR(updated_at, 'YYYY-MM-DD HH24:MI') AS updated_at
        FROM inspection_internal_standards
        WHERE UPPER(internal_standard_id) = %s
          AND is_active = TRUE
        LIMIT 1;
        """,
        (text,),
    )
    row = cur.fetchone()
    if not row:
        return None, fields, []
    link_map = fetch_internal_standard_links(cur, [row["id"]])
    tag_map = fetch_internal_standard_custom_tags(cur, [row["id"]])
    return (
        serialize_internal_standard(
            row,
            link_map.get(row["id"], []),
            fields,
            tag_map.get(row["id"], []),
        ),
        fields,
        link_map.get(row["id"], []),
    )


def prepare_issue_registration_targets(
    cur,
    station_id,
    inspector_id,
    external_standards,
    batch_id,
    today,
):
    targets = []
    period = get_inspection_completion_config(cur)["record_uniqueness_period"]
    for external in external_standards:
        inspection_table_id = str(external.get("inspection_table_id") or "").strip()
        standard_id = str(external.get("external_standard_id") or external.get("standard_id") or "").strip()
        if not inspection_table_id or not standard_id:
            raise ValueError("内部规范挂载的外部规范数据不完整。")

        inspection_table = get_inspection_table_record(cur, inspection_table_id)
        if not inspection_table:
            raise ValueError(f"外部规范ID【{standard_id}】对应的检查表不存在。")
        if not inspection_table["is_active"]:
            raise ValueError(f"外部规范ID【{standard_id}】对应的检查表未启用。")

        lock_inspection_period_scope(cur, station_id, inspection_table_id, today, period)
        existing_inspection = find_period_inspection(
            cur,
            station_id,
            inspection_table_id,
            today,
            period,
        )
        if existing_inspection and existing_inspection.get("inspector_completion_status") == INSPECTION_COMPLETION_DONE:
            raise ValueError(
                f"站点{inspection_period_scope_text(period)}【{inspection_table['table_name']}】已确认完成，不能继续登记。"
            )

        targets.append(
            {
                "inspection_table_id": int(inspection_table_id),
                "inspection_table_name": inspection_table["table_name"],
                "standard_id": int(standard_id),
                "standard_detail_text": external.get("standard_detail_text") or "",
                "inspection_id": existing_inspection["id"] if existing_inspection else None,
            }
        )
    return targets


def resolve_issue_standard_edit_target(
    cur,
    usage_mode,
    standard_id=None,
    internal_standard_id=None,
    preferred_external_standard_id=None,
):
    mode = normalize_inspection_standard_usage_mode(usage_mode)
    if mode == "internal":
        internal_code = str(internal_standard_id or "").strip().upper()
        if not internal_code:
            raise ValueError("请选择内部规范ID。")
        internal_standard, _internal_fields, linked_externals = fetch_internal_standard_by_code(
            cur,
            internal_code,
        )
        if not internal_standard:
            raise ValueError("所选内部规范不存在或未启用。")
        if not linked_externals:
            raise ValueError("所选内部规范尚未挂载外部规范，不能用于巡检问题。")

        sorted_links = sorted(
            linked_externals,
            key=lambda item: int(item.get("external_standard_id") or 0),
        )
        preferred_external_id = str(preferred_external_standard_id or "").strip()
        first_link = next(
            (
                link
                for link in sorted_links
                if str(link.get("external_standard_id") or "").strip() == preferred_external_id
            ),
            sorted_links[0],
        )
        external_map = fetch_external_standard_map(cur, [first_link["external_standard_id"]])
        external = external_map.get(int(first_link["external_standard_id"]))
        if not external:
            raise ValueError(f"内部规范挂载的外部规范ID【{first_link['external_standard_id']}】不存在。")

        return {
            "inspection_table_id": int(external["inspection_table_id"]),
            "standard_id": int(external["external_standard_id"]),
            "standard_detail_text": external.get("standard_detail_text") or "",
            "internal_standard_id": internal_standard["internal_standard_id"],
            "internal_standard_detail_text": internal_standard.get("content") or "",
        }

    external_id = normalize_import_standard_id(standard_id)
    external_map = fetch_external_standard_map(cur, [external_id])
    external = external_map.get(external_id)
    if not external:
        raise ValueError("所选外部规范不存在或对应检查表未启用。")

    linked_internal = fetch_internal_links_by_external_ids(cur, [external_id]).get(external_id)
    return {
        "inspection_table_id": int(external["inspection_table_id"]),
        "standard_id": external_id,
        "standard_detail_text": external.get("standard_detail_text") or "",
        "internal_standard_id": linked_internal.get("internal_standard_id") if linked_internal else None,
        "internal_standard_detail_text": linked_internal.get("content") if linked_internal else None,
    }


def get_or_create_issue_edit_inspection(cur, issue, target_inspection_table_id, target_inspector_id=None):
    current_table_id = int(issue["inspection_table_id"])
    target_table_id = int(target_inspection_table_id)
    if current_table_id == target_table_id:
        return int(issue["inspection_id"])

    next_inspector_id = int(target_inspector_id or issue["inspector_id"])
    period = get_inspection_completion_config(cur)["record_uniqueness_period"]
    lock_inspection_period_scope(
        cur,
        issue["station_id"],
        target_table_id,
        issue["inspection_date"],
        period,
    )
    existing_inspection = find_period_inspection(
        cur,
        issue["station_id"],
        target_table_id,
        issue["inspection_date"],
        period,
    )
    if existing_inspection:
        if existing_inspection.get("inspector_completion_status") == INSPECTION_COMPLETION_DONE:
            raise ValueError(f"目标检查表{inspection_period_scope_text(period)}已确认完成，不能把问题改挂到该检查表。")
        return int(existing_inspection["id"])

    cur.execute(
        """
        INSERT INTO inspections (
            station_id,
            inspector_id,
            inspection_table_id,
            inspection_date,
            batch_id
        )
        VALUES (%s, %s, %s, %s, %s)
        RETURNING id;
        """,
        (
            issue["station_id"],
            next_inspector_id,
            target_table_id,
            issue["inspection_date"],
            issue.get("batch_id"),
        ),
    )
    return int(cur.fetchone()["id"])


def sync_inspection_primary_inspector(cur, inspection_id):
    if not inspection_id:
        return
    cur.execute(
        """
        SELECT inspector_id
        FROM issues
        WHERE inspection_id = %s
          AND inspector_id IS NOT NULL
        ORDER BY created_at ASC, id ASC
        LIMIT 1;
        """,
        (inspection_id,),
    )
    row = cur.fetchone()
    if not row or not row.get("inspector_id"):
        return
    cur.execute(
        """
        UPDATE inspections
        SET inspector_id = %s
        WHERE id = %s
          AND inspector_id <> %s;
        """,
        (row["inspector_id"], inspection_id, row["inspector_id"]),
    )


def cleanup_empty_inspection_after_issue_move(cur, inspection_id):
    if not inspection_id:
        return
    cur.execute(
        """
        SELECT id, batch_id, sign_status, inspector_completion_status
        FROM inspections
        WHERE id = %s
        LIMIT 1;
        """,
        (inspection_id,),
    )
    inspection = cur.fetchone()
    if (
        not inspection
        or inspection.get("sign_status") == "已签名确认"
        or inspection.get("inspector_completion_status") == INSPECTION_COMPLETION_DONE
    ):
        return

    cur.execute("SELECT COUNT(*) AS issue_count FROM issues WHERE inspection_id = %s;", (inspection_id,))
    if int(cur.fetchone()["issue_count"] or 0) > 0:
        return

    cur.execute(
        """
        UPDATE inspection_plan_station_items
        SET completion_status = 'pending',
            completed_inspection_id = NULL,
            completed_at = NULL,
            updated_at = CURRENT_TIMESTAMP
        WHERE completed_inspection_id = %s;
        """,
        (inspection_id,),
    )
    cur.execute("DELETE FROM inspections WHERE id = %s;", (inspection_id,))
    if inspection.get("batch_id"):
        cur.execute(
            """
            DELETE FROM inspection_batches b
            WHERE b.id = %s
              AND NOT EXISTS (
                SELECT 1
                FROM inspections ins
                WHERE ins.batch_id = b.id
              );
            """,
            (inspection["batch_id"],),
        )


def normalize_internal_standards_backup_payload(payload):
    if not isinstance(payload, dict):
        raise ValueError("备份文件格式不正确：根内容必须是对象。")
    if payload.get("backup_type") != "ywddzx_internal_standards":
        raise ValueError("备份文件类型不匹配，请选择巡检规范库备份文件。")
    standards = payload.get("standards")
    if not isinstance(standards, list):
        raise ValueError("备份文件缺少 standards 数组。")

    raw_fields = payload.get("fields")
    if not isinstance(raw_fields, list):
        max_path_length = 0
        has_content = False
        for item in standards:
            if not isinstance(item, dict):
                continue
            path_values = item.get("path_values") or []
            if isinstance(path_values, list):
                max_path_length = max(max_path_length, len(path_values))
            if str(item.get("content") or "").strip():
                has_content = True
        inferred_fields = [
            {"field_label": f"层级{i}", "is_filterable": True}
            for i in range(1, max_path_length + 1)
        ]
        if has_content:
            inferred_fields.append({"field_label": "规范内容", "is_filterable": False})
        raw_fields = inferred_fields

    fields = normalize_internal_standard_field_rows(raw_fields or [], allow_empty=True)
    raw_tag_groups = payload.get("tag_groups")
    tag_groups = normalize_internal_tag_group_rows(raw_tag_groups) if isinstance(raw_tag_groups, list) else []
    source_tag_id_map = {}
    if isinstance(raw_tag_groups, list):
        for raw_group in raw_tag_groups:
            if not isinstance(raw_group, dict):
                continue
            group_name = normalize_text(raw_group.get("group_name"), 50)
            for raw_tag in raw_group.get("tags") or []:
                if not isinstance(raw_tag, dict):
                    continue
                raw_id = raw_tag.get("id") or raw_tag.get("tag_id")
                tag_name = normalize_text(raw_tag.get("tag_name"), 60)
                if raw_id and group_name and tag_name:
                    source_tag_id_map[str(raw_id)] = {
                        "group_name": group_name,
                        "tag_name": tag_name,
                    }
    normalized = []
    seen_internal_ids = set()
    seen_external_ids = set()
    for index, item in enumerate(standards, start=1):
        if not isinstance(item, dict):
            raise ValueError(f"第 {index} 条内部规范数据格式不正确。")
        internal_standard_id = str(item.get("internal_standard_id") or "").strip().upper()
        if not internal_standard_id:
            raise ValueError(f"第 {index} 条内部规范缺少内部规范ID。")
        if not re.match(r"^[A-Z0-9]+[0-9]$", internal_standard_id):
            raise ValueError(f"内部规范ID【{internal_standard_id}】格式不正确。")
        if internal_standard_id in seen_internal_ids:
            raise ValueError(f"备份文件内内部规范ID【{internal_standard_id}】重复。")
        seen_internal_ids.add(internal_standard_id)

        raw_field_values = item.get("field_values")
        if not isinstance(raw_field_values, dict):
            raw_field_values = {}
            legacy_path_values = item.get("path_values") or []
            if isinstance(legacy_path_values, list):
                for field_index, field in enumerate(fields):
                    if field_index < len(legacy_path_values):
                        raw_field_values[field["field_key"]] = legacy_path_values[field_index]
            content_text = normalize_text(item.get("content"), 1000)
            if content_text and fields:
                raw_field_values[fields[-1]["field_key"]] = raw_field_values.get(fields[-1]["field_key"]) or content_text
        content = normalize_text(item.get("content"), 3000)
        field_values = normalize_internal_field_values(raw_field_values, fields) if fields else {}
        if not content and fields:
            content = build_internal_standard_summary(fields, field_values)
        if not content:
            legacy_path_values = item.get("path_values") or []
            if isinstance(legacy_path_values, list):
                content = normalize_text("\n".join(str(value or "").strip() for value in legacy_path_values if str(value or "").strip()), 3000)
        path_values = [content] if content else build_internal_path_values_from_field_values(fields, field_values)
        links = normalize_external_link_rows(
            item.get("external_links") or item.get("linked_externals") or []
        )
        for link in links:
            external_standard_id = int(link["external_standard_id"])
            if external_standard_id in seen_external_ids:
                raise ValueError(
                    f"备份文件内外部规范ID【{external_standard_id}】被多条内部规范重复挂载。"
                )
            seen_external_ids.add(external_standard_id)

        custom_tag_refs = []
        seen_tag_refs = set()
        raw_custom_tags = item.get("custom_tags") or item.get("tags") or []
        if isinstance(raw_custom_tags, list):
            for raw_tag in raw_custom_tags:
                if not isinstance(raw_tag, dict):
                    continue
                if raw_tag.get("is_system") or raw_tag.get("group_type") in ("external_standard", "inspection_table"):
                    continue
                group_name = normalize_text(raw_tag.get("group_name"), 50)
                tag_name = normalize_text(raw_tag.get("tag_name"), 60)
                if not group_name or not tag_name:
                    continue
                key = (normalize_internal_tag_key(group_name), normalize_internal_tag_key(tag_name))
                if key in seen_tag_refs:
                    continue
                seen_tag_refs.add(key)
                custom_tag_refs.append({"group_name": group_name, "tag_name": tag_name})
        raw_tag_ids = item.get("tag_ids") or item.get("custom_tag_ids") or []
        if isinstance(raw_tag_ids, list):
            for raw_id in raw_tag_ids:
                tag_ref = source_tag_id_map.get(str(raw_id))
                if not tag_ref:
                    continue
                key = (normalize_internal_tag_key(tag_ref["group_name"]), normalize_internal_tag_key(tag_ref["tag_name"]))
                if key in seen_tag_refs:
                    continue
                seen_tag_refs.add(key)
                custom_tag_refs.append(tag_ref)

        normalized.append(
            {
                "internal_standard_id": internal_standard_id,
                "path_values": path_values,
                "field_values": field_values,
                "content": content,
                "notes": "",
                "is_active": bool(item.get("is_active", True)),
                "external_links": links,
                "custom_tag_refs": custom_tag_refs,
            }
        )
    return {
        "fields": fields,
        "tag_groups": tag_groups,
        "standards": normalized,
    }


def parse_internal_standards_backup_file(file_storage):
    if not file_storage or not file_storage.filename:
        raise ValueError("请选择需要导入的巡检规范库备份文件。")
    if not file_storage.filename.lower().endswith(".json"):
        raise ValueError("巡检规范库备份文件只能导入 JSON 格式。")
    file_bytes = file_storage.read()
    if not file_bytes:
        raise ValueError("备份文件内容为空。")
    try:
        payload = json.loads(file_bytes.decode("utf-8-sig"))
    except Exception:
        raise ValueError("备份文件不是有效的 JSON。")
    return normalize_internal_standards_backup_payload(payload)


def build_external_inspection_standard_ai_catalog(cur):
    external_map = fetch_external_standard_map(cur)
    internal_links = fetch_internal_links_by_external_ids(cur, external_map.keys())
    ai_catalog = []
    full_standards = []

    for standard in external_map.values():
        standard_id = str(standard.get("external_standard_id") or standard.get("standard_id") or "").strip()
        detail_text = str(standard.get("standard_detail_text") or "").strip()
        if not standard_id or not detail_text:
            continue
        linked_internal = internal_links.get(int(standard["external_standard_id"]))
        enriched_standard = {
            **standard,
            "standard_id": standard_id,
            "external_standard_id": standard["external_standard_id"],
            "inspection_table_id": standard.get("inspection_table_id") or "",
            "inspection_table_name": standard.get("inspection_table_name") or "",
            "standard_detail_text": detail_text,
            "linked_internal": linked_internal,
            "linked_internal_standard_id": linked_internal.get("internal_standard_id") if linked_internal else "",
        }
        full_standards.append(enriched_standard)
        ai_catalog.append(
            {
                "standard_id": standard_id,
                "inspection_table_name": enriched_standard["inspection_table_name"],
                "detail_text": detail_text,
            }
        )

    return ai_catalog, full_standards


def build_internal_inspection_standard_ai_catalog(cur):
    cur.execute(
        """
        SELECT
            id,
            internal_standard_id,
            path_values,
            field_values,
            content,
            notes,
            is_active,
            TO_CHAR(created_at, 'YYYY-MM-DD HH24:MI') AS created_at,
            TO_CHAR(updated_at, 'YYYY-MM-DD HH24:MI') AS updated_at
        FROM inspection_internal_standards
        WHERE is_active = TRUE
        ORDER BY internal_standard_id ASC;
        """
    )
    internal_rows = cur.fetchall()
    fields = [dict(field) for field in get_internal_standard_fields(cur)]
    link_map = fetch_internal_standard_links(cur, [row["id"] for row in internal_rows])
    ai_catalog = []
    full_standards = []

    for row in internal_rows:
        standard = serialize_internal_standard(row, link_map.get(row["id"], []), fields)
        if not standard.get("linked_externals"):
            continue

        detail_text = str(standard.get("content") or "").strip()
        if not standard.get("internal_standard_id") or not detail_text:
            continue

        linked_table_names = []
        for link in standard.get("linked_externals") or []:
            table_name = str(link.get("inspection_table_name") or "").strip()
            if table_name and table_name not in linked_table_names:
                linked_table_names.append(table_name)
        table_summary = "、".join(linked_table_names) or "未命名检查表"

        enriched_standard = {
            **standard,
            "standard_id": standard["internal_standard_id"],
            "internal_standard_id": standard["internal_standard_id"],
            "inspection_table_id": "",
            "inspection_table_name": table_summary,
            "standard_detail_text": detail_text,
        }
        full_standards.append(enriched_standard)
        ai_catalog.append(
            {
                "standard_id": str(standard["internal_standard_id"]),
                "inspection_table_name": table_summary,
                "detail_text": detail_text,
            }
        )

    return ai_catalog, full_standards


def build_inspection_standard_ai_catalog(cur, usage_mode=None):
    mode = normalize_inspection_standard_usage_mode(usage_mode)
    if mode == "external":
        return build_external_inspection_standard_ai_catalog(cur)
    return build_internal_inspection_standard_ai_catalog(cur)


@app.route("/")
def home():
    return jsonify({"message": "业务督导中心数智管理平台后端运行中"})


@app.route("/api/health")
def health():
    return jsonify({"status": "ok"})


@app.route("/api/version")
def get_frontend_version():
    response = jsonify({"version": FRONTEND_APP_VERSION})
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


@app.route("/api/page-visibility", methods=["GET"])
def get_page_visibility():
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_system_page_visibility_schema(cur)
        conn.commit()
        return jsonify({"success": True, "settings": fetch_page_visibility_settings(cur)})
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/page-visibility", methods=["GET"])
def get_management_page_visibility():
    user_id = str(request.args.get("user_id", "")).strip()
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_user_security_schema(cur)
        ensure_system_page_visibility_schema(cur)
        conn.commit()
        require_management_user(cur, user_id, "manage_users")
        return jsonify({"success": True, "settings": fetch_page_visibility_settings(cur)})
    except PermissionError as exc:
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        return jsonify({"success": False, "error": str(exc)}), 404
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/page-visibility", methods=["PUT"])
def update_management_page_visibility():
    data = request.get_json(silent=True) or {}
    user_id = str(data.get("user_id", "")).strip()
    conn = None
    cur = None
    try:
        page_items = normalize_page_visibility_items(data.get("pages"))
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_user_security_schema(cur)
        ensure_system_page_visibility_schema(cur)
        actor = require_management_user(cur, user_id, "manage_users")
        for item in page_items:
            cur.execute(
                """
                INSERT INTO system_page_visibility (
                    page_key,
                    is_visible,
                    updated_by,
                    created_at,
                    updated_at
                )
                VALUES (%s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                ON CONFLICT (page_key)
                DO UPDATE SET
                    is_visible = EXCLUDED.is_visible,
                    updated_by = EXCLUDED.updated_by,
                    updated_at = CURRENT_TIMESTAMP;
                """,
                (item["page_key"], item["is_visible"], actor["id"]),
            )
        conn.commit()
        return jsonify(
            {
                "success": True,
                "message": "页面显示设置已保存。",
                "settings": fetch_page_visibility_settings(cur),
            }
        )
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 404
    except ValueError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/system/resources")
def get_system_resources():
    response = jsonify({"success": True, **build_server_resource_snapshot()})
    response.headers["Cache-Control"] = "no-store"
    return response


@app.route("/api/db-test")
def db_test():
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT current_database() AS db_name, version() AS version;")
        row = cur.fetchone()
        cur.close()
        conn.close()
        return jsonify(
            {
                "success": True,
                "database": row["db_name"],
                "version": row["version"],
            }
        )
    except Exception as e:
        return (
            jsonify(
                {
                    "success": False,
                    "error": str(e),
                }
            ),
            500,
        )


@app.route("/api/login", methods=["POST"])
def login():
    data = request.get_json(silent=True) or {}
    username = str(data.get("username", "")).strip()
    password = str(data.get("password", "")).strip()

    if not username or not password:
        return (
            jsonify(
                {
                    "success": False,
                    "error": "用户名和密码不能为空。",
                }
            ),
            400,
        )

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_user_security_schema(cur)
        conn.commit()
        cur.execute(
            """
            SELECT
                u.id,
                u.username,
                u.password,
                u.role,
                u.real_name,
                u.phone,
                u.station_id,
                s.station_name,
                s.region,
                s.address
            FROM users u
            LEFT JOIN stations s ON u.station_id = s.id
            WHERE u.username = %s AND u.password = %s
            LIMIT 1;
        """,
            (username, password),
        )
        user = cur.fetchone()

        if not user:
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "用户名或密码错误。",
                    }
                ),
                401,
            )

        token = create_auth_token(user)
        touch_online_user_presence(user)
        return jsonify(
            {
                "success": True,
                "token": token,
                "expires_in": get_auth_token_ttl_seconds(user),
                "user": build_auth_user_payload(cur, user),
            }
        )
    except Exception as e:
        return (
            jsonify(
                {
                    "success": False,
                    "error": str(e),
                }
            ),
            500,
        )
    finally:
        close_db_resources(cur, conn)


@app.route("/api/auth/logout", methods=["POST"])
def logout_authenticated_user():
    remove_online_user_presence(g.current_user.get("id"))
    return jsonify({"success": True})


@app.route("/api/auth/me", methods=["GET"])
def get_authenticated_user():
    token = extract_bearer_token()
    cached_payload = get_cached_auth_me_payload(g.current_user["id"], token)
    if cached_payload:
        cached_payload["expires_in"] = get_auth_payload_expires_in(getattr(g, "auth_payload", {}))
        return jsonify(cached_payload)

    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_user_security_schema(cur)
        conn.commit()
        user = fetch_auth_user_by_id(cur, g.current_user["id"])
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404

        response_payload = {
            "success": True,
            "token": token,
            "expires_in": get_auth_payload_expires_in(getattr(g, "auth_payload", {})),
            "user": build_auth_user_payload(cur, user),
        }
        set_cached_auth_me_payload(user["id"], token, response_payload)
        return jsonify(response_payload)
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/users/change-password", methods=["POST"])
def change_own_password():
    data = request.get_json(silent=True) or {}

    try:
        user_id = int(get_authenticated_request_user_id(data.get("user_id")) or 0)
    except (TypeError, ValueError):
        user_id = 0

    current_password = normalize_text(data.get("current_password"), 120)
    new_password = normalize_text(data.get("new_password"), 120)
    confirm_password = normalize_text(data.get("confirm_password"), 120)

    if user_id <= 0:
        return jsonify({"success": False, "error": "缺少用户信息。"}), 400
    if not current_password:
        return jsonify({"success": False, "error": "请填写当前密码。"}), 400

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_user_security_schema(cur)
        conn.commit()

        cur.execute(
            """
            SELECT id, username, password
            FROM users
            WHERE id = %s
            LIMIT 1;
            """,
            (user_id,),
        )
        user = cur.fetchone()

        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404
        if user["password"] != current_password:
            return jsonify({"success": False, "error": "当前密码不正确。"}), 400

        validated_password = validate_new_login_password(
            user["username"], new_password, confirm_password
        )
        if validated_password == current_password:
            return jsonify({"success": False, "error": "新密码不能与当前密码相同。"}), 400

        cur.execute(
            """
            UPDATE users
            SET password = %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = %s;
            """,
            (validated_password, user_id),
        )
        conn.commit()
        invalidate_auth_caches_for_user(user_id)

        updated_user = fetch_auth_user_by_id(cur, user_id)
        token = create_auth_token(updated_user)
        return jsonify(
            {
                "success": True,
                "token": token,
                "expires_in": get_auth_token_ttl_seconds(updated_user),
                "must_change_password": False,
                "user": build_auth_user_payload(cur, updated_user),
            }
        )
    except ValueError as e:
        return jsonify({"success": False, "error": str(e)}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


# === 检查表级签名确认 API ===
@app.route("/api/inspections/sign-pending-count", methods=["GET"])
def get_inspection_sign_pending_count():
    user = get_current_request_user()
    conn = None
    cur = None
    if not is_station_manager(user) or not user.get("station_id"):
        return jsonify({"success": True, "pending_count": 0})

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_issue_inspector_schema(cur)
        ensure_inspection_completion_schema(cur)
        if not can_sign_inspection_records(cur, user):
            conn.commit()
            return jsonify({"success": True, "pending_count": 0})
        cur.execute(
            """
            SELECT COUNT(*) AS pending_count
            FROM inspections ins
            WHERE ins.station_id = %s
              AND COALESCE(ins.inspector_completion_status, '待检查人确认') = %s
              AND COALESCE(ins.sign_status, '待签名确认') <> '已签名确认'
              AND NOT EXISTS (
                  SELECT 1
                  FROM issues i
                  WHERE i.inspection_id = ins.id
                    AND COALESCE(i.audit_status, 'pending') = 'pending'
              );
            """,
            (user["station_id"], INSPECTION_COMPLETION_DONE),
        )
        pending_count = int(cur.fetchone()["pending_count"] or 0)
        conn.commit()
        return jsonify({"success": True, "pending_count": pending_count})
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/inspections/<int:inspection_id>/sign", methods=["POST"])
def sign_inspection_record(inspection_id):
    user_id = get_authenticated_request_user_id(request.form.get("user_id"))
    signed_name = str(request.form.get("signed_name", "")).strip()
    signature_file = request.files.get("signature")

    if not user_id:
        return jsonify({"success": False, "error": "缺少用户信息。"}), 400

    if not signed_name:
        return jsonify({"success": False, "error": "请填写站经理姓名。"}), 400

    if not signature_file or not signature_file.filename:
        return jsonify({"success": False, "error": "请提交站经理签名。"}), 400

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_issue_inspector_schema(cur)
        ensure_inspection_completion_schema(cur)
        sync_signed_inspections_completion(cur)
        auto_complete_overdue_inspections(cur)
        conn.commit()

        user = get_user_by_id(cur, user_id)

        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404

        if not is_station_manager(user):
            return jsonify({"success": False, "error": "只有站点账号可以完成站经理签名验收。"}), 403
        if not can_sign_inspection_records(cur, user):
            return jsonify({"success": False, "error": "当前账号无权完成站经理签名验收。"}), 403

        cur.execute(
            """
            SELECT
                ins.id,
                ins.station_id,
                ins.inspection_date,
                ins.sign_status,
                ins.inspector_completion_status,
                ins.inspection_table_id,
                s.station_name,
                t.table_name
            FROM inspections ins
            JOIN stations s ON ins.station_id = s.id
            JOIN inspection_tables t ON ins.inspection_table_id = t.id
            WHERE ins.id = %s
            LIMIT 1;
            """,
            (inspection_id,),
        )
        inspection = cur.fetchone()

        if not inspection:
            return jsonify({"success": False, "error": "巡检记录不存在。"}), 404

        if inspection["sign_status"] == "已签名确认":
            return jsonify({"success": False, "error": "该检查表已完成签名验收。"}), 400

        if inspection.get("inspector_completion_status") != INSPECTION_COMPLETION_DONE:
            return jsonify({"success": False, "error": "该检查表仍在等待检查人完成确认，暂不能签字验收。"}), 400

        if str(user.get("station_id") or "") != str(inspection.get("station_id") or ""):
            return jsonify({"success": False, "error": "只能签署本账号所属站点的巡检记录。"}), 403

        cur.execute(
            """
            SELECT COUNT(*) AS pending_audit_count
            FROM issues
            WHERE inspection_id = %s
              AND COALESCE(audit_status, 'pending') = 'pending';
            """,
            (inspection_id,),
        )
        pending_audit_count = int(cur.fetchone()["pending_audit_count"] or 0)
        if pending_audit_count > 0:
            return jsonify({"success": False, "error": f"该检查表仍有 {pending_audit_count} 条问题待审核，暂不能签字确认。"}), 400

        signature_path = save_signature_file(signature_file)
        signed_at = beijing_now()

        cur.execute(
            """
            UPDATE inspections
            SET sign_status = '已签名确认',
                station_manager_signed_name = %s,
                station_manager_signature_path = %s,
                station_manager_signed_at = %s,
                inspector_completion_status = %s,
                inspector_completed_by = CASE
                    WHEN inspector_completion_status = %s THEN inspector_completed_by
                    ELSE %s
                END,
                inspector_completed_at = CASE
                    WHEN inspector_completion_status = %s THEN inspector_completed_at
                    ELSE %s
                END,
                inspector_completion_source = CASE
                    WHEN inspector_completion_status = %s THEN inspector_completion_source
                    ELSE 'signature'
                END,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = %s;
            """,
            (
                signed_name,
                signature_path,
                signed_at,
                INSPECTION_COMPLETION_DONE,
                INSPECTION_COMPLETION_DONE,
                user_id,
                INSPECTION_COMPLETION_DONE,
                signed_at,
                INSPECTION_COMPLETION_DONE,
                inspection_id,
            ),
        )

        conn.commit()
        return jsonify(
            {
                "success": True,
                "message": "本检查表签名验收成功。",
                "inspection_id": inspection_id,
                "signature_path": signature_path,
            }
        )
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/inspections/<int:inspection_id>/flow/reset", methods=["POST"])
def reset_inspection_record_flow(inspection_id):
    data = request.get_json(silent=True) or {}
    user_id = get_authenticated_request_user_id(data.get("user_id"))
    if not user_id:
        return jsonify({"success": False, "error": "缺少用户信息。"}), 400

    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_issue_inspector_schema(cur)
        ensure_inspection_completion_schema(cur)

        user = get_user_by_id(cur, user_id)
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404
        if not can_reset_inspection_signature(cur, user):
            return jsonify({"success": False, "error": "当前账号无权重置巡检记录流程。"}), 403

        cur.execute(
            """
            SELECT
                ins.id,
                ins.station_id,
                s.region AS station_region,
                ins.inspection_table_id,
                ins.inspection_date,
                ins.sign_status,
                ins.station_manager_signature_path,
                ins.station_manager_signed_at,
                ins.station_manager_signed_name,
                ins.inspector_completion_status,
                COUNT(i.id) AS total_issue_count,
                COUNT(i.id) FILTER (WHERE COALESCE(i.audit_status, 'pending') = 'pending') AS pending_audit_count,
                COUNT(i.id) FILTER (
                    WHERE NULLIF(TRIM(COALESCE(i.rectification_result, '')), '') IS NOT NULL
                       OR NULLIF(TRIM(COALESCE(i.rectification_note, '')), '') IS NOT NULL
                       OR NULLIF(TRIM(COALESCE(i.rectification_photo_path, '')), '') IS NOT NULL
                ) AS rectified_issue_count
            FROM inspections ins
            JOIN stations s ON s.id = ins.station_id
            LEFT JOIN issues i ON i.inspection_id = ins.id
            WHERE ins.id = %s
            GROUP BY
                ins.id,
                ins.station_id,
                s.region,
                ins.inspection_table_id,
                ins.inspection_date,
                ins.sign_status,
                ins.station_manager_signature_path,
                ins.station_manager_signed_at,
                ins.station_manager_signed_name,
                ins.inspector_completion_status
            LIMIT 1;
            """,
            (inspection_id,),
        )
        inspection = cur.fetchone()
        if not inspection:
            return jsonify({"success": False, "error": "巡检记录不存在。"}), 404

        if not is_inspection_table_allowed_for_user(
            cur,
            user,
            inspection["inspection_table_id"],
            "limit_record_inspection_table_scope",
        ):
            return jsonify({"success": False, "error": "当前账号无权操作该检查表的巡检记录。"}), 403
        if not is_station_region_allowed_for_user(
            cur,
            user,
            inspection.get("station_region"),
            "limit_record_station_region_scope",
        ):
            return jsonify({"success": False, "error": "当前账号无权操作该片区的巡检记录。"}), 403

        is_signed = (
            inspection.get("sign_status") == "已签名确认"
            or bool(inspection.get("station_manager_signature_path"))
            or bool(inspection.get("station_manager_signed_at"))
            or bool(inspection.get("station_manager_signed_name"))
        )
        is_completed = inspection.get("inspector_completion_status") == INSPECTION_COMPLETION_DONE
        total_issue_count = int(inspection.get("total_issue_count") or 0)
        pending_audit_count = int(inspection.get("pending_audit_count") or 0)
        rectified_issue_count = int(inspection.get("rectified_issue_count") or 0)

        if not is_completed and not is_signed:
            return jsonify({"success": False, "error": "该巡检记录仍在等待检查人确认，无需重置。"}), 400

        if is_signed:
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "该巡检记录已完成站经理签名验收，不能重置流程。",
                    }
                ),
                400,
            )

        cur.execute(
            """
            UPDATE issues
            SET audit_status = 'pending',
                audited_by = NULL,
                audited_at = NULL,
                is_excellent = FALSE
            WHERE inspection_id = %s;
            """,
            (inspection_id,),
        )
        reset_issue_count = cur.rowcount

        reset_to_confirmation = (
            is_completed
            and not is_signed
            and (pending_audit_count > 0 or total_issue_count == 0)
        )
        affected_plan_config_ids = []
        if reset_to_confirmation:
            cur.execute(
                """
                SELECT DISTINCT plan_config_id
                FROM inspection_plan_station_items
                WHERE completed_inspection_id = %s;
                """,
                (inspection_id,),
            )
            affected_plan_config_ids = [row["plan_config_id"] for row in cur.fetchall()]
            reopen_inspection_record(cur, inspection_id)
            for plan_config_id in affected_plan_config_ids:
                sync_plan_station_items_completion_by_history(cur, plan_config_id)

        conn.commit()

        if reset_to_confirmation:
            message = "巡检记录已回退到等待检查人确认，关联问题审核记录已清空。"
        else:
            message = "巡检记录已回退到等待问题审核，关联问题审核记录已清空。"

        return jsonify(
            {
                "success": True,
                "message": message,
                "reset_issue_count": reset_issue_count,
                "reset_to": "inspector_confirmation" if reset_to_confirmation else "issue_audit",
            }
        )
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/inspections/<int:inspection_id>/signature/reset", methods=["POST"])
def reset_inspection_signature(inspection_id):
    data = request.get_json(silent=True) or {}
    user_id = get_authenticated_request_user_id(data.get("user_id"))
    if not user_id:
        return jsonify({"success": False, "error": "缺少用户信息。"}), 400

    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_inspection_completion_schema(cur)

        user = get_user_by_id(cur, user_id)
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404
        if not can_reset_inspection_signature(cur, user):
            return jsonify({"success": False, "error": "当前账号无权重置巡检记录流程。"}), 403

        cur.execute(
            """
            SELECT
                ins.id,
                ins.station_id,
                s.region AS station_region,
                ins.inspection_table_id,
                ins.sign_status,
                ins.station_manager_signed_name,
                ins.station_manager_signature_path,
                ins.station_manager_signed_at,
                ins.inspector_completion_status,
                ins.inspector_completion_source
            FROM inspections ins
            JOIN stations s ON s.id = ins.station_id
            WHERE ins.id = %s
            LIMIT 1;
            """,
            (inspection_id,),
        )
        inspection = cur.fetchone()
        if not inspection:
            return jsonify({"success": False, "error": "巡检记录不存在。"}), 404
        if not is_inspection_table_allowed_for_user(
            cur,
            user,
            inspection["inspection_table_id"],
            "limit_record_inspection_table_scope",
        ):
            return jsonify({"success": False, "error": "当前账号无权操作该检查表的巡检记录。"}), 403
        if not is_station_region_allowed_for_user(
            cur,
            user,
            inspection.get("station_region"),
            "limit_record_station_region_scope",
        ):
            return jsonify({"success": False, "error": "当前账号无权操作该片区的巡检记录。"}), 403

        return jsonify({"success": False, "error": "站经理已签名验收的巡检记录不能重置。"}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/stations")
def get_stations():
    conn = None
    cur = None
    try:
        scope = str(request.args.get("scope", "")).strip().lower()
        light = str(request.args.get("light", "")).strip().lower() in {"1", "true", "yes"}
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_station_management_columns(cur)
        conn.commit()
        if light or scope in {"plan", "light"}:
            cur.execute(
                """
                SELECT
                    id,
                    station_name,
                    region,
                    COALESCE(monitoring_status, '运行中') AS monitoring_status
                FROM stations
                ORDER BY id;
                """
            )
            return jsonify(cur.fetchall())

        cur.execute(
            """
            SELECT
                s.id,
                s.station_name,
                s.region,
                s.address,
                s.longitude,
                s.latitude,
                s.station_manager_name,
                s.station_manager_phone,
                s.station_type,
                CASE
                    WHEN s.asset_type LIKE '%股权%' OR s.asset_type LIKE '%控股%' OR s.asset_type LIKE '%参股%' THEN '股权'
                    ELSE '全资'
                END AS asset_type,
                COALESCE(s.is_consolidated, '否') AS is_consolidated,
                COALESCE(s.online_3_status, '未上线') AS online_3_status,
                COALESCE(s.monitoring_status, '运行中') AS monitoring_status,
                s.hos_station_code,
                s.landline_phone,
                s.status,
                s.operating_hours,
                COALESCE(station_accounts.station_usernames, '') AS station_usernames
            FROM stations s
            LEFT JOIN (
                SELECT
                    station_id,
                    STRING_AGG(username, ' ' ORDER BY username) AS station_usernames
                FROM users
                WHERE role = 'station_manager'
                  AND station_id IS NOT NULL
                GROUP BY station_id
            ) station_accounts ON station_accounts.station_id = s.id
            ORDER BY s.id;
        """
        )
        rows = cur.fetchall()
        return jsonify(rows)
    except Exception as e:
        return (
            jsonify(
                {
                    "success": False,
                    "error": str(e),
                }
            ),
            500,
        )
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/users", methods=["GET"])
def get_management_users():
    user_id = str(request.args.get("user_id", "")).strip()
    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_inspection_checklist_management_schema(cur)
        ensure_user_security_schema(cur)
        conn.commit()
        require_management_user(cur, user_id, "manage_users")

        cur.execute(
            """
            SELECT
                u.id,
                u.username,
                u.role,
                u.real_name,
                u.phone,
                u.station_id,
                s.station_name,
                s.region AS station_region,
                TO_CHAR(u.created_at, 'YYYY-MM-DD HH24:MI') AS created_at,
                TO_CHAR(u.updated_at, 'YYYY-MM-DD HH24:MI') AS updated_at
            FROM users u
            LEFT JOIN stations s ON u.station_id = s.id
            ORDER BY
                CASE u.role
                    WHEN 'root' THEN 1
                    WHEN 'supervisor' THEN 2
                    WHEN 'quality_safety' THEN 3
                    WHEN 'development_plan' THEN 4
                    WHEN 'oil_gas' THEN 5
                    WHEN 'non_oil' THEN 6
                    WHEN 'finance' THEN 7
                    WHEN 'area_account' THEN 8
                    ELSE 9
                END,
                u.id ASC;
            """
        )
        rows = cur.fetchall()
        user_ids = [row["id"] for row in rows]

        permission_override_map = {user_id: {} for user_id in user_ids}
        if user_ids:
            cur.execute(
                """
                SELECT user_id, permission_key, is_allowed
                FROM user_permissions
                WHERE user_id = ANY(%s);
                """,
                (user_ids,),
            )
            for permission_row in cur.fetchall():
                target = permission_override_map.setdefault(permission_row["user_id"], {})
                permission_key = permission_row["permission_key"]
                if permission_key == "limit_inspection_table_scope":
                    for scope_key in INSPECTION_TABLE_SCOPE_PERMISSION_KEYS:
                        target.setdefault(scope_key, bool(permission_row["is_allowed"]))
                else:
                    target[permission_key] = bool(permission_row["is_allowed"])

        inspection_scope_override_map = {
            user_id: {scope_key: [] for scope_key in INSPECTION_TABLE_SCOPE_PERMISSION_KEYS}
            for user_id in user_ids
        }
        if user_ids:
            cur.execute(
                """
                SELECT user_id, scope_key, inspection_table_id
                FROM user_inspection_table_scopes
                WHERE user_id = ANY(%s)
                ORDER BY user_id ASC, scope_key ASC, inspection_table_id ASC;
                """,
                (user_ids,),
            )
            for scope_row in cur.fetchall():
                user_scope_map = inspection_scope_override_map.setdefault(
                    scope_row["user_id"],
                    {scope_key: [] for scope_key in INSPECTION_TABLE_SCOPE_PERMISSION_KEYS},
                )
                scope_key = scope_row["scope_key"]
                if scope_key in user_scope_map:
                    user_scope_map[scope_key].append(scope_row["inspection_table_id"])

        station_region_scope_override_map = {
            user_id: {scope_key: [] for scope_key in STATION_REGION_SCOPE_PERMISSION_KEYS}
            for user_id in user_ids
        }
        if user_ids:
            cur.execute(
                """
                SELECT user_id, scope_key, station_region
                FROM user_station_region_scopes
                WHERE user_id = ANY(%s)
                ORDER BY user_id ASC, scope_key ASC, station_region ASC;
                """,
                (user_ids,),
            )
            for scope_row in cur.fetchall():
                user_scope_map = station_region_scope_override_map.setdefault(
                    scope_row["user_id"],
                    {scope_key: [] for scope_key in STATION_REGION_SCOPE_PERMISSION_KEYS},
                )
                scope_key = scope_row["scope_key"]
                if scope_key in user_scope_map:
                    user_scope_map[scope_key].append(
                        normalize_station_region_value(scope_row["station_region"])
                    )

        role_permission_overrides = get_role_permission_overrides_map(cur)
        role_effective_permissions = {
            role: build_role_effective_permissions(cur, role)
            for role in ROLE_OPTIONS
            if role != "root"
        }
        role_inspection_table_scope_ids = build_role_effective_inspection_table_scope_map_all(cur)
        role_station_region_scope_values = build_role_effective_station_region_scope_map_all(cur)
        role_default_table_id_map = {
            role: set(get_role_default_inspection_table_ids(cur, role))
            for role in ROLE_DEFAULT_CHECKLIST_SCOPES
        }

        users = []
        for row in rows:
            overrides = permission_override_map.get(row["id"], {})
            if is_root_user(row):
                permissions = {item["key"]: True for item in PERMISSION_CATALOG}
            else:
                role_permissions = role_effective_permissions.get(row["role"], {})
                permissions = {
                    item["key"]: overrides.get(item["key"], role_permissions.get(item["key"], False))
                    for item in PERMISSION_CATALOG
                }
                permissions = enforce_exclusive_permissions(permissions, row["role"])
            inspection_scope_overrides = inspection_scope_override_map.get(row["id"], {})
            station_region_scope_overrides = station_region_scope_override_map.get(row["id"], {})
            has_personalized_config = bool(
                overrides
                or any(inspection_scope_overrides.get(scope_key) for scope_key in INSPECTION_TABLE_SCOPE_PERMISSION_KEYS)
                or any(station_region_scope_overrides.get(scope_key) for scope_key in STATION_REGION_SCOPE_PERMISSION_KEYS)
            )
            inspection_table_scope_ids = {}
            for scope_key in INSPECTION_TABLE_SCOPE_PERMISSION_KEYS:
                if not permissions.get(scope_key):
                    inspection_table_scope_ids[scope_key] = []
                elif inspection_scope_overrides.get(scope_key):
                    inspection_table_scope_ids[scope_key] = list(inspection_scope_overrides[scope_key])
                elif role_inspection_table_scope_ids.get(row["role"], {}).get(scope_key):
                    inspection_table_scope_ids[scope_key] = list(role_inspection_table_scope_ids[row["role"]][scope_key])
                elif has_role_default_checklist_scope(row):
                    inspection_table_scope_ids[scope_key] = list(role_default_table_id_map.get(row["role"], set()))
                else:
                    inspection_table_scope_ids[scope_key] = []
            station_region_scope_values = {}
            for scope_key in STATION_REGION_SCOPE_PERMISSION_KEYS:
                if not permissions.get(scope_key):
                    station_region_scope_values[scope_key] = []
                elif station_region_scope_overrides.get(scope_key):
                    station_region_scope_values[scope_key] = sorted(station_region_scope_overrides[scope_key])
                elif role_station_region_scope_values.get(row["role"], {}).get(scope_key):
                    station_region_scope_values[scope_key] = sorted(role_station_region_scope_values[row["role"]][scope_key])
                else:
                    station_region_scope_values[scope_key] = []
            users.append(
                {
                    "id": row["id"],
                    "username": row["username"],
                    "role": row["role"],
                    "real_name": row["real_name"],
                    "phone": row["phone"],
                    "station_id": row["station_id"],
                    "station_name": row["station_name"],
                    "station_region": row["station_region"],
                    "created_at": row["created_at"],
                    "updated_at": row["updated_at"],
                    "permission_overrides": overrides,
                    "has_permission_overrides": has_personalized_config,
                    "permissions": permissions,
                    "inspection_table_scope_ids": inspection_table_scope_ids,
                    "station_region_scope_values": station_region_scope_values,
                }
            )

        cur.execute(
            """
            SELECT id, station_name, region, hos_station_code
            FROM stations
            ORDER BY region ASC NULLS LAST, station_name ASC, id ASC;
            """
        )
        stations = cur.fetchall()

        cur.execute(
            """
            SELECT id, table_name, checklist_mode
            FROM inspection_tables
            WHERE is_active = TRUE
            ORDER BY id ASC;
            """
        )
        inspection_table_rows = cur.fetchall()
        inspection_tables = [
            {
                **dict(row),
                "checklist_mode_label": "视频检查"
                if normalize_checklist_mode(row.get("checklist_mode")) == "online"
                else "现场检查",
                "is_quality_safety_default": row["id"] in role_default_table_id_map.get("quality_safety", set()),
                "is_development_plan_default": row["id"] in role_default_table_id_map.get("development_plan", set()),
                "is_oil_gas_default": row["id"] in role_default_table_id_map.get("oil_gas", set()),
                "is_non_oil_default": row["id"] in role_default_table_id_map.get("non_oil", set()),
                "is_finance_default": row["id"] in role_default_table_id_map.get("finance", set()),
                "default_scope_role_labels": [
                    ROLE_LABELS.get(role, role)
                    for role, table_ids in role_default_table_id_map.items()
                    if row["id"] in table_ids
                ],
            }
            for row in inspection_table_rows
        ]
        role_default_inspection_table_scope_ids = {
            role: sorted(table_ids)
            for role, table_ids in role_default_table_id_map.items()
        }
        return jsonify(
            {
                "success": True,
                "users": users,
                "stations": stations,
                "inspection_tables": inspection_tables,
                "quality_safety_default_inspection_table_ids": [
                    row["id"] for row in inspection_tables if row["is_quality_safety_default"]
                ],
                "development_plan_default_inspection_table_ids": [
                    row["id"] for row in inspection_tables if row["is_development_plan_default"]
                ],
                "oil_gas_default_inspection_table_ids": [
                    row["id"] for row in inspection_tables if row["is_oil_gas_default"]
                ],
                "non_oil_default_inspection_table_ids": [
                    row["id"] for row in inspection_tables if row["is_non_oil_default"]
                ],
                "finance_default_inspection_table_ids": [
                    row["id"] for row in inspection_tables if row["is_finance_default"]
                ],
                "role_default_inspection_table_scope_ids": role_default_inspection_table_scope_ids,
                "role_permission_overrides": role_permission_overrides,
                "role_effective_permissions": role_effective_permissions,
                "role_inspection_table_scope_ids": role_inspection_table_scope_ids,
                "role_station_region_scope_values": role_station_region_scope_values,
                "roles": [{"value": key, "label": label} for key, label in ROLE_LABELS.items()],
                "permissions": PERMISSION_CATALOG,
            }
        )
    except PermissionError as exc:
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        return jsonify({"success": False, "error": str(exc)}), 404
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/user-birthdays", methods=["GET"])
def get_management_user_birthdays():
    user_id = str(request.args.get("user_id", "")).strip()
    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_user_security_schema(cur)
        conn.commit()
        require_management_user(cur, user_id, "manage_users")

        cur.execute(
            """
            SELECT
                b.id,
                b.real_name,
                b.birthday_month,
                b.birthday_day,
                TO_CHAR(b.updated_at, 'YYYY-MM-DD HH24:MI') AS updated_at,
                matched_user.id AS matched_user_id,
                matched_user.username AS matched_username,
                matched_user.role AS matched_role,
                matched_user.phone AS matched_phone
            FROM user_birthdays b
            LEFT JOIN LATERAL (
                SELECT id, username, role, phone
                FROM users u
                WHERE TRIM(COALESCE(u.real_name, '')) = TRIM(b.real_name)
                ORDER BY
                    CASE u.role WHEN 'root' THEN 1 WHEN 'supervisor' THEN 2 ELSE 3 END,
                    u.id ASC
                LIMIT 1
            ) matched_user ON TRUE
            ORDER BY b.birthday_month ASC, b.birthday_day ASC, b.real_name ASC;
            """
        )
        rows = cur.fetchall()
        birthdays = []
        for row in rows:
            month = int(row["birthday_month"])
            day = int(row["birthday_day"])
            birthday_row = {
                "real_name": row["real_name"],
                "birthday_month": month,
                "birthday_day": day,
            }
            birthdays.append(
                {
                    "id": row["id"],
                    "real_name": row["real_name"],
                    "birthday_month": month,
                    "birthday_day": day,
                    "birthday_label": format_birthday_label(month, day),
                    "updated_at": row["updated_at"],
                    "matched_user_id": row["matched_user_id"],
                    "matched_username": row["matched_username"],
                    "matched_role": row["matched_role"],
                    "matched_phone": row["matched_phone"],
                    "birthday_event": build_birthday_event_payload(
                        {"real_name": row["real_name"]},
                        birthday_row,
                        force=True,
                    ),
                }
            )

        return jsonify(
            {
                "success": True,
                "birthdays": birthdays,
                "work_start_date": WORK_ANNIVERSARY_START_DATE.isoformat(),
            }
        )
    except PermissionError as exc:
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        return jsonify({"success": False, "error": str(exc)}), 404
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/user-birthdays", methods=["PUT"])
def update_management_user_birthdays():
    data = request.get_json(silent=True) or {}
    user_id = str(data.get("user_id", "")).strip()
    birthday_items = data.get("birthdays")
    conn = None
    cur = None

    if not isinstance(birthday_items, list):
        return jsonify({"success": False, "error": "生日信息格式不正确。"}), 400

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_user_security_schema(cur)
        actor = require_management_user(cur, user_id, "manage_users")

        seen_names = set()
        saved_count = 0
        for item in birthday_items:
            if not isinstance(item, dict):
                raise ValueError("生日列表中存在无效记录。")
            real_name = normalize_text(item.get("real_name"), 80)
            if not real_name:
                raise ValueError("生日名单中的姓名不能为空。")
            if real_name in seen_names:
                raise ValueError(f"生日名单中存在重复姓名：{real_name}")
            seen_names.add(real_name)
            month, day = validate_birthday_month_day(
                item.get("birthday_month"),
                item.get("birthday_day"),
            )
            cur.execute(
                """
                INSERT INTO user_birthdays (
                    real_name,
                    birthday_month,
                    birthday_day,
                    updated_by,
                    created_at,
                    updated_at
                )
                VALUES (%s, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                ON CONFLICT (real_name)
                DO UPDATE SET
                    birthday_month = EXCLUDED.birthday_month,
                    birthday_day = EXCLUDED.birthday_day,
                    updated_by = EXCLUDED.updated_by,
                    updated_at = CURRENT_TIMESTAMP;
                """,
                (real_name, month, day, actor["id"]),
            )
            saved_count += 1

        conn.commit()
        invalidate_auth_caches_for_user()
        return jsonify(
            {
                "success": True,
                "message": f"生日信息已保存，共维护 {saved_count} 条记录。",
                "count": saved_count,
            }
        )
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 404
    except ValueError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/role-permissions/<role>", methods=["PUT"])
def update_management_role_permissions(role):
    data = request.get_json(silent=True) or {}
    user_id = str(data.get("user_id", "")).strip()
    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_user_security_schema(cur)
        actor = require_management_user(cur, user_id, "manage_users")
        normalized_role = normalize_user_role(role)
        if normalized_role == "root":
            return jsonify({"success": False, "error": "root 固定拥有全部权限，不需要配置角色通用权限。"}), 400

        applied_permissions = apply_role_permission_updates(
            cur,
            normalized_role,
            data.get("permissions"),
            actor["id"],
        )
        apply_role_inspection_table_scope_updates(
            cur,
            normalized_role,
            data.get("inspection_table_scope_ids"),
            actor["id"],
        )
        apply_role_station_region_scope_updates(
            cur,
            normalized_role,
            data.get("station_region_scope_values"),
            actor["id"],
        )
        conn.commit()
        return jsonify(
            {
                "success": True,
                "message": f"{ROLE_LABELS.get(normalized_role, normalized_role)}通用权限已保存；用户个性化权限不会被覆盖。",
                "role": normalized_role,
                "permissions": applied_permissions,
                "role_permission_overrides": get_role_permission_overrides(cur, normalized_role),
                "role_effective_permissions": build_role_effective_permissions(cur, normalized_role),
                "role_inspection_table_scope_ids": build_role_effective_inspection_table_scope_map(
                    cur,
                    normalized_role,
                    applied_permissions,
                ),
                "role_station_region_scope_values": build_role_effective_station_region_scope_map(
                    cur,
                    normalized_role,
                    applied_permissions,
                ),
            }
        )
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 404
    except ValueError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/users/reset-permissions-to-role-defaults", methods=["POST"])
def reset_all_user_permissions_to_role_defaults():
    data = request.get_json(silent=True) or {}
    user_id = str(data.get("user_id", "")).strip()
    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_user_security_schema(cur)
        require_management_user(cur, user_id, "manage_users")

        cur.execute(
            """
            SELECT COUNT(DISTINCT user_id) AS affected_user_count
            FROM (
                SELECT user_id FROM user_permissions
                UNION
                SELECT user_id FROM user_inspection_table_scopes
                UNION
                SELECT user_id FROM user_station_region_scopes
            ) affected_users;
            """
        )
        affected_user_count = int(cur.fetchone()["affected_user_count"] or 0)

        cur.execute("DELETE FROM user_permissions;")
        deleted_permission_count = cur.rowcount
        cur.execute("DELETE FROM user_inspection_table_scopes;")
        deleted_table_scope_count = cur.rowcount
        cur.execute("DELETE FROM user_station_region_scopes;")
        deleted_region_scope_count = cur.rowcount
        conn.commit()

        return jsonify(
            {
                "success": True,
                "message": f"已重置 {affected_user_count} 个用户的个性化权限，所有用户将按角色通用权限生效。",
                "affected_user_count": affected_user_count,
                "deleted_permission_count": deleted_permission_count,
                "deleted_table_scope_count": deleted_table_scope_count,
                "deleted_region_scope_count": deleted_region_scope_count,
            }
        )
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 404
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/users/export", methods=["GET"])
def export_management_users():
    user_id = str(request.args.get("user_id", "")).strip()
    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_inspection_checklist_management_schema(cur)
        ensure_user_security_schema(cur)
        conn.commit()
        require_management_user(cur, user_id, "manage_users")

        cur.execute(
            """
            SELECT
                u.id,
                u.username,
                u.password,
                u.role,
                u.real_name,
                u.phone,
                u.station_id,
                s.station_name,
                s.hos_station_code,
                TO_CHAR(u.created_at, 'YYYY-MM-DD HH24:MI:SS') AS created_at,
                TO_CHAR(u.updated_at, 'YYYY-MM-DD HH24:MI:SS') AS updated_at
            FROM users u
            LEFT JOIN stations s ON u.station_id = s.id
            ORDER BY
                CASE u.role
                    WHEN 'root' THEN 1
                    WHEN 'supervisor' THEN 2
                    WHEN 'quality_safety' THEN 3
                    WHEN 'development_plan' THEN 4
                    WHEN 'oil_gas' THEN 5
                    WHEN 'non_oil' THEN 6
                    WHEN 'finance' THEN 7
                    WHEN 'area_account' THEN 8
                    ELSE 9
                END,
                u.id ASC;
            """
        )
        rows = cur.fetchall()
        users = []
        for row in rows:
            users.append(
                {
                    "id": row["id"],
                    "username": row["username"],
                    "password": row["password"],
                    "role": row["role"],
                    "real_name": row["real_name"],
                    "phone": row["phone"],
                    "station_id": row["station_id"],
                    "station_name": row["station_name"],
                    "hos_station_code": row["hos_station_code"],
                    "created_at": row["created_at"],
                    "updated_at": row["updated_at"],
                    "must_change_password": row["password"] == DEFAULT_INITIAL_PASSWORD,
                    "permission_overrides": get_permission_overrides(cur, row["id"]),
                    "permissions": get_effective_permissions(cur, row),
                    "inspection_table_scope_ids": get_user_inspection_table_scope_overrides(cur, row["id"]),
                    "station_region_scope_values": get_user_station_region_scope_overrides(cur, row["id"]),
                }
            )

        cur.execute(
            """
            SELECT real_name, birthday_month, birthday_day
            FROM user_birthdays
            ORDER BY birthday_month ASC, birthday_day ASC, real_name ASC;
            """
        )
        user_birthdays = [dict(row) for row in cur.fetchall()]

        now = beijing_now()
        response = jsonify(
            {
                "backup_type": "ywddzx_users",
                "version": 1,
                "exported_at": now.isoformat(),
                "includes_passwords": True,
                "password_backup_mode": "database_current_value",
                "default_password_policy": "password == 123456 triggers forced password change on next login",
                "role_permission_overrides": get_role_permission_overrides_map(cur),
                "role_inspection_table_scope_ids": get_role_inspection_table_scope_overrides_map(cur),
                "role_station_region_scope_values": get_role_station_region_scope_overrides_map(cur),
                "user_birthdays": user_birthdays,
                "users": users,
            }
        )
        filename = f"ywddzx_users_backup_{now.strftime('%Y%m%d_%H%M%S')}.json"
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        response.headers["Cache-Control"] = "no-store"
        return response
    except PermissionError as exc:
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        return jsonify({"success": False, "error": str(exc)}), 404
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/users/import", methods=["POST"])
def import_management_users():
    user_id = str(request.form.get("user_id", "")).strip()
    backup_file = request.files.get("file")
    conn = None
    cur = None

    try:
        backup_payload = parse_user_backup_json(backup_file)
        raw_users = backup_payload.get("users") or []
        raw_role_permissions = backup_payload.get("role_permission_overrides") or {}
        raw_role_inspection_scopes = backup_payload.get("role_inspection_table_scope_ids") or {}
        raw_role_region_scopes = backup_payload.get("role_station_region_scope_values") or {}
        raw_birthdays = backup_payload.get("user_birthdays") or []
        user_payloads = []
        skipped_builtin_count = 0
        for raw_user in raw_users:
            if not isinstance(raw_user, dict):
                raise ValueError("用户备份文件中存在无效的用户记录。")

            raw_username = normalize_text(raw_user.get("username"), 80)
            raw_role = normalize_text(raw_user.get("role"))
            if (raw_username == "root") != (raw_role == "root"):
                raise ValueError("root 账号备份记录必须同时使用用户名 root 和角色 root。")

            user_data = build_management_user_payload(raw_user, is_create=True)

            raw_permissions = raw_user.get("permissions")
            if raw_permissions in (None, ""):
                raw_permissions = raw_user.get("permission_overrides")
            user_data["user_id"] = normalize_user_backup_id(raw_user.get("id"))
            user_data["permissions"] = normalize_permission_updates(raw_permissions)
            user_data["inspection_table_scope_ids"] = normalize_inspection_table_scope_updates(
                raw_user.get("inspection_table_scope_ids")
            )
            user_data["station_region_scope_values"] = normalize_station_region_scope_updates(
                raw_user.get("station_region_scope_values")
            )
            user_data["created_at"] = normalize_text(raw_user.get("created_at")) or None
            user_data["updated_at"] = normalize_text(raw_user.get("updated_at")) or None
            user_payloads.append(user_data)

        if not user_payloads:
            raise ValueError("用户备份文件中没有可导入的非 root 用户。")

        conn = get_db_connection()
        cur = conn.cursor()
        ensure_inspection_checklist_management_schema(cur)
        ensure_user_security_schema(cur)
        actor = require_management_user(cur, user_id, "manage_users")

        if raw_birthdays:
            if not isinstance(raw_birthdays, list):
                raise ValueError("用户备份文件中的生日信息格式不正确。")
            for raw_birthday in raw_birthdays:
                if not isinstance(raw_birthday, dict):
                    raise ValueError("用户备份文件中存在无效的生日记录。")
                real_name = normalize_text(raw_birthday.get("real_name"), 80)
                if not real_name:
                    continue
                birthday_month, birthday_day = validate_birthday_month_day(
                    raw_birthday.get("birthday_month"),
                    raw_birthday.get("birthday_day"),
                )
                cur.execute(
                    """
                    INSERT INTO user_birthdays (
                        real_name,
                        birthday_month,
                        birthday_day,
                        updated_by,
                        created_at,
                        updated_at
                    )
                    VALUES (%s, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                    ON CONFLICT (real_name)
                    DO UPDATE SET
                        birthday_month = EXCLUDED.birthday_month,
                        birthday_day = EXCLUDED.birthday_day,
                        updated_by = EXCLUDED.updated_by,
                        updated_at = CURRENT_TIMESTAMP;
                    """,
                    (real_name, birthday_month, birthday_day, actor["id"]),
                )

        if isinstance(raw_role_permissions, dict):
            for role, role_permissions in raw_role_permissions.items():
                normalized_role = normalize_text(role)
                if normalized_role in ROLE_OPTIONS and normalized_role != "root":
                    apply_role_permission_updates(cur, normalized_role, role_permissions, actor["id"])
                    apply_role_inspection_table_scope_updates(
                        cur,
                        normalized_role,
                        raw_role_inspection_scopes.get(normalized_role) if isinstance(raw_role_inspection_scopes, dict) else None,
                        actor["id"],
                    )
                    apply_role_station_region_scope_updates(
                        cur,
                        normalized_role,
                        raw_role_region_scopes.get(normalized_role) if isinstance(raw_role_region_scopes, dict) else None,
                        actor["id"],
                    )

        imported_count = 0
        restored_root_count = 0
        auth_invalidated = False
        for user_data in user_payloads:
            if user_data["role"] == "root":
                if user_data["username"] != "root":
                    raise ValueError("root 账号备份记录的用户名必须为 root。")

                cur.execute(
                    """
                    SELECT id, password
                    FROM users
                    WHERE username = 'root'
                    LIMIT 1;
                    """
                )
                root_user = cur.fetchone()
                if not root_user:
                    raise ValueError("系统内置 root 账号不存在，请先检查用户表初始化状态。")

                if root_user["id"] == actor["id"] and root_user["password"] != user_data["password"]:
                    auth_invalidated = True

                cur.execute(
                    """
                    UPDATE users
                    SET password = %(password)s,
                        role = 'root',
                        real_name = %(real_name)s,
                        phone = %(phone)s,
                        station_id = NULL,
                        updated_at = COALESCE(NULLIF(%(updated_at)s, '')::timestamp, CURRENT_TIMESTAMP)
                    WHERE id = %(root_user_id)s
                    RETURNING id;
                    """,
                    {**user_data, "root_user_id": root_user["id"]},
                )
                target_user_id = cur.fetchone()["id"]
                cur.execute("DELETE FROM user_permissions WHERE user_id = %s;", (target_user_id,))
                cur.execute("DELETE FROM user_inspection_table_scopes WHERE user_id = %s;", (target_user_id,))
                cur.execute("DELETE FROM user_station_region_scopes WHERE user_id = %s;", (target_user_id,))
                imported_count += 1
                restored_root_count += 1
                continue

            if user_data["role"] == "station_manager":
                cur.execute(
                    """
                    SELECT id
                    FROM stations
                    WHERE id = %s
                    LIMIT 1;
                    """,
                    (user_data["station_id"],),
                )
                if not cur.fetchone():
                    raise ValueError("用户备份文件中的站点账号关联了不存在的站点ID，请先导入站点备份。")

            cur.execute(
                """
                SELECT id, role
                FROM users
                WHERE username = %s
                LIMIT 1;
                """,
                (user_data["username"],),
            )
            existing_by_username = cur.fetchone()

            if existing_by_username and existing_by_username["role"] == "root":
                skipped_builtin_count += 1
                continue

            target_user_id = None
            if existing_by_username:
                target_user_id = existing_by_username["id"]
                cur.execute(
                    """
                    UPDATE users
                    SET password = %(password)s,
                        role = %(role)s,
                        real_name = %(real_name)s,
                        phone = %(phone)s,
                        station_id = %(station_id)s,
                        updated_at = COALESCE(NULLIF(%(updated_at)s, '')::timestamp, CURRENT_TIMESTAMP)
                    WHERE id = %(target_user_id)s
                    RETURNING id;
                    """,
                    {**user_data, "target_user_id": target_user_id},
                )
                target_user_id = cur.fetchone()["id"]
            else:
                backup_user_id = user_data["user_id"]
                if backup_user_id:
                    cur.execute(
                        """
                        SELECT id, role
                        FROM users
                        WHERE id = %s
                        LIMIT 1;
                        """,
                        (backup_user_id,),
                    )
                    existing_by_id = cur.fetchone()
                    if existing_by_id:
                        backup_user_id = None

                if backup_user_id:
                    cur.execute(
                        """
                        INSERT INTO users (
                            id,
                            username,
                            password,
                            role,
                            real_name,
                            phone,
                            station_id,
                            created_at,
                            updated_at
                        )
                        VALUES (
                            %(user_id)s,
                            %(username)s,
                            %(password)s,
                            %(role)s,
                            %(real_name)s,
                            %(phone)s,
                            %(station_id)s,
                            COALESCE(NULLIF(%(created_at)s, '')::timestamp, CURRENT_TIMESTAMP),
                            COALESCE(NULLIF(%(updated_at)s, '')::timestamp, CURRENT_TIMESTAMP)
                        )
                        RETURNING id;
                        """,
                        {**user_data, "user_id": backup_user_id},
                    )
                    target_user_id = cur.fetchone()["id"]
                else:
                    cur.execute(
                        """
                        INSERT INTO users (
                            username,
                            password,
                            role,
                            real_name,
                            phone,
                            station_id,
                            created_at,
                            updated_at
                        )
                        VALUES (
                            %(username)s,
                            %(password)s,
                            %(role)s,
                            %(real_name)s,
                            %(phone)s,
                            %(station_id)s,
                            COALESCE(NULLIF(%(created_at)s, '')::timestamp, CURRENT_TIMESTAMP),
                            COALESCE(NULLIF(%(updated_at)s, '')::timestamp, CURRENT_TIMESTAMP)
                        )
                        RETURNING id;
                        """,
                        user_data,
                    )
                    target_user_id = cur.fetchone()["id"]

            target_user = {
                "id": target_user_id,
                "role": user_data["role"],
            }
            cur.execute("DELETE FROM user_permissions WHERE user_id = %s;", (target_user_id,))
            apply_user_permission_updates(cur, target_user, user_data["permissions"], actor["id"])
            apply_user_inspection_table_scope_updates(
                cur,
                target_user,
                user_data["inspection_table_scope_ids"],
                actor["id"],
            )
            apply_user_station_region_scope_updates(
                cur,
                target_user,
                user_data["station_region_scope_values"],
                actor["id"],
            )
            imported_count += 1

        cur.execute(
            """
            SELECT setval(
                pg_get_serial_sequence('users', 'id'),
                GREATEST(COALESCE((SELECT MAX(id) FROM users), 1), 1),
                TRUE
            );
            """
        )
        conn.commit()
        invalidate_auth_caches_for_user()

        skipped_text = f"，已跳过 {skipped_builtin_count} 条内置 root 记录" if skipped_builtin_count else ""
        return jsonify(
            {
                "success": True,
                "message": f"用户数据导入完成，共处理 {imported_count} 条记录{skipped_text}。",
                "count": imported_count,
                "skipped_builtin_count": skipped_builtin_count,
                "restored_root_count": restored_root_count,
                "auth_invalidated": auth_invalidated,
            }
        )
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 404
    except ValueError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        if getattr(e, "pgcode", "") == "23505":
            return jsonify({"success": False, "error": "导入失败：备份文件中的用户名与现有用户存在冲突。"}), 400
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/users", methods=["POST"])
def create_management_user():
    data = request.get_json(silent=True) or {}
    user_id = str(data.get("user_id", "")).strip()
    conn = None
    cur = None

    try:
        user_data = build_management_user_payload(data, is_create=True)
        if user_data["role"] == "root":
            return jsonify({"success": False, "error": "系统管理员账号为内置账号，不能通过页面新增。"}), 400
        permissions = normalize_permission_updates(data.get("permissions"))
        scope_ids = normalize_inspection_table_scope_updates(data.get("inspection_table_scope_ids"))
        region_scope_values = normalize_station_region_scope_updates(data.get("station_region_scope_values"))
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_inspection_checklist_management_schema(cur)
        ensure_user_security_schema(cur)
        actor = require_management_user(cur, user_id, "manage_users")

        cur.execute(
            """
            INSERT INTO users (
                username,
                password,
                role,
                real_name,
                phone,
                station_id,
                created_at,
                updated_at
            )
            VALUES (%s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            RETURNING id, username, role;
            """,
            (
                user_data["username"],
                user_data["password"],
                user_data["role"],
                user_data["real_name"],
                user_data["phone"],
                user_data["station_id"],
            ),
        )
        created_user = cur.fetchone()
        apply_user_permission_updates(cur, created_user, permissions, actor["id"])
        apply_user_inspection_table_scope_updates(cur, created_user, scope_ids, actor["id"])
        apply_user_station_region_scope_updates(cur, created_user, region_scope_values, actor["id"])
        conn.commit()
        return jsonify({"success": True, "message": "用户已新增。", "id": created_user["id"]})
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 404
    except ValueError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        if getattr(e, "pgcode", "") == "23505":
            return jsonify({"success": False, "error": "用户名已存在。"}), 400
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/users/<int:target_user_id>", methods=["PUT"])
def update_management_user(target_user_id):
    data = request.get_json(silent=True) or {}
    user_id = str(data.get("user_id", "")).strip()
    conn = None
    cur = None

    try:
        user_data = build_management_user_payload(data, is_create=False)
        permissions = normalize_permission_updates(data.get("permissions"))
        scope_ids = normalize_inspection_table_scope_updates(data.get("inspection_table_scope_ids"))
        region_scope_values = normalize_station_region_scope_updates(data.get("station_region_scope_values"))
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_inspection_checklist_management_schema(cur)
        ensure_user_security_schema(cur)
        actor = require_management_user(cur, user_id, "manage_users")

        cur.execute(
            """
            SELECT id, username, role
            FROM users
            WHERE id = %s
            LIMIT 1;
            """,
            (target_user_id,),
        )
        target_user = cur.fetchone()
        if not target_user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404

        if target_user["role"] == "root" and user_data["role"] != "root":
            return jsonify({"success": False, "error": "root 系统管理员角色不可修改。"}), 400

        if target_user["role"] != "root" and user_data["role"] == "root":
            return jsonify({"success": False, "error": "不能将普通用户升级为 root 系统管理员。"}), 400

        if user_data["password"]:
            cur.execute(
                """
                UPDATE users
                SET username = %s,
                    password = %s,
                    role = %s,
                    real_name = %s,
                    phone = %s,
                    station_id = %s,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
                RETURNING id, username, role;
                """,
                (
                    user_data["username"],
                    user_data["password"],
                    user_data["role"],
                    user_data["real_name"],
                    user_data["phone"],
                    user_data["station_id"],
                    target_user_id,
                ),
            )
        else:
            cur.execute(
                """
                UPDATE users
                SET username = %s,
                    role = %s,
                    real_name = %s,
                    phone = %s,
                    station_id = %s,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
                RETURNING id, username, role;
                """,
                (
                    user_data["username"],
                    user_data["role"],
                    user_data["real_name"],
                    user_data["phone"],
                    user_data["station_id"],
                    target_user_id,
                ),
            )
        updated_user = cur.fetchone()
        apply_user_permission_updates(cur, updated_user, permissions, actor["id"])
        apply_user_inspection_table_scope_updates(cur, updated_user, scope_ids, actor["id"])
        apply_user_station_region_scope_updates(cur, updated_user, region_scope_values, actor["id"])
        conn.commit()
        invalidate_auth_caches_for_user(target_user_id)
        return jsonify({"success": True, "message": "用户已更新。"})
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 404
    except ValueError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        if getattr(e, "pgcode", "") == "23505":
            return jsonify({"success": False, "error": "用户名已存在。"}), 400
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/users/<int:target_user_id>", methods=["DELETE"])
def delete_management_user(target_user_id):
    data = request.get_json(silent=True) or {}
    user_id = str(data.get("user_id") or request.args.get("user_id", "")).strip()
    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_user_security_schema(cur)
        actor = require_management_user(cur, user_id, "manage_users")
        if str(actor["id"]) == str(target_user_id):
            return jsonify({"success": False, "error": "不能删除当前登录账号。"}), 400

        cur.execute(
            """
            SELECT id, role
            FROM users
            WHERE id = %s
            LIMIT 1;
            """,
            (target_user_id,),
        )
        target_user = cur.fetchone()
        if not target_user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404

        if target_user["role"] == "root":
            return jsonify({"success": False, "error": "root 系统管理员账号不可删除。"}), 400

        cur.execute("DELETE FROM users WHERE id = %s;", (target_user_id,))
        conn.commit()
        return jsonify({"success": True, "message": "用户已删除。"})
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 404
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/feedbacks/unread-count", methods=["GET"])
def get_system_feedback_unread_count():
    return frontend_version_expired_response()


@app.route("/api/feedbacks/mark-read", methods=["POST"])
def mark_system_feedbacks_read():
    current_user = get_current_request_user()
    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_feedback_schema(cur)
        mark_feedbacks_read(cur, current_user["id"])
        conn.commit()
        return jsonify({"success": True, "unread_count": 0})
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/feedbacks", methods=["GET"])
def get_system_feedbacks():
    current_user = get_current_request_user()
    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_feedback_schema(cur)
        conn.commit()

        cur.execute(
            """
            SELECT
                id,
                feedback_type,
                module,
                title,
                description,
                created_by,
                author_name,
                author_phone,
                author_role,
                accepted_by,
                TO_CHAR(accepted_at, 'YYYY-MM-DD HH24:MI') AS accepted_at,
                TO_CHAR(created_at, 'YYYY-MM-DD HH24:MI') AS created_at
            FROM system_feedbacks
            ORDER BY id DESC
            LIMIT 200;
            """
        )
        feedback_rows = cur.fetchall()
        feedback_ids = [row["id"] for row in feedback_rows]

        screenshots_by_feedback = {feedback_id: [] for feedback_id in feedback_ids}
        comments_by_feedback = {feedback_id: [] for feedback_id in feedback_ids}

        if feedback_ids:
            cur.execute(
                """
                SELECT
                    id,
                    feedback_id,
                    file_path,
                    sort_order
                FROM system_feedback_screenshots
                WHERE feedback_id = ANY(%s)
                ORDER BY feedback_id ASC, sort_order ASC, id ASC;
                """,
                (feedback_ids,),
            )
            for row in cur.fetchall():
                screenshots_by_feedback.setdefault(row["feedback_id"], []).append(dict(row))

            cur.execute(
                """
                SELECT
                    id,
                    feedback_id,
                    comment_text,
                    created_by,
                    author_name,
                    author_phone,
                    author_role,
                    TO_CHAR(created_at, 'YYYY-MM-DD HH24:MI') AS created_at
                FROM system_feedback_comments
                WHERE feedback_id = ANY(%s)
                ORDER BY feedback_id ASC, id ASC;
                """,
                (feedback_ids,),
            )
            for row in cur.fetchall():
                comments_by_feedback.setdefault(row["feedback_id"], []).append(
                    serialize_feedback_comment(row, is_root_user(current_user))
                )

        return jsonify(
            {
                "success": True,
                "items": [
                    serialize_feedback_row(
                        row,
                        screenshots_by_feedback.get(row["id"], []),
                        comments_by_feedback.get(row["id"], []),
                        is_root_user(current_user),
                    )
                    for row in feedback_rows
                ],
                "options": {
                    "feedback_types": sorted(FEEDBACK_TYPE_OPTIONS),
                    "modules": sorted(FEEDBACK_MODULE_OPTIONS),
                },
                "can_delete": is_root_user(current_user),
            }
        )
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/feedbacks", methods=["POST"])
def create_system_feedback():
    current_user = get_current_request_user()
    saved_files = []
    conn = None
    cur = None

    try:
        feedback_type = normalize_feedback_type(request.form.get("feedback_type"))
        module = normalize_feedback_module(request.form.get("module"))
        description = normalize_text(request.form.get("description"), 3000)
        screenshot_files = collect_feedback_files(request.files)

        if not description:
            return jsonify({"success": False, "error": "请填写详细说明。"}), 400
        if len(screenshot_files) > 6:
            return jsonify({"success": False, "error": "最多上传 6 张截图。"}), 400

        title_result = generate_feedback_title(feedback_type, module, description)
        title = normalize_text(title_result.get("title"), 120) or "系统反馈事项"

        for file_storage in screenshot_files:
            saved_files.append(save_uploaded_file(file_storage, "feedback_screenshots"))

        conn = get_db_connection()
        cur = conn.cursor()
        ensure_feedback_schema(cur)
        author = build_feedback_author_snapshot(current_user)

        cur.execute(
            """
            INSERT INTO system_feedbacks (
                feedback_type,
                module,
                title,
                description,
                created_by,
                author_name,
                author_phone,
                author_role
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id;
            """,
            (
                feedback_type,
                module,
                title,
                description,
                current_user["id"],
                author["author_name"],
                author["author_phone"],
                author["author_role"],
            ),
        )
        feedback = cur.fetchone()
        feedback_id = feedback["id"]

        for index, file_path in enumerate(saved_files, start=1):
            cur.execute(
                """
                INSERT INTO system_feedback_screenshots (
                    feedback_id,
                    file_path,
                    sort_order
                )
                VALUES (%s, %s, %s);
                """,
                (feedback_id, file_path, index),
            )

        record_ai_usage_log(
            cur,
            current_user,
            title_result,
            "系统反馈",
            "自动生成反馈标题",
            f"{feedback_type} · {module} · {description[:160]}",
        )

        conn.commit()
        return jsonify(
            {
                "success": True,
                "message": "反馈已发布，所有用户均可查看并参与讨论。",
                "id": feedback_id,
                "title": title,
                "ai_title_generated": bool(title_result.get("generated")),
                "ai_title_message": title_result.get("message") or "",
            }
        )
    except ValueError as e:
        if conn:
            conn.rollback()
        for file_path in saved_files:
            remove_storage_file(file_path)
        return jsonify({"success": False, "error": str(e)}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        for file_path in saved_files:
            remove_storage_file(file_path)
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/feedbacks/<int:feedback_id>/comments", methods=["POST"])
def create_system_feedback_comment(feedback_id):
    current_user = get_current_request_user()
    data = request.get_json(silent=True) or {}
    comment_text = normalize_text(data.get("comment_text"), 1200)
    conn = None
    cur = None

    if not comment_text:
        return jsonify({"success": False, "error": "请填写讨论内容。"}), 400

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_feedback_schema(cur)

        cur.execute("SELECT id FROM system_feedbacks WHERE id = %s LIMIT 1;", (feedback_id,))
        if not cur.fetchone():
            return jsonify({"success": False, "error": "反馈不存在。"}), 404

        author = build_feedback_author_snapshot(current_user)
        cur.execute(
            """
            INSERT INTO system_feedback_comments (
                feedback_id,
                comment_text,
                created_by,
                author_name,
                author_phone,
                author_role
            )
            VALUES (%s, %s, %s, %s, %s, %s)
            RETURNING
                id,
                feedback_id,
                comment_text,
                created_by,
                author_name,
                author_phone,
                author_role,
                TO_CHAR(created_at, 'YYYY-MM-DD HH24:MI') AS created_at;
            """,
            (
                feedback_id,
                comment_text,
                current_user["id"],
                author["author_name"],
                author["author_phone"],
                author["author_role"],
            ),
        )
        comment = cur.fetchone()
        conn.commit()
        return jsonify(
            {
                "success": True,
                "message": "讨论已发布。",
                "comment": serialize_feedback_comment(comment, is_root_user(current_user)),
            }
        )
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/feedbacks/<int:feedback_id>/acceptance", methods=["PATCH"])
def update_system_feedback_acceptance(feedback_id):
    current_user = get_current_request_user()
    if not is_root_user(current_user):
        return jsonify({"success": False, "error": "只有 root 账号可以设置反馈采纳状态。"}), 403

    data = request.get_json(silent=True) or {}
    accepted = bool(data.get("accepted"))
    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_feedback_schema(cur)

        cur.execute(
            """
            UPDATE system_feedbacks
            SET accepted_at = CASE WHEN %s THEN CURRENT_TIMESTAMP ELSE NULL END,
                accepted_by = CASE WHEN %s THEN %s ELSE NULL END
            WHERE id = %s
            RETURNING
                id,
                accepted_by,
                TO_CHAR(accepted_at, 'YYYY-MM-DD HH24:MI') AS accepted_at;
            """,
            (accepted, accepted, current_user["id"], feedback_id),
        )
        feedback = cur.fetchone()
        if not feedback:
            conn.rollback()
            return jsonify({"success": False, "error": "反馈不存在。"}), 404

        conn.commit()
        return jsonify(
            {
                "success": True,
                "message": "反馈已标记为已采纳。" if accepted else "反馈已取消采纳。",
                "feedback": {
                    "id": feedback["id"],
                    "accepted_by": feedback["accepted_by"],
                    "accepted_at": feedback["accepted_at"],
                    "is_accepted": bool(feedback["accepted_at"]),
                },
            }
        )
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/feedbacks/<int:feedback_id>", methods=["DELETE"])
def delete_system_feedback(feedback_id):
    current_user = get_current_request_user()
    if not is_root_user(current_user):
        return jsonify({"success": False, "error": "只有 root 账号可以删除反馈。"}), 403

    conn = None
    cur = None
    files_to_remove = []

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_feedback_schema(cur)

        cur.execute("SELECT id FROM system_feedbacks WHERE id = %s LIMIT 1;", (feedback_id,))
        if not cur.fetchone():
            conn.rollback()
            return jsonify({"success": False, "error": "反馈不存在。"}), 404

        cur.execute(
            """
            SELECT file_path
            FROM system_feedback_screenshots
            WHERE feedback_id = %s;
            """,
            (feedback_id,),
        )
        files_to_remove = [row["file_path"] for row in cur.fetchall() if row.get("file_path")]

        cur.execute("DELETE FROM system_feedback_comments WHERE feedback_id = %s;", (feedback_id,))
        cur.execute("DELETE FROM system_feedback_screenshots WHERE feedback_id = %s;", (feedback_id,))
        cur.execute("DELETE FROM system_feedbacks WHERE id = %s;", (feedback_id,))

        conn.commit()
        for file_path in files_to_remove:
            remove_storage_file(file_path)
        return jsonify({"success": True, "message": "反馈已删除。"})
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/feedback-comments/<int:comment_id>", methods=["DELETE"])
def delete_system_feedback_comment(comment_id):
    current_user = get_current_request_user()
    if not is_root_user(current_user):
        return jsonify({"success": False, "error": "只有 root 账号可以删除讨论。"}), 403

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_feedback_schema(cur)
        cur.execute("DELETE FROM system_feedback_comments WHERE id = %s;", (comment_id,))
        if cur.rowcount == 0:
            return jsonify({"success": False, "error": "讨论不存在。"}), 404
        conn.commit()
        return jsonify({"success": True, "message": "讨论已删除。"})
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/ai-usage", methods=["GET"])
def get_management_ai_usage():
    user_id = str(request.args.get("user_id", "")).strip()
    date_from = str(request.args.get("date_from", "")).strip()
    date_to = str(request.args.get("date_to", "")).strip()
    usage_module = str(request.args.get("usage_module", "")).strip()
    model = str(request.args.get("model", "")).strip()
    status = str(request.args.get("status", "")).strip()
    keyword = str(request.args.get("keyword", "")).strip()
    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_user_security_schema(cur)
        ensure_ai_usage_schema(cur)
        require_management_user(cur, user_id, "manage_ai_usage")

        where_clauses = []
        params = []

        if date_from:
            datetime.strptime(date_from, "%Y-%m-%d")
            where_clauses.append("created_at >= %s::date")
            params.append(date_from)

        if date_to:
            datetime.strptime(date_to, "%Y-%m-%d")
            where_clauses.append("created_at < (%s::date + INTERVAL '1 day')")
            params.append(date_to)

        if usage_module:
            where_clauses.append("usage_module = %s")
            params.append(usage_module)

        if model:
            where_clauses.append("model = %s")
            params.append(model)

        if status == "success":
            where_clauses.append("success = TRUE")
        elif status == "fallback":
            where_clauses.append("fallback_used = TRUE")
        elif status == "called":
            where_clauses.append("ai_called = TRUE")
        elif status == "not_called":
            where_clauses.append("ai_called = FALSE")

        if keyword:
            where_clauses.append(
                """
                (
                    username ILIKE %s OR
                    real_name ILIKE %s OR
                    usage_module ILIKE %s OR
                    usage_action ILIKE %s OR
                    request_summary ILIKE %s
                )
                """
            )
            like_keyword = f"%{keyword}%"
            params.extend([like_keyword] * 5)

        where_sql = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""
        cur.execute(
            f"""
            SELECT
                id,
                user_id,
                username,
                real_name,
                role,
                usage_module,
                usage_action,
                model,
                base_url,
                ai_called,
                ai_generated,
                success,
                fallback_used,
                status_code,
                prompt_chars,
                prompt_chinese_chars,
                prompt_other_chars,
                completion_chars,
                completion_chinese_chars,
                completion_other_chars,
                total_chars,
                input_tokens_est,
                output_tokens_est,
                total_tokens_est,
                input_cost_est,
                output_cost_est,
                total_cost_est,
                message,
                request_summary,
                TO_CHAR(created_at, 'YYYY-MM-DD HH24:MI') AS created_at
            FROM ai_usage_logs
            {where_sql}
            ORDER BY id DESC
            LIMIT 2000;
            """,
            params,
        )
        rows = [serialize_ai_usage_row(row) for row in cur.fetchall()]

        cur.execute(
            """
            SELECT DISTINCT usage_module
            FROM ai_usage_logs
            WHERE usage_module IS NOT NULL AND usage_module <> ''
            ORDER BY usage_module ASC;
            """
        )
        modules = [row["usage_module"] for row in cur.fetchall()]

        cur.execute(
            """
            SELECT DISTINCT model
            FROM ai_usage_logs
            WHERE model IS NOT NULL AND model <> ''
            ORDER BY model ASC;
            """
        )
        models = [row["model"] for row in cur.fetchall()]
        conn.commit()

        return jsonify(
            {
                "success": True,
                "summary": build_ai_usage_summary(rows),
                "by_user": build_ai_usage_aggregate(
                    rows,
                    ["user_id", "username"],
                    lambda row: row.get("real_name")
                    or row.get("username")
                    or f"用户{row.get('user_id') or '-'}",
                ),
                "by_context": build_ai_usage_aggregate(
                    rows,
                    ["usage_module", "usage_action"],
                    lambda row: f"{row.get('usage_module') or '-'} · {row.get('usage_action') or '-'}",
                ),
                "items": rows,
                "options": {
                    "modules": modules,
                    "models": models,
                    "pricing": get_ai_pricing_table(),
                    "status": [
                        {"value": "", "label": "全部状态"},
                        {"value": "success", "label": "AI 成功"},
                        {"value": "fallback", "label": "使用回退"},
                        {"value": "called", "label": "已请求 AI"},
                        {"value": "not_called", "label": "未请求 AI"},
                    ],
                },
            }
        )
    except PermissionError as exc:
        return jsonify({"success": False, "error": str(exc)}), 403
    except ValueError:
        return jsonify({"success": False, "error": "日期格式不正确。"}), 400
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/backups", methods=["GET"])
def get_management_backups():
    user_id = str(request.args.get("user_id", "")).strip()
    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        require_management_user(cur, user_id, "manage_backups")
        config = read_backup_config()
        cos_status, latest_cos_backups = get_cos_backup_overview(config)
        return jsonify(
            {
                "success": True,
                "config": {
                    **config,
                    "next_run_at": get_backup_next_run_at(config),
                    "cos": cos_status,
                },
                "frequency_options": [
                    {"value": "off", "label": "关闭自动备份"},
                    {"value": "hourly", "label": "每小时"},
                    {"value": "daily", "label": "每天"},
                    {"value": "weekly", "label": "每周"},
                    {"value": "monthly", "label": "每月"},
                ],
                "latest_backups": list_backup_files(config.get("destination_path")),
                "latest_cos_backups": latest_cos_backups,
            }
        )
    except PermissionError as exc:
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        return jsonify({"success": False, "error": str(exc)}), 404
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/backups/config", methods=["PUT"])
def update_management_backup_config():
    data = request.get_json(silent=True) or {}
    user_id = str(data.get("user_id", "")).strip()
    conn = None
    cur = None

    try:
        frequency = str(data.get("frequency") or "off").strip()
        if frequency not in BACKUP_FREQUENCY_INTERVALS:
            raise ValueError("自动备份频率不正确。")

        scheduled_time = normalize_backup_scheduled_time(data.get("scheduled_time"))
        destination_path = normalize_backup_destination_path(data.get("destination_path"))
        conn = get_db_connection()
        cur = conn.cursor()
        require_management_user(cur, user_id, "manage_backups")

        os.makedirs(destination_path, exist_ok=True)
        test_path = os.path.join(destination_path, ".ywddzx_backup_write_test")
        with open(test_path, "w", encoding="utf-8") as f:
            f.write("ok")
        os.remove(test_path)

        current_config = read_backup_config()
        previous_destination_path = current_config.get("destination_path")
        config = write_backup_config(
            {
                **current_config,
                "destination_path": destination_path,
                "frequency": frequency,
                "scheduled_time": scheduled_time,
                "next_auto_run_at": calculate_next_auto_run_at(
                    {
                        **current_config,
                        "frequency": frequency,
                        "scheduled_time": scheduled_time,
                    }
                ),
                "last_status": current_config.get("last_status") or "idle",
                "last_error": current_config.get("last_error") or "",
            }
        )
        if (
            previous_destination_path
            and os.path.abspath(previous_destination_path) != os.path.abspath(destination_path)
        ):
            cleanup_local_backup_files(previous_destination_path)
        cos_status, latest_cos_backups = get_cos_backup_overview(config)
        return jsonify(
            {
                "success": True,
                "message": "备份设置已保存。",
                "config": {
                    **config,
                    "next_run_at": get_backup_next_run_at(config),
                    "cos": cos_status,
                },
                "latest_backups": list_backup_files(config.get("destination_path")),
                "latest_cos_backups": latest_cos_backups,
            }
        )
    except PermissionError as exc:
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        return jsonify({"success": False, "error": str(exc)}), 404
    except ValueError as exc:
        return jsonify({"success": False, "error": str(exc)}), 400
    except OSError as exc:
        return jsonify({"success": False, "error": f"备份目录不可写：{exc}"}), 400
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/backups/export", methods=["POST"])
def export_management_full_backup():
    data = request.get_json(silent=True) or {}
    user_id = str(data.get("user_id", "")).strip()
    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        require_management_user(cur, user_id, "manage_backups")
        close_db_resources(cur, conn)
        cur = None
        conn = None

        if not backup_job_lock.acquire(blocking=False):
            return jsonify({"success": False, "error": "已有备份或恢复任务正在执行，请稍后再试。"}), 409
        try:
            config = read_backup_config()
            result = create_full_backup_archive(config.get("destination_path"), reason="manual")
            cos_result = result.get("cos") or {}
            cos_status = cos_result.get("status") or "not_configured"
            config.update(
                {
                    "last_backup_path": result["path"],
                    "last_backup_size": result["size"],
                    "last_cos_status": cos_status,
                    "last_cos_key": cos_result.get("key"),
                    "last_cos_uploaded_at": cos_result.get("uploaded_at"),
                    "last_cos_retained_count": cos_result.get("retained_count") or 0,
                    "last_cos_error": cos_result.get("message") if cos_status == "error" else "",
                    "last_status": "warning" if cos_status == "error" else "success",
                    "last_error": cos_result.get("message") if cos_status == "error" else "",
                }
            )
            write_backup_config(config)
        finally:
            backup_job_lock.release()

        response = send_file(
            result["path"],
            as_attachment=True,
            download_name=result["download_filename"],
            mimetype="application/zip",
        )
        response.headers["X-Backup-Path"] = result["path"]
        response.headers["X-Backup-Size"] = str(result["size"])
        response.headers["X-COS-Backup-Status"] = cos_status
        response.headers["X-COS-Backup-Key"] = cos_result.get("key") or ""
        response.headers["Cache-Control"] = "no-store"
        return response
    except PermissionError as exc:
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        return jsonify({"success": False, "error": str(exc)}), 404
    except RuntimeError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/backups/import", methods=["POST"])
def import_management_full_backup():
    user_id = str(request.form.get("user_id", "")).strip()
    backup_file = request.files.get("file")
    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        require_management_user(cur, user_id, "manage_backups")
        close_db_resources(cur, conn)
        cur = None
        conn = None

        if not backup_job_lock.acquire(blocking=False):
            return jsonify({"success": False, "error": "已有备份或恢复任务正在执行，请稍后再试。"}), 409
        try:
            manifest = restore_full_backup_archive(backup_file)
        finally:
            backup_job_lock.release()

        config = read_backup_config()
        config.update(
            {
                "last_status": "success",
                "last_error": "",
            }
        )
        write_backup_config(config)
        return jsonify(
            {
                "success": True,
                "message": "完整备份已导入，数据库与上传文件目录已恢复。",
                "manifest": manifest,
            }
        )
    except PermissionError as exc:
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        return jsonify({"success": False, "error": str(exc)}), 404
    except ValueError as exc:
        return jsonify({"success": False, "error": str(exc)}), 400
    except RuntimeError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/stations", methods=["GET"])
def get_management_stations():
    user_id = str(request.args.get("user_id", "")).strip()
    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_station_management_columns(cur)
        conn.commit()
        require_management_user(cur, user_id, "manage_stations")

        cur.execute(
            """
            SELECT
                s.id,
                s.station_name,
                s.region,
                s.address,
                s.longitude,
                s.latitude,
                s.station_manager_name,
                s.station_manager_phone,
                s.station_type,
                CASE
                    WHEN s.asset_type LIKE '%股权%' OR s.asset_type LIKE '%控股%' OR s.asset_type LIKE '%参股%' THEN '股权'
                    ELSE '全资'
                END AS asset_type,
                COALESCE(s.is_consolidated, '否') AS is_consolidated,
                COALESCE(s.online_3_status, '未上线') AS online_3_status,
                COALESCE(s.monitoring_status, '运行中') AS monitoring_status,
                s.hos_station_code,
                s.landline_phone,
                s.status,
                s.operating_hours,
                COALESCE(station_accounts.station_usernames, '') AS station_usernames,
                TO_CHAR(s.created_at, 'YYYY-MM-DD HH24:MI') AS created_at,
                TO_CHAR(s.updated_at, 'YYYY-MM-DD HH24:MI') AS updated_at
            FROM stations s
            LEFT JOIN (
                SELECT
                    station_id,
                    STRING_AGG(username, ' ' ORDER BY username) AS station_usernames
                FROM users
                WHERE role = 'station_manager'
                  AND station_id IS NOT NULL
                GROUP BY station_id
            ) station_accounts ON station_accounts.station_id = s.id
            ORDER BY s.id ASC;
            """
        )
        rows = cur.fetchall()
        return jsonify(
            {
                "success": True,
                "stations": rows,
                "options": {
                    "station_types": [DISPLAY_OIL_STATION_TYPE, "充电站"],
                    "asset_types": ["全资", "股权"],
                    "is_consolidated": ["是", "否"],
                    "online_3_statuses": ["上线", "上线参股模式", "未上线"],
                    "monitoring_statuses": ["运行中", "未运行"],
                    "statuses": ["营业中", "停业"],
                },
            }
        )
    except PermissionError as exc:
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        return jsonify({"success": False, "error": str(exc)}), 404
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/stations/export", methods=["GET"])
def export_management_stations():
    user_id = str(request.args.get("user_id", "")).strip()
    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_station_management_columns(cur)
        conn.commit()
        require_management_user(cur, user_id, "manage_stations")

        cur.execute(
            """
            SELECT
                id,
                station_name,
                region,
                address,
                longitude::TEXT AS longitude,
                latitude::TEXT AS latitude,
                station_manager_name,
                station_manager_phone,
                station_type,
                CASE
                    WHEN asset_type LIKE '%股权%' OR asset_type LIKE '%控股%' OR asset_type LIKE '%参股%' THEN '股权'
                    ELSE '全资'
                END AS asset_type,
                COALESCE(is_consolidated, '否') AS is_consolidated,
                COALESCE(online_3_status, '未上线') AS online_3_status,
                COALESCE(monitoring_status, '运行中') AS monitoring_status,
                hos_station_code,
                landline_phone,
                COALESCE(status, '营业中') AS status,
                COALESCE(operating_hours, '24小时') AS operating_hours,
                TO_CHAR(created_at, 'YYYY-MM-DD HH24:MI:SS') AS created_at,
                TO_CHAR(updated_at, 'YYYY-MM-DD HH24:MI:SS') AS updated_at
            FROM stations
            ORDER BY id ASC;
            """
        )
        rows = cur.fetchall()
        now = beijing_now()
        response = jsonify(
            {
                "backup_type": "ywddzx_stations",
                "version": 1,
                "exported_at": now.isoformat(),
                "stations": rows,
            }
        )
        filename = f"ywddzx_stations_backup_{now.strftime('%Y%m%d_%H%M%S')}.json"
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        response.headers["Cache-Control"] = "no-store"
        return response
    except PermissionError as exc:
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        return jsonify({"success": False, "error": str(exc)}), 404
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/stations/export-data", methods=["POST"])
def export_management_stations_data():
    data = request.get_json(silent=True) or {}
    user_id = str(data.get("user_id", "")).strip()
    conn = None
    cur = None

    try:
        selected_field_keys = normalize_station_export_field_keys(data.get("field_keys"))
        station_ids = normalize_station_export_ids(data.get("station_ids"))
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_station_management_columns(cur)
        conn.commit()
        require_management_user(cur, user_id, "manage_stations")

        rows = fetch_station_export_rows(cur, station_ids)
        if not rows:
            return jsonify({"success": False, "error": "当前筛选结果为空，不能导出。"}), 400

        output = build_station_data_export_workbook(rows, selected_field_keys)
        now = beijing_now()
        filename = f"站点数据导出_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response.headers["Cache-Control"] = "no-store"
        return response
    except PermissionError as exc:
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        return jsonify({"success": False, "error": str(exc)}), 404
    except ValueError as exc:
        return jsonify({"success": False, "error": str(exc)}), 400
    except RuntimeError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/stations/import", methods=["POST"])
def import_management_stations():
    user_id = str(request.form.get("user_id", "")).strip()
    backup_file = request.files.get("file")
    conn = None
    cur = None

    try:
        raw_stations = parse_station_backup_json(backup_file)
        station_payloads = []
        for raw_station in raw_stations:
            if not isinstance(raw_station, dict):
                raise ValueError("站点备份文件中存在无效的站点记录。")
            station_data = build_station_payload(raw_station)
            station_data["station_id"] = normalize_station_backup_id(raw_station.get("id"))
            station_data["created_at"] = normalize_text(raw_station.get("created_at")) or None
            station_data["updated_at"] = normalize_text(raw_station.get("updated_at")) or None
            station_payloads.append(station_data)

        conn = get_db_connection()
        cur = conn.cursor()
        ensure_station_management_columns(cur)
        require_management_user(cur, user_id, "manage_stations")

        imported_count = 0
        for station_data in station_payloads:
            if station_data["station_id"]:
                cur.execute(
                    """
                    INSERT INTO stations (
                        id,
                        station_name,
                        region,
                        address,
                        longitude,
                        latitude,
                        station_manager_name,
                        station_manager_phone,
                        station_type,
                        asset_type,
                        is_consolidated,
                        online_3_status,
                        monitoring_status,
                        hos_station_code,
                        landline_phone,
                        status,
                        operating_hours,
                        created_at,
                        updated_at
                    )
                    VALUES (
                        %(station_id)s,
                        %(station_name)s,
                        %(region)s,
                        %(address)s,
                        %(longitude)s,
                        %(latitude)s,
                        %(station_manager_name)s,
                        %(station_manager_phone)s,
                        %(station_type)s,
                        %(asset_type)s,
                        %(is_consolidated)s,
                        %(online_3_status)s,
                        %(monitoring_status)s,
                        %(hos_station_code)s,
                        %(landline_phone)s,
                        %(status)s,
                        %(operating_hours)s,
                        COALESCE(NULLIF(%(created_at)s, '')::timestamp, CURRENT_TIMESTAMP),
                        COALESCE(NULLIF(%(updated_at)s, '')::timestamp, CURRENT_TIMESTAMP)
                    )
                    ON CONFLICT (id)
                    DO UPDATE SET
                        station_name = EXCLUDED.station_name,
                        region = EXCLUDED.region,
                        address = EXCLUDED.address,
                        longitude = EXCLUDED.longitude,
                        latitude = EXCLUDED.latitude,
                        station_manager_name = EXCLUDED.station_manager_name,
                        station_manager_phone = EXCLUDED.station_manager_phone,
                        station_type = EXCLUDED.station_type,
                        asset_type = EXCLUDED.asset_type,
                        is_consolidated = EXCLUDED.is_consolidated,
                        online_3_status = EXCLUDED.online_3_status,
                        monitoring_status = EXCLUDED.monitoring_status,
                        hos_station_code = EXCLUDED.hos_station_code,
                        landline_phone = EXCLUDED.landline_phone,
                        status = EXCLUDED.status,
                        operating_hours = EXCLUDED.operating_hours,
                        updated_at = EXCLUDED.updated_at;
                    """,
                    station_data,
                )
            else:
                cur.execute(
                    """
                    INSERT INTO stations (
                        station_name,
                        region,
                        address,
                        longitude,
                        latitude,
                        station_manager_name,
                        station_manager_phone,
                        station_type,
                        asset_type,
                        is_consolidated,
                        online_3_status,
                        monitoring_status,
                        hos_station_code,
                        landline_phone,
                        status,
                        operating_hours,
                        created_at,
                        updated_at
                    )
                    VALUES (
                        %(station_name)s,
                        %(region)s,
                        %(address)s,
                        %(longitude)s,
                        %(latitude)s,
                        %(station_manager_name)s,
                        %(station_manager_phone)s,
                        %(station_type)s,
                        %(asset_type)s,
                        %(is_consolidated)s,
                        %(online_3_status)s,
                        %(monitoring_status)s,
                        %(hos_station_code)s,
                        %(landline_phone)s,
                        %(status)s,
                        %(operating_hours)s,
                        COALESCE(NULLIF(%(created_at)s, '')::timestamp, CURRENT_TIMESTAMP),
                        COALESCE(NULLIF(%(updated_at)s, '')::timestamp, CURRENT_TIMESTAMP)
                    )
                    ON CONFLICT (station_name)
                    DO UPDATE SET
                        region = EXCLUDED.region,
                        address = EXCLUDED.address,
                        longitude = EXCLUDED.longitude,
                        latitude = EXCLUDED.latitude,
                        station_manager_name = EXCLUDED.station_manager_name,
                        station_manager_phone = EXCLUDED.station_manager_phone,
                        station_type = EXCLUDED.station_type,
                        asset_type = EXCLUDED.asset_type,
                        is_consolidated = EXCLUDED.is_consolidated,
                        online_3_status = EXCLUDED.online_3_status,
                        monitoring_status = EXCLUDED.monitoring_status,
                        hos_station_code = EXCLUDED.hos_station_code,
                        landline_phone = EXCLUDED.landline_phone,
                        status = EXCLUDED.status,
                        operating_hours = EXCLUDED.operating_hours,
                        updated_at = EXCLUDED.updated_at;
                    """,
                    station_data,
                )
            imported_count += 1

        cur.execute(
            """
            SELECT setval(
                pg_get_serial_sequence('stations', 'id'),
                GREATEST(COALESCE((SELECT MAX(id) FROM stations), 1), 1),
                TRUE
            );
            """
        )
        conn.commit()
        return jsonify(
            {
                "success": True,
                "message": f"站点数据导入完成，共处理 {imported_count} 条记录。",
                "count": imported_count,
            }
        )
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 404
    except ValueError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        if getattr(e, "pgcode", "") == "23505":
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "导入失败：备份文件中的站点名称或 HOS编码 与现有站点存在冲突。",
                    }
                ),
                400,
            )
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/stations/<int:station_id>/reset-password", methods=["POST"])
def reset_management_station_account_password(station_id):
    data = request.get_json(silent=True) or {}
    user_id = str(data.get("user_id", "")).strip()
    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_station_management_columns(cur)
        require_management_user(cur, user_id, "reset_station_account_password")

        cur.execute(
            """
            SELECT id, station_name
            FROM stations
            WHERE id = %s;
            """,
            (station_id,),
        )
        station = cur.fetchone()
        if not station:
            return jsonify({"success": False, "error": "站点不存在。"}), 404

        cur.execute(
            """
            UPDATE users
            SET password = %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE role = 'station_manager'
              AND station_id = %s
            RETURNING id, username, real_name;
            """,
            (DEFAULT_INITIAL_PASSWORD, station_id),
        )
        accounts = cur.fetchall()
        if not accounts:
            return jsonify({"success": False, "error": "该站点暂无绑定站点账号。"}), 400

        conn.commit()
        for account in accounts:
            invalidate_auth_caches_for_user(account.get("id"))
        usernames = [account["username"] for account in accounts if account.get("username")]
        return jsonify(
            {
                "success": True,
                "message": f"已将【{station['station_name']}】绑定站点账号密码重置为 123456，下次登录需重新设置密码。",
                "reset_count": len(accounts),
                "usernames": usernames,
            }
        )
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 404
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/stations", methods=["POST"])
def create_management_station():
    data = request.get_json(silent=True) or {}
    user_id = str(data.get("user_id", "")).strip()
    conn = None
    cur = None

    try:
        station_data = build_station_payload(data)
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_station_management_columns(cur)
        require_management_user(cur, user_id, "manage_stations")

        cur.execute(
            """
            INSERT INTO stations (
                station_name,
                region,
                address,
                longitude,
                latitude,
                station_manager_name,
                station_manager_phone,
                station_type,
                asset_type,
                is_consolidated,
                online_3_status,
                monitoring_status,
                hos_station_code,
                landline_phone,
                status,
                operating_hours,
                created_at,
                updated_at
            )
            VALUES (
                %(station_name)s,
                %(region)s,
                %(address)s,
                %(longitude)s,
                %(latitude)s,
                %(station_manager_name)s,
                %(station_manager_phone)s,
                %(station_type)s,
                %(asset_type)s,
                %(is_consolidated)s,
                %(online_3_status)s,
                %(monitoring_status)s,
                %(hos_station_code)s,
                %(landline_phone)s,
                %(status)s,
                %(operating_hours)s,
                CURRENT_TIMESTAMP,
                CURRENT_TIMESTAMP
            )
            RETURNING id;
            """,
            station_data,
        )
        row = cur.fetchone()
        conn.commit()
        return jsonify({"success": True, "message": "站点已新增。", "id": row["id"]})
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 404
    except ValueError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        if getattr(e, "pgcode", "") == "23505":
            return jsonify({"success": False, "error": "站点名称或 HOS编码 已存在，请检查是否重复添加。"}), 400
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/stations/<int:station_id>", methods=["PUT"])
def update_management_station(station_id):
    data = request.get_json(silent=True) or {}
    user_id = str(data.get("user_id", "")).strip()
    conn = None
    cur = None

    try:
        station_data = build_station_payload(data)
        station_data["station_id"] = station_id
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_station_management_columns(cur)
        require_management_user(cur, user_id, "manage_stations")

        cur.execute(
            """
            UPDATE stations
            SET station_name = %(station_name)s,
                region = %(region)s,
                address = %(address)s,
                longitude = %(longitude)s,
                latitude = %(latitude)s,
                station_manager_name = %(station_manager_name)s,
                station_manager_phone = %(station_manager_phone)s,
                station_type = %(station_type)s,
                asset_type = %(asset_type)s,
                is_consolidated = %(is_consolidated)s,
                online_3_status = %(online_3_status)s,
                monitoring_status = %(monitoring_status)s,
                hos_station_code = %(hos_station_code)s,
                landline_phone = %(landline_phone)s,
                status = %(status)s,
                operating_hours = %(operating_hours)s,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = %(station_id)s
            RETURNING id;
            """,
            station_data,
        )
        row = cur.fetchone()
        if not row:
            return jsonify({"success": False, "error": "站点不存在。"}), 404

        conn.commit()
        return jsonify({"success": True, "message": "站点已更新。"})
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 404
    except ValueError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        if getattr(e, "pgcode", "") == "23505":
            return jsonify({"success": False, "error": "站点名称或 HOS编码 已存在，请检查是否重复添加。"}), 400
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/stations/<int:station_id>", methods=["DELETE"])
def delete_management_station(station_id):
    data = request.get_json(silent=True) or {}
    user_id = str(data.get("user_id") or request.args.get("user_id", "")).strip()
    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_station_management_columns(cur)
        require_management_user(cur, user_id, "manage_stations")

        cur.execute("DELETE FROM stations WHERE id = %s RETURNING id;", (station_id,))
        row = cur.fetchone()
        if not row:
            return jsonify({"success": False, "error": "站点不存在。"}), 404

        conn.commit()
        return jsonify({"success": True, "message": "站点已删除。"})
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 404
    except Exception as e:
        if conn:
            conn.rollback()
        if getattr(e, "pgcode", "") == "23503":
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "该站点已有用户、巡检、问题或证照等业务数据引用，暂不能删除。",
                    }
                ),
                400,
            )
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/checklists", methods=["GET"])
def get_management_checklists():
    user_id = str(request.args.get("user_id", "")).strip()
    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_inspection_checklist_management_schema(cur)
        conn.commit()
        require_management_user(cur, user_id, "manage_checklists")

        cur.execute(
            """
            SELECT
                id,
                table_code,
                table_name,
                checklist_mode,
                standard_id_base,
                description,
                is_active,
                TO_CHAR(created_at, 'YYYY-MM-DD HH24:MI') AS created_at,
                TO_CHAR(updated_at, 'YYYY-MM-DD HH24:MI') AS updated_at
            FROM inspection_tables
            ORDER BY id ASC;
            """
        )
        rows = cur.fetchall()
        return jsonify(
            {
                "success": True,
                "checklists": [serialize_management_checklist(cur, row) for row in rows],
            }
        )
    except PermissionError as exc:
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        return jsonify({"success": False, "error": str(exc)}), 404
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/checklists", methods=["POST"])
def create_management_checklist():
    data = request.get_json(silent=True) or {}
    user_id = str(data.get("user_id", "")).strip()
    conn = None
    cur = None

    try:
        table_code = normalize_checklist_code(data.get("table_code"))
        table_name = sanitize_display_string(normalize_text(data.get("table_name"), 120))
        if not table_name:
            raise ValueError("请填写检查表名称。")
        checklist_mode = normalize_checklist_mode(data.get("checklist_mode"))
        description = normalize_text(data.get("description"), 300)
        fields = normalize_checklist_field_rows(data.get("fields"), table_code)
        physical_table_name = get_physical_table_name_by_code(table_code)

        conn = get_db_connection()
        cur = conn.cursor()
        ensure_inspection_checklist_management_schema(cur)
        require_management_user(cur, user_id, "manage_checklists")

        cur.execute(
            """
            SELECT id
            FROM inspection_tables
            WHERE table_code = %s OR table_name = %s
            LIMIT 1;
            """,
            (table_code, table_name),
        )
        if cur.fetchone():
            raise ValueError("检查表编码或名称已存在。")
        ensure_unique_checklist_field_keys(cur, fields)

        cur.execute(
            """
            INSERT INTO inspection_tables (
                table_code,
                table_name,
                checklist_mode,
                description,
                is_active,
                created_at,
                updated_at
            )
            VALUES (%s, %s, %s, %s, TRUE, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            RETURNING id;
            """,
            (table_code, table_name, checklist_mode, description),
        )
        row = cur.fetchone()
        standard_id_base = get_default_checklist_standard_id_base(row["id"])
        cur.execute(
            """
            UPDATE inspection_tables
            SET standard_id_base = %s
            WHERE id = %s;
            """,
            (standard_id_base, row["id"]),
        )
        ensure_checklist_field_columns(cur, physical_table_name, fields)
        upsert_checklist_fields(cur, row["id"], fields)
        conn.commit()
        return jsonify(
            {
                "success": True,
                "message": "检查表已创建，字段结构已生成。",
                "id": row["id"],
                "standard_id_base": standard_id_base,
                "physical_table_name": physical_table_name,
            }
        )
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 404
    except ValueError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        if getattr(e, "pgcode", "") == "23505":
            return jsonify({"success": False, "error": "检查表编码、名称或字段系统标识已存在。"}), 400
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/checklists/<int:inspection_table_id>", methods=["PUT"])
def update_management_checklist(inspection_table_id):
    data = request.get_json(silent=True) or {}
    user_id = str(data.get("user_id", "")).strip()
    conn = None
    cur = None

    try:
        table_name = sanitize_display_string(normalize_text(data.get("table_name"), 120))
        if not table_name:
            raise ValueError("请填写检查表名称。")
        checklist_mode = normalize_checklist_mode(data.get("checklist_mode"))
        description = normalize_text(data.get("description"), 300)
        is_active = bool(data.get("is_active", True))

        conn = get_db_connection()
        cur = conn.cursor()
        ensure_inspection_checklist_management_schema(cur)
        require_management_user(cur, user_id, "manage_checklists")

        current_row = fetch_management_checklist(cur, inspection_table_id)
        if not current_row:
            return jsonify({"success": False, "error": "检查表不存在。"}), 404
        fields = normalize_checklist_field_rows(data.get("fields"), current_row["table_code"])
        ensure_unique_checklist_field_keys(cur, fields, inspection_table_id)
        physical_table_name = get_physical_table_name_by_code(current_row["table_code"])
        ensure_checklist_field_columns(cur, physical_table_name, fields)
        upsert_checklist_fields(cur, inspection_table_id, fields)
        cur.execute(
            """
            UPDATE inspection_tables
            SET table_name = %s,
                checklist_mode = %s,
                description = %s,
                is_active = %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = %s
            RETURNING id;
            """,
            (table_name, checklist_mode, description, is_active, inspection_table_id),
        )
        conn.commit()
        return jsonify({"success": True, "message": "检查表信息和字段结构已保存。"})
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 404
    except ValueError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        if getattr(e, "pgcode", "") == "23505":
            return jsonify({"success": False, "error": "检查表名称或字段系统标识已存在。"}), 400
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/checklists/<int:inspection_table_id>", methods=["DELETE"])
def delete_management_checklist(inspection_table_id):
    data = request.get_json(silent=True) or {}
    user_id = str(data.get("user_id") or request.args.get("user_id", "")).strip()
    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_inspection_checklist_management_schema(cur)
        require_management_user(cur, user_id, "manage_checklists")

        current_row = fetch_management_checklist(cur, inspection_table_id)
        if not current_row:
            return jsonify({"success": False, "error": "检查表不存在。"}), 404

        blockers = get_blocking_checklist_references(cur, [inspection_table_id])
        if blockers:
            raise ValueError(
                "无法删除：该检查表仍被"
                f"{'、'.join(blockers)}引用。请先处理相关业务数据后再删除。"
            )

        physical_table_name = get_physical_table_name_by_code(current_row["table_code"])
        cur.execute("DELETE FROM inspection_table_fields WHERE inspection_table_id = %s;", (inspection_table_id,))
        cur.execute("DELETE FROM inspection_tables WHERE id = %s;", (inspection_table_id,))
        if physical_table_name:
            cur.execute(sql.SQL("DROP TABLE IF EXISTS {};").format(sql.Identifier(physical_table_name)))
        conn.commit()
        return jsonify({"success": True, "message": "检查表已删除。"})
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 404
    except ValueError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        if getattr(e, "pgcode", "") == "23503":
            return jsonify({"success": False, "error": "该检查表已有业务数据引用，暂不能删除。"}), 400
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/checklists/export", methods=["GET"])
def export_management_checklists():
    user_id = str(request.args.get("user_id", "")).strip()
    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_inspection_checklist_management_schema(cur)
        conn.commit()
        require_management_user(cur, user_id, "manage_checklists")

        cur.execute(
            """
            SELECT
                id,
                table_code,
                table_name,
                checklist_mode,
                standard_id_base,
                description,
                is_active,
                TO_CHAR(created_at, 'YYYY-MM-DD HH24:MI:SS') AS created_at,
                TO_CHAR(updated_at, 'YYYY-MM-DD HH24:MI:SS') AS updated_at
            FROM inspection_tables
            ORDER BY id ASC;
            """
        )
        exported_checklists = []
        for row in cur.fetchall():
            fields = [dict(field) for field in get_management_checklist_fields(cur, row["id"])]
            physical_table_name = get_physical_table_name_by_code(row["table_code"])
            ensure_checklist_field_columns(cur, physical_table_name, fields)
            exported_checklists.append(
                {
                    "table_code": row["table_code"],
                    "table_name": row["table_name"],
                    "checklist_mode": normalize_checklist_mode(row.get("checklist_mode")),
                    "standard_id_base": normalize_checklist_standard_id_base(
                        row.get("standard_id_base"),
                        row["id"],
                    ),
                    "description": row["description"],
                    "is_active": row["is_active"],
                    "physical_table_name": physical_table_name,
                    "created_at": row["created_at"],
                    "updated_at": row["updated_at"],
                    "fields": fields,
                    "standards": fetch_checklist_standard_rows(cur, physical_table_name, fields),
                }
            )

        conn.commit()
        now = beijing_now()
        response = jsonify(
            {
                "backup_type": "ywddzx_checklists",
                "version": 2,
                "exported_at": now.isoformat(),
                "checklists": exported_checklists,
            }
        )
        filename = f"ywddzx_checklists_backup_{now.strftime('%Y%m%d_%H%M%S')}.json"
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        response.headers["Cache-Control"] = "no-store"
        return response
    except PermissionError as exc:
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        return jsonify({"success": False, "error": str(exc)}), 404
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/checklists/import", methods=["POST"])
def import_management_checklists_backup():
    user_id = str(request.form.get("user_id", "")).strip()
    backup_file = request.files.get("file")
    conn = None
    cur = None

    try:
        parsed_backup = parse_checklist_backup_file(backup_file)
        imported_checklists = parsed_backup["checklists"]
        imported_codes = {item["table_code"] for item in imported_checklists}
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_inspection_checklist_management_schema(cur)
        require_management_user(cur, user_id, "manage_checklists")

        cur.execute(
            """
            SELECT id, table_code, standard_id_base
            FROM inspection_tables
            ORDER BY id ASC;
            """
        )
        existing_rows = cur.fetchall()
        existing_by_code = {row["table_code"]: row for row in existing_rows}
        rows_to_remove = [row for row in existing_rows if row["table_code"] not in imported_codes]
        blockers = get_blocking_checklist_references(cur, [row["id"] for row in rows_to_remove])
        if blockers:
            raise ValueError(
                "无法覆盖导入：当前系统中有检查表不在备份文件内，但仍被"
                f"{'、'.join(blockers)}引用。请先处理这些业务数据，或导入包含这些检查表的备份文件。"
            )

        removed_count = 0
        for row in rows_to_remove:
            physical_table_name = get_physical_table_name_by_code(row["table_code"])
            cur.execute("DELETE FROM inspection_table_fields WHERE inspection_table_id = %s;", (row["id"],))
            cur.execute("DELETE FROM inspection_tables WHERE id = %s;", (row["id"],))
            if physical_table_name:
                cur.execute(sql.SQL("DROP TABLE IF EXISTS {};").format(sql.Identifier(physical_table_name)))
            removed_count += 1

        for item in imported_checklists:
            current_row = existing_by_code.get(item["table_code"])
            if current_row:
                inspection_table_id = current_row["id"]
                target_standard_id_base = normalize_checklist_standard_id_base(
                    current_row.get("standard_id_base"),
                    inspection_table_id,
                )
                cur.execute(
                    """
                    UPDATE inspection_tables
                    SET table_name = %s,
                        checklist_mode = %s,
                        description = %s,
                        is_active = %s,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = %s;
                    """,
                    (
                        item["table_name"],
                        item["checklist_mode"],
                        item["description"],
                        item["is_active"],
                        inspection_table_id,
                    ),
                )
            else:
                cur.execute(
                    """
                    INSERT INTO inspection_tables (
                        table_code,
                        table_name,
                        checklist_mode,
                        description,
                        is_active,
                        created_at,
                        updated_at
                    )
                    VALUES (%s, %s, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                    RETURNING id;
                    """,
                    (
                        item["table_code"],
                        item["table_name"],
                        item["checklist_mode"],
                        item["description"],
                        item["is_active"],
                    ),
                )
                inspection_table_id = cur.fetchone()["id"]
                target_standard_id_base = get_default_checklist_standard_id_base(inspection_table_id)
                cur.execute(
                    """
                    UPDATE inspection_tables
                    SET standard_id_base = %s
                    WHERE id = %s;
                    """,
                    (target_standard_id_base, inspection_table_id),
                )

            physical_table_name = get_physical_table_name_by_code(item["table_code"])
            cur.execute(sql.SQL("DROP TABLE IF EXISTS {};").format(sql.Identifier(physical_table_name)))
            ensure_checklist_field_columns(cur, physical_table_name, item["fields"])
            upsert_checklist_fields(cur, inspection_table_id, item["fields"])
            standards = rebase_checklist_standard_rows(
                item["standards"],
                item["standard_id_base"],
                target_standard_id_base,
            )
            insert_checklist_standard_rows(
                cur,
                physical_table_name,
                item["fields"],
                standards,
            )

        conn.commit()
        return jsonify(
            {
                "success": True,
                "message": f"巡检表备份已导入，覆盖 {len(imported_checklists)} 张检查表，移除 {removed_count} 张旧检查表。",
                "imported_count": len(imported_checklists),
                "removed_count": removed_count,
            }
        )
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 404
    except ValueError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": f"巡检表备份导入失败：{str(e)}"}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/checklists/<int:inspection_table_id>/standards", methods=["GET"])
def get_management_checklist_standards(inspection_table_id):
    user_id = str(request.args.get("user_id", "")).strip()
    keyword = str(request.args.get("keyword", "")).strip()
    page, page_size = normalize_page_args(
        request.args.get("page"),
        request.args.get("page_size"),
    )
    offset = (page - 1) * page_size
    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_inspection_checklist_management_schema(cur)
        require_management_user(cur, user_id, "manage_checklists")
        _checklist, fields, physical_table_name = get_management_checklist_standard_context(
            cur, inspection_table_id, require_fields=False
        )

        where_sql, where_params = build_standard_search_sql(fields, keyword)
        cur.execute(
            sql.SQL("SELECT COUNT(*) AS total FROM {}{};").format(
                sql.Identifier(physical_table_name),
                where_sql,
            ),
            where_params,
        )
        total = int(cur.fetchone()["total"] or 0)
        total_pages = max((total + page_size - 1) // page_size, 1)
        if page > total_pages:
            page = total_pages
            offset = (page - 1) * page_size

        select_columns = [
            sql.Identifier("id"),
            sql.Identifier("standard_id"),
            sql.SQL("TO_CHAR(created_at, 'YYYY-MM-DD HH24:MI') AS created_at"),
            *[sql.Identifier(field["field_key"]) for field in fields],
        ]
        cur.execute(
            sql.SQL(
                """
                SELECT {}
                FROM {}{}
                ORDER BY standard_id ASC
                LIMIT %s OFFSET %s;
                """
            ).format(
                sql.SQL(", ").join(select_columns),
                sql.Identifier(physical_table_name),
                where_sql,
            ),
            [*where_params, page_size, offset],
        )
        field_meta = [(field["field_key"], field["field_label"]) for field in fields]
        items = [serialize_standard_row(field_meta, row) for row in cur.fetchall()]
        conn.commit()
        return jsonify(
            {
                "success": True,
                "items": items,
                "fields": fields,
                "page": page,
                "page_size": page_size,
                "total": total,
                "total_pages": total_pages,
            }
        )
    except PermissionError as exc:
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        return jsonify({"success": False, "error": str(exc)}), 404
    except ValueError as exc:
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/checklists/<int:inspection_table_id>/standards", methods=["POST"])
def create_management_checklist_standard(inspection_table_id):
    data = request.get_json(silent=True) or {}
    user_id = str(data.get("user_id", "")).strip()
    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_inspection_checklist_management_schema(cur)
        require_management_user(cur, user_id, "manage_checklists")
        checklist, fields, physical_table_name = get_management_checklist_standard_context(
            cur, inspection_table_id
        )
        row = normalize_checklist_standard_payload(data, fields)
        row["standard_id"] = get_next_checklist_standard_id(cur, checklist, physical_table_name)

        columns = ["standard_id", *[field["field_key"] for field in fields]]
        cur.execute(
            sql.SQL(
                """
                INSERT INTO {} ({})
                VALUES ({});
                """
            ).format(
                sql.Identifier(physical_table_name),
                sql.SQL(", ").join(sql.Identifier(column) for column in columns),
                sql.SQL(", ").join(sql.Placeholder() for _ in columns),
            ),
            [row.get(column) for column in columns],
        )
        cur.execute(
            """
            UPDATE inspection_tables
            SET updated_at = CURRENT_TIMESTAMP
            WHERE id = %s;
            """,
            (inspection_table_id,),
        )
        conn.commit()
        return jsonify(
            {
                "success": True,
                "message": f"外部规范数据已新增，系统生成外部规范ID {row['standard_id']}。",
                "standard_id": row["standard_id"],
            }
        )
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 404
    except ValueError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        if getattr(e, "pgcode", "") == "23505":
            return jsonify({"success": False, "error": "外部规范ID自动生成冲突，请重试。"}), 400
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/checklists/<int:inspection_table_id>/standards/<standard_id>", methods=["PUT"])
def update_management_checklist_standard(inspection_table_id, standard_id):
    data = request.get_json(silent=True) or {}
    user_id = str(data.get("user_id", "")).strip()
    conn = None
    cur = None

    try:
        old_standard_id = normalize_import_standard_id(standard_id)
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_inspection_checklist_management_schema(cur)
        require_management_user(cur, user_id, "manage_checklists")
        _checklist, fields, physical_table_name = get_management_checklist_standard_context(
            cur, inspection_table_id
        )
        row = normalize_checklist_standard_payload(data, fields)

        cur.execute(
            sql.SQL("SELECT id FROM {} WHERE standard_id = %s LIMIT 1;").format(
                sql.Identifier(physical_table_name)
            ),
            (old_standard_id,),
        )
        if not cur.fetchone():
            return jsonify({"success": False, "error": "规范数据不存在。"}), 404

        columns = [field["field_key"] for field in fields]
        cur.execute(
            sql.SQL(
                """
                UPDATE {}
                SET {}
                WHERE standard_id = %s
                RETURNING standard_id;
                """
            ).format(
                sql.Identifier(physical_table_name),
                sql.SQL(", ").join(
                    sql.SQL("{} = %s").format(sql.Identifier(column))
                    for column in columns
                ),
            ),
            [row.get(column) for column in columns] + [old_standard_id],
        )
        if not cur.fetchone():
            return jsonify({"success": False, "error": "规范数据不存在。"}), 404

        standard_detail_text = build_standard_detail_text(
            [(field["field_key"], field["field_label"]) for field in fields],
            row,
        )
        if not standard_detail_text:
            raise ValueError("规范详情生成失败，请至少保留一项规范内容。")
        synced_count = sync_referenced_standard_detail_text(
            cur,
            inspection_table_id,
            old_standard_id,
            standard_detail_text,
        )

        cur.execute(
            """
            UPDATE inspection_tables
            SET updated_at = CURRENT_TIMESTAMP
            WHERE id = %s;
            """,
            (inspection_table_id,),
        )
        conn.commit()
        message = "规范数据已保存。"
        if synced_count:
            message = f"规范数据已保存，并同步更新 {synced_count} 条已引用问题。"
        return jsonify({"success": True, "message": message, "synced_issue_count": synced_count})
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 404
    except ValueError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/checklists/<int:inspection_table_id>/standards/<standard_id>", methods=["DELETE"])
def delete_management_checklist_standard(inspection_table_id, standard_id):
    data = request.get_json(silent=True) or {}
    user_id = str(data.get("user_id") or request.args.get("user_id", "")).strip()
    conn = None
    cur = None

    try:
        standard_id_value = normalize_import_standard_id(standard_id)
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_inspection_checklist_management_schema(cur)
        require_management_user(cur, user_id, "manage_checklists")
        _checklist, fields, physical_table_name = get_management_checklist_standard_context(
            cur, inspection_table_id, require_fields=False
        )
        ensure_checklist_field_columns(cur, physical_table_name, fields)

        reference_counts = get_checklist_standard_reference_counts(
            cur,
            inspection_table_id,
            standard_id_value,
        )
        reference_message = format_checklist_standard_reference_message(reference_counts)
        if reference_message:
            raise ValueError(
                f"无法删除：该规范已被{reference_message}引用。"
                "可以编辑规范内容，系统会同步更新已引用该规范的业务数据，但不能删除已被使用的规范。"
            )

        cur.execute(
            sql.SQL("DELETE FROM {} WHERE standard_id = %s RETURNING standard_id;").format(
                sql.Identifier(physical_table_name)
            ),
            (standard_id_value,),
        )
        if not cur.fetchone():
            return jsonify({"success": False, "error": "规范数据不存在。"}), 404

        cur.execute(
            """
            UPDATE inspection_tables
            SET updated_at = CURRENT_TIMESTAMP
            WHERE id = %s;
            """,
            (inspection_table_id,),
        )
        conn.commit()
        return jsonify({"success": True, "message": "规范数据已删除。"})
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 404
    except ValueError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/station-certificates", methods=["GET"])
def get_station_certificates():
    user_id = str(request.args.get("user_id", "")).strip()
    if not user_id:
        return jsonify({"success": False, "error": "缺少用户信息。"}), 400

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_station_certificates_table(cur)
        conn.commit()

        user = get_user_by_id(cur, user_id)
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404

        if not can_view_certificates(cur, user):
            return (
                jsonify({"success": False, "error": "当前账号无权访问证照管理。"}),
                403,
            )

        scope_all = can_edit_all_certificates(user) or can_view_all_certificates(cur, user)
        scope_region = (not scope_all) and can_view_region_certificates(cur, user)
        can_edit_own = can_edit_own_certificates(cur, user)
        empty_scope_response = {
            "success": True,
            "certificate_types": CERTIFICATE_TYPES,
            "stations": [],
            "records": [],
            "can_view_all": scope_all,
            "can_view_region": scope_region,
            "can_edit_all": can_edit_all_certificates(user),
            "can_edit_own": can_edit_own,
        }

        station_params = []
        station_where_clauses = []
        if scope_all:
            pass
        elif scope_region:
            if not append_station_region_scope_filter(
                cur,
                user,
                station_where_clauses,
                station_params,
                "region",
                "limit_certificate_station_region_scope",
            ):
                return jsonify(empty_scope_response)
        else:
            if not user["station_id"]:
                return jsonify(empty_scope_response)
            station_where_clauses.append("id = %s")
            station_params.append(user["station_id"])
        station_where = f"WHERE {' AND '.join(station_where_clauses)}" if station_where_clauses else ""

        cur.execute(
            f"""
            SELECT
                id,
                station_name,
                region,
                address,
                station_type,
                asset_type,
                status
            FROM stations
            {station_where}
            ORDER BY region NULLS LAST, station_name ASC, id ASC;
            """,
            tuple(station_params),
        )
        stations = cur.fetchall()

        record_params = []
        record_where_clauses = []
        if scope_all:
            pass
        elif scope_region:
            if not append_station_region_scope_filter(
                cur,
                user,
                record_where_clauses,
                record_params,
                "s.region",
                "limit_certificate_station_region_scope",
            ):
                return jsonify(empty_scope_response)
        else:
            record_where_clauses.append("sc.station_id = %s")
            record_params.append(user["station_id"])
        record_where = f"WHERE {' AND '.join(record_where_clauses)}" if record_where_clauses else ""

        cur.execute(
            f"""
            SELECT
                sc.id,
                sc.station_id,
                s.station_name,
                s.region,
                s.asset_type,
                sc.certificate_type,
                sc.certificate_name,
                TO_CHAR(sc.start_date, 'YYYY-MM-DD') AS start_date,
                TO_CHAR(sc.expiry_date, 'YYYY-MM-DD') AS expiry_date,
                sc.remark,
                TO_CHAR(sc.created_at, 'YYYY-MM-DD HH24:MI') AS created_at,
                TO_CHAR(sc.updated_at, 'YYYY-MM-DD HH24:MI') AS updated_at
            FROM station_certificates sc
            JOIN stations s ON sc.station_id = s.id
            {record_where}
            ORDER BY sc.expiry_date ASC, s.station_name ASC, sc.certificate_name ASC;
            """,
            tuple(record_params),
        )
        records = cur.fetchall()

        for record in records:
            type_meta = CERTIFICATE_TYPE_BY_CODE.get(record["certificate_type"], {})
            if type_meta:
                record["certificate_name"] = type_meta.get(
                    "name", record["certificate_name"]
                )
            record["recommended_reminder_days"] = type_meta.get(
                "recommended_reminder_days", 30
            )
            record["legal_reminder_days"] = type_meta.get("legal_reminder_days", 7)
            record["recommended_label"] = type_meta.get("recommended_label")
            record["legal_label"] = type_meta.get("legal_label")
            record["rule"] = type_meta.get("rule")

        return jsonify(
            {
                "success": True,
                "certificate_types": CERTIFICATE_TYPES,
                "stations": stations,
                "records": records,
                "can_view_all": scope_all,
                "can_view_region": scope_region,
                "can_edit_all": can_edit_all_certificates(user),
                "can_edit_own": can_edit_own,
            }
        )
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/station-certificates", methods=["POST"])
def save_station_certificate():
    data = request.get_json(silent=True) or {}
    user_id = str(data.get("user_id", "")).strip()

    if not user_id:
        return jsonify({"success": False, "error": "缺少用户信息。"}), 400

    station_id = data.get("station_id")
    certificate_type = str(data.get("certificate_type", "")).strip()
    remark = str(data.get("remark", "")).strip() or None

    if not station_id:
        return jsonify({"success": False, "error": "请选择站点。"}), 400

    if certificate_type not in CERTIFICATE_TYPE_BY_CODE:
        return jsonify({"success": False, "error": "请选择有效的证照类型。"}), 400

    try:
        start_date = normalize_optional_date(data.get("start_date"))
        expiry_date = normalize_optional_date(data.get("expiry_date"))
    except ValueError as exc:
        return jsonify({"success": False, "error": str(exc)}), 400

    if not expiry_date:
        return jsonify({"success": False, "error": "到期时间必须录入。"}), 400

    if start_date and start_date > expiry_date:
        return jsonify({"success": False, "error": "起始日期不能晚于到期时间。"}), 400

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_station_certificates_table(cur)

        user = get_user_by_id(cur, user_id)
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404

        can_edit_all = can_edit_all_certificates(user)
        can_edit_own = can_edit_own_certificates(cur, user)
        if not can_edit_all and not can_edit_own:
            return (
                jsonify(
                    {"success": False, "error": "当前账号无权维护证照有效期。"}
                ),
                403,
            )

        if not can_edit_all and str(station_id) != str(user.get("station_id") or ""):
            return jsonify({"success": False, "error": "当前账号只能维护本站证照有效期。"}), 403

        cur.execute(
            """
            SELECT id
            FROM stations
            WHERE id = %s
            LIMIT 1;
            """,
            (station_id,),
        )
        if not cur.fetchone():
            return jsonify({"success": False, "error": "站点不存在。"}), 404

        type_meta = CERTIFICATE_TYPE_BY_CODE[certificate_type]
        cur.execute(
            """
            INSERT INTO station_certificates (
                station_id,
                certificate_type,
                certificate_name,
                start_date,
                expiry_date,
                remark,
                created_by,
                updated_by,
                created_at,
                updated_at
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            ON CONFLICT (station_id, certificate_type)
            DO UPDATE SET
                certificate_name = EXCLUDED.certificate_name,
                start_date = EXCLUDED.start_date,
                expiry_date = EXCLUDED.expiry_date,
                remark = EXCLUDED.remark,
                updated_by = EXCLUDED.updated_by,
                updated_at = CURRENT_TIMESTAMP
            RETURNING id;
            """,
            (
                station_id,
                certificate_type,
                type_meta["name"],
                start_date,
                expiry_date,
                remark,
                user["id"],
                user["id"],
            ),
        )
        row = cur.fetchone()
        conn.commit()
        return jsonify(
            {
                "success": True,
                "message": "证照有效期已保存。",
                "id": row["id"],
            }
        )
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/station-certificates/<int:certificate_id>", methods=["DELETE"])
def delete_station_certificate(certificate_id):
    data = request.get_json(silent=True) or {}
    user_id = str(data.get("user_id") or request.args.get("user_id", "")).strip()

    if not user_id:
        return jsonify({"success": False, "error": "缺少用户信息。"}), 400

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_station_certificates_table(cur)

        user = get_user_by_id(cur, user_id)
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404

        can_edit_all = can_edit_all_certificates(user)
        can_edit_own = can_edit_own_certificates(cur, user)
        if not can_edit_all and not can_edit_own:
            return (
                jsonify(
                    {"success": False, "error": "当前账号无权删除证照记录。"}
                ),
                403,
            )

        cur.execute(
            """
            SELECT id, station_id
            FROM station_certificates
            WHERE id = %s
            LIMIT 1;
            """,
            (certificate_id,),
        )
        certificate = cur.fetchone()
        if not certificate:
            return jsonify({"success": False, "error": "证照记录不存在。"}), 404

        if not can_edit_all and str(certificate["station_id"]) != str(user.get("station_id") or ""):
            return jsonify({"success": False, "error": "当前账号只能删除本站证照记录。"}), 403

        cur.execute(
            """
            DELETE FROM station_certificates
            WHERE id = %s
            RETURNING id;
            """,
            (certificate_id,),
        )

        conn.commit()
        return jsonify({"success": True, "message": "证照记录已删除。"})
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/station-map")
def get_station_map():
    user_id = str(request.args.get("user_id", "")).strip()
    if not user_id:
        return jsonify({"success": False, "error": "缺少用户信息。"}), 400

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()

        user = get_user_by_id(cur, user_id)
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404
        if not has_permission(cur, user, "view_station_map"):
            return jsonify({"success": False, "error": "当前账号无权查看站点地图。"}), 403

        cur.execute(
            """
            WITH issue_counts AS (
                SELECT
                    station_id,
                    COUNT(*) FILTER (WHERE TRIM(COALESCE(status, '')) IN ('待整改', '未整改', '站经无法整改')) AS pending_rectification_count,
                    COUNT(*) FILTER (WHERE TRIM(COALESCE(status, '')) = '待复核') AS pending_review_count,
                    COUNT(*) FILTER (WHERE TRIM(COALESCE(status, '')) IN ('已闭环', '已整改')) AS closed_count
                FROM issues
                WHERE COALESCE(audit_status, 'pending') <> 'rejected'
                  AND TRIM(COALESCE(status, '')) IN (
                      '待整改',
                      '未整改',
                      '站经无法整改',
                      '待复核',
                      '已闭环',
                      '已整改'
                  )
                GROUP BY station_id
            ),
            inspection_latest AS (
                SELECT
                    station_id,
                    MAX(inspection_date) AS latest_inspection_date
                FROM inspections
                GROUP BY station_id
            )
            SELECT
                s.id AS station_id,
                s.station_name,
                s.region,
                s.address,
                s.longitude,
                s.latitude,
                s.station_manager_name,
                s.station_manager_phone,
                s.station_type,
                s.asset_type,
                s.status,
                COALESCE(ic.pending_rectification_count, 0) AS pending_rectification_count,
                COALESCE(ic.pending_review_count, 0) AS pending_review_count,
                COALESCE(ic.closed_count, 0) AS closed_count,
                il.latest_inspection_date
            FROM stations s
            LEFT JOIN issue_counts ic ON ic.station_id = s.id
            LEFT JOIN inspection_latest il ON il.station_id = s.id
            WHERE s.longitude IS NOT NULL
              AND s.latitude IS NOT NULL
            ORDER BY s.id;
            """
        )
        rows = cur.fetchall()
        response = jsonify(rows)
        response.headers["Cache-Control"] = "no-store"
        return response
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/event-feed")
def get_event_feed():
    user_id = str(request.args.get("user_id", "")).strip()
    if not user_id:
        return jsonify({"success": False, "error": "缺少用户信息。"}), 400

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()

        user = get_user_by_id(cur, user_id)
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404
        if not has_permission(cur, user, "view_station_map"):
            return jsonify({"success": False, "error": "当前账号无权查看站点地图动态。"}), 403

        cur.execute(
            """
            SELECT
                CONCAT('issue-create-', i.id) AS id,
                s.id AS "stationId",
                s.station_name AS "stationName",
                CONCAT(
                    '【',
                    t.table_name,
                    '】新增问题，规范ID：',
                    i.standard_id,
                    '，当前状态：',
                    CASE
                        WHEN TRIM(COALESCE(i.status, '')) IN ('已闭环', '已整改') THEN '已闭环'
                        WHEN TRIM(COALESCE(i.status, '')) = '未整改' THEN '待整改'
                        ELSE i.status
                    END,
                    '。'
                ) AS text,
                TO_CHAR(COALESCE(i.created_at, NOW()), 'HH24:MI') AS time,
                CASE
                    WHEN TRIM(COALESCE(i.status, '')) IN ('待整改', '未整改', '站经无法整改') THEN 'danger'
                    WHEN TRIM(COALESCE(i.status, '')) = '待复核' THEN 'warning'
                    ELSE 'info'
                END AS level,
                COALESCE(i.created_at, NOW()) AS sort_time
            FROM issues i
            JOIN stations s ON s.id = i.station_id
            JOIN inspection_tables t ON t.id = i.inspection_table_id
            WHERE COALESCE(i.audit_status, 'pending') <> 'rejected'
            ORDER BY sort_time DESC
            LIMIT 5;
            """
        )
        rows = cur.fetchall()
        response = jsonify(rows)
        response.headers["Cache-Control"] = "no-store"
        return response
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/users")
def get_users():
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            """
            SELECT
                u.id,
                u.username,
                u.role,
                u.real_name,
                u.phone,
                u.station_id,
                s.station_name
            FROM users u
            LEFT JOIN stations s ON u.station_id = s.id
            ORDER BY u.id;
        """
        )
        rows = cur.fetchall()
        return jsonify(rows)
    except Exception as e:
        return (
            jsonify(
                {
                    "success": False,
                    "error": str(e),
                }
            ),
            500,
        )
    finally:
        close_db_resources(cur, conn)


@app.route("/api/training-materials", methods=["GET"])
def get_training_materials():
    user_id = str(request.args.get("user_id", "")).strip()
    if not user_id:
        return jsonify({"success": False, "error": "缺少用户信息。"}), 400

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_training_materials_table(cur)
        conn.commit()

        user = get_user_by_id(cur, user_id)
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404

        if not has_permission(cur, user, "view_training_materials"):
            return jsonify({"success": False, "error": "当前账号无权访问培训材料库。"}), 403

        cur.execute(
            """
            SELECT
                tm.id,
                tm.title,
                tm.file_type,
                tm.file_path,
                tm.original_filename,
                tm.file_size,
                tm.uploaded_by,
                TO_CHAR(tm.created_at, 'YYYY-MM-DD HH24:MI') AS created_at,
                TO_CHAR(tm.updated_at, 'YYYY-MM-DD HH24:MI') AS updated_at,
                uploader.username AS uploaded_by_username,
                uploader.real_name AS uploaded_by_real_name
            FROM training_materials tm
            LEFT JOIN users uploader ON uploader.id = tm.uploaded_by
            ORDER BY tm.updated_at DESC, tm.id DESC;
            """
        )
        rows = cur.fetchall()
        items = []
        for row in rows:
            file_path = row.get("file_path")
            items.append(
                {
                    "id": row["id"],
                    "title": row["title"],
                    "file_type": row["file_type"],
                    "file_path": file_path,
                    "file_url": f"/storage{file_path}" if file_path else "",
                    "original_filename": row["original_filename"],
                    "file_size": row["file_size"],
                    "uploaded_by": row["uploaded_by"],
                    "created_at": row["created_at"],
                    "updated_at": row["updated_at"],
                    "uploaded_by_username": row["uploaded_by_username"],
                    "uploaded_by_real_name": row["uploaded_by_real_name"],
                    "can_edit": can_edit_training_material(cur, user, row),
                    "can_delete": can_delete_training_material(cur, user, row),
                }
            )

        return jsonify(
            {
                "success": True,
                "can_upload": can_upload_training_materials(cur, user),
                "can_delete_any": can_delete_any_training_material(cur, user),
                "items": items,
            }
        )
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/training-materials", methods=["POST"])
def create_training_material():
    user_id = str(request.form.get("user_id", "")).strip()
    title = str(request.form.get("title", "")).strip()
    material_file = request.files.get("file")

    if not user_id:
        return jsonify({"success": False, "error": "缺少用户信息。"}), 400
    if not title:
        return jsonify({"success": False, "error": "请填写培训材料标题。"}), 400
    if len(title) > 120:
        return jsonify({"success": False, "error": "标题不能超过 120 个字符。"}), 400

    new_file_path = None
    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_training_materials_table(cur)

        user = get_user_by_id(cur, user_id)
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404
        if not can_upload_training_materials(cur, user):
            return jsonify({"success": False, "error": "只有督导组账号可以上传培训材料。"}), 403

        new_file_path, original_filename, file_size, file_type = save_training_material_file(material_file)
        cur.execute(
            """
            INSERT INTO training_materials (
                title,
                file_type,
                file_path,
                original_filename,
                file_size,
                uploaded_by,
                created_at,
                updated_at
            )
            VALUES (%s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            RETURNING id;
            """,
            (title, file_type, new_file_path, original_filename, file_size, user["id"]),
        )
        row = cur.fetchone()
        conn.commit()
        return jsonify(
            {
                "success": True,
                "message": "培训材料已上传。",
                "id": row["id"],
            }
        )
    except ValueError as exc:
        if conn:
            conn.rollback()
        if new_file_path:
            remove_storage_file(new_file_path)
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        if new_file_path:
            remove_storage_file(new_file_path)
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/training-materials/<int:material_id>", methods=["PUT"])
def update_training_material(material_id):
    user_id = str(request.form.get("user_id", "")).strip()
    title = str(request.form.get("title", "")).strip()
    material_file = request.files.get("file")

    if not user_id:
        return jsonify({"success": False, "error": "缺少用户信息。"}), 400
    if not title:
        return jsonify({"success": False, "error": "请填写培训材料标题。"}), 400
    if len(title) > 120:
        return jsonify({"success": False, "error": "标题不能超过 120 个字符。"}), 400

    new_file_path = None
    old_file_path = None
    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_training_materials_table(cur)

        user = get_user_by_id(cur, user_id)
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404
        if not can_upload_training_materials(cur, user):
            return jsonify({"success": False, "error": "当前账号无权维护培训材料。"}), 403

        cur.execute(
            """
            SELECT id, uploaded_by, file_path
            FROM training_materials
            WHERE id = %s
            LIMIT 1;
            """,
            (material_id,),
        )
        material = cur.fetchone()
        if not material:
            return jsonify({"success": False, "error": "培训材料不存在。"}), 404
        if not can_edit_training_material(cur, user, material):
            return jsonify({"success": False, "error": "只能编辑自己上传的培训材料。"}), 403

        old_file_path = material["file_path"]
        if material_file and material_file.filename:
            new_file_path, original_filename, file_size, file_type = save_training_material_file(material_file)
            cur.execute(
                """
                UPDATE training_materials
                SET title = %s,
                    file_type = %s,
                    file_path = %s,
                    original_filename = %s,
                    file_size = %s,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = %s;
                """,
                (title, file_type, new_file_path, original_filename, file_size, material_id),
            )
        else:
            cur.execute(
                """
                UPDATE training_materials
                SET title = %s,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = %s;
                """,
                (title, material_id),
            )

        conn.commit()
        if new_file_path and old_file_path and old_file_path != new_file_path:
            remove_storage_file(old_file_path)

        return jsonify({"success": True, "message": "培训材料已更新。"})
    except ValueError as exc:
        if conn:
            conn.rollback()
        if new_file_path:
            remove_storage_file(new_file_path)
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        if new_file_path:
            remove_storage_file(new_file_path)
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/training-materials/<int:material_id>", methods=["DELETE"])
def delete_training_material(material_id):
    data = request.get_json(silent=True) or {}
    user_id = str(data.get("user_id") or request.args.get("user_id", "")).strip()

    if not user_id:
        return jsonify({"success": False, "error": "缺少用户信息。"}), 400

    old_file_path = None
    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_training_materials_table(cur)

        user = get_user_by_id(cur, user_id)
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404
        if not can_upload_training_materials(cur, user):
            return jsonify({"success": False, "error": "当前账号无权删除培训材料。"}), 403

        cur.execute(
            """
            SELECT id, uploaded_by, file_path
            FROM training_materials
            WHERE id = %s
            LIMIT 1;
            """,
            (material_id,),
        )
        material = cur.fetchone()
        if not material:
            return jsonify({"success": False, "error": "培训材料不存在。"}), 404
        if not can_delete_training_material(cur, user, material):
            return jsonify({"success": False, "error": "只能删除自己上传的培训材料。"}), 403

        old_file_path = material["file_path"]
        cur.execute("DELETE FROM training_materials WHERE id = %s;", (material_id,))
        conn.commit()
        remove_storage_file(old_file_path)
        return jsonify({"success": True, "message": "培训材料已删除。"})
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/inspection-table-originals", methods=["GET"])
def get_inspection_table_originals():
    user_id = str(request.args.get("user_id", "")).strip()
    if not user_id:
        return jsonify({"success": False, "error": "缺少用户信息。"}), 400

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_inspection_table_original_files_table(cur)
        conn.commit()

        user = get_user_by_id(cur, user_id)
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404

        if not can_view_checklist_originals(cur, user):
            return jsonify({"success": False, "error": "当前账号无权查看检查表原件库。"}), 403

        cur.execute(
            """
            SELECT
                it.id AS inspection_table_id,
                it.table_code,
                it.table_name,
                it.description,
                it.is_active,
                doc.id AS document_id,
                doc.file_path,
                doc.original_filename,
                doc.file_size,
                TO_CHAR(doc.created_at, 'YYYY-MM-DD HH24:MI') AS created_at,
                TO_CHAR(doc.updated_at, 'YYYY-MM-DD HH24:MI') AS updated_at,
                uploader.username AS uploaded_by_username,
                uploader.real_name AS uploaded_by_real_name
            FROM inspection_tables it
            LEFT JOIN inspection_table_original_files doc
              ON doc.inspection_table_id = it.id
            LEFT JOIN users uploader
              ON uploader.id = doc.uploaded_by
            WHERE it.is_active = TRUE
            ORDER BY it.id ASC;
            """
        )
        rows = cur.fetchall()

        items = []
        for row in rows:
            file_path = row.get("file_path")
            items.append(
                {
                    "inspection_table_id": row["inspection_table_id"],
                    "table_code": row["table_code"],
                    "table_name": row["table_name"],
                    "description": row["description"],
                    "is_active": row["is_active"],
                    "document_id": row["document_id"],
                    "has_pdf": bool(file_path),
                    "file_path": file_path,
                    "file_url": f"/storage{file_path}" if file_path else "",
                    "original_filename": row["original_filename"],
                    "file_size": row["file_size"],
                    "created_at": row["created_at"],
                    "updated_at": row["updated_at"],
                    "uploaded_by_username": row["uploaded_by_username"],
                    "uploaded_by_real_name": row["uploaded_by_real_name"],
                }
            )

        return jsonify(
            {
                "success": True,
                "can_manage": can_manage_checklist_originals(cur, user),
                "items": items,
            }
        )
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route(
    "/api/inspection-table-originals/<int:inspection_table_id>/pdf",
    methods=["POST"],
)
def upload_inspection_table_original_pdf(inspection_table_id):
    user_id = str(request.form.get("user_id", "")).strip()
    if not user_id:
        return jsonify({"success": False, "error": "缺少用户信息。"}), 400

    pdf_file = request.files.get("pdf") or request.files.get("file")
    new_file_path = None
    old_file_path = None
    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_inspection_table_original_files_table(cur)

        user = get_user_by_id(cur, user_id)
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404

        if not can_manage_checklist_originals(cur, user):
            return jsonify({"success": False, "error": "当前账号无权上传检查表原件。"}), 403

        cur.execute(
            """
            SELECT id, table_name
            FROM inspection_tables
            WHERE id = %s AND is_active = TRUE
            LIMIT 1;
            """,
            (inspection_table_id,),
        )
        inspection_table = cur.fetchone()
        if not inspection_table:
            return jsonify({"success": False, "error": "检查表不存在或未启用。"}), 404

        cur.execute(
            """
            SELECT file_path
            FROM inspection_table_original_files
            WHERE inspection_table_id = %s
            LIMIT 1;
            """,
            (inspection_table_id,),
        )
        old_row = cur.fetchone()
        old_file_path = old_row["file_path"] if old_row else None

        new_file_path, original_filename, file_size = save_inspection_original_pdf(pdf_file)
        cur.execute(
            """
            INSERT INTO inspection_table_original_files (
                inspection_table_id,
                file_path,
                original_filename,
                file_size,
                uploaded_by,
                created_at,
                updated_at
            )
            VALUES (%s, %s, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            ON CONFLICT (inspection_table_id)
            DO UPDATE SET
                file_path = EXCLUDED.file_path,
                original_filename = EXCLUDED.original_filename,
                file_size = EXCLUDED.file_size,
                uploaded_by = EXCLUDED.uploaded_by,
                updated_at = CURRENT_TIMESTAMP
            RETURNING id;
            """,
            (
                inspection_table_id,
                new_file_path,
                original_filename,
                file_size,
                user["id"],
            ),
        )
        row = cur.fetchone()
        conn.commit()

        if old_file_path and old_file_path != new_file_path:
            remove_storage_file(old_file_path)

        return jsonify(
            {
                "success": True,
                "message": f"【{inspection_table['table_name']}】原件 PDF 已更新。",
                "document_id": row["id"],
                "file_path": new_file_path,
                "file_url": f"/storage{new_file_path}",
            }
        )
    except ValueError as exc:
        if conn:
            conn.rollback()
        if new_file_path:
            remove_storage_file(new_file_path)
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        if new_file_path:
            remove_storage_file(new_file_path)
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


# 新增 inspection-tables API
@app.route("/api/inspection-tables")
def get_inspection_tables():
    user_id = str(request.args.get("user_id", "")).strip()
    scope_context = str(request.args.get("scope", "")).strip()
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_inspection_checklist_management_schema(cur)
        conn.commit()

        where_clauses = ["is_active = TRUE"]
        params = []
        if user_id:
            user = get_user_by_id(cur, user_id)
            if not user:
                return jsonify({"success": False, "error": "用户不存在。"}), 404
            scope_permission_key = INSPECTION_TABLE_SCOPE_CONTEXT_MAP.get(scope_context)
            if scope_permission_key:
                if not append_inspection_table_scope_filter(
                    cur,
                    user,
                    where_clauses,
                    params,
                    "id",
                    scope_permission_key,
                ):
                    return jsonify([])

        cur.execute(
            sql.SQL(
                """
            SELECT
                id,
                table_code,
                table_name,
                checklist_mode,
                standard_id_base,
                description,
                is_active
            FROM inspection_tables
            WHERE {where_clause}
            ORDER BY id;
            """
            ).format(where_clause=sql.SQL(" AND ").join(sql.SQL(part) for part in where_clauses)),
            params,
        )
        rows = cur.fetchall()
        return jsonify(rows)
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/inspection-table-fields")
def get_inspection_table_fields():
    table_id = str(request.args.get("table_id", "")).strip()

    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_inspection_checklist_management_schema(cur)
        conn.commit()
        if not table_id:
            return jsonify([])

        inspection_table = get_inspection_table_record(cur, table_id)
        if not inspection_table:
            return jsonify({"success": False, "error": "检查表不存在。"}), 404
        if not inspection_table["is_active"]:
            return jsonify({"success": False, "error": "检查表未启用。"}), 400

        return jsonify(
            [dict(field) for field in get_management_checklist_fields(cur, table_id, include_public=True)]
        )
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/inspection-standard-export-template", methods=["GET"])
def get_inspection_standard_export_template():
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_inspection_checklist_management_schema(cur)
        conn.commit()

        user = getattr(g, "current_user", None)
        if not (can_view_inspection_standards(cur, user) or has_permission(cur, user, "submit_inspections")):
            return jsonify({"success": False, "error": "当前账号无权访问巡检规范库。"}), 403

        template = get_standard_export_template_config(cur)
        response = jsonify(
            {
                "success": True,
                "has_saved": template["has_saved"],
                "tables": template["tables"],
            }
        )
        response.headers["Cache-Control"] = "no-store"
        return response
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/inspection-standard-export-template", methods=["PUT"])
def update_inspection_standard_export_template():
    data = request.get_json(silent=True) or {}
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_inspection_checklist_management_schema(cur)

        user = getattr(g, "current_user", None)
        if not can_view_inspection_standards(cur, user):
            return jsonify({"success": False, "error": "当前账号无权访问巡检规范库。"}), 403

        normalized = save_standard_export_template_config(cur, data.get("tables") or {})
        conn.commit()
        return jsonify(
            {
                "success": True,
                "message": "导出规范公共模板已保存。",
                "has_saved": True,
                "tables": normalized,
            }
        )
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


# 新增：获取检查表规范（标准）API
@app.route("/api/inspection-table-standards")
def get_inspection_table_standards():
    table_id = str(request.args.get("table_id", "")).strip()
    keyword = str(request.args.get("keyword", "")).strip().lower()

    if not table_id:
        return jsonify({"success": False, "error": "缺少检查表信息。"}), 400

    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_inspection_checklist_management_schema(cur)
        conn.commit()

        inspection_table = get_inspection_table_record(cur, table_id)
        if not inspection_table:
            return jsonify({"success": False, "error": "检查表不存在。"}), 404

        if not inspection_table["is_active"]:
            return jsonify({"success": False, "error": "检查表未启用。"}), 400

        physical_table_name = get_physical_table_name_by_code(
            inspection_table["table_code"]
        )
        fields = [dict(field) for field in get_management_checklist_fields(cur, inspection_table["id"], include_public=True)]
        field_meta = [(field["field_key"], field["field_label"]) for field in fields]
        register_field_meta = build_register_display_field_meta(fields)
        if not physical_table_name or not checklist_physical_table_exists(cur, physical_table_name):
            return jsonify({"success": False, "error": "检查表未配置物理表映射。"}), 400

        cur.execute(
            sql.SQL("SELECT * FROM {} ORDER BY standard_id ASC;").format(
                sql.Identifier(physical_table_name)
            )
        )
        rows = cur.fetchall()

        filters = {
            key: str(value).strip().lower()
            for key, value in request.args.items()
            if key not in {"table_id", "keyword"} and str(value).strip()
        }

        result = []
        for row in rows:
            item = serialize_standard_row(field_meta, row, register_field_meta)
            haystack = "\n".join(
                [str(v).strip() for v in item.values() if v is not None]
            ).lower()
            if keyword and keyword not in haystack:
                continue

            matched = True
            for field_key, field_value in filters.items():
                raw_value = row.get(field_key)
                raw_text = "" if raw_value is None else str(raw_value).strip().lower()
                if field_value not in raw_text:
                    matched = False
                    break

            if matched:
                result.append(item)

        internal_links = fetch_internal_links_by_external_ids(
            cur,
            [item["standard_id"] for item in result],
        )
        for item in result:
            linked_internal = internal_links.get(int(item["standard_id"]))
            item["linked_internal"] = linked_internal
            item["linked_internal_standard_id"] = (
                linked_internal.get("internal_standard_id") if linked_internal else ""
            )

        return jsonify(result)
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/inspection-plan-overview")
def get_inspection_plan_overview():
    user_id = str(request.args.get("user_id", "")).strip()
    inspection_table_id = str(request.args.get("inspection_table_id", "")).strip()
    coverage_type = str(request.args.get("coverage_type", "")).strip()
    period_key = str(request.args.get("period_key", "")).strip()

    if not inspection_table_id:
        return jsonify({"success": False, "error": "缺少检查表信息。"}), 400

    if coverage_type not in {"monthly", "quarterly", "yearly"}:
        return jsonify({"success": False, "error": "覆盖要求参数不合法。"}), 400

    if not period_key:
        return jsonify({"success": False, "error": "缺少周期标识。"}), 400

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()

        user = get_user_by_id(cur, user_id) if user_id else None
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404
        if not has_permission(cur, user, "view_inspection_plans"):
            return jsonify({"success": False, "error": "当前账号无权查看巡检计划。"}), 403

        cur.execute(
            """
            SELECT
                pc.id,
                pc.inspection_table_id,
                it.table_name AS inspection_table_name,
                it.checklist_mode,
                pc.coverage_type,
                pc.period_key,
                pc.status,
                pc.remark,
                creator.username AS created_by_username,
                updater.username AS updated_by_username,
                pc.created_at,
                pc.updated_at
            FROM inspection_plan_configs pc
            JOIN inspection_tables it ON pc.inspection_table_id = it.id
            JOIN users creator ON pc.created_by = creator.id
            LEFT JOIN users updater ON pc.updated_by = updater.id
            WHERE pc.inspection_table_id = %s
              AND pc.coverage_type = %s
              AND pc.period_key = %s
            LIMIT 1;
            """,
            (inspection_table_id, coverage_type, period_key),
        )
        config_row = cur.fetchone()

        if not config_row:
            return (
                jsonify({"success": False, "error": "未找到对应的巡检计划配置。"}),
                404,
            )
        if not is_inspection_table_allowed_for_user(
            cur,
            user,
            config_row["inspection_table_id"],
            "limit_plan_inspection_table_scope",
        ):
            return jsonify({"success": False, "error": "当前账号无权查看该检查表的巡检计划。"}), 403

        ensure_inspection_completion_schema(cur)
        ensure_inspection_plan_assignment_schema(cur)

        station_where_clauses = ["psi.plan_config_id = %s"]
        station_params = [config_row["id"]]
        if not append_station_region_scope_filter(
            cur,
            user,
            station_where_clauses,
            station_params,
            "s.region",
            "limit_plan_station_region_scope",
        ):
            station_rows = []
        else:
            station_where_sql = " AND ".join(station_where_clauses)
            cur.execute(
                sql.SQL(
                    """
            SELECT
                psi.id,
                psi.station_id,
                s.station_name,
                s.region,
                s.address,
                COALESCE(s.monitoring_status, '运行中') AS monitoring_status,
                psi.is_included,
                psi.assigned_inspector_id,
                COALESCE(assigned_group.assigned_inspector_ids, '[]'::jsonb) AS assigned_inspector_ids,
                COALESCE(assigned_group.assigned_inspectors, '[]'::jsonb) AS assigned_inspectors,
                assigned_group.assigned_inspector_usernames AS assigned_inspector_username,
                assigned_group.assigned_inspector_names AS assigned_inspector_name,
                assigned_group.assigned_inspector_phones AS assigned_inspector_phone,
                TO_CHAR(psi.assigned_at, 'YYYY-MM-DD HH24:MI') AS assigned_at,
                psi.completion_status,
                psi.completed_inspection_id,
                TO_CHAR(psi.completed_at, 'YYYY-MM-DD HH24:MI') AS completed_at,
                psi.note
            FROM inspection_plan_station_items psi
            JOIN stations s ON psi.station_id = s.id
            LEFT JOIN LATERAL (
                SELECT
                    JSONB_AGG(assigned_user.id ORDER BY COALESCE(assigned_user.real_name, assigned_user.username, assigned_user.id::text)) AS assigned_inspector_ids,
                    JSONB_AGG(
                        JSONB_BUILD_OBJECT(
                            'id', assigned_user.id,
                            'username', assigned_user.username,
                            'real_name', assigned_user.real_name,
                            'phone', assigned_user.phone,
                            'display_name', COALESCE(assigned_user.real_name, assigned_user.username, assigned_user.id::text)
                        )
                        ORDER BY COALESCE(assigned_user.real_name, assigned_user.username, assigned_user.id::text)
                    ) AS assigned_inspectors,
                    STRING_AGG(assigned_user.username, '、' ORDER BY COALESCE(assigned_user.real_name, assigned_user.username, assigned_user.id::text)) AS assigned_inspector_usernames,
                    STRING_AGG(COALESCE(assigned_user.real_name, assigned_user.username, assigned_user.id::text), '、' ORDER BY COALESCE(assigned_user.real_name, assigned_user.username, assigned_user.id::text)) AS assigned_inspector_names,
                    STRING_AGG(assigned_user.phone, '、' ORDER BY COALESCE(assigned_user.real_name, assigned_user.username, assigned_user.id::text)) AS assigned_inspector_phones
                FROM (
                    SELECT DISTINCT inspector_id
                    FROM (
                        SELECT jsonb_array_elements_text(COALESCE(psi.assigned_inspector_ids, '[]'::jsonb))::integer AS inspector_id
                        UNION ALL
                        SELECT psi.assigned_inspector_id
                        WHERE psi.assigned_inspector_id IS NOT NULL
                    ) raw_assigned_ids
                ) assigned_ids
                JOIN users assigned_user ON assigned_user.id = assigned_ids.inspector_id
            ) assigned_group ON TRUE
            WHERE {station_where_sql}
            ORDER BY s.id ASC;
            """,
                ).format(station_where_sql=sql.SQL(station_where_sql)),
                station_params,
            )
            station_rows = cur.fetchall()

        included_count = sum(1 for row in station_rows if row["is_included"])
        completed_count = sum(
            1
            for row in station_rows
            if row["is_included"] and row["completion_status"] == "completed"
        )
        pending_count = sum(
            1
            for row in station_rows
            if row["is_included"] and row["completion_status"] == "pending"
        )
        completion_rate = (
            round((completed_count / included_count) * 100) if included_count > 0 else 0
        )

        return jsonify(
            {
                "success": True,
                "item": {
                    "id": config_row["id"],
                    "inspection_table_id": config_row["inspection_table_id"],
                    "inspection_table_name": config_row["inspection_table_name"],
                    "checklist_mode": normalize_checklist_mode(config_row.get("checklist_mode")),
                    "checklist_mode_label": "视频检查" if normalize_checklist_mode(config_row.get("checklist_mode")) == "online" else "现场检查",
                    "coverage_type": config_row["coverage_type"],
                    "coverage_type_label": COVERAGE_TYPE_LABELS.get(
                        config_row["coverage_type"], config_row["coverage_type"]
                    ),
                    "period_key": config_row["period_key"],
                    "status": config_row["status"],
                    "remark": config_row["remark"],
                    "created_by_username": config_row["created_by_username"],
                    "updated_by_username": config_row["updated_by_username"],
                    "included_station_count": included_count,
                    "completed_station_count": completed_count,
                    "pending_station_count": pending_count,
                    "completion_rate": completion_rate,
                    "created_at": config_row["created_at"],
                    "updated_at": config_row["updated_at"],
                    "stations": station_rows,
                },
            }
        )
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/external-standards", methods=["GET"])
def get_external_standards():
    keyword = str(request.args.get("keyword", "")).strip().lower()
    table_id = str(request.args.get("table_id", "")).strip()
    conn = None
    cur = None
    try:
        user = get_current_request_user()
        conn = get_db_connection()
        cur = conn.cursor()
        if not (
            can_view_checklist_originals(cur, user)
            or can_view_inspection_standards(cur, user)
            or has_permission(cur, user, "submit_inspections")
            or has_permission(cur, user, "manage_internal_standards")
        ):
            return jsonify({"success": False, "error": "当前账号无权查看规范库。"}), 403

        external_map = fetch_external_standard_map(cur)
        conn.commit()
        internal_links = fetch_internal_links_by_external_ids(cur, external_map.keys())
        items = []
        for item in external_map.values():
            if table_id and str(item.get("inspection_table_id")) != table_id:
                continue
            haystack = "\n".join(
                [
                    str(item.get("external_standard_id") or ""),
                    str(item.get("inspection_table_name") or ""),
                    str(item.get("standard_detail_text") or ""),
                ]
            ).lower()
            if keyword and keyword not in haystack:
                continue
            linked_internal = internal_links.get(int(item["external_standard_id"]))
            items.append(
                {
                    **item,
                    "linked_internal": linked_internal,
                    "linked_internal_standard_id": linked_internal.get("internal_standard_id") if linked_internal else "",
                }
            )
        return jsonify({"success": True, "items": items})
    except PermissionError as exc:
        return jsonify({"success": False, "error": str(exc)}), 401
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/inspection-internal-standards", methods=["GET"])
def get_inspection_internal_standards():
    keyword = str(request.args.get("keyword", "")).strip().lower()
    conn = None
    cur = None
    try:
        user = get_current_request_user()
        conn = get_db_connection()
        cur = conn.cursor()
        if not can_view_inspection_standards(cur, user):
            return jsonify({"success": False, "error": "当前账号无权访问巡检规范库。"}), 403

        cur.execute(
            """
            SELECT
                id,
                internal_standard_id,
                path_values,
                field_values,
                content,
                notes,
                is_active,
                TO_CHAR(created_at, 'YYYY-MM-DD HH24:MI') AS created_at,
                TO_CHAR(updated_at, 'YYYY-MM-DD HH24:MI') AS updated_at
            FROM inspection_internal_standards
            WHERE is_active = TRUE
            ORDER BY internal_standard_id ASC;
            """
        )
        rows = cur.fetchall()
        fields = [dict(field) for field in get_internal_standard_fields(cur)]
        conn.commit()
        link_map = fetch_internal_standard_links(cur, [row["id"] for row in rows])
        tag_map = fetch_internal_standard_custom_tags(cur, [row["id"] for row in rows])
        tag_groups = get_internal_standard_tag_groups(cur)
        items = []
        for row in rows:
            item = serialize_internal_standard(
                row,
                link_map.get(row["id"], []),
                fields,
                tag_map.get(row["id"], []),
            )
            haystack = "\n".join(
                [
                    item["internal_standard_id"],
                    " ".join(str(value or "") for value in item["field_values"].values()),
                    item["content"],
                    " ".join(str(tag.get("tag_name") or "") for tag in item["tags"]),
                    " ".join(str(link.get("external_standard_id") or "") for link in item["linked_externals"]),
                ]
            ).lower()
            if keyword and keyword not in haystack:
                continue
            items.append(item)
        return jsonify({"success": True, "fields": fields, "tag_groups": tag_groups, "items": items})
    except PermissionError as exc:
        return jsonify({"success": False, "error": str(exc)}), 401
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/inspection-standard-usage-mode", methods=["GET"])
def get_public_inspection_standard_usage_mode():
    conn = None
    cur = None
    try:
        user = get_current_request_user()
        conn = get_db_connection()
        cur = conn.cursor()
        if not has_permission(cur, user, "submit_inspections"):
            return jsonify({"success": False, "error": "当前账号无权使用巡检登记功能。"}), 403
        return jsonify(
            {
                "success": True,
                "usage_mode": get_inspection_standard_usage_mode(cur),
            }
        )
    except PermissionError as exc:
        return jsonify({"success": False, "error": str(exc)}), 401
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/internal-standards", methods=["GET"])
def get_management_internal_standards():
    user_id = str(request.args.get("user_id", "")).strip()
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        require_management_user(cur, user_id, "manage_internal_standards")
        cur.execute(
            """
            SELECT
                id,
                internal_standard_id,
                path_values,
                field_values,
                content,
                notes,
                is_active,
                TO_CHAR(created_at, 'YYYY-MM-DD HH24:MI') AS created_at,
                TO_CHAR(updated_at, 'YYYY-MM-DD HH24:MI') AS updated_at
            FROM inspection_internal_standards
            ORDER BY internal_standard_id ASC;
            """
        )
        rows = cur.fetchall()
        fields = [dict(field) for field in get_internal_standard_fields(cur)]
        conn.commit()
        link_map = fetch_internal_standard_links(cur, [row["id"] for row in rows])
        tag_map = fetch_internal_standard_custom_tags(cur, [row["id"] for row in rows])
        tag_groups = get_internal_standard_tag_groups(cur)
        usage_mode = get_inspection_standard_usage_mode(cur)
        return jsonify(
            {
                "success": True,
                "usage_mode": usage_mode,
                "fields": fields,
                "tag_groups": tag_groups,
                "items": [
                    serialize_internal_standard(
                        row,
                        link_map.get(row["id"], []),
                        fields,
                        tag_map.get(row["id"], []),
                    )
                    for row in rows
                ],
            }
        )
    except PermissionError as exc:
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        return jsonify({"success": False, "error": str(exc)}), 404
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/internal-standards/usage-mode", methods=["GET"])
def get_management_internal_standard_usage_mode():
    user_id = str(request.args.get("user_id", "")).strip()
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        require_management_user(cur, user_id, "manage_internal_standards")
        return jsonify(
            {
                "success": True,
                "usage_mode": get_inspection_standard_usage_mode(cur),
            }
        )
    except PermissionError as exc:
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        return jsonify({"success": False, "error": str(exc)}), 404
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/internal-standards/usage-mode", methods=["PUT"])
def update_management_internal_standard_usage_mode():
    data = request.get_json(silent=True) or {}
    user_id = str(data.get("user_id", "")).strip()
    mode = str(data.get("mode", "")).strip()
    conn = None
    cur = None
    try:
        if mode not in INSPECTION_STANDARD_USAGE_MODES:
            return jsonify({"success": False, "error": "巡检登记规范来源参数不合法。"}), 400
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_inspection_standard_usage_settings_schema(cur)
        conn.commit()
        require_management_user(cur, user_id, "manage_internal_standards")
        usage_mode = save_inspection_standard_usage_mode(cur, mode, user_id)
        conn.commit()
        return jsonify(
            {
                "success": True,
                "message": f"巡检登记已切换为使用{usage_mode['mode_label']}。",
                "usage_mode": usage_mode,
            }
        )
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 404
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/internal-standards/fields", methods=["GET"])
def get_management_internal_standard_fields():
    user_id = str(request.args.get("user_id", "")).strip()
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        require_management_user(cur, user_id, "manage_internal_standards")
        return jsonify(
            {
                "success": True,
                "fields": [dict(field) for field in get_internal_standard_fields(cur)],
            }
        )
    except PermissionError as exc:
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        return jsonify({"success": False, "error": str(exc)}), 404
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/internal-standards/fields", methods=["PUT"])
def update_management_internal_standard_fields():
    data = request.get_json(silent=True) or {}
    user_id = str(data.get("user_id", "")).strip()
    conn = None
    cur = None
    try:
        fields = normalize_internal_standard_field_rows(data.get("fields") or [], allow_empty=True)
        conn = get_db_connection()
        cur = conn.cursor()
        require_management_user(cur, user_id, "manage_internal_standards")
        upsert_internal_standard_fields(cur, fields)
        conn.commit()
        return jsonify(
            {
                "success": True,
                "message": "内部规范字段配置已保存。",
                "fields": [dict(field) for field in get_internal_standard_fields(cur)],
            }
        )
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 404
    except ValueError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        if getattr(e, "pgcode", "") == "23505":
            return jsonify({"success": False, "error": "内部规范字段名称或系统标识已存在。"}), 400
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/internal-standards/tag-groups", methods=["GET"])
def get_management_internal_standard_tag_groups():
    user_id = str(request.args.get("user_id", "")).strip()
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        require_management_user(cur, user_id, "manage_internal_standards")
        return jsonify(
            {
                "success": True,
                "tag_groups": get_internal_standard_tag_groups(cur),
            }
        )
    except PermissionError as exc:
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        return jsonify({"success": False, "error": str(exc)}), 404
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/internal-standards/tag-groups", methods=["PUT"])
def update_management_internal_standard_tag_groups():
    data = request.get_json(silent=True) or {}
    user_id = str(data.get("user_id", "")).strip()
    conn = None
    cur = None
    try:
        groups = normalize_internal_tag_group_rows(data.get("tag_groups") or [])
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_internal_standard_schema(cur)
        conn.commit()
        require_management_user(cur, user_id, "manage_internal_standards")
        upsert_internal_standard_tag_groups(cur, groups)
        conn.commit()
        return jsonify(
            {
                "success": True,
                "message": "内部规范标签群组已保存。",
                "tag_groups": get_internal_standard_tag_groups(cur),
            }
        )
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 404
    except ValueError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        if getattr(e, "pgcode", "") == "23505":
            return jsonify({"success": False, "error": "标签群组或标签名称已存在。"}), 400
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/internal-standards", methods=["POST"])
def create_management_internal_standard():
    data = request.get_json(silent=True) or {}
    user_id = str(data.get("user_id", "")).strip()
    conn = None
    cur = None
    try:
        links = normalize_external_link_rows(data.get("external_links") or [])
        tag_ids = normalize_internal_custom_tag_ids(data.get("tag_ids") or data.get("custom_tag_ids") or [])
        content = normalize_text(data.get("content"), 3000)
        if not content:
            return jsonify({"success": False, "error": "请填写内部规范内容。"}), 400
        if not links:
            return jsonify({"success": False, "error": "内部规范必须至少绑定一个外部规范ID。"}), 400

        conn = get_db_connection()
        cur = conn.cursor()
        ensure_inspection_checklist_management_schema(cur)
        ensure_internal_standard_schema(cur)
        conn.commit()
        require_management_user(cur, user_id, "manage_internal_standards")
        path_values = [content]
        field_values = {}
        internal_standard_id = generate_internal_standard_id(cur, content)
        cur.execute(
            """
            INSERT INTO inspection_internal_standards (
                internal_standard_id,
                path_values,
                field_values,
                content,
                notes,
                is_active,
                created_at,
                updated_at
            )
            VALUES (%s, %s::jsonb, %s::jsonb, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            RETURNING id;
            """,
            (
                internal_standard_id,
                json.dumps(path_values, ensure_ascii=False),
                json.dumps(field_values, ensure_ascii=False),
                content,
                "",
                bool(data.get("is_active", True)),
            ),
        )
        row = cur.fetchone()
        replace_internal_standard_links(cur, row["id"], links)
        replace_internal_standard_custom_tags(cur, row["id"], tag_ids)
        conn.commit()
        return jsonify(
            {
                "success": True,
                "message": f"内部规范已创建，系统生成内部规范ID {internal_standard_id}。",
                "id": row["id"],
                "internal_standard_id": internal_standard_id,
            }
        )
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except (LookupError, ValueError) as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/internal-standards/<int:standard_id>", methods=["PUT"])
def update_management_internal_standard(standard_id):
    data = request.get_json(silent=True) or {}
    user_id = str(data.get("user_id", "")).strip()
    conn = None
    cur = None
    try:
        links = normalize_external_link_rows(data.get("external_links") or [])
        tag_ids = normalize_internal_custom_tag_ids(data.get("tag_ids") or data.get("custom_tag_ids") or [])
        content = normalize_text(data.get("content"), 3000)
        if not content:
            return jsonify({"success": False, "error": "请填写内部规范内容。"}), 400
        if not links:
            return jsonify({"success": False, "error": "内部规范必须至少绑定一个外部规范ID。"}), 400

        conn = get_db_connection()
        cur = conn.cursor()
        ensure_inspection_checklist_management_schema(cur)
        ensure_internal_standard_schema(cur)
        conn.commit()
        require_management_user(cur, user_id, "manage_internal_standards")
        path_values = [content]
        field_values = {}
        cur.execute(
            "SELECT id, internal_standard_id FROM inspection_internal_standards WHERE id = %s LIMIT 1;",
            (standard_id,),
        )
        existing_standard = cur.fetchone()
        if not existing_standard:
            return jsonify({"success": False, "error": "内部规范不存在。"}), 404
        cur.execute(
            """
            UPDATE inspection_internal_standards
            SET path_values = %s::jsonb,
                field_values = %s::jsonb,
                content = %s,
                notes = %s,
                is_active = %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = %s;
            """,
            (
                json.dumps(path_values, ensure_ascii=False),
                json.dumps(field_values, ensure_ascii=False),
                content,
                "",
                bool(data.get("is_active", True)),
                standard_id,
            ),
        )
        replace_internal_standard_links(cur, standard_id, links)
        replace_internal_standard_custom_tags(cur, standard_id, tag_ids)
        sync_referenced_internal_standard_detail_text(
            cur,
            existing_standard["internal_standard_id"],
            content,
        )
        conn.commit()
        return jsonify({"success": True, "message": "内部规范已保存。"})
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except (LookupError, ValueError) as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/internal-standards/<int:standard_id>", methods=["DELETE"])
def delete_management_internal_standard(standard_id):
    user_id = str(request.args.get("user_id", "") or (request.get_json(silent=True) or {}).get("user_id", "")).strip()
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_internal_standard_schema(cur)
        conn.commit()
        require_management_user(cur, user_id, "manage_internal_standards")
        cur.execute(
            "DELETE FROM inspection_internal_standards WHERE id = %s RETURNING internal_standard_id;",
            (standard_id,),
        )
        row = cur.fetchone()
        if not row:
            return jsonify({"success": False, "error": "内部规范不存在。"}), 404
        conn.commit()
        return jsonify({"success": True, "message": f"内部规范 {row['internal_standard_id']} 已删除。"})
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/internal-standards/export", methods=["GET"])
def export_management_internal_standards():
    user_id = str(request.args.get("user_id", "")).strip()
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_inspection_checklist_management_schema(cur)
        ensure_internal_standard_schema(cur)
        conn.commit()
        require_management_user(cur, user_id, "manage_internal_standards")
        cur.execute(
            """
            SELECT
                id,
                internal_standard_id,
                path_values,
                field_values,
                content,
                notes,
                is_active,
                TO_CHAR(created_at, 'YYYY-MM-DD HH24:MI') AS created_at,
                TO_CHAR(updated_at, 'YYYY-MM-DD HH24:MI') AS updated_at
            FROM inspection_internal_standards
            ORDER BY internal_standard_id ASC;
            """
        )
        rows = cur.fetchall()
        fields = [dict(field) for field in get_internal_standard_fields(cur)]
        link_map = fetch_internal_standard_links(cur, [row["id"] for row in rows])
        tag_map = fetch_internal_standard_custom_tags(cur, [row["id"] for row in rows])
        tag_groups = get_internal_standard_tag_groups(cur, include_system=False)
        standards = []
        for row in rows:
            item = serialize_internal_standard(
                row,
                link_map.get(row["id"], []),
                fields,
                tag_map.get(row["id"], []),
            )
            standards.append(
                {
                    "internal_standard_id": item["internal_standard_id"],
                    "content": item["content"],
                    "field_values": item["field_values"],
                    "is_active": item["is_active"],
                    "custom_tags": [
                        {
                            "group_name": tag.get("group_name"),
                            "tag_name": tag.get("tag_name"),
                        }
                        for tag in item["custom_tags"]
                    ],
                    "external_links": [
                        {
                            "external_standard_id": link["external_standard_id"],
                            "external_inspection_table_id": link.get("external_inspection_table_id"),
                        }
                        for link in item["linked_externals"]
                    ],
                }
            )

        now = beijing_now()
        response = jsonify(
            {
                "backup_type": "ywddzx_internal_standards",
                "version": 3,
                "exported_at": now.isoformat(),
                "fields": fields,
                "tag_groups": tag_groups,
                "standards": standards,
            }
        )
        filename = f"ywddzx_internal_standards_backup_{now.strftime('%Y%m%d_%H%M%S')}.json"
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        response.headers["Cache-Control"] = "no-store"
        return response
    except PermissionError as exc:
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        return jsonify({"success": False, "error": str(exc)}), 404
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/internal-standards/import", methods=["POST"])
def import_management_internal_standards():
    user_id = str(request.form.get("user_id", "")).strip()
    file_storage = request.files.get("file")
    conn = None
    cur = None
    try:
        backup_data = parse_internal_standards_backup_file(file_storage)
        fields = backup_data["fields"]
        tag_groups = backup_data.get("tag_groups", [])
        standards = backup_data["standards"]
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_inspection_checklist_management_schema(cur)
        ensure_internal_standard_schema(cur)
        conn.commit()
        require_management_user(cur, user_id, "manage_internal_standards")

        external_ids = [
            link["external_standard_id"]
            for item in standards
            for link in item["external_links"]
        ]
        external_map = fetch_external_standard_map(cur, external_ids)
        for external_id in external_ids:
            if int(external_id) not in external_map:
                raise ValueError(f"外部规范ID【{external_id}】不存在，无法导入。")

        cur.execute("DELETE FROM inspection_internal_standards;")
        upsert_internal_standard_fields(cur, fields)
        upsert_internal_standard_tag_groups(cur, tag_groups)
        restored_tag_groups = get_internal_standard_tag_groups(cur, include_system=False)
        restored_tag_id_map = {}
        for group in restored_tag_groups:
            group_key = normalize_internal_tag_key(group.get("group_name"))
            for tag in group.get("tags") or []:
                restored_tag_id_map[(group_key, normalize_internal_tag_key(tag.get("tag_name")))] = tag.get("id")
        for item in standards:
            cur.execute(
                """
                INSERT INTO inspection_internal_standards (
                    internal_standard_id,
                    path_values,
                    field_values,
                    content,
                    notes,
                    is_active,
                    created_at,
                    updated_at
                )
                VALUES (%s, %s::jsonb, %s::jsonb, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                RETURNING id;
                """,
                (
                    item["internal_standard_id"],
                    json.dumps(item["path_values"], ensure_ascii=False),
                    json.dumps(item["field_values"], ensure_ascii=False),
                    item["content"],
                    "",
                    item["is_active"],
                ),
            )
            row = cur.fetchone()
            replace_internal_standard_links(cur, row["id"], item["external_links"])
            restored_tag_ids = [
                restored_tag_id_map.get((
                    normalize_internal_tag_key(tag_ref.get("group_name")),
                    normalize_internal_tag_key(tag_ref.get("tag_name")),
                ))
                for tag_ref in item.get("custom_tag_refs", [])
            ]
            replace_internal_standard_custom_tags(
                cur,
                row["id"],
                [tag_id for tag_id in restored_tag_ids if tag_id],
            )
            sync_referenced_internal_standard_detail_text(
                cur,
                item["internal_standard_id"],
                item["content"],
            )

        conn.commit()
        return jsonify(
            {
                "success": True,
                "message": f"巡检规范库备份已导入，共恢复 {len(standards)} 条内部规范。",
            }
        )
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 404
    except ValueError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/inspection-standards/ai-recommend", methods=["POST"])
def recommend_inspection_standard_by_ai():
    data = request.get_json(silent=True) or {}
    description = str(data.get("description") or "").strip()

    if len(description) < 4:
        return jsonify({"success": False, "error": "请先填写更具体的实际问题描述。"}), 400

    conn = None
    cur = None
    try:
        current_user = get_current_request_user()
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_inspection_checklist_management_schema(cur)
        ensure_internal_standard_schema(cur)
        conn.commit()

        if not has_permission(cur, current_user, "submit_inspections"):
            return jsonify({"success": False, "error": "当前账号无权使用巡检登记功能。"}), 403

        usage_mode = get_inspection_standard_usage_mode(cur)
        ai_catalog, full_standards = build_inspection_standard_ai_catalog(cur, usage_mode["mode"])
        conn.commit()

        recommendation_result = generate_standard_recommendations(
            description,
            ai_catalog,
        )
        record_ai_usage_log(
            cur,
            current_user,
            recommendation_result,
            "巡检登记",
            "AI引用规范",
            description[:200],
        )
        conn.commit()
        standard_map = {
            str(item.get("standard_id")): item
            for item in full_standards
            if item.get("standard_id") is not None
        }
        items = []
        for recommendation in recommendation_result.get("recommendations") or []:
            standard = standard_map.get(str(recommendation.get("standard_id") or ""))
            if not standard:
                continue
            items.append(
                {
                    **standard,
                    "confidence": recommendation.get("confidence") or "中",
                    "reason": recommendation.get("reason") or "",
                }
            )

        no_related = bool(recommendation_result.get("no_related")) or not items
        return jsonify(
            {
                "success": True,
                "ai_generated": bool(recommendation_result.get("generated")),
                "message": recommendation_result.get("message") or "",
                "summary": recommendation_result.get("summary") or "",
                "no_related": no_related,
                "items": [] if no_related else items,
                "catalog_count": len(ai_catalog),
                "usage_mode": usage_mode,
            }
        )
    except PermissionError as e:
        return jsonify({"success": False, "error": str(e)}), 401
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/inspection-register", methods=["POST"])
def inspection_register():
    inspector_id = get_authenticated_request_user_id(request.form.get("inspector_id"))
    station_id = str(request.form.get("station_id", "")).strip()
    inspection_table_id = str(request.form.get("inspection_table_id", "")).strip()
    has_issue = str(request.form.get("has_issue", "yes")).strip().lower()
    standard_id = str(request.form.get("standard_id", "")).strip()
    internal_standard_id = str(request.form.get("internal_standard_id", "")).strip().upper()
    description = str(request.form.get("description", "")).strip()
    photo = request.files.get("photo")

    if not inspector_id:
        return jsonify({"success": False, "error": "缺少巡检人信息。"}), 400

    if not station_id:
        return jsonify({"success": False, "error": "请选择站点名称。"}), 400

    if has_issue != "yes" and not inspection_table_id:
        return jsonify({"success": False, "error": "请选择检查表。"}), 400

    if has_issue not in {"yes", "no"}:
        return jsonify({"success": False, "error": "是否发现问题参数不合法。"}), 400

    if has_issue == "yes" and not standard_id and not internal_standard_id:
        return jsonify({"success": False, "error": "请选择规范。"}), 400

    if has_issue == "yes" and standard_id and not internal_standard_id and not inspection_table_id:
        return jsonify({"success": False, "error": "请选择检查表。"}), 400

    if has_issue == "yes" and not description:
        return jsonify({"success": False, "error": "请填写实际问题描述。"}), 400

    if has_issue == "yes" and (not photo or not photo.filename):
        return jsonify({"success": False, "error": "请上传问题照片。"}), 400

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_inspection_checklist_management_schema(cur)
        ensure_internal_standard_schema(cur)
        ensure_issue_inspector_schema(cur)
        ensure_inspection_completion_schema(cur)
        auto_complete_overdue_inspections(cur)
        conn.commit()

        cur.execute(
            """
            SELECT id, username, role, real_name
            FROM users
            WHERE id = %s
            LIMIT 1;
            """,
            (inspector_id,),
        )
        inspector = cur.fetchone()

        if not inspector:
            return jsonify({"success": False, "error": "巡检人不存在。"}), 404

        if not has_permission(cur, inspector, "submit_inspections"):
            return (
                jsonify(
                    {"success": False, "error": "只有督导组账号可以提交巡检登记。"}
                ),
                403,
            )

        cur.execute(
            """
            SELECT id, station_name
            FROM stations
            WHERE id = %s
            LIMIT 1;
            """,
            (station_id,),
        )
        station = cur.fetchone()

        if not station:
            return jsonify({"success": False, "error": "站点不存在。"}), 404

        today = beijing_today()
        batch_id = get_or_create_inspection_batch(cur, station_id, inspector_id, today)
        usage_mode = get_inspection_standard_usage_mode(cur)

        if has_issue == "yes" and usage_mode["mode"] == "external" and internal_standard_id:
            return jsonify({"success": False, "error": "当前巡检登记已切换为外部规范库，请选择外部规范ID后提交。"}), 400

        if has_issue == "yes" and usage_mode["mode"] == "internal" and standard_id and not internal_standard_id:
            return jsonify({"success": False, "error": "当前巡检登记已切换为内部规范库，请选择内部规范ID后提交。"}), 400

        if has_issue == "yes" and internal_standard_id:
            internal_standard, _internal_fields, linked_externals = fetch_internal_standard_by_code(
                cur,
                internal_standard_id,
            )
            if not internal_standard:
                return jsonify({"success": False, "error": "所选内部规范不存在或未启用。"}), 404
            if not linked_externals:
                return jsonify({"success": False, "error": "所选内部规范尚未挂载外部规范，不能登记问题。"}), 400

            external_map = fetch_external_standard_map(
                cur,
                [link["external_standard_id"] for link in linked_externals],
            )
            external_standards = []
            for link in linked_externals:
                external = external_map.get(int(link["external_standard_id"]))
                if not external:
                    return jsonify(
                        {
                            "success": False,
                            "error": f"内部规范挂载的外部规范ID【{link['external_standard_id']}】不存在。",
                        }
                    ), 400
                external_standards.append(external)

            targets = prepare_issue_registration_targets(
                cur,
                station_id,
                inspector_id,
                external_standards,
                batch_id,
                today,
            )
            photo_path = save_uploaded_file(photo, "issues")
            inspection_id_by_table = {}
            created_issue_ids = []
            internal_detail_text = internal_standard.get("content") or ""

            for target in targets:
                table_key = str(target["inspection_table_id"])
                inspection_id = target["inspection_id"] or inspection_id_by_table.get(table_key)
                if not inspection_id:
                    inspection_id = create_inspection_record(
                        cur,
                        station_id,
                        inspector_id,
                        target["inspection_table_id"],
                        batch_id,
                    )
                    inspection_id_by_table[table_key] = inspection_id

                cur.execute(
                    """
                    INSERT INTO issues (
                        inspection_id,
                        inspector_id,
                        station_id,
                        inspection_table_id,
                        standard_id,
                        standard_detail_text,
                        internal_standard_id,
                        internal_standard_detail_text,
                        description,
                        photo_path,
                        status
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING id;
                    """,
                    (
                        inspection_id,
                        inspector_id,
                        station_id,
                        target["inspection_table_id"],
                        target["standard_id"],
                        target["standard_detail_text"],
                        internal_standard["internal_standard_id"],
                        internal_detail_text,
                        description,
                        photo_path,
                        "待整改",
                    ),
                )
                created_issue_ids.append(cur.fetchone()["id"])
                mark_related_plan_items_completed(
                    cur,
                    station_id,
                    target["inspection_table_id"],
                    inspection_id,
                    today,
                )

            conn.commit()
            return jsonify(
                {
                    "success": True,
                    "message": f"巡检问题登记成功，已按内部规范生成 {len(created_issue_ids)} 条外部规范问题。",
                    "inspection_id": next(iter(inspection_id_by_table.values()), targets[0]["inspection_id"]),
                    "issue_id": created_issue_ids[0] if created_issue_ids else None,
                    "issue_ids": created_issue_ids,
                }
            )

        inspection_table = get_inspection_table_record(cur, inspection_table_id)
        if not inspection_table:
            return jsonify({"success": False, "error": "检查表不存在。"}), 404

        if not inspection_table["is_active"]:
            return jsonify({"success": False, "error": "检查表未启用。"}), 400

        physical_table_name = get_physical_table_name_by_code(
            inspection_table["table_code"]
        )
        fields = [dict(field) for field in get_management_checklist_fields(cur, inspection_table["id"], include_public=True)]
        field_meta = [(field["field_key"], field["field_label"]) for field in fields]
        if not physical_table_name or not checklist_physical_table_exists(cur, physical_table_name):
            return jsonify({"success": False, "error": "检查表未配置物理表映射。"}), 400
        ensure_checklist_field_columns(cur, physical_table_name, fields)

        today = beijing_today()

        inspection_id = get_or_create_period_inspection(
            cur,
            station_id,
            inspector_id,
            inspection_table_id,
            today,
        )

        if has_issue == "no":
            mark_related_plan_items_completed(
                cur,
                station_id,
                inspection_table_id,
                inspection_id,
                today,
            )
            conn.commit()
            return jsonify(
                {
                    "success": True,
                    "message": "巡检记录提交成功，未发现问题。",
                    "inspection_id": inspection_id,
                    "issue_id": None,
                }
            )

        standard = fetch_standard_from_table(
            cur,
            physical_table_name,
            standard_id,
        )
        if not standard:
            return jsonify({"success": False, "error": "所选规范不存在。"}), 404

        standard_detail_text = build_standard_detail_text(
            field_meta,
            standard,
        )
        if not standard_detail_text:
            return jsonify({"success": False, "error": "规范详情生成失败。"}), 400

        linked_internal = fetch_internal_links_by_external_ids(cur, [standard_id]).get(int(standard_id))
        linked_internal_standard_id = linked_internal.get("internal_standard_id") if linked_internal else None
        linked_internal_detail_text = linked_internal.get("content") if linked_internal else None
        photo_path = save_uploaded_file(photo, "issues")

        cur.execute(
            """
            INSERT INTO issues (
                inspection_id,
                inspector_id,
                station_id,
                inspection_table_id,
                standard_id,
                standard_detail_text,
                internal_standard_id,
                internal_standard_detail_text,
                description,
                photo_path,
                status
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id;
            """,
            (
                inspection_id,
                inspector_id,
                station_id,
                inspection_table_id,
                standard_id,
                standard_detail_text,
                linked_internal_standard_id,
                linked_internal_detail_text,
                description,
                photo_path,
                "待整改",
            ),
        )
        issue = cur.fetchone()

        mark_related_plan_items_completed(
            cur,
            station_id,
            inspection_table_id,
            inspection_id,
            today,
        )
        conn.commit()
        return jsonify(
            {
                "success": True,
                "message": "巡检问题登记成功。",
                "inspection_id": inspection_id,
                "issue_id": issue["id"],
            }
        )
    except ValueError as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


# ===== 新增：我的待整改/待复核问题相关API =====


@app.route("/api/my-issues")
def get_my_issues():
    user_id = str(request.args.get("user_id", "")).strip()

    if not user_id:
        return jsonify({"success": False, "error": "缺少用户信息。"}), 400

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_issue_inspector_schema(cur)
        ensure_inspection_completion_schema(cur)
        sync_signed_inspections_completion(cur)
        auto_complete_overdue_inspections(cur)
        conn.commit()

        cur.execute(
            """
            SELECT id, role, station_id
            FROM users
            WHERE id = %s
            LIMIT 1;
            """,
            (user_id,),
        )
        user = cur.fetchone()

        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404

        if is_station_manager(user):
            if not user["station_id"]:
                return jsonify([])

            cur.execute(
                """
                SELECT
                    i.id,
                    i.station_id,
                    COALESCE(i.inspector_id, ins.inspector_id) AS inspector_id,
                    TO_CHAR(i.created_at, 'YYYY-MM') AS month,
                    TO_CHAR(i.created_at, 'YYYY-MM-DD HH24:MI') AS time,
                    s.region,
                    s.station_name AS station,
                    t.table_name AS inspection_table_name,
                    i.standard_id,
                    i.standard_detail_text,
                    i.internal_standard_id,
                    i.internal_standard_detail_text,
                    i.description,
                    i.photo_path AS issue_photo,
                    i.rectification_result,
                    i.rectification_note,
                    TO_CHAR(i.rectification_at, 'YYYY-MM-DD HH24:MI') AS rectification_at,
                    i.rectification_photo_path AS rectification_photo,
                    i.review_result,
                    i.review_note,
                    TO_CHAR(i.review_at, 'YYYY-MM-DD HH24:MI') AS review_at,
                    i.review_photo_path AS review_photo,
                    i.status,
                    COALESCE(i.audit_status, 'pending') AS audit_status,
                    COALESCE(i.is_excellent, FALSE) AS is_excellent,
                    i.audited_by,
                    TO_CHAR(i.audited_at, 'YYYY-MM-DD HH24:MI') AS audited_at,
                    ins.sign_status AS inspection_sign_status,
                    ins.station_manager_signed_name,
                    ins.station_manager_signature_path,
                    TO_CHAR(ins.station_manager_signed_at, 'YYYY-MM-DD HH24:MI') AS station_manager_signed_at,
                    ins.inspector_completion_status AS inspection_completion_status
                FROM issues i
                JOIN inspections ins ON i.inspection_id = ins.id
                JOIN stations s ON i.station_id = s.id
                JOIN inspection_tables t ON i.inspection_table_id = t.id
                WHERE i.station_id = %s
                  AND i.status = '待整改'
                  AND ins.sign_status = '已签名确认'
                  AND COALESCE(i.audit_status, 'pending') = 'approved'
                ORDER BY i.id DESC;
                """,
                (user["station_id"],),
            )
            rows = cur.fetchall()
            can_explicit_edit = can_edit_inspection_issues(cur, user)
            can_explicit_delete = can_delete_inspection_issues(cur, user)
            can_explicit_change_inspector = can_change_issue_inspector(cur, user)
            return jsonify(
                [
                    normalize_issue_row_for_response(
                        row,
                        user,
                        can_explicit_edit,
                        can_explicit_delete,
                        False,
                        can_explicit_change_inspector,
                    )
                    for row in rows
                ]
            )

        if is_root_user(user) or user.get("role") == "supervisor":
            cur.execute(
                """
                SELECT
                    i.id,
                    i.station_id,
                    COALESCE(i.inspector_id, ins.inspector_id) AS inspector_id,
                    TO_CHAR(i.created_at, 'YYYY-MM') AS month,
                    TO_CHAR(i.created_at, 'YYYY-MM-DD HH24:MI') AS time,
                    s.region,
                    s.station_name AS station,
                    t.table_name AS inspection_table_name,
                    i.standard_id,
                    i.standard_detail_text,
                    i.internal_standard_id,
                    i.internal_standard_detail_text,
                    i.description,
                    i.photo_path AS issue_photo,
                    i.rectification_result,
                    i.rectification_note,
                    TO_CHAR(i.rectification_at, 'YYYY-MM-DD HH24:MI') AS rectification_at,
                    i.rectification_photo_path AS rectification_photo,
                    i.review_result,
                    i.review_note,
                    TO_CHAR(i.review_at, 'YYYY-MM-DD HH24:MI') AS review_at,
                    i.review_photo_path AS review_photo,
                    i.status,
                    COALESCE(i.audit_status, 'pending') AS audit_status,
                    COALESCE(i.is_excellent, FALSE) AS is_excellent,
                    i.audited_by,
                    TO_CHAR(i.audited_at, 'YYYY-MM-DD HH24:MI') AS audited_at,
                    ins.sign_status AS inspection_sign_status,
                    ins.inspector_completion_status AS inspection_completion_status
                FROM issues i
                JOIN inspections ins ON i.inspection_id = ins.id
                JOIN stations s ON i.station_id = s.id
                JOIN inspection_tables t ON i.inspection_table_id = t.id
                WHERE i.status = '待复核'
                  AND COALESCE(i.audit_status, 'pending') <> 'rejected'
                ORDER BY i.id DESC;
                """
            )
            rows = cur.fetchall()
            can_explicit_edit = can_edit_inspection_issues(cur, user)
            can_explicit_delete = can_delete_inspection_issues(cur, user)
            can_explicit_change_inspector = can_change_issue_inspector(cur, user)
            return jsonify(
                [
                    normalize_issue_row_for_response(
                        row,
                        user,
                        can_explicit_edit,
                        can_explicit_delete,
                        False,
                        can_explicit_change_inspector,
                    )
                    for row in rows
                ]
            )

        return jsonify([])
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/my-issues/pending-rectification-count", methods=["GET"])
def get_my_pending_rectification_count():
    return frontend_version_expired_response()


@app.route("/api/issues")
def get_issues():
    user_id = str(request.args.get("user_id", "")).strip()
    issue_description_keyword = str(request.args.get("issue_description", "")).strip()

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_issue_inspector_schema(cur)
        ensure_inspection_completion_schema(cur)
        auto_complete_overdue_inspections(cur)
        conn.commit()

        user = None
        where_clauses = []
        params = []

        if user_id:
            cur.execute(
                """
                SELECT id, role, station_id
                FROM users
                WHERE id = %s
                LIMIT 1;
                """,
                (user_id,),
            )
            user = cur.fetchone()

        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404

        can_view_all = can_view_all_inspection_issues(cur, user) or can_view_region_inspection_issues(cur, user)
        can_view_own = can_view_own_inspection_issues(cur, user)
        if can_view_all:
            pass
        elif can_view_own:
            if not user["station_id"]:
                where_clauses.append("COALESCE(i.inspector_id, ins.inspector_id) = %s")
                params.append(user["id"])
            else:
                where_clauses.append("(i.station_id = %s OR COALESCE(i.inspector_id, ins.inspector_id) = %s)")
                params.extend([user["station_id"], user["id"]])
        else:
            where_clauses.append("COALESCE(i.inspector_id, ins.inspector_id) = %s")
            params.append(user["id"])

        if not append_inspection_table_scope_filter(
            cur,
            user,
            where_clauses,
            params,
            "i.inspection_table_id",
            "limit_issue_inspection_table_scope",
        ):
            return jsonify([])
        if not append_station_region_scope_filter(
            cur,
            user,
            where_clauses,
            params,
            "s.region",
            "limit_issue_station_region_scope",
        ):
            return jsonify([])
        append_pending_audit_issue_visibility_filter(user, where_clauses)
        if issue_description_keyword:
            where_clauses.append("COALESCE(i.description, '') ILIKE %s")
            params.append(f"%{issue_description_keyword}%")

        where_clause = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""

        can_explicit_edit = can_edit_inspection_issues(cur, user)
        can_explicit_delete = can_delete_inspection_issues(cur, user)
        can_explicit_audit = can_audit_inspection_issues(cur, user)
        can_explicit_change_inspector = can_change_issue_inspector(cur, user)
        hide_inspector_contact = should_hide_inspector_contact_info(cur, user)

        cur.execute(
            sql.SQL(
                """
                SELECT
                    i.id,
                    i.station_id,
                    COALESCE(i.inspector_id, ins.inspector_id) AS inspector_id,
                    TO_CHAR(i.created_at, 'YYYY-MM') AS month,
                    TO_CHAR(i.created_at, 'YYYY-MM-DD HH24:MI') AS time,
                    s.region,
                    s.station_name AS station,
                    s.station_manager_name AS station_manager,
                    s.station_manager_phone AS station_manager_phone,
                    issue_inspector.real_name AS inspector,
                    issue_inspector.phone AS inspector_phone,
                    t.table_name AS inspection_table_name,
                    i.standard_id,
                    i.standard_detail_text,
                    i.internal_standard_id,
                    i.internal_standard_detail_text,
                    i.description,
                    i.photo_path AS issue_photo,
                    i.rectification_result,
                    i.rectification_note,
                    TO_CHAR(i.rectification_at, 'YYYY-MM-DD HH24:MI') AS rectification_at,
                    i.rectification_photo_path AS rectification_photo,
                    i.review_result,
                    i.review_note,
                    TO_CHAR(i.review_at, 'YYYY-MM-DD HH24:MI') AS review_at,
                    i.review_photo_path AS review_photo,
                    i.status,
                    COALESCE(i.audit_status, 'pending') AS audit_status,
                    COALESCE(i.is_excellent, FALSE) AS is_excellent,
                    i.audited_by,
                    audit_user.real_name AS audited_by_name,
                    TO_CHAR(i.audited_at, 'YYYY-MM-DD HH24:MI') AS audited_at,
                    ins.sign_status AS inspection_sign_status,
                    ins.station_manager_signed_name,
                    ins.station_manager_signature_path,
                    TO_CHAR(ins.station_manager_signed_at, 'YYYY-MM-DD HH24:MI') AS station_manager_signed_at,
                    ins.inspector_completion_status AS inspection_completion_status
                FROM issues i
                JOIN inspections ins ON i.inspection_id = ins.id
                JOIN stations s ON i.station_id = s.id
                JOIN inspection_tables t ON i.inspection_table_id = t.id
                JOIN users issue_inspector ON COALESCE(i.inspector_id, ins.inspector_id) = issue_inspector.id
                LEFT JOIN users audit_user ON audit_user.id = i.audited_by
                {where_clause}
                ORDER BY i.id DESC;
                """
            ).format(where_clause=sql.SQL(where_clause)),
            params,
        )
        rows = cur.fetchall()
        attach_internal_standard_tags_to_issue_rows(cur, rows)
        conn.commit()
        return jsonify(
            [
                normalize_issue_row_for_response(
                    row,
                    user,
                    can_explicit_edit,
                    can_explicit_delete,
                    can_explicit_audit,
                    can_explicit_change_inspector,
                    hide_inspector_contact,
                )
                for row in rows
            ]
        )
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/issues/<int:issue_id>/audit", methods=["POST"])
def audit_issue(issue_id):
    data = request.get_json(silent=True) or {}
    user_id = get_authenticated_request_user_id(data.get("user_id"))
    action = str(data.get("action") or "").strip().lower()
    status_map = {
        "approve": "approved",
        "approved": "approved",
        "reject": "rejected",
        "rejected": "rejected",
        "reset": "pending",
        "pending": "pending",
    }
    audit_status = status_map.get(action)
    if not audit_status:
        return jsonify({"success": False, "error": "审核操作不正确。"}), 400

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_issue_inspector_schema(cur)
        ensure_inspection_completion_schema(cur)
        auto_complete_overdue_inspections(cur)
        user = get_user_by_id(cur, user_id)
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404
        can_explicit_audit = can_audit_inspection_issues(cur, user)
        if not can_explicit_audit:
            return jsonify({"success": False, "error": "当前账号无权审核巡检问题。"}), 403

        cur.execute(
            """
            SELECT
                i.id,
                i.station_id,
                s.region AS station_region,
                i.inspection_id,
                i.inspection_table_id,
                i.status,
                COALESCE(i.audit_status, 'pending') AS audit_status,
                ins.sign_status AS inspection_sign_status,
                ins.station_manager_signed_name,
                ins.station_manager_signature_path,
                ins.station_manager_signed_at,
                ins.inspector_completion_status AS inspection_completion_status
            FROM issues i
            JOIN inspections ins ON ins.id = i.inspection_id
            JOIN stations s ON s.id = i.station_id
            WHERE i.id = %s
            LIMIT 1;
            """,
            (issue_id,),
        )
        issue = cur.fetchone()
        if not issue:
            return jsonify({"success": False, "error": "巡检问题不存在。"}), 404
        if is_issue_inspection_signed(issue):
            return jsonify({"success": False, "error": "该问题所属检查表已完成站经理签字确认，不能继续审核。"}), 403
        if not is_issue_inspection_completion_done(issue):
            return jsonify({"success": False, "error": "该问题所属巡检记录仍在等待检查人确认，暂不能审核。"}), 400

        can_view_all = can_view_all_inspection_issues(cur, user) or can_view_region_inspection_issues(cur, user)
        can_view_own = can_view_own_inspection_issues(cur, user)
        if not can_view_all and not (
            can_view_own
            and user.get("station_id")
            and issue["station_id"] == user["station_id"]
        ):
            return jsonify({"success": False, "error": "当前账号无权操作该巡检问题。"}), 403
        if not is_inspection_table_allowed_for_user(
            cur,
            user,
            issue["inspection_table_id"],
            "limit_issue_inspection_table_scope",
        ):
            return jsonify({"success": False, "error": "当前账号无权操作该检查表的问题。"}), 403
        if not is_station_region_allowed_for_user(
            cur,
            user,
            issue.get("station_region"),
            "limit_issue_station_region_scope",
        ):
            return jsonify({"success": False, "error": "当前账号无权操作该片区的问题。"}), 403

        if audit_status == "pending":
            cur.execute(
                """
                UPDATE issues
                SET audit_status = 'pending',
                    audited_by = NULL,
                    audited_at = NULL
                WHERE id = %s;
                """,
                (issue_id,),
            )
            message = "该问题已恢复为待审核。"
        else:
            cur.execute(
                """
                UPDATE issues
                SET audit_status = %s,
                    audited_by = %s,
                    audited_at = CURRENT_TIMESTAMP,
                    is_excellent = CASE WHEN %s = 'rejected' THEN FALSE ELSE is_excellent END
                WHERE id = %s;
                """,
                (audit_status, user["id"], audit_status, issue_id),
            )
            message = "该问题已审核通过。" if audit_status == "approved" else "该问题已审核否决，后续不参与记录统计和问题流转。"

        can_explicit_edit = can_edit_inspection_issues(cur, user)
        can_explicit_delete = can_delete_inspection_issues(cur, user)
        can_explicit_change_inspector = can_change_issue_inspector(cur, user)
        hide_inspector_contact = should_hide_inspector_contact_info(cur, user)
        updated_issue = fetch_issue_row_for_response(
            cur,
            issue_id,
            user,
            can_explicit_edit,
            can_explicit_delete,
            can_explicit_audit,
            can_explicit_change_inspector,
            hide_inspector_contact,
        )

        conn.commit()
        return jsonify(
            {
                "success": True,
                "message": message,
                "audit_status": audit_status,
                "audit_status_label": ISSUE_AUDIT_STATUS_LABELS.get(audit_status, "待审核"),
                "issue": updated_issue,
            }
        )
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/issues/<int:issue_id>/excellent", methods=["POST"])
def update_issue_excellent(issue_id):
    data = request.get_json(silent=True) or {}
    user_id = get_authenticated_request_user_id(data.get("user_id"))
    raw_is_excellent = data.get("is_excellent")
    if isinstance(raw_is_excellent, bool):
        is_excellent = raw_is_excellent
    else:
        is_excellent = str(raw_is_excellent).strip().lower() in {"1", "true", "yes", "on"}

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_issue_inspector_schema(cur)
        user = get_user_by_id(cur, user_id)
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404
        if not can_audit_inspection_issues(cur, user):
            return jsonify({"success": False, "error": "当前账号无权标记优秀问题。"}), 403

        cur.execute(
            """
            SELECT
                i.id,
                i.station_id,
                s.region AS station_region,
                i.inspection_table_id,
                COALESCE(i.audit_status, 'pending') AS audit_status
            FROM issues i
            JOIN stations s ON s.id = i.station_id
            WHERE i.id = %s
            LIMIT 1;
            """,
            (issue_id,),
        )
        issue = cur.fetchone()
        if not issue:
            return jsonify({"success": False, "error": "巡检问题不存在。"}), 404

        if normalize_issue_audit_status(issue.get("audit_status")) == "rejected" and is_excellent:
            return jsonify({"success": False, "error": "审核否决的问题不能标记为优秀问题。"}), 400

        can_view_all = can_view_all_inspection_issues(cur, user) or can_view_region_inspection_issues(cur, user)
        can_view_own = can_view_own_inspection_issues(cur, user)
        if not can_view_all and not (
            can_view_own
            and user.get("station_id")
            and issue["station_id"] == user["station_id"]
        ):
            return jsonify({"success": False, "error": "当前账号无权操作该巡检问题。"}), 403
        if not is_inspection_table_allowed_for_user(
            cur,
            user,
            issue["inspection_table_id"],
            "limit_issue_inspection_table_scope",
        ):
            return jsonify({"success": False, "error": "当前账号无权操作该检查表的问题。"}), 403
        if not is_station_region_allowed_for_user(
            cur,
            user,
            issue.get("station_region"),
            "limit_issue_station_region_scope",
        ):
            return jsonify({"success": False, "error": "当前账号无权操作该片区的问题。"}), 403

        cur.execute(
            """
            UPDATE issues
            SET is_excellent = %s
            WHERE id = %s
            RETURNING COALESCE(is_excellent, FALSE) AS is_excellent;
            """,
            (is_excellent, issue_id),
        )
        updated = cur.fetchone()
        conn.commit()
        return jsonify(
            {
                "success": True,
                "message": "已点亮优秀问题。" if updated["is_excellent"] else "已取消优秀问题标记。",
                "is_excellent": bool(updated["is_excellent"]),
            }
        )
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/issues/export-tasks", methods=["POST"])
def create_issue_export_task():
    data = request.get_json(silent=True) or {}
    user_id = get_authenticated_request_user_id(data.get("user_id"))
    conn = None
    cur = None

    try:
        issue_ids = normalize_issue_export_ids(data.get("issue_ids"))
        filter_summary = normalize_issue_export_filter_summary(data.get("filter_summary"))
        export_options = normalize_issue_export_options(data.get("export_options"))
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_issue_export_schema(cur)
        cleanup_expired_issue_exports(cur)

        user = get_user_by_id(cur, user_id)
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404
        if not selected_issue_export_field_keys(export_options):
            raise ValueError("请至少选择一个导出字段。")
        if issue_export_includes_photos(export_options) and not can_export_issue_photos(cur, user):
            raise PermissionError("当前账号无权导出巡检照片，请联系 root 授权。")
        visible_issue_rows = fetch_issue_export_rows(cur, user, issue_ids)
        if len({row["id"] for row in visible_issue_rows}) != len(set(issue_ids)):
            raise PermissionError("导出数据中包含当前账号无权查看的问题，请刷新列表后重新选择。")

        task_id = uuid.uuid4().hex
        now = beijing_now()
        download_filename = f"巡检问题列表_{now.strftime('%Y%m%d_%H%M%S')}_{task_id[:8]}.xlsx"
        cur.execute(
            """
            INSERT INTO issue_export_tasks (
                task_id,
                created_by,
                status,
                selected_count,
                exported_count,
                filter_summary,
                export_options,
                download_filename,
                expires_at
            )
            VALUES (%s, %s, 'pending', %s, 0, %s::jsonb, %s::jsonb, %s, CURRENT_TIMESTAMP + INTERVAL '7 days')
            RETURNING
                task_id,
                created_by,
                status,
                selected_count,
                exported_count,
                filter_summary,
                export_options,
                file_path,
                download_filename,
                error_message,
                TO_CHAR(created_at, 'YYYY-MM-DD HH24:MI') AS created_at,
                TO_CHAR(updated_at, 'YYYY-MM-DD HH24:MI') AS updated_at,
                TO_CHAR(completed_at, 'YYYY-MM-DD HH24:MI') AS completed_at,
                TO_CHAR(expires_at, 'YYYY-MM-DD HH24:MI') AS expires_at;
            """,
            (
                task_id,
                user["id"],
                len(issue_ids),
                json.dumps(filter_summary, ensure_ascii=False),
                json.dumps(export_options, ensure_ascii=False),
                download_filename,
            ),
        )
        task = cur.fetchone()
        conn.commit()
        start_issue_export_task(task_id, issue_ids, user["id"], export_options)
        return jsonify(
            {
                "success": True,
                "message": "导出任务已提交，系统正在后台生成 Excel 文件。",
                "task": serialize_issue_export_task(task),
            }
        )
    except ValueError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except Exception as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/issues/export-tasks/<task_id>")
def get_issue_export_task(task_id):
    user_id = get_authenticated_request_user_id(request.args.get("user_id"))
    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_issue_export_schema(cur)
        cleanup_expired_issue_exports(cur)
        user = get_user_by_id(cur, user_id)
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404
        task = get_issue_export_task_for_user(cur, task_id, user)
        conn.commit()
        return jsonify({"success": True, "task": serialize_issue_export_task(task)})
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 404
    except Exception as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/issues/export-tasks/<task_id>/download")
def download_issue_export_task(task_id):
    user_id = get_authenticated_request_user_id(request.args.get("user_id"))
    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_issue_export_schema(cur)
        cleanup_expired_issue_exports(cur)
        user = get_user_by_id(cur, user_id)
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404
        task = get_issue_export_task_for_user(cur, task_id, user)
        conn.commit()

        if task["status"] != "completed":
            return jsonify({"success": False, "error": "导出文件尚未生成完成。"}), 400
        if not task.get("file_path"):
            return jsonify({"success": False, "error": "导出文件不存在或已过期。"}), 404

        abs_path = resolve_storage_abs_path(task["file_path"])
        if not abs_path or not os.path.isfile(abs_path):
            return jsonify({"success": False, "error": "导出文件不存在或已过期。"}), 404

        response = send_file(
            abs_path,
            as_attachment=True,
            download_name=task.get("download_filename") or f"inspection_issues_{task_id[:8]}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response.headers["Cache-Control"] = "no-store"
        return response
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 404
    except Exception as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/issues/<int:issue_id>", methods=["PUT"])
def update_issue(issue_id):
    data = request.form if request.form else (request.get_json(silent=True) or {})
    user_id = str(data.get("user_id", "")).strip()
    description = str(data.get("description", "")).strip()
    status = canonical_issue_status(data.get("status"))
    rectification_note = str(data.get("rectification_note", "")).strip() or None
    review_note = str(data.get("review_note", "")).strip() or None
    standard_id = str(data.get("standard_id", "")).strip()
    internal_standard_id = str(data.get("internal_standard_id", "")).strip().upper()
    target_inspector_id_param = str(data.get("target_inspector_id", "")).strip()
    issue_photo = request.files.get("issue_photo")

    if not user_id:
        return jsonify({"success": False, "error": "缺少用户信息。"}), 400

    if not description:
        return jsonify({"success": False, "error": "请填写问题描述。"}), 400

    if status not in ISSUE_STATUS_OPTIONS:
        return jsonify({"success": False, "error": "问题状态参数不合法。"}), 400

    try:
        rectification_result = normalize_optional_issue_result(
            data.get("rectification_result"), "站经理整改结果"
        )
        review_result = normalize_optional_issue_result(
            data.get("review_result"), "督导组复核结果"
        )
    except ValueError as error:
        return jsonify({"success": False, "error": str(error)}), 400

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_issue_inspector_schema(cur)
        ensure_inspection_completion_schema(cur)

        user = get_user_by_id(cur, user_id)
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404

        cur.execute(
            """
            SELECT
                i.id,
                i.inspection_id,
                i.station_id,
                s.region AS station_region,
                i.inspection_table_id,
                i.standard_id,
                i.standard_detail_text,
                i.internal_standard_id,
                i.internal_standard_detail_text,
                i.description,
                i.status,
                COALESCE(i.audit_status, 'pending') AS audit_status,
                i.rectification_result,
                i.rectification_note,
                i.rectification_at,
                i.review_result,
                i.review_note,
                i.review_at,
                COALESCE(i.inspector_id, ins.inspector_id) AS inspector_id,
                ins.inspection_date,
                ins.batch_id,
                ins.sign_status AS inspection_sign_status,
                ins.station_manager_signed_name,
                ins.station_manager_signature_path,
                ins.station_manager_signed_at,
                ins.inspector_completion_status AS inspection_completion_status
            FROM issues i
            JOIN inspections ins ON i.inspection_id = ins.id
            JOIN stations s ON s.id = i.station_id
            WHERE i.id = %s
            LIMIT 1;
            """,
            (issue_id,),
        )
        issue = cur.fetchone()

        if not issue:
            return jsonify({"success": False, "error": "巡检问题不存在。"}), 404

        if is_issue_inspection_signed(issue):
            return jsonify({"success": False, "error": "该问题所属检查表已完成站经理签字确认，不能继续编辑。"}), 403

        if is_closed_issue_status(issue["status"]) and not is_root_user(user):
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "该问题已闭环，只有 root 账号可以编辑。",
                    }
                ),
                403,
            )

        can_explicit_edit = can_edit_inspection_issues(cur, user)
        can_explicit_change_inspector = can_change_issue_inspector(cur, user)
        creator_can_modify = can_user_use_creator_issue_controls(user, issue)
        creator_audit_locked = issue_created_by_user(user, issue) and not is_issue_audit_pending(issue)
        if creator_audit_locked and not can_explicit_edit and not can_explicit_change_inspector:
            return jsonify({"success": False, "error": "该问题已审核，需审核人员重新判定为待审核后才能编辑。"}), 403
        if not can_explicit_edit and not creator_can_modify and not can_explicit_change_inspector:
            return jsonify({"success": False, "error": "当前账号无权编辑巡检问题。"}), 403

        if creator_can_modify and not can_explicit_edit and not is_root_user(user):
            if (
                status != "待整改"
                or rectification_result
                or rectification_note
                or review_result
                or review_note
            ):
                return (
                    jsonify(
                        {
                            "success": False,
                            "error": "上传者只能在站点整改前修改问题描述和问题照片，不能调整流转状态或整改复核信息。",
                        }
                    ),
                    400,
                )

        can_view_all = can_view_all_inspection_issues(cur, user) or can_view_region_inspection_issues(cur, user)
        can_view_own = can_view_own_inspection_issues(cur, user)
        if not can_view_all and not (
            can_view_own
            and user.get("station_id")
            and issue["station_id"] == user["station_id"]
        ) and not issue_created_by_user(user, issue):
            return jsonify({"success": False, "error": "当前账号无权操作该巡检问题。"}), 403
        if not is_inspection_table_allowed_for_user(
            cur,
            user,
            issue["inspection_table_id"],
            "limit_issue_inspection_table_scope",
        ):
            return jsonify({"success": False, "error": "当前账号无权操作该检查表的问题。"}), 403
        if not is_station_region_allowed_for_user(
            cur,
            user,
            issue.get("station_region"),
            "limit_issue_station_region_scope",
        ):
            return jsonify({"success": False, "error": "当前账号无权操作该片区的问题。"}), 403

        target_inspector_id = int(issue["inspector_id"])
        inspector_changed = False
        if target_inspector_id_param:
            try:
                requested_inspector_id = int(target_inspector_id_param)
            except (TypeError, ValueError):
                return jsonify({"success": False, "error": "检查人参数不合法。"}), 400

            if requested_inspector_id != int(issue["inspector_id"]):
                if not can_explicit_change_inspector:
                    return jsonify({"success": False, "error": "当前账号无权修改问题检查人归属。"}), 403
                target_inspector = get_user_by_id(cur, requested_inspector_id)
                if not target_inspector:
                    return jsonify({"success": False, "error": "目标检查人不存在。"}), 404
                if not is_supervisor_like(target_inspector) and not has_permission(
                    cur, target_inspector, "submit_inspections"
                ):
                    return jsonify({"success": False, "error": "目标检查人必须具备巡检登记权限。"}), 400
                target_inspector_id = requested_inspector_id
                inspector_changed = True
        elif can_explicit_change_inspector and not (can_explicit_edit or creator_can_modify):
            return jsonify({"success": False, "error": "请选择要调整到的检查人。"}), 400

        can_edit_issue_content = bool(can_explicit_edit or creator_can_modify)
        if not can_edit_issue_content:
            if not inspector_changed:
                return jsonify({"success": False, "error": "当前账号只能调整检查人归属，请选择新的检查人后保存。"}), 400
            description = issue.get("description") or ""
            status = canonical_issue_status(issue.get("status"))
            rectification_result = normalize_issue_result_for_response(issue.get("rectification_result"))
            rectification_note = issue.get("rectification_note")
            review_result = normalize_issue_result_for_response(issue.get("review_result"))
            review_note = issue.get("review_note")
            standard_id = str(issue.get("standard_id") or "")
            internal_standard_id = str(issue.get("internal_standard_id") or "").strip().upper()

        current_standard = {
            "inspection_table_id": int(issue["inspection_table_id"]),
            "standard_id": int(issue["standard_id"]),
            "standard_detail_text": issue.get("standard_detail_text") or "",
            "internal_standard_id": issue.get("internal_standard_id"),
            "internal_standard_detail_text": issue.get("internal_standard_detail_text"),
        }
        usage_mode = get_inspection_standard_usage_mode(cur)
        if not can_edit_issue_content:
            target_standard = current_standard
        elif usage_mode["mode"] == "internal":
            if internal_standard_id or issue.get("internal_standard_id"):
                target_standard = resolve_issue_standard_edit_target(
                    cur,
                    "internal",
                    internal_standard_id=internal_standard_id or issue.get("internal_standard_id"),
                    preferred_external_standard_id=issue.get("standard_id"),
                )
            else:
                target_standard = current_standard
        else:
            target_standard = resolve_issue_standard_edit_target(
                cur,
                "external",
                standard_id=standard_id or issue.get("standard_id"),
            )
        if not is_inspection_table_allowed_for_user(
            cur,
            user,
            target_standard["inspection_table_id"],
            "limit_issue_inspection_table_scope",
        ):
            return jsonify({"success": False, "error": "当前账号无权把问题调整到该检查表。"}), 403

        target_inspection_id = get_or_create_issue_edit_inspection(
            cur,
            issue,
            target_standard["inspection_table_id"],
            target_inspector_id,
        )
        old_inspection_id = int(issue["inspection_id"])
        new_photo_path = (
            save_uploaded_file(issue_photo, "issues")
            if can_edit_issue_content and issue_photo and issue_photo.filename
            else None
        )
        previous_has_rectification_record = bool(
            normalize_issue_result_for_response(issue.get("rectification_result"))
            or issue.get("rectification_note")
        )
        previous_has_review_record = bool(
            normalize_issue_result_for_response(issue.get("review_result"))
            or issue.get("review_note")
        )
        has_rectification_record = bool(rectification_result or rectification_note)
        has_review_record = bool(review_result or review_note)
        should_start_rectification_at = has_rectification_record and not previous_has_rectification_record
        should_start_review_at = has_review_record and not previous_has_review_record

        cur.execute(
            """
            UPDATE issues
            SET description = %s,
                inspector_id = %s,
                inspection_id = %s,
                inspection_table_id = %s,
                standard_id = %s,
                standard_detail_text = %s,
                internal_standard_id = %s,
                internal_standard_detail_text = %s,
                status = %s,
                rectification_result = %s,
                rectification_note = %s,
                rectification_at = CASE
                    WHEN NOT %s THEN NULL
                    WHEN %s THEN CURRENT_TIMESTAMP
                    ELSE rectification_at
                END,
                review_result = %s,
                review_note = %s,
                review_at = CASE
                    WHEN NOT %s THEN NULL
                    WHEN %s THEN CURRENT_TIMESTAMP
                    ELSE review_at
                END,
                photo_path = COALESCE(%s, photo_path)
            WHERE id = %s;
            """,
            (
                description,
                target_inspector_id,
                target_inspection_id,
                target_standard["inspection_table_id"],
                target_standard["standard_id"],
                target_standard["standard_detail_text"],
                target_standard.get("internal_standard_id"),
                target_standard.get("internal_standard_detail_text"),
                status,
                rectification_result,
                rectification_note,
                has_rectification_record,
                should_start_rectification_at,
                review_result,
                review_note,
                has_review_record,
                should_start_review_at,
                new_photo_path,
                issue_id,
            ),
        )
        mark_related_plan_items_completed(
            cur,
            issue["station_id"],
            target_standard["inspection_table_id"],
            target_inspection_id,
            issue["inspection_date"],
        )
        sync_inspection_primary_inspector(cur, target_inspection_id)
        if old_inspection_id != int(target_inspection_id):
            sync_inspection_primary_inspector(cur, old_inspection_id)
            cleanup_empty_inspection_after_issue_move(cur, old_inspection_id)

        can_explicit_delete = can_delete_inspection_issues(cur, user)
        can_explicit_audit = can_audit_inspection_issues(cur, user)
        hide_inspector_contact = should_hide_inspector_contact_info(cur, user)
        updated_issue = fetch_issue_row_for_response(
            cur,
            issue_id,
            user,
            can_explicit_edit,
            can_explicit_delete,
            can_explicit_audit,
            can_explicit_change_inspector,
            hide_inspector_contact,
        )

        conn.commit()
        return jsonify({"success": True, "message": "巡检问题已保存。", "issue": updated_issue})
    except ValueError as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/issues/<int:issue_id>", methods=["DELETE"])
def delete_issue(issue_id):
    data = request.get_json(silent=True) or {}
    user_id = str(data.get("user_id") or request.args.get("user_id", "")).strip()

    if not user_id:
        return jsonify({"success": False, "error": "缺少用户信息。"}), 400

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_issue_inspector_schema(cur)
        ensure_inspection_completion_schema(cur)
        auto_complete_overdue_inspections(cur)
        conn.commit()

        user = get_user_by_id(cur, user_id)
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404

        cur.execute(
            """
            SELECT
                i.id,
                i.inspection_id,
                i.station_id,
                s.region AS station_region,
                i.inspection_table_id,
                i.status,
                COALESCE(i.audit_status, 'pending') AS audit_status,
                i.rectification_result,
                COALESCE(i.inspector_id, ins.inspector_id) AS inspector_id,
                ins.sign_status AS inspection_sign_status,
                ins.station_manager_signed_name,
                ins.station_manager_signature_path,
                ins.station_manager_signed_at,
                ins.inspector_completion_status AS inspection_completion_status
            FROM issues i
            JOIN inspections ins ON i.inspection_id = ins.id
            JOIN stations s ON s.id = i.station_id
            WHERE i.id = %s
            LIMIT 1;
            """,
            (issue_id,),
        )
        issue = cur.fetchone()

        if not issue:
            return jsonify({"success": False, "error": "巡检问题不存在。"}), 404

        if is_issue_inspection_signed(issue):
            return jsonify({"success": False, "error": "该问题所属检查表已完成站经理签字确认，不能继续删除。"}), 403

        if is_closed_issue_status(issue["status"]) and not is_root_user(user):
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "该问题已闭环，只有 root 账号可以删除。",
                    }
                ),
                403,
            )

        can_explicit_delete = can_delete_inspection_issues(cur, user)
        creator_can_modify = can_user_use_creator_issue_controls(user, issue)
        creator_audit_locked = issue_created_by_user(user, issue) and not is_issue_audit_pending(issue)
        if creator_audit_locked and not can_explicit_delete:
            return jsonify({"success": False, "error": "该问题已审核，需审核人员重新判定为待审核后才能删除。"}), 403
        if not can_explicit_delete and not creator_can_modify:
            return jsonify({"success": False, "error": "当前账号无权删除巡检问题。"}), 403

        can_view_all = can_view_all_inspection_issues(cur, user) or can_view_region_inspection_issues(cur, user)
        can_view_own = can_view_own_inspection_issues(cur, user)
        if not can_view_all and not (
            can_view_own
            and user.get("station_id")
            and issue["station_id"] == user["station_id"]
        ) and not issue_created_by_user(user, issue):
            return jsonify({"success": False, "error": "当前账号无权操作该巡检问题。"}), 403
        if not is_inspection_table_allowed_for_user(
            cur,
            user,
            issue["inspection_table_id"],
            "limit_issue_inspection_table_scope",
        ):
            return jsonify({"success": False, "error": "当前账号无权操作该检查表的问题。"}), 403
        if not is_station_region_allowed_for_user(
            cur,
            user,
            issue.get("station_region"),
            "limit_issue_station_region_scope",
        ):
            return jsonify({"success": False, "error": "当前账号无权操作该片区的问题。"}), 403

        cur.execute("DELETE FROM issues WHERE id = %s;", (issue_id,))
        sync_inspection_primary_inspector(cur, issue["inspection_id"])

        # 巡检主记录和计划完成痕迹保留，记录结果由剩余问题数聚合自动体现。
        conn.commit()
        return jsonify({"success": True, "message": "巡检问题已删除。"})
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/inspection-plan-configs", methods=["POST"])
def create_inspection_plan_config():
    data = request.get_json(silent=True) or {}
    user_id = str(data.get("user_id", "")).strip()
    inspection_table_id = str(data.get("inspection_table_id", "")).strip()
    coverage_type = str(data.get("coverage_type", "")).strip()
    period_key = str(data.get("period_key", "")).strip()
    remark = str(data.get("remark", "")).strip() or None
    status = "active"

    if not user_id:
        return jsonify({"success": False, "error": "缺少用户信息。"}), 400

    if not inspection_table_id:
        return jsonify({"success": False, "error": "缺少检查表信息。"}), 400

    if coverage_type not in {"monthly", "quarterly", "yearly"}:
        return jsonify({"success": False, "error": "覆盖要求参数不合法。"}), 400

    if not period_key:
        return jsonify({"success": False, "error": "缺少周期标识。"}), 400

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()

        user = get_user_by_id(cur, user_id)
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404

        if not can_manage_plan(cur, user):
            return (
                jsonify({"success": False, "error": "当前账号无权维护巡检计划。"}),
                403,
            )

        cur.execute(
            """
            SELECT id, table_name, is_active
            FROM inspection_tables
            WHERE id = %s
            LIMIT 1;
            """,
            (inspection_table_id,),
        )
        inspection_table = cur.fetchone()

        if not inspection_table:
            return jsonify({"success": False, "error": "检查表不存在。"}), 404

        if not inspection_table["is_active"]:
            return jsonify({"success": False, "error": "检查表未启用。"}), 400
        if not is_inspection_table_allowed_for_user(
            cur,
            user,
            inspection_table_id,
            "limit_plan_inspection_table_scope",
        ):
            return jsonify({"success": False, "error": "当前账号无权维护该检查表的巡检计划。"}), 403

        cur.execute(
            """
            DELETE FROM inspection_plan_configs
            WHERE inspection_table_id = %s
              AND coverage_type <> %s;
            """,
            (inspection_table_id, coverage_type),
        )

        cur.execute(
            """
            SELECT id
            FROM inspection_plan_configs
            WHERE inspection_table_id = %s
              AND coverage_type = %s
              AND period_key = %s
            LIMIT 1;
            """,
            (inspection_table_id, coverage_type, period_key),
        )
        existing = cur.fetchone()

        if existing:
            conn.commit()
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "该检查表在当前覆盖要求与周期下已存在计划配置。",
                        "existing_id": existing["id"],
                    }
                ),
                409,
            )

        cur.execute(
            """
            INSERT INTO inspection_plan_configs (
                inspection_table_id,
                coverage_type,
                period_key,
                created_by,
                updated_by,
                status,
                remark,
                updated_at
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
            RETURNING id;
            """,
            (
                inspection_table_id,
                coverage_type,
                period_key,
                user_id,
                user_id,
                status,
                remark,
            ),
        )
        created_row = cur.fetchone()
        sync_plan_station_items_completion_by_history(cur, created_row["id"])
        conn.commit()

        return jsonify(
            {
                "success": True,
                "message": "巡检计划配置创建成功。",
                "plan_config_id": created_row["id"],
            }
        )
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/inspection-plan-configs/<int:plan_config_id>", methods=["PUT"])
def update_inspection_plan_config(plan_config_id):
    data = request.get_json(silent=True) or {}
    user_id = str(data.get("user_id", "")).strip()
    coverage_type = str(data.get("coverage_type", "")).strip()
    period_key = str(data.get("period_key", "")).strip()
    remark = data.get("remark")

    if not user_id:
        return jsonify({"success": False, "error": "缺少用户信息。"}), 400

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()

        user = get_user_by_id(cur, user_id)
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404

        if not can_manage_plan(cur, user):
            return (
                jsonify({"success": False, "error": "当前账号无权维护巡检计划。"}),
                403,
            )

        cur.execute(
            """
            SELECT id, inspection_table_id, coverage_type, period_key, status, remark
            FROM inspection_plan_configs
            WHERE id = %s
            LIMIT 1;
            """,
            (plan_config_id,),
        )
        current_row = cur.fetchone()

        if not current_row:
            return jsonify({"success": False, "error": "巡检计划配置不存在。"}), 404
        if not is_inspection_table_allowed_for_user(
            cur,
            user,
            current_row["inspection_table_id"],
            "limit_plan_inspection_table_scope",
        ):
            return jsonify({"success": False, "error": "当前账号无权维护该检查表的巡检计划。"}), 403

        new_coverage_type = coverage_type or current_row["coverage_type"]
        new_period_key = period_key or current_row["period_key"]
        new_status = "active"
        new_remark = (
            current_row["remark"] if remark is None else (str(remark).strip() or None)
        )

        if new_coverage_type not in {"monthly", "quarterly", "yearly"}:
            return jsonify({"success": False, "error": "覆盖要求参数不合法。"}), 400

        if not new_period_key:
            return jsonify({"success": False, "error": "缺少周期标识。"}), 400

        cur.execute(
            """
            SELECT id
            FROM inspection_plan_configs
            WHERE inspection_table_id = %s
              AND coverage_type = %s
              AND period_key = %s
              AND id <> %s
            LIMIT 1;
            """,
            (
                current_row["inspection_table_id"],
                new_coverage_type,
                new_period_key,
                plan_config_id,
            ),
        )
        duplicate_row = cur.fetchone()

        if duplicate_row:
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "更新后会与现有计划配置重复。",
                        "existing_id": duplicate_row["id"],
                    }
                ),
                409,
            )

        cur.execute(
            """
            DELETE FROM inspection_plan_configs
            WHERE inspection_table_id = %s
              AND coverage_type <> %s
              AND id <> %s;
            """,
            (current_row["inspection_table_id"], new_coverage_type, plan_config_id),
        )

        cur.execute(
            """
            UPDATE inspection_plan_configs
            SET coverage_type = %s,
                period_key = %s,
                status = %s,
                remark = %s,
                updated_by = %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = %s;
            """,
            (
                new_coverage_type,
                new_period_key,
                new_status,
                new_remark,
                user_id,
                plan_config_id,
            ),
        )

        sync_plan_station_items_completion_by_history(cur, plan_config_id)
        conn.commit()
        return jsonify({"success": True, "message": "巡检计划配置更新成功。"})
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/inspection-plan-configs/<int:plan_config_id>", methods=["DELETE"])
def delete_inspection_plan_config(plan_config_id):
    data = request.get_json(silent=True) or {}
    user_id = str(data.get("user_id") or request.args.get("user_id", "")).strip()

    if not user_id:
        return jsonify({"success": False, "error": "缺少用户信息。"}), 400

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()

        user = get_user_by_id(cur, user_id)
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404

        if not can_manage_plan(cur, user):
            return (
                jsonify({"success": False, "error": "当前账号无权维护巡检计划。"}),
                403,
            )

        cur.execute(
            """
            SELECT id, inspection_table_id
            FROM inspection_plan_configs
            WHERE id = %s
            LIMIT 1;
            """,
            (plan_config_id,),
        )
        current_row = cur.fetchone()

        if not current_row:
            return jsonify({"success": False, "error": "巡检计划配置不存在。"}), 404
        if not is_inspection_table_allowed_for_user(
            cur,
            user,
            current_row["inspection_table_id"],
            "limit_plan_inspection_table_scope",
        ):
            return jsonify({"success": False, "error": "当前账号无权维护该检查表的巡检计划。"}), 403

        cur.execute(
            "DELETE FROM inspection_plan_configs WHERE id = %s;",
            (plan_config_id,),
        )

        conn.commit()
        return jsonify({"success": True, "message": "巡检计划删除成功。"})
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/inspections/<int:inspection_id>/complete", methods=["POST"])
def complete_inspection_by_inspector(inspection_id):
    data = request.get_json(silent=True) or {}
    user_id = get_authenticated_request_user_id(data.get("user_id"))

    if not user_id:
        return jsonify({"success": False, "error": "缺少用户信息。"}), 400

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_issue_inspector_schema(cur)
        ensure_inspection_completion_schema(cur)

        user = get_user_by_id(cur, user_id)
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404

        if not is_supervisor_like(user):
            return jsonify({"success": False, "error": "只有督导组账号可以确认检查表完成。"}), 403

        cur.execute(
            """
            SELECT
                ins.id,
                ins.station_id,
                ins.inspection_table_id,
                ins.inspection_date,
                ins.inspector_completion_status,
                s.station_name,
                t.table_name
            FROM inspections ins
            JOIN stations s ON ins.station_id = s.id
            JOIN inspection_tables t ON ins.inspection_table_id = t.id
            WHERE ins.id = %s
            LIMIT 1;
            """,
            (inspection_id,),
        )
        inspection = cur.fetchone()
        if not inspection:
            return jsonify({"success": False, "error": "巡检记录不存在。"}), 404

        if inspection.get("inspector_completion_status") == INSPECTION_COMPLETION_DONE:
            return jsonify({"success": False, "error": "该检查表已确认完成。"}), 400

        participant_ids = get_inspection_participant_ids(cur, inspection_id)
        current_user_id = int(user["id"])
        if current_user_id not in participant_ids:
            return jsonify({"success": False, "error": "只有参与该检查表录入的检查人员可以确认完成。"}), 403

        progress = confirm_inspection_participant(cur, inspection_id, current_user_id)
        all_confirmed = bool(progress.get("total")) and int(progress.get("pending") or 0) == 0
        if all_confirmed:
            complete_inspection_record(cur, inspection_id, current_user_id, "manual_all")
            mark_related_plan_items_completed(
                cur,
                inspection["station_id"],
                inspection["inspection_table_id"],
                inspection_id,
                inspection["inspection_date"],
            )
            progress = fetch_inspection_completion_progress(cur, [inspection_id]).get(int(inspection_id), progress)
            message = "所有检查人均已确认，本检查表已确认完成。"
        else:
            cur.execute(
                """
                UPDATE inspections
                SET updated_at = CURRENT_TIMESTAMP
                WHERE id = %s;
                """,
                (inspection_id,),
            )
            pending_names = "、".join(progress.get("pending_names") or [])
            message = (
                f"已记录你的完成确认，仍需等待：{pending_names}。"
                if pending_names
                else "已记录你的完成确认，仍需等待其他检查人确认。"
            )
        conn.commit()
        return jsonify(
            {
                "success": True,
                "message": message,
                "inspection_id": inspection_id,
                "inspection": {
                    "id": inspection_id,
                    "inspector_completion_status": (
                        INSPECTION_COMPLETION_DONE if all_confirmed else INSPECTION_COMPLETION_PENDING
                    ),
                    "inspector_completion_source_label": (
                        inspection_completion_source_label("manual_all") if all_confirmed else ""
                    ),
                    "inspector_completion_progress": progress,
                },
            }
        )
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/inspection-completion")
def get_management_inspection_completion():
    user_id = str(request.args.get("user_id", "")).strip()
    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_issue_inspector_schema(cur)
        ensure_inspection_completion_schema(cur)
        sync_signed_inspections_completion(cur)
        actor = require_management_user(cur, user_id, "manage_inspection_completion")
        auto_completed_count = auto_complete_overdue_inspections(cur)
        config = get_inspection_completion_config(cur)

        cur.execute(
            """
            SELECT
                ins.id,
                ins.batch_id,
                TO_CHAR(ins.inspection_date, 'YYYY-MM-DD') AS inspection_date,
                TO_CHAR(ins.inspection_date, 'YYYY-MM') AS inspection_month,
                (%s::date - ins.inspection_date) AS age_days,
                s.station_name,
                s.region AS station_region,
                t.table_name AS inspection_table_name,
                COUNT(i.id) FILTER (WHERE COALESCE(i.audit_status, 'pending') <> 'rejected') AS issue_count,
                ins.sign_status,
                ins.station_manager_signed_name,
                TO_CHAR(ins.station_manager_signed_at, 'YYYY-MM-DD HH24:MI') AS station_manager_signed_at,
                ins.inspector_completion_status,
                ins.inspector_completion_source,
                TO_CHAR(ins.inspector_completed_at, 'YYYY-MM-DD HH24:MI') AS inspector_completed_at,
                COALESCE(completed_user.username, '') AS inspector_completed_by_username,
                COALESCE(completed_user.real_name, '') AS inspector_completed_by_name,
                STRING_AGG(
                    DISTINCT COALESCE(participant.real_name, participant.username),
                    '、'
                ) FILTER (WHERE participant.id IS NOT NULL) AS inspector_names,
                TO_CHAR(ins.created_at, 'YYYY-MM-DD HH24:MI') AS created_at,
                TO_CHAR(COALESCE(ins.updated_at, ins.created_at), 'YYYY-MM-DD HH24:MI') AS updated_at
            FROM inspections ins
            JOIN stations s ON s.id = ins.station_id
            JOIN inspection_tables t ON t.id = ins.inspection_table_id
            LEFT JOIN issues i ON i.inspection_id = ins.id
                AND COALESCE(i.audit_status, 'pending') <> 'rejected'
            LEFT JOIN users completed_user ON completed_user.id = ins.inspector_completed_by
            LEFT JOIN users participant ON participant.id = COALESCE(i.inspector_id, ins.inspector_id)
            GROUP BY
                ins.id,
                ins.batch_id,
                ins.inspection_date,
                s.station_name,
                s.region,
                t.table_name,
                ins.sign_status,
                ins.station_manager_signed_name,
                ins.station_manager_signed_at,
                ins.inspector_completion_status,
                ins.inspector_completion_source,
                ins.inspector_completed_at,
                completed_user.username,
                completed_user.real_name,
                ins.created_at,
                ins.updated_at
            ORDER BY ins.inspection_date DESC, ins.id DESC
            LIMIT 500;
            """,
            (beijing_today(),),
        )
        records = []
        for row in cur.fetchall():
            item = dict(row)
            item["issue_count"] = int(item.get("issue_count") or 0)
            item["age_days"] = int(item.get("age_days") or 0)
            item["inspection_period_key"] = format_inspection_period_key(
                item.get("inspection_date"),
                config["record_uniqueness_period"],
            )
            item["inspection_period_label"] = format_inspection_period_label(
                item.get("inspection_date"),
                config["record_uniqueness_period"],
            )
            item["inspector_completion_source_label"] = inspection_completion_source_label(
                item.get("inspector_completion_source")
            )
            records.append(item)

        conn.commit()
        return jsonify(
            {
                "success": True,
                "config": config,
                "records": records,
                "auto_completed_count": auto_completed_count,
                "actor": {
                    "id": actor["id"],
                    "username": actor["username"],
                    "real_name": actor["real_name"],
                },
            }
        )
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 404
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/inspection-completion/config", methods=["PUT"])
def update_management_inspection_completion_config():
    data = request.get_json(silent=True) or {}
    user_id = str(data.get("user_id", "")).strip()
    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        actor = require_management_user(cur, user_id, "manage_inspection_completion")
        config = save_inspection_completion_config(cur, data, actor["id"])
        conn.commit()
        return jsonify({"success": True, "message": "巡检完成确认规则已保存。", "config": config})
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 404
    except ValueError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/inspection-completion/<int:inspection_id>/complete", methods=["POST"])
def complete_management_inspection_record(inspection_id):
    data = request.get_json(silent=True) or {}
    user_id = str(data.get("user_id", "")).strip()
    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        actor = require_management_user(cur, user_id, "manage_inspection_completion")
        ensure_inspection_completion_schema(cur)
        updated = complete_inspection_record(cur, inspection_id, actor["id"], "admin")
        if updated == 0:
            return jsonify({"success": False, "error": "该巡检记录不存在或已确认完成。"}), 400
        conn.commit()
        return jsonify({"success": True, "message": "已在后台确认该检查表完成。"})
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 404
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/management/inspection-completion/<int:inspection_id>/reopen", methods=["POST"])
def reopen_management_inspection_record(inspection_id):
    data = request.get_json(silent=True) or {}
    user_id = str(data.get("user_id", "")).strip()
    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        require_management_user(cur, user_id, "manage_inspection_completion")
        ensure_inspection_completion_schema(cur)
        sync_signed_inspections_completion(cur)
        cur.execute(
            """
            SELECT id, sign_status
            FROM inspections
            WHERE id = %s
            LIMIT 1;
            """,
            (inspection_id,),
        )
        inspection = cur.fetchone()
        if not inspection:
            return jsonify({"success": False, "error": "该巡检记录不存在。"}), 404
        if inspection.get("sign_status") == "已签名确认":
            return jsonify({"success": False, "error": "站经理已签字确认的巡检记录不能恢复为未完成。"}), 400
        updated = reopen_inspection_record(cur, inspection_id)
        if updated == 0:
            return jsonify({"success": False, "error": "该巡检记录不存在。"}), 404
        conn.commit()
        return jsonify({"success": True, "message": "该检查表已恢复为未完成状态。"})
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 404
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/assessment/peer-reviews", methods=["GET"])
def get_assessment_peer_reviews():
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_peer_review_schema(cur)
        conn.commit()

        user = get_user_by_id(cur, g.current_user["id"])
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404
        if not can_view_peer_reviews(cur, user):
            return jsonify({"success": False, "error": "当前账号无权查看成员互评。"}), 403

        can_manage = can_manage_peer_review_tasks(cur, user)
        templates = []
        people = fetch_peer_review_people(cur) if can_manage else []
        if can_manage:
            cur.execute(
                """
                SELECT id
                FROM peer_review_templates
                WHERE is_active = TRUE
                ORDER BY updated_at DESC, id DESC;
                """
            )
            templates = [serialize_peer_review_template(cur, row["id"]) for row in cur.fetchall()]

        return jsonify(
            {
                "success": True,
                "can_manage": can_manage,
                "current_user": {
                    "id": user["id"],
                    "username": user["username"],
                    "real_name": user["real_name"],
                    "display_name": user.get("real_name") or user.get("username") or "",
                },
                "people": people,
                "templates": templates,
                "tasks": build_peer_review_dashboard(cur, user, can_manage),
            }
        )
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/assessment/peer-reviews/pending-count", methods=["GET"])
def get_assessment_peer_review_pending_count():
    return frontend_version_expired_response()


@app.route("/api/assessment/peer-reviews/templates", methods=["POST"])
def create_assessment_peer_review_template():
    data = request.get_json(silent=True) or {}
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_peer_review_schema(cur)
        user = get_user_by_id(cur, g.current_user["id"])
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404
        if not can_manage_peer_review_tasks(cur, user):
            return jsonify({"success": False, "error": "当前账号无权管理成员互评任务。"}), 403
        template = upsert_peer_review_template(cur, data, user)
        conn.commit()
        return jsonify({"success": True, "message": "成员互评模板已创建。", "template": template})
    except LookupError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 404
    except ValueError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/assessment/peer-reviews/templates/<int:template_id>", methods=["PUT"])
def update_assessment_peer_review_template(template_id):
    data = request.get_json(silent=True) or {}
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_peer_review_schema(cur)
        user = get_user_by_id(cur, g.current_user["id"])
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404
        if not can_manage_peer_review_tasks(cur, user):
            return jsonify({"success": False, "error": "当前账号无权管理成员互评任务。"}), 403
        template = upsert_peer_review_template(cur, data, user, template_id)
        conn.commit()
        return jsonify({"success": True, "message": "成员互评模板已保存。", "template": template})
    except LookupError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 404
    except ValueError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/assessment/peer-reviews/templates/<int:template_id>", methods=["DELETE"])
def delete_assessment_peer_review_template(template_id):
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_peer_review_schema(cur)
        user = get_user_by_id(cur, g.current_user["id"])
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404
        if not can_manage_peer_review_tasks(cur, user):
            return jsonify({"success": False, "error": "当前账号无权管理成员互评任务。"}), 403
        cur.execute(
            """
            DELETE FROM peer_review_templates
            WHERE id = %s
            RETURNING id;
            """,
            (template_id,),
        )
        if not cur.fetchone():
            return jsonify({"success": False, "error": "成员互评模板不存在。"}), 404
        conn.commit()
        return jsonify({"success": True, "message": "成员互评模板已删除，历史任务不受影响。"})
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/assessment/peer-reviews/tasks", methods=["POST"])
def create_assessment_peer_review_task():
    data = request.get_json(silent=True) or {}
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_peer_review_schema(cur)
        user = get_user_by_id(cur, g.current_user["id"])
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404
        if not can_manage_peer_review_tasks(cur, user):
            return jsonify({"success": False, "error": "当前账号无权发起成员互评任务。"}), 403
        task_id = create_peer_review_task_from_template(cur, data, user)
        conn.commit()
        return jsonify({"success": True, "message": "成员互评任务已发起。", "task_id": task_id})
    except LookupError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 404
    except ValueError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/assessment/peer-reviews/tasks/<int:task_id>/status", methods=["PUT"])
def update_assessment_peer_review_task_status(task_id):
    data = request.get_json(silent=True) or {}
    status = normalize_peer_review_status(data.get("status"))
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_peer_review_schema(cur)
        user = get_user_by_id(cur, g.current_user["id"])
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404
        if not can_manage_peer_review_tasks(cur, user):
            return jsonify({"success": False, "error": "当前账号无权管理成员互评任务。"}), 403
        cur.execute(
            """
            UPDATE peer_review_tasks
            SET status = %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = %s
            RETURNING id;
            """,
            (status, task_id),
        )
        if not cur.fetchone():
            return jsonify({"success": False, "error": "成员互评任务不存在。"}), 404
        conn.commit()
        return jsonify({"success": True, "message": "成员互评任务状态已更新。"})
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/assessment/peer-reviews/tasks/<int:task_id>", methods=["DELETE"])
def delete_assessment_peer_review_task(task_id):
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_peer_review_schema(cur)
        user = get_user_by_id(cur, g.current_user["id"])
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404
        if not can_manage_peer_review_tasks(cur, user):
            return jsonify({"success": False, "error": "当前账号无权删除成员互评任务。"}), 403
        cur.execute(
            """
            DELETE FROM peer_review_tasks
            WHERE id = %s
            RETURNING id;
            """,
            (task_id,),
        )
        if not cur.fetchone():
            return jsonify({"success": False, "error": "成员互评任务不存在。"}), 404
        conn.commit()
        return jsonify({"success": True, "message": "成员互评任务已删除。"})
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/assessment/peer-reviews/responses", methods=["POST"])
def submit_assessment_peer_review_response():
    data = request.get_json(silent=True) or {}
    conn = None
    cur = None
    try:
        task_id = int(data.get("task_id") or 0)
        reviewee_id = int(data.get("reviewee_id") or 0)
        answers = data.get("answers") if isinstance(data.get("answers"), list) else []
        if task_id <= 0 or reviewee_id <= 0:
            raise ValueError("评价提交参数不完整。")

        conn = get_db_connection()
        cur = conn.cursor()
        ensure_peer_review_schema(cur)
        user = get_user_by_id(cur, g.current_user["id"])
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404
        if not can_view_peer_reviews(cur, user):
            return jsonify({"success": False, "error": "当前账号无权提交成员互评。"}), 403

        cur.execute(
            """
            SELECT id, status, deadline_at
            FROM peer_review_tasks
            WHERE id = %s
            LIMIT 1;
            """,
            (task_id,),
        )
        task = cur.fetchone()
        if not task:
            return jsonify({"success": False, "error": "成员互评任务不存在。"}), 404
        if normalize_peer_review_status(task.get("status")) != "active":
            return jsonify({"success": False, "error": "该互评任务已关闭，不能继续填写。"}), 400
        if task.get("deadline_at") and task["deadline_at"] < datetime.now():
            return jsonify({"success": False, "error": "该互评任务已超过截止时间。"}), 400
        if int(user["id"]) == reviewee_id:
            return jsonify({"success": False, "error": "成员互评不支持评价自己。"}), 400

        cur.execute(
            "SELECT 1 FROM peer_review_task_participants WHERE task_id = %s AND user_id = %s;",
            (task_id, user["id"]),
        )
        if not cur.fetchone():
            return jsonify({"success": False, "error": "当前账号不是该任务的填写人员。"}), 403
        cur.execute(
            "SELECT 1 FROM peer_review_task_reviewees WHERE task_id = %s AND user_id = %s;",
            (task_id, reviewee_id),
        )
        if not cur.fetchone():
            return jsonify({"success": False, "error": "该人员不是本任务的被评人。"}), 400
        cur.execute(
            """
            SELECT id, item_type, title, max_score
            FROM peer_review_task_items
            WHERE task_id = %s
            ORDER BY sort_order ASC, id ASC;
            """,
            (task_id,),
        )
        task_items = [dict(row) for row in cur.fetchall()]
        if not task_items:
            return jsonify({"success": False, "error": "该任务没有可填写的评价项目。"}), 400
        answer_map = {}
        for answer in answers:
            if not isinstance(answer, dict):
                continue
            try:
                answer_map[int(answer.get("task_item_id") or 0)] = answer
            except (TypeError, ValueError):
                continue

        normalized_answers = []
        for item in task_items:
            answer = answer_map.get(int(item["id"])) or {}
            if normalize_peer_review_item_type(item.get("item_type")) == "score":
                try:
                    score_value = int(answer.get("score_value") or 0)
                except (TypeError, ValueError):
                    score_value = 0
                max_score = int(item.get("max_score") or 5)
                if score_value < 1 or score_value > max_score:
                    raise ValueError(f"请为“{item['title']}”选择 1-{max_score} 分。")
                normalized_answers.append(
                    {
                        "task_item_id": item["id"],
                        "item_type": "score",
                        "score_value": score_value,
                        "text_value": "",
                    }
                )
            else:
                text_value = normalize_text(answer.get("text_value"), 2000)
                if not text_value:
                    raise ValueError(f"请填写“{item['title']}”。")
                normalized_answers.append(
                    {
                        "task_item_id": item["id"],
                        "item_type": "text",
                        "score_value": None,
                        "text_value": text_value,
                    }
                )

        cur.execute(
            """
            INSERT INTO peer_review_responses (
                task_id,
                reviewer_id,
                reviewee_id,
                submitted_at,
                updated_at
            )
            VALUES (%s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            ON CONFLICT (task_id, reviewer_id, reviewee_id)
            DO UPDATE SET
                updated_at = CURRENT_TIMESTAMP
            RETURNING id;
            """,
            (task_id, user["id"], reviewee_id),
        )
        response_id = cur.fetchone()["id"]
        cur.execute("DELETE FROM peer_review_response_items WHERE response_id = %s;", (response_id,))
        for answer in normalized_answers:
            cur.execute(
                """
                INSERT INTO peer_review_response_items (
                    response_id,
                    task_item_id,
                    item_type,
                    score_value,
                    text_value,
                    created_at,
                    updated_at
                )
                VALUES (%s, %s, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP);
                """,
                (
                    response_id,
                    answer["task_item_id"],
                    answer["item_type"],
                    answer["score_value"],
                    answer["text_value"],
                ),
            )
        conn.commit()
        return jsonify({"success": True, "message": "评价已提交。", "response_id": response_id})
    except ValueError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/assessment/station-scores", methods=["GET"])
def get_assessment_station_scores():
    station_id = str(request.args.get("station_id", "")).strip()
    if not station_id:
        return jsonify({"success": False, "error": "请选择需要评分的站点。"}), 400
    try:
        station_id = int(station_id)
        month, month_start, next_month = parse_reporting_date_range(
            request.args.get("month"),
            request.args.get("date_from"),
            request.args.get("date_to"),
        )
    except ValueError as exc:
        return jsonify({"success": False, "error": str(exc)}), 400

    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_inspection_checklist_management_schema(cur)
        ensure_issue_inspector_schema(cur)
        ensure_station_score_schema(cur)
        conn.commit()

        user = get_user_by_id(cur, g.current_user["id"])
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404
        if not can_view_station_scores(cur, user):
            return jsonify({"success": False, "error": "当前账号无权查看站点评分。"}), 403

        payload = build_station_score_payload(cur, station_id, month, month_start, next_month, user)
        return jsonify({"success": True, **payload})
    except LookupError as exc:
        return jsonify({"success": False, "error": str(exc)}), 404
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/assessment/station-scores/scorable-checklists", methods=["GET"])
def get_assessment_station_score_scorable_checklists():
    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_inspection_checklist_management_schema(cur)
        user = get_user_by_id(cur, g.current_user["id"])
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404
        if not can_view_station_scores(cur, user):
            return jsonify({"success": False, "error": "当前账号无权查看站点评分。"}), 403
        checklists = []
        for checklist in fetch_scorable_checklist_rows(cur):
            checklist["checklist_mode"] = normalize_checklist_mode(checklist.get("checklist_mode"))
            checklist["checklist_mode_label"] = "视频检查" if checklist["checklist_mode"] == "online" else "现场检查"
            checklists.append(checklist)
        return jsonify({"success": True, "checklists": checklists})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/assessment/station-scores/export-tasks", methods=["POST"])
def create_assessment_station_score_export_task():
    data = request.get_json(silent=True) or {}
    mode = str(data.get("mode") or "single").strip().lower()
    if mode not in {"single", "all"}:
        return jsonify({"success": False, "error": "导出范围参数不合法。"}), 400
    try:
        month, month_start, next_month = parse_reporting_date_range(
            data.get("month"),
            data.get("date_from"),
            data.get("date_to"),
        )
        station_id = int(data.get("station_id") or 0) if mode == "single" else None
        inspection_table_id = int(data.get("inspection_table_id") or 0) if mode == "single" else None
        inspection_table_ids = normalize_station_score_table_ids(data.get("inspection_table_ids")) if mode == "all" else []
        if mode == "single" and (not station_id or not inspection_table_id):
            raise ValueError("请选择需要导出的站点和检查表。")
        export_options = normalize_station_score_export_options(data.get("export_options"))
    except ValueError as exc:
        return jsonify({"success": False, "error": str(exc)}), 400

    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_inspection_checklist_management_schema(cur)
        ensure_issue_inspector_schema(cur)
        ensure_station_score_schema(cur)
        ensure_station_score_export_schema(cur)
        cleanup_expired_station_score_exports(cur)

        user = get_user_by_id(cur, g.current_user["id"])
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404
        if not can_adjust_station_scores(cur, user):
            return jsonify({"success": False, "error": "当前账号无权导出站点评分。"}), 403

        if mode == "single":
            payload = build_station_score_payload(
                cur,
                station_id,
                month,
                month_start,
                next_month,
                user,
                inspection_table_id=inspection_table_id,
            )
            if not payload.get("tables"):
                return jsonify({"success": False, "error": "当前站点和检查表暂无可导出的评分数据。"}), 404
            station = payload.get("station") or {}
            table = payload["tables"][0]
            selected_count = int(table.get("item_count") or 0)
            filter_summary = build_station_score_export_filter_summary(mode, month, station, table)
            download_filename = f"站点评分_{station.get('station_name') or '站点'}_{table.get('table_name') or '检查表'}_{month}.xlsx"
        else:
            scorable_checklists = fetch_scorable_checklist_rows(cur)
            scorable_ids = {int(item["id"]) for item in scorable_checklists}
            if inspection_table_ids:
                invalid_ids = [table_id for table_id in inspection_table_ids if table_id not in scorable_ids]
                if invalid_ids:
                    return jsonify({"success": False, "error": "选择的检查表不存在或未配置可评分字段。"}), 400
            selected_checklists = [
                item
                for item in scorable_checklists
                if not inspection_table_ids or int(item["id"]) in set(inspection_table_ids)
            ]
            if not selected_checklists:
                return jsonify({"success": False, "error": "请选择至少一张可评分检查表。"}), 400
            cur.execute("SELECT COUNT(*) AS station_count FROM stations;")
            row = cur.fetchone()
            selected_count = int(row.get("station_count") or 0)
            if selected_count <= 0:
                return jsonify({"success": False, "error": "当前没有可导出的站点。"}), 404
            filter_summary = build_station_score_export_filter_summary(
                mode,
                month,
                station_count=selected_count,
                selected_tables=selected_checklists,
            )
            download_filename = f"站点评分全部站点_{month}.zip"

        task_id = uuid.uuid4().hex
        cur.execute(
            """
            INSERT INTO station_score_export_tasks (
                task_id,
                created_by,
                status,
                export_mode,
                selected_count,
                exported_count,
                filter_summary,
                export_options,
                download_filename,
                expires_at
            )
            VALUES (%s, %s, 'pending', %s, %s, 0, %s::jsonb, %s::jsonb, %s, CURRENT_TIMESTAMP + INTERVAL '7 days')
            RETURNING
                task_id,
                created_by,
                status,
                export_mode,
                selected_count,
                exported_count,
                filter_summary,
                export_options,
                file_path,
                download_filename,
                error_message,
                TO_CHAR(created_at, 'YYYY-MM-DD HH24:MI') AS created_at,
                TO_CHAR(updated_at, 'YYYY-MM-DD HH24:MI') AS updated_at,
                TO_CHAR(completed_at, 'YYYY-MM-DD HH24:MI') AS completed_at,
                TO_CHAR(expires_at, 'YYYY-MM-DD HH24:MI') AS expires_at;
            """,
            (
                task_id,
                user["id"],
                mode,
                selected_count,
                json.dumps(filter_summary, ensure_ascii=False),
                json.dumps(export_options, ensure_ascii=False),
                download_filename,
            ),
        )
        task = cur.fetchone()
        conn.commit()
        start_station_score_export_task(
            task_id,
            user["id"],
            mode,
            month,
            month_start,
            next_month,
            export_options,
            station_id=station_id,
            inspection_table_id=inspection_table_id,
            inspection_table_ids=inspection_table_ids,
        )
        return jsonify(
            {
                "success": True,
                "message": "导出任务已提交，系统正在后台生成文件。",
                "task": serialize_station_score_export_task(task),
            }
        )
    except LookupError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 404
    except PermissionError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 403
    except ValueError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/assessment/station-scores/export-tasks/<task_id>", methods=["GET"])
def get_assessment_station_score_export_task(task_id):
    conn = None
    cur = None
    try:
        maybe_cleanup_expired_station_score_exports()
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_station_score_export_schema(cur)
        user = get_user_by_id(cur, g.current_user["id"])
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404
        if not can_adjust_station_scores(cur, user):
            return jsonify({"success": False, "error": "当前账号无权查看站点评分导出任务。"}), 403
        task = get_station_score_export_task_for_user(cur, task_id, user)
        return jsonify({"success": True, "task": serialize_station_score_export_task(task)})
    except PermissionError as exc:
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        return jsonify({"success": False, "error": str(exc)}), 404
    except Exception as exc:
        return jsonify({"success": False, "error": str(exc)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/assessment/station-scores/export-tasks/<task_id>/download", methods=["GET"])
def download_assessment_station_score_export_task(task_id):
    conn = None
    cur = None
    try:
        maybe_cleanup_expired_station_score_exports()
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_station_score_export_schema(cur)
        user = get_user_by_id(cur, g.current_user["id"])
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404
        if not can_adjust_station_scores(cur, user):
            return jsonify({"success": False, "error": "当前账号无权下载站点评分导出文件。"}), 403
        task = get_station_score_export_task_for_user(cur, task_id, user)
        if task["status"] != "completed":
            return jsonify({"success": False, "error": "导出文件尚未生成完成。"}), 400
        if not task.get("file_path"):
            return jsonify({"success": False, "error": "导出文件不存在或已过期。"}), 404
        abs_path = resolve_storage_abs_path(task["file_path"])
        if not abs_path or not os.path.isfile(abs_path):
            return jsonify({"success": False, "error": "导出文件不存在或已过期。"}), 404
        filename = task.get("download_filename") or f"station_score_{task_id[:8]}"
        mimetype = "application/zip" if str(filename).lower().endswith(".zip") else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        response = send_file(
            abs_path,
            as_attachment=True,
            download_name=filename,
            mimetype=mimetype,
        )
        response.headers["Cache-Control"] = "no-store"
        return response
    except PermissionError as exc:
        return jsonify({"success": False, "error": str(exc)}), 403
    except LookupError as exc:
        return jsonify({"success": False, "error": str(exc)}), 404
    except Exception as exc:
        return jsonify({"success": False, "error": str(exc)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/assessment/station-scores/adjustment", methods=["PUT"])
def save_assessment_station_score_adjustment():
    data = request.get_json(silent=True) or {}
    conn = None
    cur = None
    try:
        station_id = int(data.get("station_id") or 0)
        inspection_table_id = int(data.get("inspection_table_id") or 0)
        standard_id = int(data.get("standard_id") or 0)
        month = str(data.get("score_period") or data.get("month") or "").strip()
        manual_score = float(data.get("manual_score"))
        note = normalize_text(data.get("note"), 300)
        if station_id <= 0 or inspection_table_id <= 0 or standard_id <= 0:
            raise ValueError("评分调整参数不完整。")
    except (TypeError, ValueError) as exc:
        return jsonify({"success": False, "error": str(exc) or "评分调整参数不合法。"}), 400

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_inspection_checklist_management_schema(cur)
        ensure_station_score_schema(cur)

        user = get_user_by_id(cur, g.current_user["id"])
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404
        if not can_view_station_scores(cur, user):
            return jsonify({"success": False, "error": "当前账号无权查看站点评分。"}), 403
        if not can_adjust_station_scores(cur, user):
            return jsonify({"success": False, "error": "当前账号无权手动调整站点评分。"}), 403

        context = fetch_station_score_standard_context(
            cur,
            station_id,
            month,
            inspection_table_id,
            standard_id,
            data.get("date_from"),
            data.get("date_to"),
        )
        max_score = context["max_score"]
        if manual_score < 0 or manual_score > max_score + 0.001:
            return (
                jsonify(
                    {
                        "success": False,
                        "error": f"手动评分必须在 0 到 {round_score(max_score)} 之间。",
                    }
                ),
                400,
            )

        cur.execute(
            """
            INSERT INTO station_score_adjustments (
                station_id,
                inspection_table_id,
                standard_id,
                score_month,
                manual_score,
                note,
                adjusted_by,
                adjusted_at
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
            ON CONFLICT (station_id, inspection_table_id, standard_id, score_month)
            DO UPDATE SET
                manual_score = EXCLUDED.manual_score,
                note = EXCLUDED.note,
                adjusted_by = EXCLUDED.adjusted_by,
                adjusted_at = CURRENT_TIMESTAMP
            RETURNING TO_CHAR(adjusted_at, 'YYYY-MM-DD HH24:MI') AS adjusted_at;
            """,
            (
                station_id,
                inspection_table_id,
                standard_id,
                context["month"],
                round_score(manual_score),
                note,
                user["id"],
            ),
        )
        row = cur.fetchone()
        conn.commit()
        return jsonify(
            {
                "success": True,
                "message": "站点评分已调整。",
                "manual_score": round_score(manual_score),
                "adjusted_by_name": user.get("real_name") or user.get("username") or "",
                "adjusted_at": row.get("adjusted_at") if row else "",
            }
        )
    except LookupError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 404
    except ValueError as exc:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/assessment/attendance", methods=["GET"])
def get_assessment_attendance():
    try:
        month, month_start, next_month = parse_reporting_date_range(
            request.args.get("month"),
            request.args.get("date_from"),
            request.args.get("date_to"),
        )
        mode_filter = normalize_attendance_mode(request.args.get("mode", "all"))
    except ValueError as exc:
        return jsonify({"success": False, "error": str(exc)}), 400

    conn = None
    cur = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        user = get_user_by_id(cur, g.current_user["id"])
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404
        if not can_view_attendance(cur, user):
            return jsonify({"success": False, "error": "当前账号无权查看人员出勤统计。"}), 403

        base_mode_clause = ""
        issue_mode_clause = ""
        params = [month_start, next_month]
        if mode_filter != "all":
            base_mode_clause = "AND COALESCE(NULLIF(t.checklist_mode, ''), 'online') = %s"
            params.append(mode_filter)
        params.extend([month_start, next_month])
        if mode_filter != "all":
            issue_mode_clause = "AND COALESCE(NULLIF(t.checklist_mode, ''), 'online') = %s"
            params.append(mode_filter)
        issue_stats_params = [month_start, next_month]
        issue_stats_mode_clause = ""
        if mode_filter != "all":
            issue_stats_mode_clause = "AND COALESCE(NULLIF(t.checklist_mode, ''), 'online') = %s"
            issue_stats_params.append(mode_filter)
        params.extend(issue_stats_params)

        cur.execute(
            f"""
            WITH participant_rows AS (
                SELECT
                    ins.id AS inspection_id,
                    ins.inspection_date,
                    ins.station_id,
                    s.station_name,
                    s.region AS station_region,
                    t.id AS inspection_table_id,
                    t.table_name AS inspection_table_name,
                    COALESCE(NULLIF(t.checklist_mode, ''), 'online') AS checklist_mode,
                    ins.inspector_id
                FROM inspections ins
                JOIN stations s ON s.id = ins.station_id
                JOIN inspection_tables t ON t.id = ins.inspection_table_id
                WHERE ins.inspection_date >= %s
                  AND ins.inspection_date < %s
                  {base_mode_clause}

                UNION

                SELECT
                    ins.id AS inspection_id,
                    ins.inspection_date,
                    ins.station_id,
                    s.station_name,
                    s.region AS station_region,
                    t.id AS inspection_table_id,
                    t.table_name AS inspection_table_name,
                    COALESCE(NULLIF(t.checklist_mode, ''), 'online') AS checklist_mode,
                    i.inspector_id
                FROM issues i
                JOIN inspections ins ON ins.id = i.inspection_id
                JOIN stations s ON s.id = ins.station_id
                JOIN inspection_tables t ON t.id = ins.inspection_table_id
                WHERE ins.inspection_date >= %s
                  AND ins.inspection_date < %s
                  AND i.inspector_id IS NOT NULL
                  AND COALESCE(i.audit_status, 'pending') <> 'rejected'
                  {issue_mode_clause}
            ),
            issue_stats AS (
                SELECT
                    COALESCE(i.inspector_id, ins.inspector_id) AS inspector_id,
                    COALESCE(NULLIF(t.checklist_mode, ''), 'online') AS checklist_mode,
                    COUNT(i.id) AS issue_count,
                    COUNT(i.id) FILTER (
                        WHERE COALESCE(i.audit_status, 'pending') = 'approved'
                    ) AS approved_issue_count
                FROM issues i
                JOIN inspections ins ON ins.id = i.inspection_id
                JOIN inspection_tables t ON t.id = ins.inspection_table_id
                WHERE ins.inspection_date >= %s
                  AND ins.inspection_date < %s
                  AND COALESCE(i.inspector_id, ins.inspector_id) IS NOT NULL
                  {issue_stats_mode_clause}
                GROUP BY
                    COALESCE(i.inspector_id, ins.inspector_id),
                    COALESCE(NULLIF(t.checklist_mode, ''), 'online')
            )
            SELECT DISTINCT
                p.inspection_id,
                TO_CHAR(p.inspection_date, 'YYYY-MM-DD') AS inspection_date,
                p.station_id,
                p.station_name,
                p.station_region,
                p.inspection_table_id,
                p.inspection_table_name,
                p.checklist_mode,
                p.inspector_id,
                u.username AS inspector_username,
                u.real_name AS inspector_name,
                u.phone AS inspector_phone,
                COALESCE(st.issue_count, 0) AS issue_count,
                COALESCE(st.approved_issue_count, 0) AS approved_issue_count
            FROM participant_rows p
            JOIN users u ON u.id = p.inspector_id
            LEFT JOIN issue_stats st
              ON st.inspector_id = p.inspector_id
             AND st.checklist_mode = p.checklist_mode
            WHERE p.inspector_id IS NOT NULL
              AND p.checklist_mode IN ('online', 'offline')
            ORDER BY
                p.checklist_mode,
                inspection_date,
                p.station_region,
                p.station_name,
                p.inspection_table_name,
                inspector_name;
            """,
            params,
        )
        rows = [dict(row) for row in cur.fetchall()]
        return jsonify(
            {
                "success": True,
                **build_attendance_payload(rows, month, mode_filter),
            }
        )
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


# 新增巡检记录接口
@app.route("/api/inspections")
def get_inspections():
    def parse_request_int(value, default):
        try:
            return int(value)
        except (TypeError, ValueError):
            return default

    user_id = str(request.args.get("user_id", "")).strip()
    page = max(parse_request_int(request.args.get("page"), 1), 1)
    page_size = max(1, min(parse_request_int(request.args.get("page_size"), 20), 100))
    include_options = str(request.args.get("include_options", "")).strip().lower() in ("1", "true", "yes")

    conn = None
    cur = None

    def parse_list_arg(*names):
        values = []
        for name in names:
            for raw_value in request.args.getlist(name):
                text = str(raw_value or "").strip()
                if not text:
                    continue
                parsed_items = None
                if text.startswith("["):
                    try:
                        parsed_json = json.loads(text)
                        if isinstance(parsed_json, list):
                            parsed_items = parsed_json
                    except Exception:
                        parsed_items = None
                if parsed_items is None:
                    if "|||" in text:
                        parsed_items = text.split("|||")
                    elif "," in text:
                        parsed_items = text.split(",")
                    else:
                        parsed_items = [text]
                values.extend(str(item or "").strip() for item in parsed_items)

        result = []
        seen = set()
        for value in values:
            if not value or value in seen:
                continue
            seen.add(value)
            result.append(value)
        return result

    def normalize_date_arg(value):
        text = str(value or "").strip()
        return text if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text) else ""

    def get_month_bounds(value):
        text = str(value or "").strip()
        if not re.fullmatch(r"\d{4}-\d{2}", text):
            return "", ""
        month_start = datetime.strptime(f"{text}-01", "%Y-%m-%d").date()
        next_month = (
            month_start.replace(year=month_start.year + 1, month=1)
            if month_start.month == 12
            else month_start.replace(month=month_start.month + 1)
        )
        return month_start.isoformat(), (next_month - timedelta(days=1)).isoformat()

    def empty_page_payload(filter_options=None):
        return {
            "success": True,
            "items": [],
            "total": 0,
            "page": page,
            "page_size": page_size,
            "total_pages": 1,
            "filter_options": filter_options or {},
        }

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_issue_inspector_schema(cur)
        ensure_inspection_completion_schema(cur)
        conn.commit()

        user = None
        base_where_clauses = []
        base_params = []

        if user_id:
            cur.execute(
                """
                SELECT id, role, station_id
                FROM users
                WHERE id = %s
                LIMIT 1;
                """,
                (user_id,),
            )
            user = cur.fetchone()

        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404

        can_view_all = can_view_all_inspection_records(cur, user) or can_view_region_inspection_records(cur, user)
        can_view_own = can_view_own_inspection_records(cur, user)
        can_delete_records = can_delete_inspection_records(cur, user)
        can_sign_records = can_sign_inspection_records(cur, user)
        can_reset_signature = can_reset_inspection_signature(cur, user)
        if can_view_all:
            pass
        elif can_view_own:
            if not user["station_id"]:
                return jsonify(empty_page_payload())
            base_where_clauses.append("ins.station_id = %s")
            base_params.append(user["station_id"])
        else:
            return jsonify({"success": False, "error": "当前账号无权查看巡检记录。"}), 403

        if not append_inspection_table_scope_filter(
            cur,
            user,
            base_where_clauses,
            base_params,
            "ins.inspection_table_id",
            "limit_record_inspection_table_scope",
        ):
            return jsonify(empty_page_payload())
        if not append_station_region_scope_filter(
            cur,
            user,
            base_where_clauses,
            base_params,
            "s.region",
            "limit_record_station_region_scope",
        ):
            return jsonify(empty_page_payload())
        append_pending_audit_inspection_visibility_filter(user, base_where_clauses)
        hide_inspector_contact = should_hide_inspector_contact_info(cur, user)

        where_clauses = list(base_where_clauses)
        params = list(base_params)
        month_from, month_to = get_month_bounds(request.args.get("month"))
        date_from = month_from or normalize_date_arg(request.args.get("date_from"))
        date_to = month_to or normalize_date_arg(request.args.get("date_to"))
        selected_stations = parse_list_arg("stations", "station")
        selected_tables = parse_list_arg("inspection_tables", "inspection_table_name", "tables")
        selected_inspectors = [] if hide_inspector_contact else parse_list_arg("inspectors", "inspector")
        result_filter = str(request.args.get("result", "")).strip()
        sign_filter = str(request.args.get("sign_status", "")).strip()
        completion_filter = str(request.args.get("completion_status", "")).strip()

        if date_from:
            where_clauses.append("ins.inspection_date >= %s")
            params.append(date_from)
        if date_to:
            where_clauses.append("ins.inspection_date <= %s")
            params.append(date_to)
        if selected_stations:
            where_clauses.append("s.station_name = ANY(%s)")
            params.append(selected_stations)
        if selected_tables:
            where_clauses.append("t.table_name = ANY(%s)")
            params.append(selected_tables)
        if selected_inspectors:
            where_clauses.append(
                """
                (
                    EXISTS (
                        SELECT 1
                        FROM users inspector_filter_user
                        WHERE inspector_filter_user.id = ins.inspector_id
                          AND (
                              COALESCE(inspector_filter_user.real_name, '') = ANY(%s)
                              OR COALESCE(inspector_filter_user.username, '') = ANY(%s)
                              OR COALESCE(inspector_filter_user.phone, '') = ANY(%s)
                          )
                    )
                    OR EXISTS (
                        SELECT 1
                        FROM issues inspector_filter_issue
                        JOIN users inspector_filter_issue_user
                          ON inspector_filter_issue_user.id = inspector_filter_issue.inspector_id
                        WHERE inspector_filter_issue.inspection_id = ins.id
                          AND inspector_filter_issue.inspector_id IS NOT NULL
                          AND COALESCE(inspector_filter_issue.audit_status, 'pending') <> 'rejected'
                          AND (
                              COALESCE(inspector_filter_issue_user.real_name, '') = ANY(%s)
                              OR COALESCE(inspector_filter_issue_user.username, '') = ANY(%s)
                              OR COALESCE(inspector_filter_issue_user.phone, '') = ANY(%s)
                          )
                    )
                )
                """
            )
            params.extend([selected_inspectors] * 6)
        if result_filter == "异常":
            where_clauses.append(
                """
                EXISTS (
                    SELECT 1
                    FROM issues result_issue
                    WHERE result_issue.inspection_id = ins.id
                      AND COALESCE(result_issue.audit_status, 'pending') <> 'rejected'
                )
                """
            )
        elif result_filter == "正常":
            where_clauses.append(
                """
                NOT EXISTS (
                    SELECT 1
                    FROM issues result_issue
                    WHERE result_issue.inspection_id = ins.id
                      AND COALESCE(result_issue.audit_status, 'pending') <> 'rejected'
                )
                """
            )
        if sign_filter == "signed":
            where_clauses.append("COALESCE(ins.sign_status, '待签名确认') = '已签名确认'")
        elif sign_filter == "pending":
            where_clauses.append("COALESCE(ins.sign_status, '待签名确认') <> '已签名确认'")
        if completion_filter == "completed":
            where_clauses.append("COALESCE(ins.inspector_completion_status, '待检查人确认') = %s")
            params.append(INSPECTION_COMPLETION_DONE)
        elif completion_filter == "pending":
            where_clauses.append("COALESCE(ins.inspector_completion_status, '待检查人确认') <> %s")
            params.append(INSPECTION_COMPLETION_DONE)

        where_clause = f"WHERE {' AND '.join(where_clauses)}" if where_clauses else ""

        filter_options = {}
        if include_options:
            station_option_clauses = []
            station_option_params = []
            if not can_view_all and can_view_own and user.get("station_id"):
                station_option_clauses.append("id = %s")
                station_option_params.append(user["station_id"])
            else:
                region_scope_values = get_effective_station_region_scope_values(
                    cur, user, "limit_record_station_region_scope"
                )
                if region_scope_values is not None:
                    if not region_scope_values:
                        station_option_clauses.append("FALSE")
                    else:
                        station_option_clauses.append(
                            "COALESCE(NULLIF(TRIM(region), ''), '未填写片区') = ANY(%s)"
                        )
                        station_option_params.append(list(region_scope_values))
            station_option_where = (
                f"WHERE {' AND '.join(station_option_clauses)}"
                if station_option_clauses
                else ""
            )
            cur.execute(
                f"""
                SELECT station_name
                FROM stations
                {station_option_where}
                ORDER BY station_name ASC;
                """,
                station_option_params,
            )
            filter_options["stations"] = [
                row["station_name"] for row in cur.fetchall() if row.get("station_name")
            ]

            table_option_clauses = ["COALESCE(is_active, TRUE) = TRUE"]
            table_option_params = []
            table_scope_ids = get_effective_inspection_table_scope_ids(
                cur, user, "limit_record_inspection_table_scope"
            )
            if table_scope_ids is not None:
                if not table_scope_ids:
                    table_option_clauses.append("FALSE")
                else:
                    table_option_clauses.append("id = ANY(%s)")
                    table_option_params.append(list(table_scope_ids))
            table_option_where = f"WHERE {' AND '.join(table_option_clauses)}"
            cur.execute(
                f"""
                SELECT table_name
                FROM inspection_tables
                {table_option_where}
                ORDER BY table_name ASC;
                """,
                table_option_params,
            )
            filter_options["inspection_tables"] = [
                row["table_name"] for row in cur.fetchall() if row.get("table_name")
            ]
            if hide_inspector_contact:
                filter_options["inspectors"] = []
            else:
                cur.execute(
                    """
                    SELECT DISTINCT
                        COALESCE(NULLIF(TRIM(real_name), ''), NULLIF(TRIM(username), ''), NULLIF(TRIM(phone), ''), id::text) AS label
                    FROM users
                    WHERE role IN ('root', 'supervisor')
                    ORDER BY label ASC;
                    """
                )
                filter_options["inspectors"] = [
                    row["label"] for row in cur.fetchall() if row.get("label")
                ]

        cur.execute(
            f"""
            SELECT COUNT(*) AS total
            FROM inspections ins
            JOIN stations s ON ins.station_id = s.id
            JOIN inspection_tables t ON ins.inspection_table_id = t.id
            {where_clause};
            """,
            params,
        )
        total = int((cur.fetchone() or {}).get("total") or 0)
        total_pages = max(1, (total + page_size - 1) // page_size)
        effective_page = min(page, total_pages)
        offset = (effective_page - 1) * page_size

        cur.execute(
            f"""
            WITH filtered_ids AS (
                SELECT ins.id
                FROM inspections ins
                JOIN stations s ON ins.station_id = s.id
                JOIN inspection_tables t ON ins.inspection_table_id = t.id
                {where_clause}
                ORDER BY ins.inspection_date DESC, ins.id DESC
                LIMIT %s OFFSET %s
            ),
            issue_stats AS (
                SELECT
                    i.inspection_id,
                    COUNT(i.id) FILTER (WHERE COALESCE(i.audit_status, 'pending') <> 'rejected') AS issue_count,
                    COUNT(i.id) AS total_issue_count,
                    COUNT(i.id) FILTER (WHERE COALESCE(i.audit_status, 'pending') = 'pending') AS pending_audit_count,
                    COUNT(i.id) FILTER (WHERE COALESCE(i.audit_status, 'pending') <> 'pending') AS audited_issue_count,
                    COUNT(i.id) FILTER (
                        WHERE COALESCE(i.audit_status, 'pending') <> 'rejected'
                          AND (
                              NULLIF(TRIM(COALESCE(i.rectification_result, '')), '') IS NOT NULL
                              OR NULLIF(TRIM(COALESCE(i.rectification_note, '')), '') IS NOT NULL
                              OR NULLIF(TRIM(COALESCE(i.rectification_photo_path, '')), '') IS NOT NULL
                          )
                    ) AS rectified_issue_count
                FROM issues i
                JOIN filtered_ids fid ON fid.id = i.inspection_id
                GROUP BY i.inspection_id
            ),
            participant_rows AS (
                SELECT fid.id AS inspection_id, ins.inspector_id AS inspector_id
                FROM filtered_ids fid
                JOIN inspections ins ON ins.id = fid.id
                WHERE ins.inspector_id IS NOT NULL
                UNION
                SELECT fid.id AS inspection_id, issue_part.inspector_id AS inspector_id
                FROM filtered_ids fid
                JOIN issues issue_part ON issue_part.inspection_id = fid.id
                WHERE issue_part.inspector_id IS NOT NULL
                  AND COALESCE(issue_part.audit_status, 'pending') <> 'rejected'
            ),
            participant_stats AS (
                SELECT
                    pr.inspection_id,
                    STRING_AGG(
                        DISTINCT COALESCE(participant.real_name, participant.username, participant.phone, participant.id::text),
                        '、'
                    ) AS inspector_names,
                    STRING_AGG(
                        DISTINCT CONCAT_WS(' ', participant.real_name, participant.username, participant.phone),
                        ' '
                    ) AS inspector_search_text,
                    JSONB_AGG(
                        DISTINCT JSONB_BUILD_OBJECT(
                            'id', participant.id,
                            'username', participant.username,
                            'real_name', participant.real_name,
                            'phone', participant.phone
                        )
                    ) FILTER (WHERE participant.id IS NOT NULL) AS inspectors
                FROM participant_rows pr
                JOIN users participant ON participant.id = pr.inspector_id
                GROUP BY pr.inspection_id
            )
            SELECT
                ins.id AS id,
                ins.batch_id AS batch_id,
                ins.station_id AS station_id,
                TO_CHAR(ins.inspection_date, 'YYYY-MM-DD') AS date,
                s.station_name AS station,
                t.table_name AS inspection_table_name,
                CASE
                    WHEN COALESCE(issue_stats.issue_count, 0) > 0 THEN '异常'
                    ELSE '正常'
                END AS result,
                COALESCE(issue_stats.issue_count, 0) AS issue_count,
                COALESCE(issue_stats.total_issue_count, 0) AS total_issue_count,
                COALESCE(issue_stats.pending_audit_count, 0) AS pending_audit_count,
                COALESCE(issue_stats.audited_issue_count, 0) AS audited_issue_count,
                COALESCE(issue_stats.rectified_issue_count, 0) AS rectified_issue_count,
                ins.sign_status,
                ins.station_manager_signed_name,
                ins.station_manager_signature_path,
                TO_CHAR(ins.station_manager_signed_at, 'YYYY-MM-DD HH24:MI') AS station_manager_signed_at,
                ins.inspector_completion_status,
                ins.inspector_completion_source,
                TO_CHAR(ins.inspector_completed_at, 'YYYY-MM-DD HH24:MI') AS inspector_completed_at,
                completed_user.username AS inspector_completed_by_username,
                completed_user.real_name AS inspector_completed_by_name,
                participant_stats.inspector_names,
                participant_stats.inspector_search_text,
                COALESCE(participant_stats.inspectors, '[]'::jsonb) AS inspectors,
                EXISTS (
                    SELECT 1
                    FROM participant_rows current_participant
                    WHERE current_participant.inspection_id = ins.id
                      AND current_participant.inspector_id = %s
                ) AS current_user_participated
            FROM filtered_ids fid
            JOIN inspections ins ON ins.id = fid.id
            JOIN stations s ON ins.station_id = s.id
            JOIN inspection_tables t ON ins.inspection_table_id = t.id
            LEFT JOIN users completed_user ON completed_user.id = ins.inspector_completed_by
            LEFT JOIN issue_stats ON issue_stats.inspection_id = ins.id
            LEFT JOIN participant_stats ON participant_stats.inspection_id = ins.id
            ORDER BY ins.inspection_date DESC, ins.id DESC;
            """,
            [*params, page_size, offset, user["id"]],
        )
        rows = cur.fetchall()
        items = []
        for row in rows:
            item = dict(row)
            inspectors = item.get("inspectors") or []
            if isinstance(inspectors, str):
                try:
                    inspectors = json.loads(inspectors)
                except Exception:
                    inspectors = []
            item["inspector_names"] = "" if hide_inspector_contact else item.get("inspector_names")
            item["inspector_search_text"] = "" if hide_inspector_contact else item.get("inspector_search_text")
            item["inspectors"] = [] if hide_inspector_contact else inspectors
            item["can_delete_record"] = bool(can_delete_records)
            is_record_completed = item.get("inspector_completion_status") == INSPECTION_COMPLETION_DONE
            is_record_signed = item.get("sign_status") == "已签名确认"
            item["can_reset_record_flow"] = bool(
                can_reset_signature
                and is_record_completed
                and not is_record_signed
            )
            item["can_sign_record"] = bool(
                is_station_manager(user)
                and can_sign_records
                and is_record_completed
                and item.get("sign_status") != "已签名确认"
                and user.get("station_id")
                and item.get("station_id") == user.get("station_id")
                and int(item.get("pending_audit_count") or 0) == 0
            )
            item["can_complete_record"] = bool(
                is_supervisor_like(user)
                and item.get("inspector_completion_status") != INSPECTION_COMPLETION_DONE
                and item.get("current_user_participated")
            )
            item["inspector_completion_source_label"] = inspection_completion_source_label(
                item.get("inspector_completion_source")
            )
            items.append(item)

        apply_inspection_completion_progress(cur, items, hide_inspector_contact)
        for item in items:
            progress = item.get("inspector_completion_progress") or {}
            current_user_confirmed = any(
                str(participant.get("id") or "") == str(user["id"])
                and participant.get("confirmed")
                for participant in progress.get("participants") or []
            )
            item["current_user_completion_confirmed"] = bool(current_user_confirmed)
            item["can_complete_record"] = bool(
                is_supervisor_like(user)
                and item.get("inspector_completion_status") != INSPECTION_COMPLETION_DONE
                and item.get("current_user_participated")
                and not current_user_confirmed
            )

        return jsonify(
            {
                "success": True,
                "items": items,
                "total": total,
                "page": effective_page,
                "page_size": page_size,
                "total_pages": total_pages,
                "filter_options": filter_options,
            }
        )
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

    finally:
        close_db_resources(cur, conn)


@app.route("/api/inspections/<int:inspection_id>", methods=["DELETE"])
def delete_inspection_record(inspection_id):
    data = request.get_json(silent=True) or {}
    user_id = get_authenticated_request_user_id(
        data.get("user_id") or request.args.get("user_id", "")
    )

    if not user_id:
        return jsonify({"success": False, "error": "缺少用户信息。"}), 400

    conn = None
    cur = None
    files_to_remove = set()

    try:
        conn = get_db_connection()
        cur = conn.cursor()

        user = get_user_by_id(cur, user_id)
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404

        if not can_delete_inspection_records(cur, user):
            return jsonify({"success": False, "error": "当前账号无权删除巡检记录。"}), 403

        cur.execute(
            """
            SELECT
                ins.id,
                ins.batch_id,
                ins.station_id,
                s.region AS station_region,
                ins.inspection_table_id,
                ins.inspection_date,
                ins.station_manager_signature_path,
                s.station_name,
                t.table_name AS inspection_table_name
            FROM inspections ins
            JOIN stations s ON ins.station_id = s.id
            JOIN inspection_tables t ON ins.inspection_table_id = t.id
            WHERE ins.id = %s
            LIMIT 1;
            """,
            (inspection_id,),
        )
        inspection = cur.fetchone()

        if not inspection:
            return jsonify({"success": False, "error": "巡检记录不存在。"}), 404

        can_view_all = can_view_all_inspection_records(cur, user) or can_view_region_inspection_records(cur, user)
        can_view_own = can_view_own_inspection_records(cur, user)
        if not can_view_all and not (
            can_view_own
            and user.get("station_id")
            and inspection["station_id"] == user["station_id"]
        ):
            return jsonify({"success": False, "error": "当前账号无权操作该巡检记录。"}), 403
        if not is_inspection_table_allowed_for_user(
            cur,
            user,
            inspection["inspection_table_id"],
            "limit_record_inspection_table_scope",
        ):
            return jsonify({"success": False, "error": "当前账号无权操作该检查表的巡检记录。"}), 403
        if not is_station_region_allowed_for_user(
            cur,
            user,
            inspection.get("station_region"),
            "limit_record_station_region_scope",
        ):
            return jsonify({"success": False, "error": "当前账号无权操作该片区的巡检记录。"}), 403

        if inspection.get("station_manager_signature_path"):
            files_to_remove.add(inspection["station_manager_signature_path"])

        cur.execute(
            """
            SELECT
                id,
                photo_path,
                rectification_photo_path,
                review_photo_path
            FROM issues
            WHERE inspection_id = %s;
            """,
            (inspection_id,),
        )
        issue_rows = cur.fetchall()
        for issue in issue_rows:
            for key in (
                "photo_path",
                "rectification_photo_path",
                "review_photo_path",
            ):
                if issue.get(key):
                    files_to_remove.add(issue[key])

        cur.execute(
            """
            SELECT DISTINCT plan_config_id
            FROM inspection_plan_station_items
            WHERE completed_inspection_id = %s;
            """,
            (inspection_id,),
        )
        affected_plan_config_ids = [row["plan_config_id"] for row in cur.fetchall()]

        cur.execute(
            """
            UPDATE inspection_plan_station_items
            SET completion_status = 'pending',
                completed_inspection_id = NULL,
                completed_at = NULL,
                updated_at = CURRENT_TIMESTAMP
            WHERE completed_inspection_id = %s;
            """,
            (inspection_id,),
        )
        affected_plan_item_count = cur.rowcount

        cur.execute("DELETE FROM issues WHERE inspection_id = %s;", (inspection_id,))
        deleted_issue_count = cur.rowcount

        cur.execute("DELETE FROM inspections WHERE id = %s;", (inspection_id,))
        if cur.rowcount == 0:
            raise ValueError("巡检记录删除失败，请刷新后重试。")

        if inspection.get("batch_id"):
            cur.execute(
                """
                DELETE FROM inspection_batches b
                WHERE b.id = %s
                  AND NOT EXISTS (
                    SELECT 1
                    FROM inspections ins
                    WHERE ins.batch_id = b.id
                  );
                """,
                (inspection["batch_id"],),
            )

        for plan_config_id in affected_plan_config_ids:
            sync_plan_station_items_completion_by_history(cur, plan_config_id)

        conn.commit()

        for path in files_to_remove:
            remove_storage_file(path)

        return jsonify(
            {
                "success": True,
                "message": "巡检记录已删除，关联巡检问题和计划完成状态已同步更新。",
                "deleted_issue_count": deleted_issue_count,
                "affected_plan_item_count": affected_plan_item_count,
            }
        )
    except ValueError as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


# === 新增：本检查表录入问题简要信息 API ===
@app.route("/api/inspections/<int:inspection_id>/issues")
def get_inspection_issues(inspection_id):
    user_id = str(request.args.get("user_id", "")).strip()

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_issue_inspector_schema(cur)

        user = None
        if user_id:
            cur.execute(
                """
                SELECT id, role, station_id
                FROM users
                WHERE id = %s
                LIMIT 1;
                """,
                (user_id,),
            )
            user = cur.fetchone()

        cur.execute(
            """
            SELECT
                ins.id,
                ins.station_id,
                ins.inspection_table_id,
                ins.batch_id,
                TO_CHAR(ins.inspection_date, 'YYYY-MM-DD') AS inspection_date,
                TO_CHAR(ins.inspection_date, 'YYYY-MM-DD') AS date,
                ins.inspector_id,
                ins.sign_status,
                ins.station_manager_signed_name,
                ins.station_manager_signature_path,
                TO_CHAR(ins.station_manager_signed_at, 'YYYY-MM-DD HH24:MI') AS station_manager_signed_at,
                ins.inspector_completion_status,
                ins.inspector_completion_source,
                TO_CHAR(ins.inspector_completed_at, 'YYYY-MM-DD HH24:MI') AS inspector_completed_at,
                completed_user.username AS inspector_completed_by_username,
                completed_user.real_name AS inspector_completed_by_name,
                s.station_name,
                s.region AS station_region,
                s.address AS station_address,
                s.station_manager_name,
                s.station_manager_phone,
                inspector.username AS inspector_username,
                inspector.real_name AS inspector_name,
                inspector.phone AS inspector_phone,
                t.table_name AS inspection_table_name
            FROM inspections ins
            JOIN stations s ON ins.station_id = s.id
            JOIN inspection_tables t ON ins.inspection_table_id = t.id
            JOIN users inspector ON ins.inspector_id = inspector.id
            LEFT JOIN users completed_user ON completed_user.id = ins.inspector_completed_by
            WHERE ins.id = %s
            LIMIT 1;
            """,
            (inspection_id,),
        )
        inspection = cur.fetchone()

        if not inspection:
            return jsonify({"success": False, "error": "巡检记录不存在。"}), 404

        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404

        can_view_all = can_view_all_inspection_records(cur, user) or can_view_region_inspection_records(cur, user)
        can_view_own = can_view_own_inspection_records(cur, user)
        cur.execute(
            """
            SELECT DISTINCT
                participant.id,
                participant.username,
                participant.real_name,
                participant.phone
            FROM (
                SELECT ins.inspector_id AS inspector_id
                FROM inspections ins
                WHERE ins.id = %s
                UNION
                SELECT i.inspector_id AS inspector_id
                FROM issues i
                WHERE i.inspection_id = %s
                  AND i.inspector_id IS NOT NULL
                  AND COALESCE(i.audit_status, 'pending') <> 'rejected'
            ) participant_ids
            JOIN users participant ON participant_ids.inspector_id = participant.id
            ORDER BY participant.id ASC;
            """,
            (inspection_id, inspection_id),
        )
        inspectors = [dict(row) for row in cur.fetchall()]
        inspection = dict(inspection)
        inspection["inspectors"] = inspectors
        inspection["inspector_names"] = "、".join(
            [
                (row.get("real_name") or row.get("username") or str(row.get("id")))
                for row in inspectors
            ]
        )
        inspection["inspector_completion_source_label"] = inspection_completion_source_label(
            inspection.get("inspector_completion_source")
        )

        is_inspector = any(str(item.get("id") or "") == str(user.get("id") or "") for item in inspectors)
        if (
            can_view_own
            and not can_view_all
            and inspection["station_id"] != user["station_id"]
            and not is_inspector
        ):
            return jsonify({"success": False, "error": "无权查看该检查表内容。"}), 403
        if not can_view_all and not can_view_own and not is_inspector:
            return jsonify({"success": False, "error": "无权查看该检查表内容。"}), 403
        if not is_inspection_table_allowed_for_user(
            cur,
            user,
            inspection["inspection_table_id"],
            "limit_record_inspection_table_scope",
        ):
            return jsonify({"success": False, "error": "当前账号无权查看该检查表内容。"}), 403
        if not is_station_region_allowed_for_user(
            cur,
            user,
            inspection.get("station_region"),
            "limit_record_station_region_scope",
        ):
            return jsonify({"success": False, "error": "当前账号无权查看该片区内容。"}), 403
        if should_hide_pending_audit_flow_for_user(user):
            cur.execute(
                """
                SELECT COUNT(*) AS pending_audit_count
                FROM issues
                WHERE inspection_id = %s
                  AND COALESCE(audit_status, 'pending') = 'pending';
                """,
                (inspection_id,),
            )
            pending_audit_count = int(cur.fetchone()["pending_audit_count"] or 0)
            if pending_audit_count > 0:
                return jsonify({"success": False, "error": "该巡检记录仍有待审核问题，暂不可查看。"}), 403
        hide_inspector_contact = should_hide_inspector_contact_info(cur, user)
        apply_inspection_completion_progress(cur, [inspection], hide_inspector_contact)
        if hide_inspector_contact:
            inspection["inspectors"] = []
            inspection["inspector_names"] = ""
            inspection["inspector_id"] = None
            inspection["inspector_username"] = ""
            inspection["inspector_name"] = ""
            inspection["inspector_phone"] = ""

        cur.execute(
            """
            SELECT
                i.id,
                COALESCE(i.inspector_id, ins.inspector_id) AS inspector_id,
                issue_inspector.username AS inspector_username,
                issue_inspector.real_name AS inspector_name,
                issue_inspector.phone AS inspector_phone,
                TO_CHAR(i.created_at, 'YYYY-MM-DD HH24:MI') AS created_at,
                t.table_name AS inspection_table_name,
                i.standard_id,
                i.standard_detail_text,
                i.internal_standard_id,
                i.internal_standard_detail_text,
                i.description,
                i.photo_path AS issue_photo,
                i.rectification_result,
                i.rectification_note,
                TO_CHAR(i.rectification_at, 'YYYY-MM-DD HH24:MI') AS rectification_at,
                i.rectification_photo_path AS rectification_photo,
                i.review_result,
                i.review_note,
                TO_CHAR(i.review_at, 'YYYY-MM-DD HH24:MI') AS review_at,
                i.review_photo_path AS review_photo,
                i.status,
                COALESCE(i.audit_status, 'pending') AS audit_status,
                i.audited_by,
                TO_CHAR(i.audited_at, 'YYYY-MM-DD HH24:MI') AS audited_at,
                ins.sign_status AS inspection_sign_status,
                ins.station_manager_signed_name,
                ins.station_manager_signature_path,
                TO_CHAR(ins.station_manager_signed_at, 'YYYY-MM-DD HH24:MI') AS station_manager_signed_at,
                ins.inspector_completion_status AS inspection_completion_status
            FROM issues i
            JOIN inspections ins ON i.inspection_id = ins.id
            JOIN inspection_tables t ON i.inspection_table_id = t.id
            JOIN users issue_inspector ON COALESCE(i.inspector_id, ins.inspector_id) = issue_inspector.id
            WHERE i.inspection_id = %s
              AND COALESCE(i.audit_status, 'pending') <> 'rejected'
            ORDER BY i.id ASC;
            """,
            (inspection_id,),
        )
        issues = [
            normalize_issue_row_for_response(
                row,
                user,
                hide_inspector_contact_info=hide_inspector_contact,
            )
            for row in cur.fetchall()
        ]

        return jsonify(
            {
                "success": True,
                "inspection": inspection,
                "issues": issues,
            }
        )
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/issues/<int:issue_id>/rectification", methods=["POST"])
def submit_rectification(issue_id):
    user_id = str(request.form.get("user_id", "")).strip()
    rectification_result = canonical_issue_result(
        request.form.get("rectification_result", "")
    )
    rectification_note = str(request.form.get("rectification_note", "")).strip()
    rectification_photo = request.files.get("rectification_photo")

    if not user_id:
        return jsonify({"success": False, "error": "缺少用户信息。"}), 400

    if not rectification_result:
        return jsonify({"success": False, "error": "请选择整改结果。"}), 400

    if rectification_result not in ISSUE_RESULT_OPTIONS:
        return (
            jsonify(
                {
                    "success": False,
                    "error": "整改结果只能选择已整改或站经无法整改。",
                }
            ),
            400,
        )

    if not rectification_note:
        return jsonify({"success": False, "error": "请填写整改说明。"}), 400

    rectification_photo_required = rectification_result == "已整改"
    if rectification_photo_required and (
        not rectification_photo or not rectification_photo.filename
    ):
        return jsonify({"success": False, "error": "请上传整改照片。"}), 400

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_issue_inspector_schema(cur)

        cur.execute(
            """
            SELECT id, role, station_id
            FROM users
            WHERE id = %s
            LIMIT 1;
            """,
            (user_id,),
        )
        user = cur.fetchone()

        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404

        if user["role"] != "station_manager":
            return (
                jsonify({"success": False, "error": "只有站点账号可以提交整改。"}),
                403,
            )

        cur.execute(
            """
            SELECT
                i.id,
                i.station_id,
                i.status,
                COALESCE(i.audit_status, 'pending') AS audit_status,
                i.inspection_id,
                ins.sign_status AS inspection_sign_status
            FROM issues i
            JOIN inspections ins ON i.inspection_id = ins.id
            WHERE i.id = %s
            LIMIT 1;
            """,
            (issue_id,),
        )
        issue = cur.fetchone()

        if not issue:
            return jsonify({"success": False, "error": "问题不存在。"}), 404

        if issue["station_id"] != user["station_id"]:
            return (
                jsonify({"success": False, "error": "只能整改本账号所属站点的问题。"}),
                403,
            )

        if is_issue_audit_rejected(issue):
            return jsonify({"success": False, "error": "该问题已被审核否决，不再参与整改流转。"}), 400

        if issue["inspection_sign_status"] != "已签名确认":
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "当前问题所属检查表尚未完成站经理签名验收，暂不可提交整改。",
                    }
                ),
                400,
            )

        if issue["status"] != "待整改":
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "当前问题状态不是待整改，不能重复提交整改。",
                    }
                ),
                400,
            )

        rectification_photo_path = None
        if rectification_photo and rectification_photo.filename:
            rectification_photo_path = save_uploaded_file(
                rectification_photo, "rectifications"
            )

        if rectification_photo_path:
            cur.execute(
                """
                UPDATE issues
                SET rectification_result = %s,
                    rectification_note = %s,
                    rectification_at = CURRENT_TIMESTAMP,
                    rectification_photo_path = %s,
                    status = '待复核'
                WHERE id = %s;
                """,
                (
                    rectification_result,
                    rectification_note,
                    rectification_photo_path,
                    issue_id,
                ),
            )
        else:
            cur.execute(
                """
                UPDATE issues
                SET rectification_result = %s,
                    rectification_note = %s,
                    rectification_at = CURRENT_TIMESTAMP,
                    status = '待复核'
                WHERE id = %s;
                """,
                (
                    rectification_result,
                    rectification_note,
                    issue_id,
                ),
            )

        conn.commit()
        return jsonify({"success": True, "message": "整改提交成功，已转入待复核。"})
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/issues/<int:issue_id>/rectification-photo", methods=["POST"])
def update_rectification_photo(issue_id):
    user_id = str(request.form.get("user_id", "")).strip()
    rectification_photo = request.files.get("rectification_photo")

    if not user_id:
        return jsonify({"success": False, "error": "缺少用户信息。"}), 400

    if not rectification_photo or not rectification_photo.filename:
        return jsonify({"success": False, "error": "请上传新的整改照片。"}), 400

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()

        user = get_user_by_id(cur, user_id)
        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404

        cur.execute(
            """
            SELECT
                id,
                station_id,
                status,
                COALESCE(audit_status, 'pending') AS audit_status,
                rectification_result
            FROM issues
            WHERE id = %s
            LIMIT 1;
            """,
            (issue_id,),
        )
        issue = cur.fetchone()

        if not issue:
            return jsonify({"success": False, "error": "问题不存在。"}), 404

        can_explicit_edit = can_edit_inspection_issues(cur, user)
        if not can_user_update_rectification_photo(user, issue, can_explicit_edit):
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "当前账号无权更新该问题的整改照片，或问题已不在可修改阶段。",
                    }
                ),
                403,
            )

        rectification_photo_path = save_uploaded_file(
            rectification_photo, "rectifications"
        )
        cur.execute(
            """
            UPDATE issues
            SET rectification_photo_path = %s
            WHERE id = %s;
            """,
            (rectification_photo_path, issue_id),
        )
        conn.commit()

        return jsonify(
            {
                "success": True,
                "message": "整改照片已更新。",
                "rectification_photo": rectification_photo_path,
            }
        )
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


@app.route("/api/issues/<int:issue_id>/review", methods=["POST"])
def submit_review(issue_id):
    user_id = str(request.form.get("user_id", "")).strip()
    review_result = canonical_issue_result(request.form.get("review_result", ""))
    review_note = str(request.form.get("review_note", "")).strip()
    review_photo = request.files.get("review_photo")

    if not user_id:
        return jsonify({"success": False, "error": "缺少用户信息。"}), 400

    if not review_result:
        return jsonify({"success": False, "error": "请选择督导组复核结果。"}), 400

    if review_result not in ISSUE_RESULT_OPTIONS:
        return (
            jsonify(
                {
                    "success": False,
                    "error": "督导组复核结果只能选择已整改或站经无法整改。",
                }
            ),
            400,
        )

    if not review_note:
        return jsonify({"success": False, "error": "请填写复核说明。"}), 400

    review_photo_required = review_result == "已整改"
    if review_photo_required and (not review_photo or not review_photo.filename):
        return jsonify({"success": False, "error": "请上传复核照片。"}), 400

    conn = None
    cur = None

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        ensure_issue_inspector_schema(cur)

        cur.execute(
            """
            SELECT id, role
            FROM users
            WHERE id = %s
            LIMIT 1;
            """,
            (user_id,),
        )
        user = cur.fetchone()

        if not user:
            return jsonify({"success": False, "error": "用户不存在。"}), 404

        if not (is_root_user(user) or user.get("role") == "supervisor"):
            return (
                jsonify({"success": False, "error": "只有督导组账号可以提交复核。"}),
                403,
            )

        cur.execute(
            """
            SELECT id, status
                , COALESCE(audit_status, 'pending') AS audit_status
            FROM issues
            WHERE id = %s
            LIMIT 1;
            """,
            (issue_id,),
        )
        issue = cur.fetchone()

        if not issue:
            return jsonify({"success": False, "error": "问题不存在。"}), 404

        if is_issue_audit_rejected(issue):
            return jsonify({"success": False, "error": "该问题已被审核否决，不再参与复核流转。"}), 400

        if issue["status"] != "待复核":
            return (
                jsonify(
                    {
                        "success": False,
                        "error": "当前问题状态不是待复核，不能提交复核。",
                    }
                ),
                400,
            )

        new_status = "已闭环" if review_result == "已整改" else "站经无法整改"
        review_photo_path = None
        if review_photo and review_photo.filename:
            review_photo_path = save_uploaded_file(review_photo, "rectifications")

        cur.execute(
            """
            UPDATE issues
            SET review_result = %s,
                review_note = %s,
                review_at = CURRENT_TIMESTAMP,
                review_photo_path = %s,
                status = %s
            WHERE id = %s;
            """,
            (
                review_result,
                review_note,
                review_photo_path,
                new_status,
                issue_id,
            ),
        )

        conn.commit()
        return jsonify(
            {
                "success": True,
                "message": f"督导组复核提交成功，问题状态已更新为{new_status}。",
            }
        )
    except Exception as e:
        if conn:
            conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        close_db_resources(cur, conn)


# ===== 原有静态文件上传API =====


@app.route("/storage/<path:filename>")
def uploaded_file(filename):
    return send_from_directory(STORAGE_ROOT, filename)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False, use_reloader=False)
